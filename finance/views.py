from django.shortcuts import render
from rest_framework import viewsets , permissions , serializers
from url_filter.integrations.drf import DjangoFilterBackend
from .serializers import *
from API.permissions import *
from .models import *
from django.db.models import Q, F , Sum
from openpyxl import load_workbook,Workbook
from openpyxl.styles import PatternFill , Font , Alignment
import string
from rest_framework.views import APIView
from rest_framework.renderers import JSONRenderer
from rest_framework.response import Response
from io import BytesIO
from rest_framework.views import APIView
# from projects.models import ProjectPettyExpense , project
import datetime
from dateutil.relativedelta import relativedelta
import calendar
from django.http import HttpResponse
from django.core.urlresolvers import reverse
from reportlab import *
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.units import cm, mm
from reportlab.lib import colors , utils
from reportlab.platypus import Paragraph, Table, TableStyle, Image, Frame, Spacer, PageBreak, BaseDocTemplate, PageTemplate, SimpleDocTemplate, Flowable,ListItem,ListFlowable,NextPageTemplate
from PIL import Image
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet, TA_CENTER
from reportlab.graphics import barcode , renderPDF
from reportlab.graphics.shapes import *
from reportlab.graphics.barcode.qr import QrCodeWidget
from reportlab.lib.pagesizes import letter,A5,A4,A3,A2
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import ParagraphStyle,getSampleStyleSheet
from reportlab.lib.enums import TA_LEFT, TA_CENTER,TA_RIGHT
from reportlab.lib.colors import *
from reportlab.lib.units import inch, cm
from num2words import num2words
from django.core.mail import send_mail, EmailMessage
from django.db.models import CharField,FloatField, Value , Func
from clientRelationships.models import Contract, Contact
from clientRelationships.models import *
from openpyxl import load_workbook,Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import PatternFill , Font
from ERP.models import appSettingsField, service, address
from projects.models import *
from payroll.models import PayrollReport
# Create your views here.
import ast
import json
from clientRelationships.models import *
from marketing.models import TourPlanStop
from svglib.svglib import svg2rlg
import cv2
import numpy as np


class TermsAndConditionsViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,  )
    serializer_class = TermsAndConditionsSerializer
    def get_queryset(self):
        divsn = self.request.user.designation.division
        return TermsAndConditions.objects.filter(division = divsn)

def getProducts(request, id):
    data = Inventory.objects.filter(division = request.user.designation.division.pk, category__id = int(id))
    user = request.user
    return render(request,'app.finance.products.html',{'data':data,'user':user})


class AccountViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated, )
    serializer_class = AccountSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['number','personal','contactPerson','title']
    def get_queryset(self):
        divsn = self.request.user.designation.division
        accountObj = Account.objects.filter(division = divsn)
        if 'only' in self.request.GET:
            if self.request.GET['only'] == 'pettyCash':
                accountObj = accountObj.filter(personal=True)
            if self.request.GET['only'] == 'accounts':
                accountObj = accountObj.filter(group=None).exclude(personal=True)
            else:
                data1 = accountObj.filter(personal=True)
                data2 = accountObj.filter(group=None).exclude(personal=True)
                accountObj = data1 | data2
        if 'search__in' in self.request.GET:
            search = self.request.GET['search__in']
            accountObj = accountObj.filter(Q(title__icontains=search) |  Q(number__icontains=search))
        if 'filters' in self.request.GET:
            count = 0
            for i in self.request.GET['filters'].split(','):
                count +=1
                if i == 'pettycash':
                    data = accountObj.filter(personal=True)
                elif i == 'account':
                    data = accountObj.filter(group=None).exclude(personal=True)
                else:
                    data = accountObj.filter(group = i).exclude(personal=True)
                if count == 1:
                    toReturn =  data
                else:
                    toReturn =  toReturn | data
            return toReturn
        else:
            return accountObj

class GetAccountsTotalAPIView(APIView):
    permission_classes = (permissions.IsAuthenticated ,)
    def get(self , request , format = None):
        divsn = self.request.user.designation.division
        accountObj = Account.objects.filter(division = divsn)
        if 'type' in self.request.GET:
            if self.request.GET['type'] == 'pettyCash':
                accountObj = accountObj.filter(personal=True)
            if self.request.GET['type'] == 'accounts':
                accountObj = accountObj.filter(group=None).exclude(personal=True)
            gTot = 0
            total = accountObj.aggregate(tot = Sum('balance'))
            if total['tot'] is not None:
                gTot = total['tot']

        return Response(gTot, status=status.HTTP_200_OK)


class AccountLiteViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated, )
    serializer_class = AccountLiteSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['title','group','heading','personal']
    def get_queryset(self):
        divsn = self.request.user.designation.division
        return Account.objects.filter(division = divsn)


class InflowViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    serializer_class = InflowSerializer
    queryset = Inflow.objects.all()
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['amount','referenceID']
    def get_queryset(self):
        if 'search' in self.request.GET:
            try:
                toRet = Inflow.objects.filter(amount__gte=int(self.request.GET['search']))
                return toRet
            except:
                return Inflow.objects.all()
        else:
            return Inflow.objects.all()

class CostCenterViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    serializer_class = CostCenterSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['name']
    def get_queryset(self):
        divsn = self.request.user.designation.division
        if self.request.method in ['DELETE' ,'PATCH']:
            return CostCenter.objects.filter(division = divsn)
        else:
            return CostCenter.objects.filter(division = divsn , parent__isnull = True)


class TransactionViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    serializer_class = TransactionSerializer
    # queryset = Transaction.objects.all()
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['toAcc','outBound']
    def get_queryset(self):
        if 'filterBoth' in self.request.GET:
            toRet = Transaction.objects.filter(Q(fromAcc__number=self.request.GET['filterBoth'])|Q(toAcc__number=self.request.GET['filterBoth']))
            return toRet
        elif 'account__filter' in self.request.GET:
            toRet = Transaction.objects.filter(toAcc__group=None , toAcc__personal=False)
            return toRet
        else:
            return Transaction.objects.all()

class ExpenseSheetViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    serializer_class = ExpenseSheetSerializer
    queryset = ExpenseSheet.objects.all()
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['user','notes','stage','approved']

class ExpenseSheetLiteViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    serializer_class = ExpenseSheetLiteSerializer
    queryset = ExpenseSheet.objects.all()
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['user','notes','stage']

class ExpenseHeadingViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    serializer_class = ExpenseHeadingSerializer
    queryset = ExpenseHeading.objects.all()
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['title']

class ExpenseViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    serializer_class = ExpenseSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['description', 'dated','sheet']
    def get_queryset(self):
        u = self.request.user
        if 'sheet__isnull' in self.request.GET:
            return Expense.objects.filter(sheet__isnull = True)
        return Expense.objects.all()

class VendorProfileViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    serializer_class = VendorProfileSerializer
    queryset = VendorProfile.objects.all()
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['service']

class VendorServiceViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    serializer_class = VendorServiceSerializer
    queryset = VendorService.objects.all()
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['vendorProfile']





class CreateSalesTransactionAPI(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny ,)
    def post(self,request , format= None):
        data = request.data
        outBondObj = Sale.objects.get(pk = int(data['purchase']))
        name = outBondObj.personName
        accountObj = Account.objects.get(pk = int(data['account']))
        fromAccObj = Account.objects.get(pk = int(data['frmaccount']))
        accountObj.balance = accountObj.balance + float(data['amount'])
        accountObj.save()
        fromAccObj.balance = fromAccObj.balance + float(data['amount'])
        fromAccObj.save()
        transObj = Transaction.objects.create(toAcc = accountObj , credit = float(data['amount']) , dated = data['dated'] , user = request.user, outBound = outBondObj)
        if 'externalReferenceID' in data:
            transObj.externalReferenceID = data['externalReferenceID']
            transObj.groupId = data['externalReferenceID']
        transObj.balance = accountObj.balance
        transObj.narration = 'Invoice Payment'
        transObj.save()
        retailAcc = Account.objects.get(title = 'Accounts receivable')
        retailAcc.balance = retailAcc.balance - float(data['amount'])
        retailAcc.save()
        transObj1 = Transaction.objects.create(fromAcc = retailAcc , debit = float(data['amount']) , dated = data['dated'] , user = request.user, outBound = outBondObj)
        if 'externalReferenceID' in data:
            transObj1.externalReferenceID = data['externalReferenceID']
            transObj1.groupId = data['externalReferenceID']
        transObj1.balance = accountObj.balance
        transObj1.narration = 'Invoice Payment'
        transObj1.save()
        try:
            venAcc = Account.objects.get(title = name)
            venAcc.balance = venAcc.balance - float(data['amount'])
            venAcc.save()
            transObj2= Transaction.objects.create(fromAcc = venAcc , debit = float(data['amount']) , dated = data['dated'] , user = request.user, outBound = outBondObj)
            if 'externalReferenceID' in data:
                transObj2.externalReferenceID = data['externalReferenceID']
                transObj2.groupId = data['externalReferenceID']
                transObj2.balance = accountObj.balance
                transObj2.narration = 'Invoice Payment'
                transObj2.save()
                transObj3 = Transaction.objects.create(toAcc = fromAccObj , credit = float(data['amount']) , dated = data['dated'] , user = request.user, outBound = outBondObj)
                if 'externalReferenceID' in data:
                    transObj3.externalReferenceID = data['externalReferenceID']
                    transObj3.groupId = data['externalReferenceID']
                    transObj3.balance = accountObj.balance
                    transObj3.narration = 'Invoice Payment'
                    transObj3.save()
        except:
            pass
        outBondObj.paidAmount = float(outBondObj.paidAmount) + float(data['amount'])
        outBondObj.balanceAmount = float(outBondObj.total) - float(outBondObj.paidAmount)
        outBondObj.save()
        data = TransactionSerializer(transObj,many=False).data
        return Response(data)

class SaleViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.AllowAny ,)
    serializer_class = SaleSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['poNumber','status','id','isInvoice','phone','isPerforma','cancelled']
    def get_queryset(self):
        outBondObj = Sale.objects.all().order_by('-created')
        if 'search__in' in self.request.GET:
            search = self.request.GET['search__in']
            outBondObj = outBondObj.filter(Q(poNumber__icontains=search) | Q(name__icontains=search) | Q(personName__icontains=search)
            | Q(phone__icontains=search) | Q(email__icontains=search) | Q(address__icontains=search) )
        if 'filters' in self.request.GET:
            status = self.request.GET['filters']
            if status == 'Cancelled':
                data = outBondObj.filter(cancelled=True)
            else:
                data = outBondObj.filter(status = status).exclude(cancelled=True)
            toReturn =  data
            if 'sort' in self.request.GET:
                toReturn = toReturn.order_by('total')
            if 'haveparent' in self.request.GET:
                toReturn = toReturn.filter(parent__isnull = True)
            return toReturn
        elif 'sort' in self.request.GET:
                outBondObj = outBondObj.order_by('total')
        if 'haveparent' in self.request.GET:
            outBondObj1 = outBondObj.filter(parent__isnull = True)
            return  outBondObj1
        return outBondObj



class GetUpdatedInvoiceAmountAPI(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny ,)
    def get(self,request , format= None):
        data = {}
        if 'id' in request.GET:
            obj = Sale.objects.get(pk = int(request.GET['id']))
            objData = Sale.objects.filter(parent = obj)
            total = objData.aggregate(tot = Sum('paidAmount'))
            paid = 0
            if total['tot'] is not None:
                paid = total['tot']
            data = {'total' : obj.total , 'paidAmount' : obj.paidAmount , 'balanceAmount' : obj.balanceAmount, 'receivedAmount': paid}
        return Response(data, status = status.HTTP_200_OK)

class SaleAllViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.AllowAny ,)
    serializer_class = SaleAllSerializer
    queryset = Sale.objects.all()
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['poNumber','status','id','isInvoice','phone']

class SalesQtyViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.AllowAny ,)
    serializer_class = SalesQtySerializer
    queryset = SalesQty.objects.all()
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['outBound','product']

class CategoryViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.AllowAny ,)
    serializer_class = CategorySerializer
    def get_queryset(self):
        divsn = self.request.user.designation.division
        return Category.objects.filter(division = divsn)

class RateListViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.AllowAny ,)
    serializer_class = RateListSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['sellable','category','name']
    def get_queryset(self):
        divsn = self.request.user.designation.division
        return Inventory.objects.filter(division = divsn)

class InventoryLogViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    serializer_class = InventoryLogSerializer
    queryset = InventoryLog.objects.all()

from paypal.standard.forms import PayPalPaymentsForm

def paypal_return_view(request):
    orderObj = Sale.objects.last()
    outbound = orderObj.id
    outBoundQty = SalesQty.objects.filter(outBound = int(outbound))
    total = 0
    for i in outBoundQty:
        total  += round(i.total,2)
    return updateAndProcessOrder(request,outbound, total)

def paypal_cancel_view(request):
    return redirect('http://192.168.0.111:8000/checkout')

def paypalPaymentInitiate(request):
    # What you want the button to do.
    name = []
    product = ''
    if 'outbound' in  request.GET:
        total = 0
        orderid = request.GET['outbound']
        orderObj = Sale.objects.get(pk=orderid)
        outBound = SalesQty.objects.filter(outBound = orderid)
        count = 0
        for i in outBound:
            count += 1
            total  +=  i.total
            item = i.product  +  ' + '
            print i,'iiiiiiii'
            if count==len(outBound):
                name.append(i.product)
            else:
                name.append(item)
    for i in name:
        product += i

    paypal_dict = {
        "business": globalSettings.PAYPAL_RECEIVER_EMAIL,
        "amount": str(total),
        "item_name": product,
        "invoice": orderid,
        'currency_code': 'INR',
        "notify_url": request.build_absolute_uri(reverse('paypal-ipn')),
        "return": request.build_absolute_uri(reverse('paypal_return_view')),
        "cancel_return": request.build_absolute_uri(reverse('paypal_cancel_view')),
        "custom": "premium_plan",  # Custom command to correlate to some function later (optional)
    }
    form = PayPalPaymentsForm(initial=paypal_dict)
    context = {"form": form}
    print 'hhhhhhhhhhh'
    return render(request, "paypal.payment.html", context)



from django.template.loader import get_template
def updateAndProcessOrder(request,orderID , amnt, referenceId=None):
    print 'in updateAndProcessOrderupdateAndProcessOrder'
    orderObj = Sale.objects.get(id = orderID)
    value = SalesQty.objects.filter(outBound = orderID)
    orderObj.isInvoice = True
    orderObj.save()
    if orderObj.email:
        ctx = {
            'heading' : "Invoice Details",
            'recieverName' : orderObj.personName  ,
            'sendersAddress' : 'AquaSpa',
            'grandTotal':amnt,
            'value':value,
            'data':orderObj,

        }
        email_body = get_template('app.finance.emailDetail.html').render(ctx)
        email_subject = 'Order Placed'
        email_to = []
        email_to.append(str(orderObj.email))
        msg = EmailMessage(email_subject, email_body, to=email_to ,  )
        msg.content_subtype = 'html'
        msg.send()
    if globalSettings.G_ADMIN:
        ctx = {
            'heading' : "Invoice Details",
            'recieverName' : orderObj.personName  ,
            'sendersAddress' : 'AquaSpa',
            'grandTotal':amnt,
            'value':value,
            'data':orderObj,
        }
        email_body = get_template('app.finance.toAdminEmail.html').render(ctx)
        email_subject = 'Order Placed in AquaSpa'
        email_to = []
        for i in globalSettings.G_ADMIN:
            email_to.append(str(i))
        msg = EmailMessage(email_subject, email_body, to=email_to ,  )
        msg.content_subtype = 'html'
        msg.send()
    return render(request, 'paypal.sucess.html' ,{'amt':amnt})


class PaymentPageAPI(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny ,)
    def post(self,request , format= None):
        requestData = request.data
        name = request.data['personName']
        phone = request.data['phone']
        email = request.data['email']
        address = request.data['address']
        city = request.data['city']
        state = request.data['state']
        country = request.data['country']
        pincode = request.data['pincode']
        pin_status = request.data['pin_status']
        print name,phone,email,address,city,state,country,pincode,pin_status,'tttttttttttt'
        data = {'personName':name,'phone':phone,'email':email,'address':address,'city':city,'state':state,'country':country,'pincode':pincode,'pin_status':pin_status}
        outboundObj = Sale.objects.create(**data)
        try:
            outboundObj.division = request.user.designation.division
            outboundObj.save()
        except:
            pass
        products = request.data['item']
        outbound = outboundObj.id
        for p in products:
            productMeta = p['item']['productMeta']
            tax = productMeta['taxRate']
            hsn =   productMeta['pk']
            getProductmeta = ProductMeta.objects.get(pk=int(hsn))
            print hsn ,'lllllllll'
            qty = p['quantity']
            price = p['item']['rate']
            product = p['item']['name']
            getOutboundobj = Sale.objects.get(pk=int(outbound))
            total = ((p['rate']*tax)/100)+p['rate']
            productObj = {'outBound':getOutboundobj,'product':product,'price':price,'qty':qty,'tax':tax,'hsn':getProductmeta,'total':total}
            productObjCreate = SalesQty.objects.create(**productObj)
            try:
                productObjCreate.division = request.user.designation.division
                productObjCreate.save()
            except:
                pass
        SalesQtyObjs = SalesQty.objects.filter(outBound = int(outbound))
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment;filename="Invoicedownload.pdf"'
        invoice(response , outboundObj , SalesQtyObjs , 'outbound', request)
        f = open(os.path.join(globalSettings.BASE_DIR, 'media_root/Invoicedownload.pdf'), 'wb')
        f.write(response.content)
        f.close()
        email_subject = "Invoice"
        msgBody = "Hi "  + name + ",\n\n\t\t Find the attachment for the Invoice Details"
        email = email
        contactData = tuple(email.split(','))
        print contactData
        msg = EmailMessage(email_subject, msgBody,  to= contactData )
        a = str(f).split('media_root/')[1]
        b = str(a).split(', mode')[0]
        c = str(b).split("'")[0]
        msg.attach_file(os.path.join(globalSettings.MEDIA_ROOT,str(c)))
        msg.send()
        return Response({'orderpk':outbound})

class UplodInflowDataAPI(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.IsAuthenticated ,)
    def post(self , request , format = None):
        print 'entered','*******************'
        requestData = request.data
        print requestData
        tosend={}
        if 'exFile' in requestData:
            excelFile = request.FILES['exFile']
            if request.data['verified'] in ['true','1','yes',]:
                verified = True
            else:
                verified = False
            currency = request.data['currency']
            toAcc = Account.objects.get(pk=int(request.data['toAcc']))
            user = request.user
            print verified,type(verified),toAcc,currency,user
            wb = load_workbook(filename = BytesIO(excelFile.read()))
            for idx,ws in enumerate(wb.worksheets):
                wsTitle = ws.title
                print wsTitle,idx
                if idx==0:
                    storeData = []
                    print 'savingggggggg'
                    for i in range(2,ws.max_row+1):
                        try:
                            amount = int(ws['A'+str(i)].value) if ws['A'+str(i)].value else 0
                            referenceID = str(ws['B'+str(i)].value) if ws['B'+str(i)].value else None
                            dated = ws['C'+str(i)].value.date() if ws['C'+str(i)].value else None
                            description = str(ws['D'+str(i)].value) if ws['D'+str(i)].value else None
                            chequeNo = str(ws['E'+str(i)].value) if ws['E'+str(i)].value else None
                            mode = str(ws['F'+str(i)].value) if ws['F'+str(i)].value else 'cash'
                            mode = mode.lower()
                            gstCollected = float(ws['G'+str(i)].value) if ws['G'+str(i)].value else 0
                            try:
                                division = request.user.designation.division
                            except:
                                division = None
                            infObj = Inflow(user=user,toAcc=toAcc,currency=currency,verified=verified,amount=amount,referenceID=referenceID,dated=dated,description=description,chequeNo=chequeNo,mode=mode,gstCollected=gstCollected,division=division)
                            storeData.append(infObj)
                        except:
                            print 'row number {0} in Excel - {1} is not created'.format(i,wsTitle)
                    Inflow.objects.bulk_create(storeData)
                    print 'total {0} Objects has been created'.format(len(storeData))
        return Response(tosend, status=status.HTTP_200_OK)

class GetExpenseDataAPI(APIView):
    permission_classes = (permissions.IsAuthenticated ,)
    def get(self , request , format = None):
        print 'entered','*******************'
        print request.GET
        accountsList = list(request.user.accountsManaging.all().values_list('pk',flat=True).distinct())
        print accountsList,'user accountssssss Listttttttt'
        tosend=[]
        expObj = ProjectPettyExpense.objects.filter(account__in=accountsList).values('project__pk','project__title','project__budget').distinct()
        print expObj
        for i in expObj:
            expTotal = ProjectPettyExpense.objects.filter(project__id=int(i['project__pk']),account__in=accountsList).aggregate(tot=Sum('amount'))
            expTotal = expTotal['tot'] if expTotal['tot'] else 0
            print expTotal
            data = {'projectPk':i['project__pk'],'projectName':i['project__title'],'expTotal':expTotal,'budget':i['project__budget']}
            tosend.append(data)
        if 'limit' in request.GET:
            try:
                lmt = int(request.GET['limit'])
                tosend = tosend[0:lmt]
            except:
                pass
        return Response(tosend, status=status.HTTP_200_OK)

class GetInventoryAPI(APIView):
    permission_classes = (permissions.IsAuthenticated ,)
    def get(self , request , format = None):
        params = request.GET
        invObj = Inventory.objects.all()
        data = ''
        offset=int(params['offset'])
        limit=int(params['limit'])
        if 'warehouse' in params:
            checkinObj = Checkin.objects.filter(unit__pk = int(params['warehouse']),checkout = False)
            invObjPks = checkinObj.values_list('inventory__pk' ,flat = True).distinct()
            invObj = Inventory.objects.filter(pk__in = invObjPks)
            if 'search' in params:
                invObj = invObj.filter(Q(name__icontains = params['search'])|Q(sku__icontains = params['search']))
            data = RateListSerializer(invObj[offset:limit], many = True).data
        return Response(data, status=status.HTTP_200_OK)

class ExpensesGraphDataAPI(APIView):
    permission_classes = (permissions.IsAuthenticated ,)
    def get(self , request , format = None):
        print 'entered','*******************'
        print request.GET
        tosend={'labels':[],'datasets':[]}
        allExpenses = ProjectPettyExpense.objects.all()
        projLists = allExpenses.values('project__pk','project__title').distinct()
        accountList = allExpenses.values('account__pk','account__title').distinct()
        for idx,i in enumerate(accountList):
            data = {'label':i['account__title'],'data':[]}
            for j in projLists:
                if idx==0:
                    tosend['labels'].append(j['project__title'])
                expTotal  = {'tot' : 0}
                if j['project__pk']!=None:
                    expTotal = allExpenses.filter(project__id=int(j['project__pk']),account__id=int(i['account__pk'])).aggregate(tot=Sum('amount'))
                expTotal = expTotal['tot'] if expTotal['tot'] else 0
                data['data'].append(expTotal)
            tosend['datasets'].append(data)

        return Response(tosend, status=status.HTTP_200_OK)

class MonthsExpensesDataAPI(APIView):
    permission_classes = (permissions.IsAuthenticated ,)
    def get(self , request , format = None):
        print 'entered','*******************'
        tosend={'labels':[],'datasets':[]}
        today = datetime.today()
        lstYear = today + relativedelta(years=-1,months=-1)
        print today , lstYear
        allExpenses = ProjectPettyExpense.objects.all()
        for i in range(13):
            lstYear += relativedelta(months=1)
            mth = lstYear.month
            yr = lstYear.year
            label = str(calendar.month_abbr[mth]) + '-' + str(yr)
            tosend['labels'].append(label)
            expTotal = allExpenses.filter(created__year=yr,created__month=mth).aggregate(tot=Sum('amount'))
            expTotal = expTotal['tot'] if expTotal['tot'] else 0
            tosend['datasets'].append(expTotal)

        return Response(tosend, status=status.HTTP_200_OK)

class DownloadExpenseSummaryAPI(APIView):
    permission_classes = (permissions.IsAuthenticated ,)
    def get(self , request , format = None):
        print 'entered','*******************'

        hdFont = Font(size=12,bold=True)
        alphaChars = list(string.ascii_uppercase)
        workbook = Workbook()
        Sheet1 = workbook.active
        Sheet1.title = 'Expense Summary'
        hd = ["Project", "Budget",'Expense']
        hdWidth = [10,10,10]
        Sheet1.append(hd)
        projectsList = project.objects.filter(projectClosed=False)
        print projectsList.count()
        for i in projectsList:
            data = [i.title,i.budget]
            try:
                expTotal = ProjectPettyExpense.objects.filter(project=i).aggregate(tot=Sum('amount'))
                expTotal = expTotal['tot'] if expTotal['tot'] else 0
            except:
                expTotal = 0
            data.append(expTotal)
            Sheet1.append(data)
            for idx,j in enumerate(data):
                if (len(str(j))+5) > hdWidth[idx]:
                    hdWidth[idx] = len(str(j)) + 5
        for idx,i in enumerate(hd):
            cl = str(alphaChars[idx])+'1'
            Sheet1[cl].fill = PatternFill(start_color="48dce0", end_color="48dce0", fill_type = "solid")
            Sheet1[cl].font = hdFont
            Sheet1.column_dimensions[str(alphaChars[idx])].width = hdWidth[idx]

        allExpenses = ProjectPettyExpense.objects.all()
        projLists = allExpenses.values('project__pk','project__title').distinct()
        for i in projLists:
            Sheet = workbook.create_sheet(i['project__title'])
            hd = ['Title','Amount','Account No.','User Name','Description','Dated']
            hdWidth = [10,10,10,10,10,10]
            Sheet.append(hd)

            ptObjs = allExpenses.filter(project__id=int(i['project__pk'])).order_by('created')
            for j in ptObjs:
                uName = j.createdUser.first_name
                if j.createdUser.last_name:
                    uName += ' ' + j.createdUser.last_name
                data = [j.heading.title,j.amount,j.account.number,uName,j.description,j.created.date()]
                Sheet.append(data)
                for idx,k in enumerate(data):
                    if (len(str(k))+5) > hdWidth[idx]:
                        hdWidth[idx] = len(str(k)) + 5
            for idx,j in enumerate(hd):
                cl = str(alphaChars[idx])+'1'
                Sheet[cl].fill = PatternFill(start_color="48dce0", end_color="48dce0", fill_type = "solid")
                Sheet[cl].font = hdFont
                Sheet.column_dimensions[str(alphaChars[idx])].width = hdWidth[idx]

        response = HttpResponse(content=save_virtual_workbook(workbook),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=ExpenseSummary.xlsx'
        return response


def grn(response , project , purchaselist , request):
    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(response,pagesize=letter, topMargin=0.2*cm,leftMargin=0.1*cm,rightMargin=0.1*cm)
    doc.request = request
    elements = []

    p1 = Paragraph("<para alignment='center'fontSize=15  ><b> Goods Received Note </b></para>",styles['Normal'])
    elements.append(p1)
    elements.append(Spacer(1, 10))
    try:
        address = project.address.replace('\n', '<br />')
    except:
        address = project.address
    addrdetails = Paragraph("""
    <para >
    <b>%s</b><br/>
    %s <br/>
    </para>
    """ %(project.name ,address),styles['Normal'])
    td=[[addrdetails]]
    t=Table(td,colWidths=[4*inch])
    t.hAlign = 'LEFT'
    elements.append(t)
    elements.append(Spacer(1,10))
    p9_01 =Paragraph("<para  fontSize=11  > <b>Sl. no</b></para>",styles['Normal'])
    p9_02 =Paragraph("<para  fontSize=11  ><b>Product</b></para>",styles['Normal'])
    p9_03 =Paragraph("<para  fontSize=11  ><b>Quantity</b></para>",styles['Normal'])
    data2=[[p9_01,p9_02,p9_03]]
    id=0
    for i in purchaselist:
        id+=1
        product = i.product
        quanty = i.receivedQty
        p10_01 =Paragraph("<para  fontSize=11  >{0}</para>".format(id),styles['Normal'])
        p10_02 =Paragraph("<para  fontSize=11  >{0}</para>".format(product),styles['Normal'])
        p10_03 =Paragraph("<para  fontSize=11  >{0}</para>".format(quanty),styles['Normal'])
        data2.append([p10_01,p10_02,p10_03])
    t3=Table(data2,colWidths=[0.5*inch , 3*inch , 1*inch])
    t3.hAlign = 'LEFT'
    t3.setStyle(TableStyle([('TEXTFONT', (0, 0), (-1, -1), 'Times-Bold'),('TEXTCOLOR',(0,0),(-1,-1),black),('ALIGN',(0,0),(-1,-1),'LEFT'),('VALIGN',(0,0),(-1,-1),'TOP'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black)]))
    elements.append(t3)
    doc.build(elements)


class GrnAPIView(APIView):
    def get(self , request , format = None):
        project = InvoiceReceived.objects.get(pk = request.GET['value'])
        purchaselist = InvoiceQty.objects.filter(invoice = request.GET['value'])
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment;filename="Grndownload.pdf"'
        grn(response , project , purchaselist , request)
        return response


def poDownload(response , inv , invdetails , request):
    settingsFields = application.objects.get(name='app.CRM').settings.all()
    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(response,pagesize=letter, topMargin=0.2*cm,leftMargin=0.1*cm,rightMargin=0.1*cm)
    doc.request = request
    elements = []
    company = ''
    address = ''
    bankDetails = ''
    if inv.user!=None and inv.user.designation!=None:
        if inv.user.designation.unit!=None:
            if inv.user.designation.unit.name!=None:
                company = inv.user.designation.unit.name
            if inv.user.designation.unit.address!=None:
                address = inv.user.designation.unit.address + ' ' + inv.user.designation.unit.city + ' ' +  inv.user.designation.unit.state + ' ' + inv.user.designation.unit.country + ' - ' +   inv.user.designation.unit.pincode
            if inv.user.designation.unit.bankName!=None:
                bankDetails = bankDetails + 'Bank Name : ' + inv.user.designation.unit.bankName + '<br/> '
            if inv.user.designation.unit.bankBranch!=None:
                bankDetails = bankDetails + 'Branch : ' + inv.user.designation.unit.bankBranch + ' <br/>'
            if inv.user.designation.unit.bankAccNumber!=None:
                bankDetails = bankDetails + 'Account No : ' + inv.user.designation.unit.bankAccNumber + '<br/>'
            if inv.user.designation.unit.ifsc!=None:
                bankDetails = bankDetails + 'IFSC No : ' + inv.user.designation.unit.ifsc
    headerDetails = Paragraph("""
    <para align='center'>
    <font size ='8'>
    <b>%s</b><br/>
    %s<br/>
    Phone : 000000000<br/>
    </font>
    </para>
    """ %(company,address),styles['Normal'])
    tdheader=[['',headerDetails,'']]
    headerTitle = Paragraph("""
    <para align='center'>
    <font size ='12'>
    <b> Purchase Order</b></font>
    </para>
    """ %(),styles['Normal'])
    tdheader+=[['',headerTitle,'']]
    t2=Table(tdheader,colWidths=(54.7*mm,100*mm,54.7*mm))
    t2.setStyle(TableStyle([('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('BOX',(0,0),(-1,-1),0.25,colors.black), ('LINEABOVE', (0,1), (-1,-1), 0.25, colors.black),]))
    elements.append(t2)
    detail31 = Paragraph("""
    <para align='center'>
    <b>Vendor Address</b>
    </para>
    """ %(),styles['Normal'])
    detail32 = Paragraph("""
    <para align='center'>
    <b>Shipping Address</b>
    </para>
    """ %(),styles['Normal'])
    t2data=[detail31],[detail32]
    td2header=[t2data]
    t5=Table(td2header)
    t5.setStyle(TableStyle([('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black)]))
    elements.append(t5)
    detail41 = Paragraph("""
    <para>
    Name :
    </para>
    """ %(),styles['Normal'])
    detail42 = Paragraph("""
    <para >
    %s %s
    </para>
    """ %(inv.requester.first_name , inv.requester.last_name ),styles['Normal'])
    detail43 = Paragraph("""
    <para >
    Name :
    </para>
    """ %(),styles['Normal'])
    detail44 = Paragraph("""
    <para >
    %s
    </para>
    """ %(inv.personName),styles['Normal'])
    t4data=[detail41],[detail44],[detail43],[detail42]
    td4header=[t4data]
    t6=Table(td4header)
    t6.setStyle(TableStyle([('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black)]))
    elements.append(t6)
    try:
        shipaddr = inv.name + '<br /> ' +inv.address.replace('\n', '<br />') +', '+inv.city +'<br /> '+ inv.state +', '+ inv.country + ' - ' +inv.pincode
    except:
        shipaddr =inv.name + '<br /> '+ inv.address + inv.city +'<br /> '+ inv.state +', '+ inv.country + ' - ' +inv.pincode

    detail51 = Paragraph("""
    <para>
    Address : <br/> %s <br/>
    </para>
    """ %(shipaddr),styles['Normal'])
    detail52 = Paragraph("""
    <para>
    Address : <br/> %s <br/>
    </para>
    """ %(address),styles['Normal'])
    t5data=[detail51],[detail52]
    td5header=[t5data]
    t7=Table(td5header)
    t7.setStyle(TableStyle([('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black)]))
    elements.append(t7)
    data2=[]
    s01 =Paragraph("<para  fontSize=11  >S.No </para>",styles['Normal'])
    s02 =Paragraph("<para  fontSize=11  >Product </para>",styles['Normal'])
    s03 =Paragraph("<para  fontSize=11  >Qty </para>",styles['Normal'])
    s04 =Paragraph("<para  fontSize=11  >Rate </para>",styles['Normal'])
    data2 += [[s01,s02,s03,s04]]
    id = 0
    grandtot= 0
    for i in invdetails:
        id+=1
        total = i.price
        s21 =Paragraph("<para  fontSize=11  >{0} </para>".format(id),styles['Normal'])
        s22 =Paragraph("<para  fontSize=11  >{0} </para>".format(i.product),styles['Normal'])
        s23 =Paragraph("<para  fontSize=11 >{0} </para>".format(i.qty),styles['Normal'])
        s24 =Paragraph("<para  fontSize=11    alignment='right'>{:,} </para>".format(round(i.price,2)),styles['Normal'])
        data2.append([s21,s22,s23,s24])
    s24 =Paragraph("<para  fontSize=11  > </para>",styles['Normal'])
    s25 =Paragraph("<para  fontSize=11    alignment='right'><b>Total</b></para>",styles['Normal'])
    s26 =Paragraph("<para  fontSize=11  > </para>",styles['Normal'])
    s27 =Paragraph("<para  fontSize=11   alignment='right'><b>{:,}</b> </para>".format(round(total,2)),styles['Normal'])
    data2.append([s24,s25,s26,s27])
    t9=Table(data2)
    t9.hAlign = 'LEFT'
    t9.setStyle(TableStyle([('TEXTFONT', (0, 0), (-1, -1), 'Times-Bold'),('TEXTCOLOR',(0,0),(-1,-1),black),('ALIGN',(0,0),(-1,-1),'RIGHT'),('VALIGN',(0,0),(-1,-1),'TOP'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black)]))
    elements.append(t9)
    gtotalText = num2words(int(total), to='cardinal', lang='en_IN')
    print gtotalText
    s41 =Paragraph("<para  fontSize=11  > Rupees {0} </para>".format(gtotalText),styles['Normal'])
    datawords =[[s41]]
    t10=Table(datawords)
    t10.hAlign = 'LEFT'
    t10.setStyle(TableStyle([('TEXTFONT', (0, 0), (-1, -1), 'Times-Bold'),('TEXTCOLOR',(0,0),(-1,-1),black),('ALIGN',(0,0),(-1,-1),'RIGHT'),('VALIGN',(0,0),(-1,-1),'TOP'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black)]))
    elements.append(t10)
    try:
        invtrms = inv.invoiceTerms.split('||')
    except:
        invtrms = inv.invoiceTerms
    dataFooter = []
    s51 =Paragraph("<para  fontSize=11  > </para>".format(),styles['Normal'])
    s52 =Paragraph("<para  fontSize=11  > </para>",styles['Normal'])
    s53 =Paragraph("<para fontSize=6 alignment='center'> Certified that the particulars given above are true and correct <br/></para><para fontSize=10> {0}</para>".format(company),styles['Normal'])
    dataFooter =[[s51,s52,s53]]
    s61 =Paragraph("<para  fontSize=11  >{0}</para>".format(bankDetails),styles['Normal'])
    s62 =Paragraph("<para  fontSize=11  > </para>",styles['Normal'])
    s63 =Paragraph("<para  fontSize=11  > </para>",styles['Normal'])
    dataFooter +=[[s61,s62,s63]]
    t11=Table(dataFooter,colWidths=(80*mm,49.5*mm,80*mm))
    t11.setStyle(TableStyle([('TEXTFONT', (0, 0), (-1, -1), 'Times-Bold'),('TEXTCOLOR',(0,0),(-1,-1),black),('ALIGN',(0,0),(-1,-1),'RIGHT'),('VALIGN',(0,0),(-1,-1),'TOP'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black), ('LINEABOVE', (1,1), (-1,-1), 0.50, colors.white),]))
    elements.append(t11)
    dataFooter = []
    s51 =Paragraph("<para  fontSize=11  >Terms and Conditions : </para>",styles['Normal'])
    dataFooter =[[s51]]
    count = 0
    if invtrms is not None:
        for i in invtrms:
            count+=1
            s52 =Paragraph("<para  > {0}. {1} </para>".format(count,i),styles['Normal'])
            dataFooter.append([s52])
        t12=Table(dataFooter)
        t12.setStyle(TableStyle([('TEXTFONT', (0, 0), (-1, -1), 'Times-Bold'),('TEXTCOLOR',(0,0),(-1,-1),black),('ALIGN',(0,0),(-1,-1),'RIGHT'),('VALIGN',(0,0),(-1,-1),'TOP'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.white), ('LINEABOVE', (1,1), (-1,-1), 0.50, colors.white),]))
        elements.append(t12)
    doc.build(elements)


def invoice(response , inv , invdetails , typ, request):
    settingsFields = application.objects.get(name='app.CRM').settings.all()
    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(response,pagesize=letter, topMargin=0.2*cm,leftMargin=0.1*cm,rightMargin=0.1*cm)
    doc.request = request
    currentYear =  datetime.today().year
    elements = []
    company = ''
    address = ''
    bankDetails = ''
    dividionName = ''
    mobile = ''
    gstin = ''
    cin = ''
    cin = ''
    wesite = ''
    email = ''
    if inv.user!=None and inv.user.designation!=None:
        if inv.user.designation.unit!=None:
            if inv.user.designation.unit.name!=None:
                company = inv.user.designation.unit.name
            if inv.user.designation.unit.email!=None:
                email = inv.user.designation.unit.email
            if inv.user.designation.unit.address!=None:
                address = inv.user.designation.unit.address + ' ' + inv.user.designation.unit.city + ' ' +  inv.user.designation.unit.state + ' ' + inv.user.designation.unit.country + ' - ' +   inv.user.designation.unit.pincode
            if inv.user.designation.unit.bankName!=None:
                bankDetails = bankDetails + 'Bank Name : ' + inv.user.designation.unit.bankName + '<br/> '
            if inv.user.designation.unit.bankBranch!=None:
                bankDetails = bankDetails + 'Branch : ' + inv.user.designation.unit.bankBranch + ' <br/>'
            if inv.user.designation.unit.bankAccNumber!=None:
                bankDetails = bankDetails + 'Account No : ' + inv.user.designation.unit.bankAccNumber + '<br/>'
            if inv.user.designation.unit.ifsc!=None:
                bankDetails = bankDetails + 'IFSC No : ' + inv.user.designation.unit.ifsc + '<br/>'
            if inv.user.designation.unit.mobile!=None:
                mobile = inv.user.designation.unit.mobile
            if inv.user.designation.unit.gstin!=None:
                gstin = inv.user.designation.unit.gstin
                bankDetails = bankDetails + 'GSTIN : ' + gstin + '<br/>'
    if inv.user.designation.division!=None:
        divisionName = inv.user.designation.division.name
        if inv.user.designation.division.pan!=None:
            pan = inv.user.designation.division.pan
            bankDetails = bankDetails + 'PAN : ' + pan + '<br/>'
        if inv.user.designation.division.website!=None:
            website = inv.user.designation.division.website
        if inv.user.designation.division.cin!=None:
            cin = inv.user.designation.division.cin
            bankDetails = bankDetails + 'CIN : ' + cin
    headerDetails = Paragraph("""
    <para align='center'>
    <font size ='18'>
    <b>%s</b><br/></font>
    <font size ='8'><b>%s</b><br/>
    %s<br/>
    Phone : %s, Website : %s, Email : %s <br/>
    </font>
    </para>
    """ %(divisionName,company,address, mobile, website, email),styles['Normal'])
    from reportlab.platypus import Image
    if inv.user.designation.division is not None and inv.user.designation.division.logo is not None:
        imagePath = os.path.join(globalSettings.MEDIA_ROOT , str(inv.user.designation.division.logo))
        f = open(imagePath, 'rb')
        ima = Image(f)
        ima.drawHeight = 0.5*inch
        ima.drawWidth = 1*inch
        ima.hAlign = 'CENTER'
        tdheader=[[ima,headerDetails,'']]
    else:
        tdheader=[['',headerDetails,'']]
    title = 'SALES ORDER'
    if inv.isInvoice:
        title = 'INVOICE'
    if inv.isPerforma:
        title = 'PERFORMA'
    headerDate = Paragraph("""
    <para align='left'>
    <font size ='12'>
    <b> Date : %s</b></font>
    </para>
    """ %(inv.created.date()),styles['Normal'])
    headerTitle = Paragraph("""
    <para align='center'>
    <font size ='12'>
    <b> %s</b></font>
    </para>
    """ %(title),styles['Normal'])
    headerId = Paragraph("""
    <para align='right'>
    <font size ='12'>
    <b>ID: %s</b></font>
    </para>
    """ %('#INV-'+str(currentYear)+'-'+str(inv.pk)),styles['Normal'])
    tdheader+=[[headerDate,headerTitle,headerId]]
    t2=Table(tdheader,colWidths=(54.7*mm,100*mm,54.7*mm))
    t2.setStyle(TableStyle([('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('BOX',(0,0),(-1,-1),0.25,colors.black), ('LINEABOVE', (0,1), (-1,-1), 0.25, colors.black),]))
    elements.append(t2)
    detail31 = Paragraph("""
    <para align='center'>
    <b>Billing Address</b>
    </para>
    """ %(),styles['Normal'])
    detail32 = Paragraph("""
    <para align='center'>
    <b>Shipping Address</b>
    </para>
    """ %(),styles['Normal'])
    t2data=[detail31],[detail32]
    td2header=[t2data]
    t5=Table(td2header)
    t5.setStyle(TableStyle([('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black)]))
    elements.append(t5)
    detail41 = Paragraph("""
    <para>
    Name :
    </para>
    """ %(),styles['Normal'])
    detail42 = Paragraph("""
    <para >
    %s
    </para>
    """ %(inv.personName),styles['Normal'])
    detail43 = Paragraph("""
    <para >
    Name :
    </para>
    """ %(),styles['Normal'])
    detail44 = Paragraph("""
    <para >
    %s
    </para>
    """ %(inv.personName),styles['Normal'])
    t4data=[detail41],[detail42],[detail43],[detail44]
    td4header=[t4data]
    t6=Table(td4header)
    t6.setStyle(TableStyle([('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black)]))
    elements.append(t6)
    detailPh1 = Paragraph("""
    <para>
    Contact Number :
    </para>
    """ %(),styles['Normal'])
    detailPh2 = Paragraph("""
    <para >
    %s
    </para>
    """ %(inv.phone),styles['Normal'])

    t4Ph1data=[detailPh1],[detailPh2],[],[]
    td4Phheader=[t4Ph1data]
    tPH6=Table(td4Phheader)
    tPH6.setStyle(TableStyle([('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (1,1), 0.25, colors.black)]))
    elements.append(tPH6)

    try:
        billaddr = inv.address.replace('\n', '<br />')
    except:
        billaddr = inv.address
    try:
        shipaddr = inv.address.replace('\n', '<br />')
    except:
        shipaddr = inv.address
    if inv.city is not None:
        billaddr+= ' ' +inv.city
        shipaddr+= ' ' +inv.city
    if inv.state is not None:
        billaddr+= ' ' +inv.state
        shipaddr+= ' ' +inv.state
    if inv.country is not None:
        billaddr+= ' ' +inv.country
        shipaddr+= ' ' +inv.country
    if inv.name is not None:
        company = inv.name
    else:
        company = 'NA'
    if inv.gstIn is not None:
        gstIn = inv.gstIn
    else:
        gstIn = 'NA'


    detail51 = Paragraph("""
    <para>
     <br/>  %s<br/>  Address : <br/> %s <br/>
    %s  <br/> GSTIN : %s
    </para>
    """ %(company,billaddr,inv.pincode, gstIn),styles['Normal'])
    detail52 = Paragraph("""
    <para>
     <br/>  %s<br/> Address : <br/> %s <br/>
    %s
    </para>
    """ %(company,shipaddr,inv.pincode),styles['Normal'])
    t5data=[detail51],[detail52]
    td5header=[t5data]
    t7=Table(td5header)
    t7.setStyle(TableStyle([('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black)]))
    elements.append(t7)
    data2=[]
    s01 =Paragraph("<para  fontSize=11  >S.No </para>",styles['Normal'])
    s02 =Paragraph("<para  fontSize=11  >Product </para>",styles['Normal'])
    s03 =Paragraph("<para  fontSize=11  >Qty </para>",styles['Normal'])
    s04 =Paragraph("<para  fontSize=11  >Rate </para>",styles['Normal'])
    s05 =Paragraph("<para  fontSize=11  >HSN/SAC Code</para>",styles['Normal'])
    s06 =Paragraph("<para  fontSize=11  >Tax  </para>",styles['Normal'])
    s07 =Paragraph("<para  fontSize=11  > Total </para>",styles['Normal'])
    data2 += [[s01,s02,s03,s04,s05,s06,s07]]
    id = 0
    grandtot= 0
    for i in invdetails:
        id+=1
        grandtot += i.total
        total = i.price
        print grandtot
        s21 =Paragraph("<para  fontSize=11  >{0} </para>".format(id),styles['Normal'])
        s22 =Paragraph("<para  fontSize=11  >{0} </para>".format(i.product),styles['Normal'])
        s23 =Paragraph("<para  fontSize=11   alignment='center'>{0} </para>".format(i.qty),styles['Normal'])
        s24 =Paragraph("<para  fontSize=11    alignment='right'>{:,} </para>".format(round(total,2)),styles['Normal'])
        if i.hsn is not None:
            hsn = i.hsn
        else:
             hsn = ''
        if i.taxPer is not None:
            taxPer = i.taxPer
        else:
            taxPer = ''
        if i.tax is not None:
            tax = i.tax
        else:
             tax = ''
        s25 =Paragraph("<para  fontSize=11   alignment='right'> {0} ({1}%)</para>".format(hsn, taxPer),styles['Normal'])
        s26 =Paragraph("<para  fontSize=11   alignment='right'>{0} </para>".format(round(tax,2)),styles['Normal'])
        s27 =Paragraph("<para  fontSize=11   alignment='right'>{:,} </para>".format(round(i.total,2)),styles['Normal'])
        data2.append([s21,s22,s23,s24,s25,s26,s27])
    s21 =Paragraph("<para  fontSize=11  > </para>",styles['Normal'])
    s22 =Paragraph("<para  fontSize=11  ></para>",styles['Normal'])
    s23 =Paragraph("<para  fontSize=11  > </para>",styles['Normal'])
    s24 =Paragraph("<para  fontSize=11  > </para>",styles['Normal'])
    s25 =Paragraph("<para  fontSize=11    alignment='right'><b>Total</b></para>",styles['Normal'])
    s26 =Paragraph("<para  fontSize=11  > </para>",styles['Normal'])
    s27 =Paragraph("<para  fontSize=11   alignment='right'><b>{:,}</b> </para>".format(round(grandtot,2)),styles['Normal'])
    data2.append([s21,s22,s23,s24,s25,s26,s27])
    t9=Table(data2,colWidths=(14.5*mm,60*mm,15*mm,30*mm,30*mm,30*mm,30*mm))
    t9.hAlign = 'LEFT'
    t9.setStyle(TableStyle([('TEXTFONT', (0, 0), (-1, -1), 'Times-Bold'),('TEXTCOLOR',(0,0),(-1,-1),black),('ALIGN',(0,0),(-1,-1),'RIGHT'),('VALIGN',(0,0),(-1,-1),'TOP'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black)]))
    elements.append(t9)
    gtotalText = num2words(int(grandtot), to='cardinal', lang='en_IN')
    print gtotalText
    s41 =Paragraph("<para  fontSize=11  >Total Payable: Rupees {0} </para>".format(gtotalText),styles['Normal'])
    datawords =[[s41]]
    t10=Table(datawords)
    t10.hAlign = 'LEFT'
    t10.setStyle(TableStyle([('TEXTFONT', (0, 0), (-1, -1), 'Times-Bold'),('TEXTCOLOR',(0,0),(-1,-1),black),('ALIGN',(0,0),(-1,-1),'RIGHT'),('VALIGN',(0,0),(-1,-1),'TOP'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black)]))
    elements.append(t10)
    dataFooter = []
    s61 =Paragraph("<para  fontSize=11  >{0}</para>".format(bankDetails),styles['Normal'])
    s62 =Paragraph("<para  fontSize=11  > </para>",styles['Normal'])
    s63 =Paragraph("<para fontSize=6 alignment='center'> Certified that the particulars given above are true and correct <br/></para>",styles['Normal'])
    dataFooter +=[[s61,s62,s63]]
    t11=Table(dataFooter,colWidths=(80*mm,49.5*mm,80*mm))
    t11.setStyle(TableStyle([('TEXTFONT', (0, 0), (-1, -1), 'Times-Bold'),('TEXTCOLOR',(0,0),(-1,-1),black),('ALIGN',(0,0),(-1,-1),'RIGHT'),('VALIGN',(0,0),(-1,-1),'TOP'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black), ('LINEABOVE', (1,1), (-1,-1), 0.25, colors.white),]))
    elements.append(t11)

    try:
        invtrms = inv.terms.split('||')
    except:
        invtrms = inv.terms
    s53 =Paragraph("<para  > <b> Generated By :  </b>{0} {1} </para>".format(request.user.first_name, request.user.last_name),styles['Normal'])
    genData = [[s53]]
    t13=Table(genData)
    t13.setStyle(TableStyle([('TEXTFONT', (0, 0), (-1, -1), 'Times-Bold'),('TEXTCOLOR',(0,0),(-1,-1),black),('ALIGN',(0,0),(-1,-1),'RIGHT'),('VALIGN',(0,0),(-1,-1),'TOP'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.white), ('LINEABOVE', (1,1), (-1,-1), 0.50, colors.white),]))
    elements.append(t13)
    count = 0
    tandc = []
    if invtrms is not None:
        s52 =Paragraph("<para  > <b> Terms and Conditions : </b> </para>",styles['Normal'])
        tandc.append([s52])
        for i in invtrms:
            count+=1
            s52 =Paragraph("<para  > {0}. {1} </para>".format(count,i),styles['Normal'])
            tandc.append([s52])
        t13=Table(tandc)
        t13.setStyle(TableStyle([('TEXTFONT', (0, 0), (-1, -1), 'Times-Bold'),('TEXTCOLOR',(0,0),(-1,-1),black),('ALIGN',(0,0),(-1,-1),'RIGHT'),('VALIGN',(0,0),(-1,-1),'TOP'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.white), ('LINEABOVE', (1,1), (-1,-1), 0.50, colors.white),]))
        elements.append(t13)
    doc.build(elements)

class PODownloadAPIView(APIView):
    permission_classes = (permissions.AllowAny,)
    def get(self , request , format = None):
        inv = InvoiceReceived.objects.get(pk = request.GET['value'])
        invDetails = InvoiceQty.objects.filter(invoice = request.GET['value'])
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment;filename="podownload.pdf"'
        poDownload(response , inv , invDetails, request)
        return response

class InvoiceAPIView(APIView):
    permission_classes = (permissions.AllowAny,)
    def get(self , request , format = None):
        typ = request.GET['typ']
        inv = Sale.objects.get(pk = request.GET['value'])
        invDetails = SalesQty.objects.filter(outBound = request.GET['value'])
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment;filename="Invoicedownload.pdf"'
        invoice(response , inv , invDetails , typ, request)
        return response

class SendInvoiceAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    def post(self, request, format=None):
        contactData=[]
        print request.data,'sssssssss'
        typ = request.data['typ']
        if typ=='inbond':
            inv = InvoiceReceived.objects.get(pk = request.data['value'])
            invDetails = InvoiceQty.objects.filter(invoice = request.data['value'])
        else:
            inv = Sale.objects.get(pk = request.data['value'])
            invDetails = SalesQty.objects.filter(outBound = request.data['value'])
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment;filename="Invoicedownload.pdf"'
        invoice(response , inv , invDetails , typ, request)
        f = open(os.path.join(globalSettings.BASE_DIR, 'media_root/Invoicedownload.pdf'), 'wb')
        f.write(response.content)
        f.close()
        email_subject = "Invoice"
        msgBody = "Hi "  + inv.name + ",\n\n\t\t Find the attachment for the Invoice Details"
        email = request.data['email']
        contactData = tuple(email.split(','))
        print contactData
        msg = EmailMessage(email_subject, msgBody,  to= contactData )
        a = str(f).split('media_root/')[1]
        b = str(a).split(', mode')[0]
        c = str(b).split("'")[0]
        msg.attach_file(os.path.join(globalSettings.MEDIA_ROOT,str(c)))
        msg.send()
        return Response({}, status = status.HTTP_200_OK)

class AverageAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    def get(self, request, format=None):
        inboundobj = Inflow.objects.filter(verified=True)
        receivedInv = list(Sale.objects.filter(status='Received',isInvoice=True).values_list('pk',flat=True))
        invobj = SalesQty.objects.filter(outBound__in=receivedInv)
        crmobj = Contract.objects.filter(status='received')
        inboundtotal = inboundobj.aggregate(inboundtotal=Sum(F('amount') ,output_field=FloatField()))
        inboundtotal = inboundtotal['inboundtotal'] if inboundtotal['inboundtotal'] else 0
        inboundgst = inboundobj.aggregate(inboundgst=Sum(F('gstCollected') ,output_field=FloatField()))
        inboundgst = inboundgst['inboundgst'] if inboundgst['inboundgst'] else 0
        invtotal = invobj.aggregate(invtotal=Sum((F('price')*F('qty')) ,output_field=FloatField()))
        invtotal = invtotal['invtotal'] if invtotal['invtotal'] else 0
        invgst = invobj.aggregate(invgst=Sum(((F('total')*F('tax'))/100) ,output_field=FloatField()))
        invgst = invgst['invgst'] if invgst['invgst'] else 0
        crmtotal = crmobj.aggregate(crmtotal=Sum(F('value') ,output_field=FloatField()))
        crmtotal = crmtotal['crmtotal'] if crmtotal['crmtotal'] else 0
        crmgst = crmobj.aggregate(crmgst=Sum(F('totalTax') ,output_field=FloatField()))
        crmgst = crmgst['crmgst'] if crmgst['crmgst'] else 0
        overallTotal = inboundtotal + invtotal + crmtotal
        overallGst = inboundgst + invgst + crmgst
        return Response({'inboundtotal':inboundtotal,"inboundgst":inboundgst,'invtotal':invtotal,'invgst':invgst,'crmtotal':crmtotal,'crmgst':crmgst,'overallTotal':overallTotal,'overallGst':overallGst}, status = status.HTTP_200_OK)

class InvoiceSheetAPIView(APIView):
    def get(self, request, format=None):
        workbook = Workbook()
        Sheet1 = workbook.active
        hdFont = Font(size=12,bold=True)
        alphaChars = list(string.ascii_uppercase)
        Sheet1.title = 'External Invoices'
        hd = ["Reference Id","Bank", 'Amount','Description','GST']
        hdWidth = [10,10,10,10,10,10,10]
        Sheet1.append(hd)
        inboundobj = Inflow.objects.all()
        for i in inboundobj:
            Sheet1.append([i.referenceID,i.toAcc.title,i.amount,i.description,i.gstCollected])
        if Sheet1.max_column <= len(alphaChars):
            for character in alphaChars[0:Sheet1.max_column]:
                Sheet1.column_dimensions[character].width = 20
        for idx,i in enumerate(hd):
            cl = str(alphaChars[idx])+'1'
            Sheet1[cl].fill = PatternFill(start_color="48dce0", end_color="48dce0", fill_type = "solid")
            Sheet1[cl].font = hdFont
        Sheet2 = workbook.create_sheet('CRM Invoices')
        hd = ["Id","Dated", " Total Value",'Grand Total','Total Tax']
        Sheet2.append(hd)
        crmobj = Contract.objects.all()
        for c in crmobj:
            try:
                dated = str(c.recievedDate).split(' ')[0]
                if dated == "None":
                    dated = ''
            except:
                dated = ''
            crmId = "CRM0" + str(c.pk)
            Sheet2.append([crmId,dated,c.value,c.grandTotal,c.totalTax])
        if Sheet2.max_column <= len(alphaChars):
            for character in alphaChars[0:Sheet2.max_column]:
                Sheet2.column_dimensions[character].width = 20
        for idx,i in enumerate(hd):
            cl = str(alphaChars[idx])+'1'
            Sheet2[cl].fill = PatternFill(start_color="48dce0", end_color="48dce0", fill_type = "solid")
            Sheet2[cl].font = hdFont
        Sheet3 = workbook.create_sheet('Invoices')
        hd = ["Id", " Total",'Total Tax','Grand Total']
        Sheet3.append(hd)
        invobj = SalesQty.objects.all()
        for iv in invobj:
            print iv.total,iv.tax,'98er878euirhjhuyyy'
            if iv.tax == None:
                iv.tax = 0
                tax = ((iv.total*iv.tax)/100)
                tot = iv.total - tax
                invId = "INV0" +str(iv.pk)
                Sheet3.append([invId,tot,tax,iv.total])
            else:
                tax = ((iv.total*iv.tax)/100)
                tot = iv.total - tax
                invId = "INV0" +str(iv.pk)

                Sheet3.append([invId,tot,tax,iv.total])
        if Sheet3.max_column <= len(alphaChars):
            for character in alphaChars[0:Sheet3.max_column]:
                Sheet3.column_dimensions[character].width = 20
        for idx,i in enumerate(hd):
            cl = str(alphaChars[idx])+'1'
            Sheet3[cl].fill = PatternFill(start_color="48dce0", end_color="48dce0", fill_type = "solid")
            Sheet3[cl].font = hdFont
        response = HttpResponse(content=save_virtual_workbook(workbook),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=stockConsumed.xlsx'
        return response

class AmountCalculationAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    def get(self, request, format=None):
        print request.GET,'sssssssss'
        import datetime
        if 'stMonth' in request.GET and 'edMonth' in request.GET:
            stmonthSplitVal = str(request.GET['stMonth']).split('-')
            edmonthSplitVal = str(request.GET['edMonth']).split('-')
            print stmonthSplitVal[0], stmonthSplitVal[1],'ppppppppppppppppppppppppppp'
            stMonth = datetime.date(int(stmonthSplitVal[0]),int(stmonthSplitVal[1]),1)
            edMonth = datetime.date(int(edmonthSplitVal[0]),int(edmonthSplitVal[1]),calendar.monthrange(int(edmonthSplitVal[0]), int(edmonthSplitVal[1]))[1])
            print stMonth,edMonth
            toRet = {'pettyCash':0,'invoicing_inv':0,'exp_inv':0,'vendorAmt':0,'external_inv':0,'crmAmt':0,'claimsAmt':0,'salary':0}
            pettyCashObjs = ProjectPettyExpense.objects.filter(created__range=(stMonth,edMonth)).aggregate(tot=Sum('amount'))
            toRet['pettyCash'] = int(pettyCashObjs['tot']) if pettyCashObjs['tot'] else 0

            invInvPk = list(Sale.objects.filter(isInvoice=True,recDate__range=(stMonth,edMonth)).values_list('pk',flat=True))

            print "invInvPk" , invInvPk
            invInvsObjs = SalesQty.objects.filter(outBound__in=invInvPk).aggregate(tot=Sum((F('price')*F('qty')) ,output_field=FloatField()))

            toRet['invoicing_inv'] = int(invInvsObjs['tot']) if invInvsObjs['tot'] else 0
            print "toRet['invoicing_inv']" , toRet['invoicing_inv']

            expInvPks = list(InvoiceReceived.objects.filter(isInvoice=True,paymentDueDate__range=(stMonth,edMonth)).values_list('pk',flat=True))
            expInvObjs = InvoiceQty.objects.filter(pk__in=expInvPks).aggregate(tot=Sum((F('price')*F('receivedQty')) ,output_field=FloatField()))
            toRet['exp_inv'] = int(expInvObjs['tot']) if expInvObjs['tot'] else 0
            toRet['vendorAmt'] = 0
            externalInvObjs = Inflow.objects.filter(verified=True,dated__range=(stMonth,edMonth)).aggregate(tot=Sum('amount'))
            toRet['external_inv'] = int(externalInvObjs['tot']) if externalInvObjs['tot'] else 0

            crmDataObjs = Contract.objects.filter(recievedDate__range=(stMonth,edMonth)).aggregate(tot=Sum('value'))
            toRet['crmAmt'] = int(crmDataObjs['tot']) if crmDataObjs['tot'] else 0
            claimsInvObjs = Expense.objects.filter(dated__range=(stMonth,edMonth)).aggregate(tot=Sum('amount'))
            toRet['claimsAmt'] = int(claimsInvObjs['tot']) if claimsInvObjs['tot'] else 0

            if (edMonth-stMonth).days>=0:
                r = relativedelta(edMonth, stMonth)
                mths = r.months + (r.years*12)
                q_objects = Q()
                for i in range(mths+1):
                    dt = stMonth + relativedelta(months=i)
                    a = Q(year=dt.year)
                    a.add(Q(month=dt.month), Q.AND)
                    q_objects.add(a, Q.OR)
                payrollObjs = PayrollReport.objects.filter(q_objects).aggregate(tot=Sum('total'))
                toRet['salary'] = payrollObjs['tot'] if payrollObjs['tot'] else 0
            else:
                toRet['salary'] = 0
            return Response(toRet, status = status.HTTP_200_OK)
        else:
            return Response({}, status = status.HTTP_200_OK)

class GstCalculationAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    def get(self, request, format=None):
        print request.GET,'sssssssss'
        if 'stDate' in request.GET and 'edDate' in request.GET:
            try:
                ourGst = str(appSettingsField.objects.get(name='gstIn').value)
            except:
                ourGst = None
            toRet = {'in':0,'out':0,'external':0,'inData':[],'outData':[]}
            stDate = datetime.strptime(request.GET['stDate'],'%Y-%m-%d').date()
            edDate = datetime.strptime(request.GET['edDate'],'%Y-%m-%d').date()
            print stDate,edDate
            vendData = []
            for i in vendData:
                igst = False
                gst = ''
                if i.vendorProfile.service and i.vendorProfile.service.tin:
                    gst = i.vendorProfile.service.tin
                    if ourGst and str(i.vendorProfile.service.tin)[0:2] != ourGst[0:2]:
                        igst = True
                try:
                    toRet['out'] += int(i.gst)
                except:
                    pass
                toRet['outData'].append({'gst':gst,'inv':i.invNo,'amount':i.gst,'date':i.disbursedOn.date(),'igst':igst})
            expData = InvoiceReceived.objects.filter(isInvoice=True,paymentDueDate__range=(stDate,edDate)).order_by('paymentDueDate')
            for i in expData:
                print i.productorder.all().count(),'ccccccccccccc'
                igst = False
                gst = ''
                if i.gstIn:
                    gst = i.gstIn
                    print str(i.gstIn) , ourGst
                    if ourGst and str(i.gstIn)[0:2] != ourGst[0:2]:
                        igst = True
                for j in i.productorder.all():
                    try:
                        amt = int((j.tax*j.total)/100)
                    except:
                        amt = 0
                    toRet['out'] += amt
                    toRet['outData'].append({'gst':gst,'inv':i.invNo,'amount':amt,'date':i.paymentDueDate,'igst':igst})
            externalInv = Inflow.objects.filter(verified=True,dated__range=(stDate,edDate)).order_by('dated')
            externalInvgst = externalInv.aggregate(externalInvgst=Sum(F('gstCollected') ,output_field=FloatField()))
            externalInvgst = int(externalInvgst['externalInvgst']) if externalInvgst['externalInvgst'] else 0
            toRet['external'] = externalInvgst
            toRet['in'] += externalInvgst
            claimsInv = Expense.objects.filter(dated__range=(stDate,edDate)).order_by('dated')
            for i in claimsInv:
                igst = False
                gst = ''
                if i.gstIN:
                    gst = i.gstIN
                    if ourGst and str(i.gstIN)[0:2] != ourGst[0:2]:
                        igst = True
                try:
                    toRet['in'] += int(i.gstVal)
                except:
                    pass
                toRet['inData'].append({'gst':gst,'inv':'claim_{0}'.format(i.pk),'amount':i.gstVal,'date':i.dated,'igst':igst})
            obInv = Sale.objects.filter(isInvoice=True,recDate__range=(stDate,edDate)).order_by('recDate')
            for i in obInv:
                print i.outBoundQty.all().count(),'ccccccccccccc'
                igst = False
                gst = ''
                if i.gstIn:
                    gst = i.gstIn
                    print str(i.gstIn) , ourGst
                    if ourGst and str(i.gstIn)[0:2] != ourGst[0:2]:
                        igst = True
                for j in i.outBoundQty.all():
                    try:
                        amt = int((j.tax*j.total)/100)
                    except:
                        amt = 0
                    toRet['in'] += amt
                    toRet['inData'].append({'gst':gst,'inv':'inv_{0}'.format(j.pk),'amount':amt,'date':i.recDate,'igst':igst})
            crmData = Contract.objects.filter(recievedDate__range=(stDate,edDate)).order_by('recievedDate')
            now = datetime.now()
            for i in crmData:
                igst = False
                gst = ''
                if i.deal.company and i.deal.company.tin:
                    gst = i.deal.company.tin
                    if ourGst and str(i.deal.company.tin)[0:2] != ourGst[0:2]:
                        igst = True
                try:
                    amt = int(i.totalTax)
                except:
                    amt = 0
                toRet['in'] += amt
                try:
                    docID = '{0}{1}{2}'.format(i.deal.pk,now.year,i.pk)
                except:
                    docID = ''
                toRet['inData'].append({'gst':gst,'inv':docID,'amount':amt,'date':i.recievedDate.date(),'igst':igst})
            if 'download' in request.GET:
                print 'downloadinggggggggggg','*******************'
                hdFont = Font(size=12,bold=True)
                alphaChars = list(string.ascii_uppercase)
                workbook = Workbook()
                Sheet1 = workbook.active
                Sheet1.title = 'GST Credit'
                Sheet1.append([])
                Sheet1.append(['GST Credit = {0}'.format(toRet['in'])])
                Sheet1['A2'].font = Font(size=15,bold=True)
                Sheet1.append([])
                hd = ["GST No.", "Invoice No.",'Amount','Date','IGST']
                hdWidth = [20,20,20,20,20]
                Sheet1.append(hd)
                for i in toRet['inData']:
                    data = [i['gst'],i['inv'],i['amount'],i['date'],i['igst']]
                    Sheet1.append(data)
                    for idx,j in enumerate(data):
                        if (len(str(j))+5) > hdWidth[idx]:
                            hdWidth[idx] = len(str(j)) + 15
                for idx,i in enumerate(hd):
                    cl = str(alphaChars[idx])+'4'
                    Sheet1[cl].fill = PatternFill(start_color="48dce0", end_color="48dce0", fill_type = "solid")
                    Sheet1[cl].font = hdFont
                    Sheet1.column_dimensions[str(alphaChars[idx])].width = hdWidth[idx]
                Sheet1.append([])
                Sheet1.append(['External Invoice GST Amount = {0}'.format(toRet['external'])])
                Sheet1['A{0}'.format(Sheet1.max_row)].font = Font(size=13,bold=True)
                Sheet2 = workbook.create_sheet('GST Collected')
                Sheet2.append([])
                Sheet2.append(['GST Collected = {0}'.format(toRet['out'])])
                Sheet2['A2'].font = Font(size=15,bold=True)
                Sheet2.append([])
                hd = ["GST No.", "Invoice No.",'Amount','Date','IGST']
                hdWidth = [20,20,20,20,20]
                Sheet2.append(hd)
                for i in toRet['outData']:
                    data = [i['gst'],i['inv'],i['amount'],i['date'],i['igst']]
                    Sheet2.append(data)
                    for idx,j in enumerate(data):
                        if (len(str(j))+5) > hdWidth[idx]:
                            hdWidth[idx] = len(str(j)) + 15
                for idx,i in enumerate(hd):
                    cl = str(alphaChars[idx])+'4'
                    Sheet2[cl].fill = PatternFill(start_color="48dce0", end_color="48dce0", fill_type = "solid")
                    Sheet2[cl].font = hdFont
                    Sheet2.column_dimensions[str(alphaChars[idx])].width = hdWidth[idx]
                response = HttpResponse(content=save_virtual_workbook(workbook),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename=GSTSummary.xlsx'
                return response
            else:
                return Response(toRet, status = status.HTTP_200_OK)
        else:
            return Response({}, status = status.HTTP_200_OK)

# from datetime import datetime

def month_string_to_number(string):
    m = {
        'January': 1,
        'February': 2,
        'March': 3,
        'April':4,
         'May':5,
         'June':6,
         'July':7,
         'August':8,
         'September':9,
         'October':10,
         'November':11,
         'December':12
        }
    s = string
    try:
        out = m[s]
        return out
    except:
        raise ValueError('Not a month')

from django.db.models import Q
class GetExpensesAPIView(APIView):
    permission_classes = (permissions.IsAuthenticated , isAdmin)
    def get(self, request, format=None):
        if globalSettings.LIMIT_EXPENSE_COUNT:
            monthVal = request.GET['month']
            year = request.GET['year']
            month = month_string_to_number(monthVal)
            import datetime
            import calendar
            date  = datetime.datetime.now()
            fromDate =  date.replace(day = 1, month = int(month), year = int(year))
            fromDate = fromDate - datetime.timedelta(days=1)
            toDate = date.replace(day = calendar.monthrange(int(year), int(month))[1], month = int(month), year = int(year))
            toDate = toDate + datetime.timedelta(days=1)
            print fromDate, toDate,'ppppppppppppppppppppppppppp'
            expenseObj = ExpenseSheet.objects.filter(stage=request.GET['stage'])
            expenseObj = list(expenseObj.filter(created__range=(fromDate,toDate)).values('pk','created','user','notes','project','stage'))
            for i in expenseObj:
                invoice = Expense.objects.filter(sheet = int(i['pk'])).aggregate(totAmount=Sum('amount'),gstAmount=Sum('gstVal'))
                i['totAmount'] = 0
                i['gstAmount'] = 0
                if invoice['totAmount']!=None:
                    i['totAmount'] = invoice['totAmount']
                else:
                    i['totAmount'] = 0
                if invoice['gstAmount']!=None:
                    i['gstAmount'] = invoice['gstAmount']
                else:
                    i['gstAmount'] = 0
                i['total'] = i['totAmount'] +  i['gstAmount']
                try:
                    projectObj = project.objects.get(pk = i['project'])
                    i['project_name'] = projectObj.title
                except:
                    pass

        else:
            expenseObj = ExpenseSheet.objects.filter(stage='submitted')
            user = request.user
            expenseObj = expenseObj.filter(user__designation__reportingTo = user).values('pk','created','user','notes','project','stage')
            for i in expenseObj:
                try:
                    invoice = Expense.objects.filter(sheet = int(i['pk'])).aggregate(totAmount=Sum('amount'),gstAmount=Sum('gstVal'))
                    i['totAmount'] = 0
                    i['gstAmount'] = 0
                    i['totAmount'] = invoice['totAmount']
                    i['gstAmount'] = invoice['gstAmount']
                    i['total'] = invoice['totAmount'] +  invoice['gstAmount']
                except:
                    pass
                try:
                    projectObj = project.objects.get(pk = i['project'])
                    i['project_name'] = projectObj.title
                except:
                    pass
        return Response(expenseObj,status = status.HTTP_200_OK)

class DownloadExpensesAPIView(APIView):
    permission_classes = (permissions.IsAuthenticated , isAdmin)
    def get(self, request, format=None):
        frm = request.GET['frm']
        to = request.GET['to']
        expenseObj = ExpenseSheet.objects.all()
        if 'stage' in request.GET:
            expenseObj = expenseObj.filter(stage=request.GET['stage'])
        if 'user' in request.GET:
            expenseObj = expenseObj.filter(user__id=int(request.GET['user']))
        expenseObj = expenseObj.filter(created__range=(frm,to)).values('pk','created','user','notes','project','stage')
        workbook = Workbook()
        Sheet1 = workbook.active
        hdFont = Font(size=12,bold=True)
        alphaChars = list(string.ascii_uppercase)
        Sheet1.title = 'Reimbursement'
        hd = ["Id","Created", 'User','Title','Stage','Total Amount']
        Sheet1.append(hd)
        for i in expenseObj:
            try:
                projectObj = project.objects.get(pk = i['project'])
                i['project_name']= projectObj.title
            except:
                i['project_name'] = ""
        for idx,i in enumerate(hd):
            cl = str(alphaChars[idx])+'1'
            Sheet1[cl].font = hdFont
            Sheet1[cl].alignment = Alignment(horizontal='center')

        count = 2
        for i in expenseObj:
            user = User.objects.get(pk=int(i['user']))
            name = user.first_name + user.last_name
            expense = [i['pk'],i['created'],name,i['notes'],i['stage']]
            invoice = Expense.objects.filter(sheet = int(i['pk']))
            totAmount = 0
            gstAmount = 0
            for j in invoice:
                totAmount += int(j.amount)
            expense.append(totAmount)
            Sheet1.append(expense)
            for idx,i in enumerate(expense):
                cl = str(alphaChars[idx])+str(count)
                Sheet1[cl].alignment = Alignment(horizontal='center')
            count += 1

        if Sheet1.max_column <= len(alphaChars):
            for character in alphaChars[0:Sheet1.max_column]:
                Sheet1.column_dimensions[character].width = 20

        response = HttpResponse(content=save_virtual_workbook(workbook),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Reimbursement.xlsx'
        return response

class DownloadSheetAPIView(APIView):
    permission_classes = (permissions.IsAuthenticated , isAdmin)
    def get(self, request, format=None):
        expenseSheet = ExpenseSheet.objects.get(pk=request.GET['pkVal'])
        expenseObj =  Expense.objects.filter(sheet = expenseSheet.pk)
        workbook = Workbook()
        Sheet1 = workbook.active
        hdFont = Font(size=12,bold=True)
        alphaChars = list(string.ascii_uppercase)
        Sheet1.title = 'Reimbursement'
        hd = ["Dated", 'Particulars','Description','Amount']
        Sheet1.append(hd)
        for idx,i in enumerate(hd):
            cl = str(alphaChars[idx])+'1'
            Sheet1[cl].font = hdFont
            Sheet1[cl].alignment = Alignment(horizontal='center')

        count = 2
        for i in expenseObj:
            expense = [i.created,i.code,i.description,i.amount]
            Sheet1.append(expense)
            for idx,i in enumerate(expense):
                cl = str(alphaChars[idx])+str(count)
                Sheet1[cl].alignment = Alignment(horizontal='center')
            count += 1


        if Sheet1.max_column <= len(alphaChars):
            for character in alphaChars[0:Sheet1.max_column]:
                Sheet1.column_dimensions[character].width = 20
        try:
            fileName = str(expenseSheet.notes.replace(" ","")) + '.xlsx'
        except:
            fileName = str(expenseSheet.notes) + '.xlsx'
        print fileName
        response = HttpResponse(content=save_virtual_workbook(workbook),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename='+fileName
        return response

from django.db.models import Sum
class GetExpensesTotalAPIView(APIView):
    permission_classes = (permissions.IsAuthenticated , isAdmin)
    def get(self, request, format=None):
        user = User.objects.get(pk = request.GET['user'])
        today = datetime.today()
        expObj = ExpenseSheet.objects.filter(created__year=today.year,created__month=today.month,user=user)
        createdObj = expObj.filter(stage='created')
        createSum = 0
        for i in  createdObj:
            obj = Expense.objects.filter(sheet = i.pk).aggregate(Sum("amount"))
            if obj['amount__sum'] == None:
                obj['amount__sum'] = 0
            createSum+=obj['amount__sum']
        submdObj = expObj.filter(stage='submitted')
        submitSum = 0
        for j in  submdObj:
            objS = Expense.objects.filter(sheet = j.pk).aggregate(Sum("amount"))
            if objS['amount__sum'] == None:
                objS['amount__sum'] = 0
            submitSum+=objS['amount__sum']
        approvedmdObj = expObj.filter(stage='approved')
        approvedSum = 0
        for j in  approvedmdObj:
            objA = Expense.objects.filter(sheet = j.pk).aggregate(Sum("amount"))
            if objA['amount__sum'] == None:
                objA['amount__sum'] = 0
            approvedSum+=objA['amount__sum']
        data = {'createdCount':createdObj.count(),'submittedCount':submdObj.count(),'approvedCount':approvedmdObj.count(),'createSum':createSum,'submitSum':submitSum,'approvedSum':approvedSum}
        return Response(data,status = status.HTTP_200_OK)


from datetime import date
class CreateExpensesAPIView(APIView):
    permission_classes = (permissions.IsAuthenticated, )
    def post(self, request, format=None):
        user = request.user
        month = date.today().month
        year = date.today().year
        todayDate = date.today()
        name = str(todayDate.strftime("%B")) + '-' + str(year)
        expObj, a = ExpenseSheet.objects.get_or_create(user=user,notes = name)
        try:
            expObj.division = request.user.designation.division
            expObj.save()
        except:
            pass
        return Response({},status = status.HTTP_200_OK)


styles=getSampleStyleSheet()
styleN = styles['Normal']
styleH = styles['Heading1']


def addPageNumber(self,doc):
    """
    Add the page number
    """

    print self.pages,'229834723894712-937428374182937492137-89'
    page_num = self._pageNumber
    text = "<font size='8'>Page %s of %s</font>" %( page_num,len(self.pages)+1)
    p = Paragraph(text , styleN)
    p.wrapOn(self , 50*mm , 10*mm)
    p.drawOn(self , 100*mm , 10*mm)


class PageNumCanvas(canvas.Canvas):

    #----------------------------------------------------------------------
    def __init__(self, *args, **kwargs):
        """Constructor"""
        canvas.Canvas.__init__(self, *args, **kwargs)
        self.pages = []

    #----------------------------------------------------------------------
    def showPage(self):
        """
        On a page break, add information to the list
        """
        self.pages.append(dict(self.__dict__))
        self._startPage()

    #----------------------------------------------------------------------
    def save(self):
        """
        Add the page number to each page (page x of y)
        """
        page_count = len(self.pages)

        for page in self.pages:
            self.__dict__.update(page)
            canvas.Canvas.showPage(self)

        canvas.Canvas.save(self)


    #----------------------------------------------------------------------
    def draw_page_number(self, page_count):
        """
        Add the page number
        """

        text = "<font size='14'>Page #%s of %s</font>" % (self._pageNumber , page_count)
        p = Paragraph(text , styleN)
        p.wrapOn(self , 50*mm , 10*mm)
        p.drawOn(self , 100*mm , 10*mm)

from num2words import num2words
def genExpenses(response,expenseObj,request):


    MARGIN_SIZE = 8 * mm
    PAGE_SIZE = A4
    story = []


    pdf_doc = SimpleDocTemplate(response,pagesize = PAGE_SIZE,
        leftMargin = 2*MARGIN_SIZE, rightMargin = 2*MARGIN_SIZE,
        topMargin = 4*MARGIN_SIZE, bottomMargin = 3*MARGIN_SIZE)


    tableHeaderStyle = styles['Normal'].clone('tableHeaderStyle')
    styleN = styles['Normal']
    tableBodyStyle = styles['Normal'].clone('tableBodyStyle')
    tableHeaderStyle.textColor = colors.black;
    tableHeaderStyle.fontSize = 7
    for i in expenseObj:
        name = i.user.first_name+ ' '+i.user.last_name
        heading = str(i.created.strftime("%b")) + ' ' + str(i.created.year)
        dated = Paragraph('<para fontSize=9 align = center><strong>Dated</strong></para>' , tableHeaderStyle)
        invNo = Paragraph('<para fontSize=9 align = center><strong>Inv No</strong></para>' , tableHeaderStyle)
        particular = Paragraph('<para fontSize=9 align = center><strong>Particulars</strong></para>' , tableHeaderStyle)
        invVal = Paragraph('<para fontSize=9 align = center><strong>Invoice Values</strong></para>' , tableHeaderStyle)
        gstAm = Paragraph('<para fontSize=9 align = center><strong>GST Amount</strong></para>' , tableHeaderStyle)
        claimAm = Paragraph('<para fontSize=9 align = center><strong>Claim Amount</strong></para>' , tableHeaderStyle)
        remaRks = Paragraph('<para fontSize=9 align = center><strong>Remarks</strong></para>' , tableHeaderStyle)
        companyName = Paragraph('<para fontSize=20 align = center><strong>Essgi Infotech Private Limited</strong></para>' , tableHeaderStyle)
        city  = Paragraph('<para fontSize=11   align = center spaceb=20><strong>Bangalore</strong></para>' , tableHeaderStyle)
        status  = Paragraph('<para fontSize=11   align = center spaceb=15><strong>Monthly Expenses Claim for Service Charges / Local Conveyance</strong></para>' , tableHeaderStyle)
        claimDate  = Paragraph('<para fontSize=11   align = center spaceb=15 >Claim for the Month of ' + str(heading) + '</para>' , tableHeaderStyle)
        employee = Paragraph('<para fontSize=11   align = left spaceb=15 spacea=15 ><strong>Employees / Claimed by : '+ str(name)+' </strong></para>' , tableHeaderStyle)
        approved = Paragraph('<para fontSize=11   spaceb=25 align=left>Approved by :</para>' , tableHeaderStyle)
        claimby = Paragraph('<para fontSize=11   spaceb=25 align=right>Claimed by : '+str(name)+'</para>' , tableHeaderStyle)
        story.append(companyName)
        story.append(city)
        story.append(status)
        story.append(claimDate)
        story.append(employee)
        daata = [[dated,invNo,particular,invVal,gstAm,claimAm,remaRks]]
        invObj = Expense.objects.filter(sheet = i)
        total = 0
        invTot = 0
        gstTot = 0
        for i in invObj:
            datedVal = Paragraph('<para fontSize=9 align = center>'+ str(i.dated) + '</para>' , tableBodyStyle)
            invNoVal = Paragraph('<para fontSize=9 align = center>'+ str(i.invNo)+'</para>' , tableBodyStyle)
            particularVal = Paragraph('<para fontSize=9 align = center>'+ str(i.code) +'</para>', tableBodyStyle)
            invValVal = Paragraph('<para fontSize=9 align = center>'+ str(i.invoiceAmount)+'</para>' , tableBodyStyle)
            gstAmVal = Paragraph('<para fontSize=9 align = center>'+str(i.gstVal)+'</para>' , tableBodyStyle)
            claimAmVal = Paragraph('<para fontSize=9 align = center>'+str(i.amount)+'</para>' , tableBodyStyle)
            remaRksVal = Paragraph('<para fontSize=9 align = center>'+str(i.description)+'</para>', tableBodyStyle)

            daata+=[[datedVal,invNoVal,particularVal,invValVal,gstAmVal,claimAmVal,remaRksVal]]
            invTot+=i.invoiceAmount
            gstTot += i.gstVal
            total +=i.amount
        wordData =  num2words(round(total), lang='en_IN')
        totalVal = Paragraph('<para  fontSize=11   align = center><strong>Total Claim Amount</strong></para>' , tableHeaderStyle)
        totalinv = Paragraph('<para  fontSize=11   align = right><strong>'+str(invTot)+'</strong></para>' , tableHeaderStyle)
        totalgst = Paragraph('<para  fontSize=11   align = right><strong>'+str(gstTot)+'</strong></para>' , tableHeaderStyle)
        totalam = Paragraph('<para  fontSize=11   align = right><strong>'+str(total)+'</strong></para>' , tableHeaderStyle)
        sumofR = Paragraph('<para fontSize=11   align = center spaceb=20 spacea=20>Sum of Rupees '+str(wordData)+'</para>' , tableHeaderStyle)
        daata+=[['','',totalVal,totalinv,totalgst,totalam,'']]

        productTable = Table(daata)
        catssDetails = TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'BOTTOM'),
            ('ALIGN', (0,0), (-1, -1), 'CENTER'),
            ('ALIGN', (0,0), (-2, -1), 'RIGHT'),
            ('LINEABOVE',(0,0),(-1,-1),1,colors.gray),
            ('LINEABOVE', (-1,-1), (-1, -1), 0.25, colors.black),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
        ])
        productTable.setStyle(catssDetails)
        story.append(productTable)
        story.append(sumofR)
        bottom = [[approved,claimby]]
        botTable = Table(bottom)
        story.append(botTable)
        story.append(PageBreak())


    pdf_doc.build(story,onFirstPage=addPageNumber, onLaterPages=addPageNumber, canvasmaker=PageNumCanvas)

class DownloadExpensesPdfAPIView(APIView):
    permission_classes = (permissions.IsAuthenticated, )
    def get(self, request, format=None):
        monthVal = request.GET['month']
        year = request.GET['year']
        month = month_string_to_number(monthVal)
        import datetime
        import calendar
        date  = datetime.datetime.now()
        fromDate =  date.replace(day = 1, month = int(month), year = int(year))
        fromDate = fromDate - datetime.timedelta(days=1)
        toDate = date.replace(day = calendar.monthrange(int(year), int(month))[1], month = int(month), year = int(year))
        toDate = toDate + datetime.timedelta(days=1)
        expenseObj = ExpenseSheet.objects.filter(stage=request.GET['stage'])
        expenseObj = expenseObj.filter(created__range=(fromDate,toDate))
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment;filename="Grndownload.pdf"'
        genExpenses(response , expenseObj , request)
        return response


class CreateOrderAPI(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny ,)
    def post(self,request , format= None):
        if request.user.id == None:
            if globalSettings.ACCOUNT_KEY != request.data['key']:
                return Response({'error' : 'Unauthenticated user'}, status = status.HTTP_200_OK)
        invoiceData = request.data['invoiceData']
        invoiceQtyData = request.data['invoiceQtyData']
        invObj, a = Sale.objects.get_or_create(poNumber = invoiceData['poNumber'])
        try:
            invObj.division = request.user.designation.division
        except:
            pass
        if 'name' in invoiceData:
            invObj.personName = invoiceData['name']
        if 'mobile' in invoiceData:
            invObj.phone = invoiceData['mobile']
        if 'email' in invoiceData:
            invObj.email = invoiceData['email']
        if 'address' in invoiceData:
            invObj.address = invoiceData['address']
        if 'payment_date' in invoiceData:
            invObj.payDueDate = invoiceData['payment_date']
        invObj.save()
        try:
            SalesQty.objects.filter(outBound = invObj).delete()
        except:
            pass
        data = {'product':invoiceQtyData['productName'], 'price':invoiceQtyData['price'],'qty':invoiceQtyData['qty'],'outBound':invObj,'total':invoiceQtyData['total']}
        outboundQty = SalesQty.objects.create(**data)
        try:
            outboundQty.division = request.user.designation.division
        except:
            pass
        outboundQty.save()
        return Response({ }, status = status.HTTP_200_OK)

class AllExpensesAPI(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny ,)
    def get(self,request , format= None):
        from datetime import date, timedelta
        import calendar
        current_month_start = datetime.today().replace(day=1)
        current_month_end = datetime.today().replace(day=calendar.monthrange(datetime.today().year, datetime.today().month)[1])
        current_month = current_month_start.strftime("%B") + '-' + str(current_month_start.year)
        prev_month_start = (current_month_start - relativedelta(months = 1))
        prev_month_end = prev_month_start.replace(day=calendar.monthrange(prev_month_start.year, prev_month_start.month)[1])
        prev_month = prev_month_start.strftime("%B") + '-' + str(prev_month_start.year)
        first_month_start = (current_month_start - relativedelta(months = 2))
        first_month_end = first_month_start.replace(day=calendar.monthrange(first_month_start.year, first_month_start.month)[1])
        first_month = first_month_start.strftime("%B") + '-' + str(first_month_start.year)
        expenseObj = ExpenseSheet.objects.all()
        first_month_obj = ExpenseSheetSerializer(expenseObj.filter(created__range =(first_month_start,first_month_end)),many=True).data
        prev_month_obj = ExpenseSheetSerializer(expenseObj.filter(created__range= (prev_month_start,prev_month_end)),many=True).data
        current_month_obj = ExpenseSheetSerializer(expenseObj.filter(created__range = (current_month_start,current_month_end) ),many=True).data
        current_month_data = {'month' : current_month,'data':current_month_obj}
        prev_month_data = {'month' : prev_month,'data':prev_month_obj}
        first_month_data = {'month' : first_month,'data':first_month_obj}
        data = {'first_month':first_month_data, 'prev_month': prev_month_data,'current_month' : current_month_data}
        return Response(data,status = status.HTTP_200_OK)

class PurchaseOrderSpreadsheetAPI(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny ,)
    def get(self,request , format= None):
        workbook = Workbook()
        Sheet1 = workbook.active
        hdFont = Font(size=12,bold=True)
        alphaChars = list(string.ascii_uppercase)
        Sheet1.title = 'Purchase Orders'
        hd = ["ID","Quotation Number","Company Name", 'Person Name','Phone','Email','Pincode','Status','Delivery Date','Payment Due Date']
        hdWidth = [10,20,20,20,20,10,10,10,20,20,20,20,20]
        Sheet1.append(hd)
        purOrder = InvoiceReceived.objects.filter(isInvoice=False)
        for i in purOrder:
            Sheet1.append([i.pk,i.quoteNumber,i.name,i.personName,i.phone,i.email,i.pincode,i.status,i.deliveryDate,i.paymentDueDate])
        for idx,i in enumerate(hd):
            cl = str(alphaChars[idx])+'1'
            Sheet1[cl].fill = PatternFill(start_color="48dce0", end_color="48dce0", fill_type = "solid")
            Sheet1[cl].font = hdFont

        if Sheet1.max_column <= len(alphaChars):
            for character in alphaChars[0:Sheet1.max_column]:
                Sheet1.column_dimensions[character].width = 20
        response = HttpResponse(content=save_virtual_workbook(workbook),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=invoices.xlsx'
        return response

class InboundInvoiceExcelAPI(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny ,)
    def get(self,request , format= None):
        workbook = Workbook()
        Sheet1 = workbook.active
        hdFont = Font(size=12,bold=True)
        alphaChars = list(string.ascii_uppercase)
        Sheet1.title = 'Inbound Invoices'
        hd = ["ID","Quotation Number","Company Name", 'Person Name','Phone','Email','Pincode','Status','Delivery Date','Payment Due Date']
        hdWidth = [10,20,20,20,20,10,10,10,20,20,20,20,20]
        Sheet1.append(hd)
        purOrder = InvoiceReceived.objects.filter(isInvoice = True)
        for i in purOrder:
            Sheet1.append([i.pk,i.quoteNumber,i.name,i.personName,i.phone,i.email,i.pincode,i.status,i.deliveryDate,i.paymentDueDate])
        for idx,i in enumerate(hd):
            cl = str(alphaChars[idx])+'1'
            Sheet1[cl].fill = PatternFill(start_color="48dce0", end_color="48dce0", fill_type = "solid")
            Sheet1[cl].font = hdFont

        if Sheet1.max_column <= len(alphaChars):
            for character in alphaChars[0:Sheet1.max_column]:
                Sheet1.column_dimensions[character].width = 20
        response = HttpResponse(content=save_virtual_workbook(workbook),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=InboundInvoices.xlsx'
        return response

class InvoicingSpreadsheetAPI(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny ,)
    def get(self,request , format= None):
        workbook = Workbook()
        Sheet1 = workbook.active
        hdFont = Font(size=12,bold=True)
        alphaChars = list(string.ascii_uppercase)
        Sheet1.title = 'Outbound Invoices'
        outBondObj = Sale.objects.all()
        if 'search__in' in self.request.GET:
            search = self.request.GET['search__in']
            outBondObj = outBondObj.filter(Q(poNumber__icontains=search) | Q(name__icontains=search) | Q(personName__icontains=search)
            | Q(phone__icontains=search) | Q(email__icontains=search) | Q(address__icontains=search) )
        if 'filters' in self.request.GET:
            count = 0
            for i in self.request.GET['filters'].split(','):
                count +=1
                if i == 'Cancelled':
                    data = outBondObj.filter(cancelled=True)
                else:
                    data = outBondObj.filter(status = i).exclude(cancelled=True)
                if count == 1:
                    toReturn =  data
                else:
                    toReturn =  toReturn | data
            if 'sort' in self.request.GET:
                toReturn = toReturn.order_by('total')
            outboundInvoices = toReturn
        else:
            if 'sort' in self.request.GET:
                outBondObj = outBondObj.order_by('total')
            outboundInvoices =  outBondObj

        hd = ["ID","PO Number","Company Name", 'Person Name','Phone','Email','GSTIN','Pincode','Status','Delivery Date','Payment Due Date']
        hdWidth = [10,20,20,20,20,10,10,10,20,20,20,20,20]
        Sheet1.append(hd)
        for obInvoice in outboundInvoices:
            Sheet1.append([obInvoice.pk,obInvoice.poNumber,obInvoice.name,obInvoice.personName,obInvoice.phone,obInvoice.email,obInvoice.gstIn,obInvoice.pincode,obInvoice.status,obInvoice.deliveryDate,obInvoice.payDueDate,])
        for idx,i in enumerate(hd):
            cl = str(alphaChars[idx])+'1'
            Sheet1[cl].fill = PatternFill(start_color="48dce0", end_color="48dce0", fill_type = "solid")
            Sheet1[cl].font = hdFont

        if Sheet1.max_column <= len(alphaChars):
            for character in alphaChars[0:Sheet1.max_column]:
                Sheet1.column_dimensions[character].width = 20
        response = HttpResponse(content=save_virtual_workbook(workbook),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Sales.xlsx'
        return response

class PageNumCanvas(canvas.Canvas):

    #----------------------------------------------------------------------
    def __init__(self, *args, **kwargs):
        """Constructor"""
        canvas.Canvas.__init__(self, *args, **kwargs)
        self.pages = []

    #----------------------------------------------------------------------
    def showPage(self):
        """
        On a page break, add information to the list
        """
        self.pages.append(dict(self.__dict__))
        self._startPage()

    #----------------------------------------------------------------------
    def save(self):
        """
        Add the page number to each page (page x of y)
        """
        page_count = len(self.pages)

        for page in self.pages:
            self.__dict__.update(page)
            self.draw_page_number(page_count)
            canvas.Canvas.showPage(self)
        canvas.Canvas.save(self)

    #----------------------------------------------------------------------
    def draw_page_number(self, page_count):
        """
        Add the page number
        """

        text = "<font size='8'>Page %s of %s</font>" % (self._pageNumber , page_count)
        p = Paragraph(text , styleN)
        p.wrapOn(self , 50*mm , 10*mm)
        p.drawOn(self , 100*mm , 10*mm)


def pettyCashacc(response,pettycash,accnt,month , year,request):
    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(response,pagesize=letter, topMargin=0.5*cm,leftMargin=1*cm,rightMargin=1*cm)
    doc.request = request
    elements = []

    s01 =Paragraph("<para  fontSize=11  >S.No </para>",styles['Heading1'])
    s02 =Paragraph("<para  fontSize=11  >Date </para>",styles['Heading1'])
    s03 =Paragraph("<para  fontSize=11  >Voucher </para>",styles['Heading1'])
    s04 =Paragraph("<para  fontSize=11  >Payee </para>",styles['Heading1'])
    s05 =Paragraph("<para  fontSize=11  >Description</para>",styles['Heading1'])
    s06 =Paragraph("<para  fontSize=11  >Amount</para>",styles['Heading1'])
    s07 =Paragraph("<para  fontSize=11  >Balance</para>",styles['Heading1'])
    s08 =Paragraph("<para  fontSize=11   spaceAfter=5>Financial Year: %s</para>"%(year),styles['Heading1'])
    s09 =Paragraph("<para  fontSize=11   spaceAfter=10>Period: %s </para>"%(month),styles['Heading1'])
    s10 = Paragraph("<para  fontSize=11   align=right>Starting Cash in Hand</para>",styles['Normal'])
    s11 = Paragraph("<para  fontSize=11  >%d</para>"%(accnt.balance),styles['Normal'])
    s12 = Paragraph("<para  fontSize=11  >Voucher Totals</para>",styles['Heading1'])
    s13 = Paragraph("<para  fontSize=11  >Cash on Hand Total</para>",styles['Heading1'])
    s14 = Paragraph("<para  fontSize=11  >Owner: <br/> %s %s </para>"%(accnt.contactPerson.first_name,accnt.contactPerson.last_name),styles['Heading1'])
    s15 = Paragraph("<para  fontSize=11  >Date </para>",styles['Heading1'])
    s16 = Paragraph("<para  fontSize=11  >Difference </para>",styles['Heading1'])
    elements.append(s08)
    elements.append(s09)

    data2 = [[s01,s02,s03,s04,s05,s06,s07],[s10,"","","","",'',s11]]
    for idx,acc in enumerate(pettycash):
        s01 =Paragraph("<para  fontSize=11  >%d </para>"%(idx+1),styles['Normal'])
        s02 =Paragraph("<para  fontSize=11  >%s </para>"%(acc.created.date()),styles['Normal'])
        s03 =Paragraph("<para  fontSize=11  > %d </para>"%(acc.pk),styles['Normal'])
        s04 =Paragraph("<para  fontSize=11  >%s %s </para>"%(accnt.contactPerson.first_name,accnt.contactPerson.last_name),styles['Normal'])
        s05 =Paragraph("<para  fontSize=11  >%s</para>"%(acc.description),styles['Normal'])
        s06 =Paragraph("<para  fontSize=11  >%d</para>"%(acc.creditAmount),styles['Normal'])
        s07 =Paragraph("<para  fontSize=11  >%d</para>"%(acc.balance),styles['Normal'])
        data2 +=  [[s01,s02,s03,s04,s05,s06,s07]]
    totalAcct = data2 + [['','','',s12,'',''],['','','',s13,'',''],[s14,s15,'',s16,'','']]
    data = Table(totalAcct )
    data.setStyle(TableStyle(
    [('TEXTFONT', (0, 0), (-1, -1), 'Times-Bold'),
    ('TEXTCOLOR',(0,0),(-1,-1),black),
    ('ALIGN',(0,0),(-1,-1),'RIGHT'),
    ('VALIGN',(0,0),(-1,-1),'TOP'),
    ('BOX',(0,0),(-1,-1),0.25,colors.black),
    ('SPAN',(0,1),(5,1)),
    ('INNERGRID', (0,0), (-1,-1), 0.25, colors.black),
    ]))
    data._argW[0] = 2*cm
    data._argW[1] = 2.5*cm
    data._argW[2] = 2*cm
    data._argW[3] = 3*cm
    data._argW[4] = 4*cm
    data._argW[5] = 2*cm
    data._argW[6] = 2*cm
    elements.append(data)
    doc.build(elements,canvasmaker=PageNumCanvas)

class PettyCashAccpdfAPI(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny ,)
    def get(self,request , format= None):
        month = request.GET['month']
        temp_year = request.GET['year']
        if 'January' in month or 'February'  in month or 'March'  in month:
            year = temp_year.split('-')[1]
            if 'Quarter 4' in month:
                startDate = date.today().replace(month=1, day=1, year=int(year))
                endDate = date.today().replace(month=3, day=31, year=int(year))
            else:
                month_no = month_string_to_number(month)
                startDate = date.today().replace(month=int(month_no), day=1, year=int(year))
                endDate = startDate.replace(day=calendar.monthrange(startDate.year, startDate.month)[1])
        elif 'FY' in month:
            year = temp_year.split('-')[0]
            year1 = temp_year.split('-')[1]
            startDate = date.today().replace(month=4, day=1, year=int(year))
            endDate = date.today().replace(month=3, day=31, year=int(year1))
        else:
            year = temp_year.split('-')[0]
            if 'Quarter 1' in month :
                startDate = date.today().replace(month=4, day=1, year=int(year))
                endDate = date.today().replace(month=6, day=31, year=int(year))
            elif 'Quarter 2' in month :
                startDate = date.today().replace(month=7, day=1, year=int(year))
                endDate = date.today().replace(month=9, day=30, year=int(year))
            elif 'Quarter 3' in month :
                startDate = date.today().replace(month=10, day=1, year=int(year))
                endDate = date.today().replace(month=12, day=31, year=int(year))
            else:
                month_no = month_string_to_number(month)
                startDate = date.today().replace(month=int(month_no), day=1, year=int(year))
                endDate = startDate.replace(day=calendar.monthrange(startDate.year, startDate.month)[1])
        accnt = Account.objects.get(pk = request.GET['id'])
        pettycash = ProjectPettyExpense.objects.filter(account = request.GET['id'])
        pettycash = pettycash.filter(created__range = (startDate,endDate))
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment;filename="PettyCashAccounts.pdf"'
        pettyCashacc(response,pettycash,accnt,month,temp_year,request)
        return response

class AccountsSpreadSheetAPI(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny ,)
    def get(self,request , format= None):
        if 'only' in request.GET:
            workbook = Workbook()
            Sheet1 = workbook.active
            hdFont = Font(size=12,bold=True)
            alphaChars = list(string.ascii_uppercase)
            Sheet1.title = request.GET['only']
            hd1 = []
            hd2 = ["","","", '','Closing Balance ']
            Sheet1.append(hd2)
            hd = ["S.No","ID","Name", 'Account Number','Balance']
            hdWidth = [5,5,40,40,30]
            Sheet1.append(hd)
            allAcc = Account.objects.all()
            if self.request.GET['only'] == 'pettyCash':
                allAcc = allAcc.filter(personal=True)
            if self.request.GET['only'] == 'accounts':
                allAcc = allAcc.filter(group=None).exclude(personal=True)
            total = 0
            for idx,ac in enumerate(allAcc):
                if ac.personal:
                    account = 'Petty Cash'
                else:
                    account = ac.number
                total +=ac.balance
                Sheet1.append([idx+1,ac.pk,ac.title,account,ac.balance])
            Sheet1.append(['','','','Grand Total',total])
            for idx,i in enumerate(hd):
                cl = str(alphaChars[idx])+'1'
                Sheet1[cl].font = hdFont
            if Sheet1.max_column <= len(alphaChars):
                for character in alphaChars[0:Sheet1.max_column]:
                    Sheet1.column_dimensions[str(alphaChars[idx])].width = hdWidth[idx]
                    Sheet1.column_dimensions[character].width = 20
        response = HttpResponse(content=save_virtual_workbook(workbook),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Accounts.xlsx'
        return response

class GetOutBoundCountAPI(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny ,)
    def get(self,request , format= None):
        toReturn = []
        outstanding = Sale.objects.filter(status = 'SaleOrder', isInvoice = True).exclude(cancelled = True).count()
        val = {'name' : 'Outstanding' , 'is_selected' : False, 'count' : outstanding ,'title' : 'Pending Payments'}
        toReturn.append(val)
        saleOrder = Sale.objects.filter(status = 'SaleOrder', isInvoice = False).exclude(cancelled = True).count()
        val = {'name' : 'SaleOrder' , 'is_selected' : False, 'count' : saleOrder ,'title' : 'Sales Order'}
        toReturn.append(val)
        received = Sale.objects.filter(status = 'Received', isInvoice = True).exclude(cancelled = True).count()
        val = {'name' : 'Received' , 'is_selected' : False, 'count' : received ,'title' : 'Paid Invoices'}
        toReturn.append(val)
        overdue = Sale.objects.filter(status = 'Overdue', isInvoice = True).exclude(cancelled = True).count()
        val = {'name' : 'Overdue' , 'is_selected' : False, 'count' : overdue, 'title' : 'Deloayed Invoices'}
        toReturn.append(val)
        cancelled = Sale.objects.filter(cancelled = True, isInvoice = True).count()
        val = {'name' : 'Cancelled' , 'is_selected' : False, 'count' : cancelled , 'title' : 'Cancelled Invoices'}
        toReturn.append(val)
        return Response({'data':toReturn }, status = status.HTTP_200_OK)


def pettyCashjournal(response,request, month , year):
    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(response,pagesize=letter, topMargin=0.5*cm,leftMargin=1*cm,rightMargin=1*cm)
    doc.request = request
    elements = []
    month = month
    temp_year = year
    if 'January' in month or 'February'  in month or 'March'  in month:
        year = temp_year.split('-')[1]
        if 'Quarter 4' in month:
            startDate = date.today().replace(month=1, day=1, year=int(year))
            endDate = date.today().replace(month=3, day=31, year=int(year))
        else:
            month_no = month_string_to_number(month)
            startDate = date.today().replace(month=int(month_no), day=1, year=int(year))
            endDate = startDate.replace(day=calendar.monthrange(startDate.year, startDate.month)[1])
    elif 'FY' in month:
        year = temp_year.split('-')[0]
        year1 = temp_year.split('-')[1]
        startDate = date.today().replace(month=4, day=1, year=int(year))
        endDate = date.today().replace(month=3, day=31, year=int(year1))
    else:
        year = temp_year.split('-')[0]
        if 'Quarter 1' in month :
            startDate = date.today().replace(month=4, day=1, year=int(year))
            endDate = date.today().replace(month=6, day=31, year=int(year))
        elif 'Quarter 2' in month :
            startDate = date.today().replace(month=7, day=1, year=int(year))
            endDate = date.today().replace(month=9, day=30, year=int(year))
        elif 'Quarter 3' in month :
            startDate = date.today().replace(month=10, day=1, year=int(year))
            endDate = date.today().replace(month=12, day=31, year=int(year))
        else:
            month_no = month_string_to_number(month)
            startDate = date.today().replace(month=int(month_no), day=1, year=int(year))
            endDate = startDate.replace(day=calendar.monthrange(startDate.year, startDate.month)[1])

    s01 =Paragraph("<para  fontSize=11   align=center>Date </para>",styles['Heading1'])
    s02 =Paragraph("<para  fontSize=11   align=center>Account Title and Explanantions </para>",styles['Heading1'])
    s03 =Paragraph("<para  fontSize=11   align=center>Ref </para>",styles['Heading1'])
    s04 =Paragraph("<para  fontSize=11   align=center>Amount(Rs.) </para>",styles['Heading1'])
    s05 = Paragraph("<para  fontSize=11  >Debit </para>",styles['Heading1'])
    s06 = Paragraph("<para  fontSize=11  >Credit </para>",styles['Heading1'])
    s07 = Paragraph("<para  fontSize=11   align=center >2016 </para>",styles['Heading1'])
    s08 = Paragraph("<para  fontSize=11  >Aug. </para>",styles['Heading1'])
    s09 = Paragraph("<para  fontSize=11  >5 </para>",styles['Heading1'])
    amTab = Table([[s04],[s05,s06]])
    s10 = Paragraph("<para  fontSize=11   >Cash </para>",styles['Normal'])
    s11 = Paragraph("<para  fontSize=11   align=left >80000 </para>",styles['Normal'])
    s12 = Paragraph("<para  fontSize=11   >Furniture </para>",styles['Normal'])
    s15 = Paragraph("<para  fontSize=11   align=right >100,000</para>",styles['Normal'])
    s16 = Paragraph("<para  fontSize=11   >(Reinvest by owner in the business)</para>",styles['Normal'])
    s17 = Paragraph("<para  fontSize=11   ></para>",styles['Normal'])
    data2 = [[s01,s02,amTab]]
    ids = Transaction.objects.filter(dated__range = (startDate , endDate)).exclude(externalRecord = True).order_by('dated').values('groupId').distinct()
    for id in ids:
        res = Transaction.objects.filter(groupId = id['groupId']).exclude(externalRecord = True)
        allData = []
        amountCol = []
        for r in res:
            if r.fromAcc is not None:
                s14 = Paragraph("<para  fontSize=11   >" + r.fromAcc.title+ "</para>",styles['Normal'])
                s13 = Paragraph("<para  fontSize=11   align=left > " + str(round(r.debit ,2))+ " </para>",styles['Normal'])
                val = [s13]
            if r.toAcc is not None:
                s14 = Paragraph("<para  fontSize=11   >" + r.toAcc.title+ "</para>",styles['Normal'])
                s13 = Paragraph("<para  fontSize=11   align=left > " + str(round(r.credit,2)) + " </para>",styles['Normal'])
                val = [s17,s13]
            allData.append(s14)
            amountCol.append(val)
        narration =  Paragraph("<para  fontSize=11   >( " +res.last().narration+ " )</para>",styles['Normal'])
        allData.append(narration)
        amountCol.append('')
        dated = str(res.last().dated.strftime("%b")) + '   ' + str(res.last().dated.day)
        datTab = Table([[dated]])
        dataTab = Table(amountCol)
        dataTab.setStyle(TableStyle(
        [('TEXTFONT', (0, 0), (-1, -1), 'Times-Bold'),
        ('TEXTCOLOR',(0,0),(-1,-1),black),
        ('ALIGN',(0,0),(-1,-1),'CENTER'),
        ('TOPPADDING',(0,0),(-1,-1),-3)
        ]))
        data2.append([datTab,allData,dataTab])

    data = Table(data2)
    data.setStyle(TableStyle(
    [('TEXTFONT', (0, 0), (-1, -1), 'Times-Bold'),
    ('TEXTCOLOR',(0,0),(-1,-1),black),
    ('VALIGN',(0,0),(-1,-1),'BOTTOM'),
    ('BOX',(0,0),(-1,-1),0.25,colors.black),
    ('INNERGRID', (0,0), (-1,-1), 0.25, colors.black),
    ]))
    elements.append(data)
    doc.build(elements,canvasmaker=PageNumCanvas)

class PettyCashJournalAPI(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny ,)
    def get(self,request , format= None):
        month = request.GET['month']
        year = request.GET['year']
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment;filename="PettycashJournal.pdf"'
        pettyCashjournal(response,request,month,year)
        return response

from dateutil.relativedelta import relativedelta
class GettTransactionsAPI(APIView):
    renderer_classes = (JSONRenderer,)
    def get(self, request, format=None):
        month = 4
        year = 2020
        start = datetime(year, month, 1)
        end = start.replace(day=calendar.monthrange(start.year, start.month)[1])
        toReturn = []
        months = []
        for i in range(start.day, end.day+1):
            dated = start.replace(day=i).date()
            val = {'date' : dated,'month':dated.month,'year':dated.year}
            dataVal = 2
            obj = []
            for j in range(1,dataVal+1):
                if j%2 == 0:
                    objVal = {'description':'Description','status':'credit','amount':'1000'}
                else:
                    objVal = {'description':'Description','status':'debit','amount':'1000'}
                obj.append(objVal)
            val['data'] = obj
            val['description'] = 'Description'
            toReturn.append(val)
        return Response({'data' : toReturn},status=status.HTTP_200_OK)


class GetAccountTransactionsAPI(APIView):
    renderer_classes = (JSONRenderer,)
    def get(self, request, format=None):
        month = int(request.GET['month'])
        year = int(request.GET['year'])
        account = int(request.GET['account'])
        accountObj = Account.objects.get(pk = account)
        start = datetime(year, month, 1)
        end = start.replace(day=calendar.monthrange(start.year, start.month)[1])
        transcationObj = Transaction.objects.filter(created__range = (start,end))
        transcationObj = transcationObj.filter(Q(fromAcc=accountObj)|Q(toAcc=accountObj))
        toReturn = TransactionSerializer(transcationObj,many=True).data
        return Response({'data' : toReturn},status=status.HTTP_200_OK)


def month_string_to_number(string):
    m = {
        'jan': 1,
        'feb': 2,
        'mar': 3,
        'apr':4,
         'may':5,
         'jun':6,
         'jul':7,
         'aug':8,
         'sep':9,
         'oct':10,
         'nov':11,
         'dec':12
        }
    s = string.strip()[:3].lower()

    try:
        out = m[s]
        return out
    except:
        raise ValueError('Not a month')

class GetPettyCashDataAPI(APIView):
    renderer_classes = (JSONRenderer,)
    def get(self, request, format=None):
        month = request.GET['month']
        temp_year = request.GET['year']
        if 'January' in month or 'February'  in month or 'March'  in month:
            year = temp_year.split('-')[1]
            if 'Quarter 4' in month:
                startDate = date.today().replace(month=1, day=1, year=int(year))
                endDate = date.today().replace(month=3, day=31, year=int(year))
            else:
                month_no = month_string_to_number(month)
                startDate = date.today().replace(month=int(month_no), day=1, year=int(year))
                endDate = startDate.replace(day=calendar.monthrange(startDate.year, startDate.month)[1])
        elif 'FY' in month:
            year = temp_year.split('-')[0]
            year1 = temp_year.split('-')[1]
            startDate = date.today().replace(month=4, day=1, year=int(year))
            endDate = date.today().replace(month=3, day=31, year=int(year1))
        else:
            year = temp_year.split('-')[0]
            if 'Quarter 1' in month :
                startDate = date.today().replace(month=4, day=1, year=int(year))
                endDate = date.today().replace(month=6, day=31, year=int(year))
            elif 'Quarter 2' in month :
                startDate = date.today().replace(month=7, day=1, year=int(year))
                endDate = date.today().replace(month=9, day=30, year=int(year))
            elif 'Quarter 3' in month :
                startDate = date.today().replace(month=10, day=1, year=int(year))
                endDate = date.today().replace(month=12, day=31, year=int(year))
            else:
                month_no = month_string_to_number(month)
                startDate = date.today().replace(month=int(month_no), day=1, year=int(year))
                endDate = startDate.replace(day=calendar.monthrange(startDate.year, startDate.month)[1])
        account = int(request.GET['account'])
        accountObj = Account.objects.get(pk = account)
        pettyObj = ProjectPettyExpense.objects.filter(account = accountObj)
        pettyObj = pettyObj.filter(created__range = (startDate,endDate))
        data = PettyCashSerializer(pettyObj, many=True).data
        return Response(data,status=status.HTTP_200_OK)



class GetAccountAPI(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny,)
    def get(self,request , format= None):
        divsn = request.user.designation.division
        accountObj = Account.objects.filter(group='profit and loss', division = divsn)
        if len(accountObj)<=0:
            dataVal = [
            {'title':'Markup Income','heading':'income','group':'profit and loss','is_lock':False},
            {'title':'Retail Sales','heading':'income','group':'profit and loss','is_lock':False},
            {'title':'Service','heading':'income','group':'profit and loss','is_lock':False},
            {'title':'Cost of Goods Sold','heading':'sale','group':'profit and loss','is_lock':False},
            {'title':'Payroll Expenses','heading':'expense','group':'profit and loss','is_lock':False},
            {'title':'Bank Service Charges','heading':'expense','group':'profit and loss','is_lock':False},
            {'title':'Insurance','heading':'expense','group':'profit and loss','is_lock':False},
            {'title':'Interest Expense','heading':'expense','group':'profit and loss','is_lock':False},
            {'title':'Job Expenses','heading':'expense','group':'profit and loss','is_lock':False},
            {'title':'Mileage Reimbursement','heading':'expense','group':'profit and loss','is_lock':False},
            {'title':'Professional Fees','heading':'expense','group':'profit and loss','is_lock':False},
            {'title':'Rent','heading':'expense','group':'profit and loss','is_lock':False},
            {'title':'Repairs','heading':'expense','group':'profit and loss','is_lock':False},
            {'title':'Tools and Misc.Equipment','heading':'expense','group':'profit and loss','is_lock':False},
            {'title':'Uncategorized Expenses','heading':'expense','group':'profit and loss','is_lock':False},
            {'title':'Utilities','heading':'expense','group':'profit and loss','is_lock':False},
            {'title':'Misc Income','heading':'otherIncome','group':'profit and loss','is_lock':False},
            ]
            for i in dataVal:
                obj = Account.objects.create(**i)
                obj.division = divsn
                obj.is_lock = True
                obj.save()
            accountObj = Account.objects.filter(group='profit and loss' , division = divsn)
        incomeObj = accountObj.filter(heading = 'income')
        expenseObj = accountObj.filter(heading = 'expense')
        salesObj = accountObj.filter(heading = 'sale')
        otherIncomeObj = accountObj.filter(heading = 'otherIncome')
        income = AccountLiteSerializer(incomeObj,many=True).data
        expense = AccountLiteSerializer(expenseObj,many=True).data
        sale = AccountLiteSerializer(salesObj,many=True).data
        otherIncome = AccountLiteSerializer(otherIncomeObj,many=True).data
        data = {'income':income,'expense':expense,'sale':sale,'otherIncome':otherIncome}
        return Response(data, status = status.HTTP_200_OK)


class CashFlowAPI(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny,)
    def get(self,request , format= None):
        periodYr = request.GET['FY'].split('-')[1]
        divsn = request.user.designation.division
        accountObj = Account.objects.filter(group='cashflow', division = divsn)
        operatingObj = accountObj.filter(heading = 'operating')
        opoeratinTotal = operatingObj.aggregate(Sum('balance'))
        if opoeratinTotal['balance__sum'] == None:
            operatingT = 0
        else:
            operatingT = opoeratinTotal['balance__sum']
        investingObj = accountObj.filter(heading = 'investing')
        investingTotal = investingObj.aggregate(Sum('balance'))
        if investingTotal['balance__sum'] == None:
            investingT = 0
        else:
            investingT = investingTotal['balance__sum']
        finanaceObj = accountObj.filter(heading = 'financing')
        financeTotal = finanaceObj.aggregate(Sum('balance'))
        if financeTotal['balance__sum'] == None:
            financeT = 0
        else:
            financeT = financeTotal['balance__sum']
        allCfoas = AccountLiteSerializer(operatingObj,many=True).data
        allcfias = AccountLiteSerializer(investingObj,many=True).data
        allcffas = AccountLiteSerializer(finanaceObj,many=True).data
        cfoa= {'title':"Cash Flow from Operating Activities ",'cfoaList':allCfoas,'totalCofa':operatingT}
        cfia =  {'title':"Cash Flow from Investing Activities ",'cfiaList':allcfias,'totalCfia':investingT}
        cffa =  {'title':"Cash Flow from Financing Activities ",'cffaList':allcffas,'totalCffa':financeT}
        try:
            division = request.user.designation.division.name
        except:
            division = ''
        CashflowData = {'CFOA':cfoa,"CFIA":cfia,"CFFA":cffa,'Netincome':'1000','startYear':'1200','endYear':'2,200' , 'compName' : division , "period" : 'March 31 ,'+periodYr}
        return Response(CashflowData, status = status.HTTP_200_OK)
class BalanceSheetAPI(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny,)
    def get(self,request , format= None):
        print type(request.GET['period']),'09909090'
        if 'period' in request.GET:
            period = request.GET['period']
        else:
            period = date.today()
        divsn = request.user.designation.division
        commonRatios = [
        {'heading':" Debt Ratio ",'subheading':' (Total Liabilities/Total Assets)','value':'0.56'},
        {'heading':" Current Ratio ",'subheading':' (Current Assets /Current Liabilities)','value':'1.06'},
        {'heading':" Working Capital ",'subheading':' (Current Assets - Current Liabilities)','value':'669'},
        {'heading':" Assets-to-Equity Ratio",'subheading':" (Total Assets / Owner's Equity)",'value':'2.27'},
        {'heading':" Debt-to-Equity Ratio",'subheading':"  (Total Liabilities / Owner's Equity)",'value':'2.27'},
        ]
        accountObj = Account.objects.filter(group='balanceSheet' , division = divsn)
        currentAssetObj = accountObj.filter(heading = 'currentAsset')
        currentAssetTotal = currentAssetObj.aggregate(tot=Sum('balance'))
        fixedAssetObj = accountObj.filter(heading = 'fixedAsset')
        fixedAssetTotal = fixedAssetObj.aggregate(tot=Sum('balance'))
        otherAssetObj = accountObj.filter(heading = 'otherAsset')
        otherAssetTotal = otherAssetObj.aggregate(tot=Sum('balance'))
        currentLiabilityObj = accountObj.filter(heading = 'currentLiability')
        currentLiabilityTotal = currentLiabilityObj.aggregate(tot=Sum('balance'))
        longTermLiabilityObj = accountObj.filter(heading = 'longTermLiability')
        longTermLiabilityTotal = longTermLiabilityObj.aggregate(tot=Sum('balance'))
        ownerEquity = accountObj.filter(heading = 'ownerEquity')
        ownerEquityTotal = ownerEquity.aggregate(tot=Sum('balance'))
        currentObj = AccountLiteSerializer(currentAssetObj,many=True).data
        if currentAssetTotal['tot'] == None:
            currenttotal = 0
        else:
            currenttotal = currentAssetTotal['tot']
        fixedObj = AccountLiteSerializer(fixedAssetObj,many=True).data
        if fixedAssetTotal['tot'] == None:
            fixedtotal = 0
        else:
            fixedtotal = fixedAssetTotal['tot']
        otherObj = AccountLiteSerializer(otherAssetObj,many=True).data
        if otherAssetTotal['tot'] == None:
            othertotal = 0
        else:
            othertotal = otherAssetTotal['tot']
        currentLiabObj = AccountLiteSerializer(currentLiabilityObj,many=True).data
        if currentLiabilityTotal['tot'] == None:
            currentLiabilitytotal = 0
        else:
            currentLiabilitytotal = currentLiabilityTotal['tot']
        longTermLiabObj = AccountLiteSerializer(longTermLiabilityObj,many=True).data
        if longTermLiabilityTotal['tot'] == None:
            longTermLiabilitytotal = 0
        else:
            longTermLiabilitytotal = longTermLiabilityTotal['tot']
        ownerEquityObj = AccountLiteSerializer(ownerEquity,many=True).data
        if ownerEquityTotal['tot'] == None:
            ownertotal = 0
        else:
            ownertotal = ownerEquityTotal['tot']
        assetTotal = currenttotal  + fixedtotal  + othertotal
        liabilityTotal = currentLiabilitytotal + longTermLiabilitytotal  + ownertotal
        assets = {'cAssets':currentObj,'fixAssets':fixedObj,'otherAssets':otherObj,'CAtotal':currenttotal,'FATotal':fixedtotal,'OAssets':'','totalAssets':assetTotal}
        liabilities = {
        'Cliabilites':currentLiabObj,
        'ltLiabilities' :longTermLiabObj,
        'ownerEquity':ownerEquityObj,
        'totalLia':currentLiabilitytotal,
        'totalCL':longTermLiabilitytotal,
        'totalLTL':ownertotal,
        'totalEquity':liabilityTotal
        }
        try:
            division = request.user.designation.division.name
        except:
            division = ''
        balanceSheetData = {'compName' : division , "period" : period,'assets':assets,'liabilities':liabilities,'commonRatios':commonRatios}

        return Response(balanceSheetData, status = status.HTTP_200_OK)
from datetime import datetime
from dateutil.relativedelta import relativedelta
class ProfitLossAPI(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny,)
    def get(self,request , format= None):
        divsn = request.user.designation.division
        if 'Year' in request.GET:
            strtYr = request.GET['Year'].split('-')[0]
            endYr = request.GET['Year'].split('-')[1]
        FYstart = None
        FYlast= None
        if 'period' in request.GET:
            FYfirst = strtYr+' '+request.GET['period']
            FYstart = datetime.strptime(FYfirst, "%Y %B").date()
            FYlast = FYstart + relativedelta(months=+11)
        accountObj = Account.objects.filter(group='profit and loss' , division = divsn)
        allCfoas = accountObj.filter(heading = 'income')
        incomeTotal = allCfoas.aggregate(Sum('balance'))
        if incomeTotal['balance__sum'] == None:
            totalIncome = 0
        else:
            totalIncome = incomeTotal['balance__sum']
        expenseObj = accountObj.filter(heading = 'expense')
        expenseTotal= expenseObj.aggregate(Sum('balance'))
        if expenseTotal['balance__sum'] == None:
            totalExpense = 0
        else:
            totalExpense = expenseTotal['balance__sum']
        salesObj = accountObj.filter(heading = 'sale')
        saleTotal= salesObj.aggregate(Sum('balance'))
        if saleTotal['balance__sum'] == None:
            totalSale = 0
        else:
            totalSale = saleTotal['balance__sum']
        otherIncomeObj = accountObj.filter(heading = 'otherIncome')
        otherIncome = otherIncomeObj.aggregate(Sum('balance'))
        if otherIncome['balance__sum'] == None:
            totalOtherIncome = 0
        else:
            totalOtherIncome = otherIncome['balance__sum']
        income = AccountLiteSerializer(allCfoas,many=True).data
        expenses = AccountLiteSerializer(expenseObj,many=True).data
        COGS = AccountLiteSerializer(salesObj,many=True).data
        otherIncome = AccountLiteSerializer(otherIncomeObj,many=True).data
        grossProfit = totalIncome - totalSale
        NetordIncome = grossProfit - totalExpense
        netIncome = NetordIncome
        try:
            division = request.user.designation.division.name
        except:
            division = ''
        proLossData = {'compName' : division , "FYStart" : FYstart, "FYend":FYlast,
        'totalIncome':totalIncome,'totalCOGS':totalSale,'totalExp':totalExpense,'NetordIncome':NetordIncome,
        'totalOtherInc':totalOtherIncome,'NetOtherInc':NetordIncome,'netIncome':netIncome,'expenses':expenses,'otherIncome':otherIncome,'COGS':COGS,'income':income,'grossProfit':grossProfit
        }

        return Response(proLossData, status = status.HTTP_200_OK)

def CashFlow(response , request):
    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(response,pagesize=letter, topMargin=1*cm,leftMargin=1*cm,rightMargin=1*cm)
    doc.request = request
    elements = []
    try:
        division = request.user.designation.division.name
    except:
        division = ''
    divsn = request.user.designation.division
    accountObj = Account.objects.filter(group = 'cashFlow', division = divsn)
    l01 = Paragraph("<para align=center   fontSize=11   spaceAfter=5>%s</para>"%(division),styles['Heading1'])
    l02 = Paragraph("<para align=center  fontSize=11   spaceAfter=5>Statement of Cash Flows</para>",styles['Heading1'])
    l03 = Paragraph("<para align=center  fontSize=11   spaceAfter=5>For the Year Ended March 31 ,2020</para>",styles['Heading1'])
    elements.append(l01)
    elements.append(l02)
    elements.append(l03)
    l04 = Paragraph("<para align=left  fontSize=11   spaceAfter=5>Cash Flow from Operating Activities </para>",styles['Heading1'])
    data = [[l04]]
    ooperatingTotal = 0
    investingTotal = 0
    financingTotal = 0
    for i in accountObj.filter(heading  = 'operating'):
        ooperatingTotal +=i.balance
        l05 = Paragraph("<para align=left  fontSize=11   lindent=10 >%s </para>"%(i.title),styles['Normal'])
        l06 = Paragraph("<para align=left  fontSize=11   align=right  >%s</para>"%(i.balance),styles['Normal'])
        data.append([l05,l06])
    l015 = Paragraph("<para align=left  fontSize=11   lindent=15>Cash provided (used) in operating activities </para>",styles['Normal'])
    l016 = Paragraph("<para align=left  fontSize=11   align=right >%s </para>"%(ooperatingTotal),styles['Normal'])
    data.append([l015,l016])
    l017 = Paragraph("<para align=left  fontSize=11   >Cash Flow Investing Activities </para>",styles['Heading1'])
    data.append([l017])
    for i in accountObj.filter(heading  = 'investing'):
        investingTotal +=i.balance
        l05 = Paragraph("<para align=left  fontSize=11   lindent=10 >%s </para>"%(i.title),styles['Normal'])
        l06 = Paragraph("<para align=left  fontSize=11   align=right  >%s</para>"%(i.balance),styles['Normal'])
        data.append([l05,l06])
    l022 = Paragraph("<para align=left  fontSize=11   lindent=15 >Cash provided (used) by Investing activities </para>",styles['Normal'])
    l023 = Paragraph("<para align=left  fontSize=11    align=right>%s </para>"%(investingTotal),styles['Normal'])
    data.append([l022,l023])
    l024 = Paragraph("<para align=left  fontSize=11    >Cash Flow from Financing Activities </para>",styles['Heading1'])
    data.append([l024])
    for i in accountObj.filter(heading  = 'financing'):
        financingTotal +=i.balance
        l05 = Paragraph("<para align=left  fontSize=11   lindent=10 >%s </para>"%(i.title),styles['Normal'])
        l06 = Paragraph("<para align=left  fontSize=11   align=right  >%s</para>"%(i.balance),styles['Normal'])
        data.append([l05,l06])
    l031 = Paragraph("<para align=left  fontSize=11   lindent=15  >Cash provided (used) by financing activites</para>",styles['Normal'])
    l032 = Paragraph("<para align=left  fontSize=11    align=right>%s </para>"%(financingTotal),styles['Normal'])
    data.append([l031,l032])
    l033 = Paragraph("<para align=left  fontSize=11    >Net increase in cash </para>",styles['Normal'])
    l034 = Paragraph("<para align=left  fontSize=11    align=right>1000 </para>",styles['Normal'])
    l035 = Paragraph("<para align=left  fontSize=11    >Cash at the beginning of the year </para>",styles['Normal'])
    l036 = Paragraph("<para align=left  fontSize=11   align=right >1200 </para>",styles['Normal'])
    l037 = Paragraph("<para align=left  fontSize=11    >Cash at the end of the year </para>",styles['Normal'])
    l038 = Paragraph("<para align=left  fontSize=11    align=right>2,200 </para>",styles['Normal'])
    flowData = Table(data,rowHeights=0.75*cm)
    flowData.setStyle(TableStyle([('TEXTFONT', (0, 0), (-1, -1), 'Times-Bold'),
    ('TEXTCOLOR',(0,0),(-1,-1),black),
    ('VALIGN',(0,0),(-1,-1),'TOP'),
    ('ALIGN',(0,0),(-1,-1),'RIGHT'),
    ]))
    elements.append(flowData)
    doc.build(elements,canvasmaker=PageNumCanvas)


class CashFlowPdfAPI(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny,)
    def get(self,request , format= None):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment;filename="cashflow.pdf"'
        CashFlow(response , request)
        return response
def ProLoss(response , request):
    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(response,pagesize=letter, topMargin=1*cm,leftMargin=1*cm,rightMargin=1*cm)
    doc.request = request
    elements = []
    pnl01 = Paragraph("<para align=center  fontSize=11    spaceAfter=5>%s</para>"%(request.user.designation.division.name),styles['Heading1'])
    pnl02 = Paragraph("<para align=center fontSize=11    spaceAfter=5>Profit and Loss</para>",styles['Heading1'])
    pnl03 = Paragraph("<para align=center fontSize=11    spaceAfter=5></para>",styles['Heading1'])
    pnlDat = Paragraph("<para align=right fontSize=11    spaceAfter=5></para>",styles['Heading1'])
    elements.append(pnl01)
    elements.append(pnl02)
    elements.append(pnl03)
    elements.append(pnlDat)
    divsn = request.user.designation.division
    accountObj = Account.objects.filter(group = "profit and loss", division = divsn)
    pnl04 = Paragraph("<para align=left fontSize=11    spaceAfter=5>Ordinary Income/Expense </para>",styles['Heading1'])
    pnl05 = Paragraph("<para align=left fontSize=11    lindent=15 ><b>Income</b> </para>",styles['Normal'])
    data = [[pnl04],[pnl05]]
    totalIncome = 0
    totalSale = 0
    grossProfit = 0
    totalExpense = 0
    totalOtherIncome = 0
    for i in accountObj.filter(heading = 'income'):
        val = []
        pnl06 = Paragraph("<para align=left fontSize=11 lindent=30 ><b>%s</b></para>"%(i.title),styles['Normal'])
        pnl061 = Paragraph("<para align=right fontSize=11  >%s</para>"%(i.balance),styles['Normal'])
        totalIncome +=i.balance
        val = [pnl06 , pnl061]
        data.append(val)
    pnl13 = Paragraph("<para align=left fontSize=11     lindent=15><b>Total Income</b></para>",styles['Normal'])
    pnl14 = Paragraph("<para align=right fontSize=11     >%s</para>"%(totalIncome),styles['Normal'])
    data.append([pnl13,pnl14])
    pnl15 = Paragraph("<para align=left fontSize=11     lindent=15><b>Cost of Goods Sold</b></para>",styles['Normal'])
    data.append([pnl15])
    for i in accountObj.filter(heading = 'sale'):
        val = []
        pnl07 = Paragraph("<para align=left fontSize=11 lindent=30 ><b>%s</b></para>"%(i.title),styles['Normal'])
        pnl071 = Paragraph("<para align=right fontSize=11  >%s</para>"%(i.balance),styles['Normal'])
        totalSale +=i.balance
        val = [pnl07 , pnl071]
        data.append(val)
    pnl17 = Paragraph("<para align=left fontSize=11     lindent=15><b>Total COGS</b></para>",styles['Normal'])
    pnl18 = Paragraph("<para align=right fontSize=11     >%s</para>"%(totalSale),styles['Normal'])
    data.append([pnl17,pnl18])
    grossProfit = totalIncome - totalSale
    pnl19 = Paragraph("<para align=left fontSize=11     ><b>Gross Profit</b></para>",styles['Normal'])
    pnl20 = Paragraph("<para align=right fontSize=11     >%s</para>"%(grossProfit),styles['Normal'])
    data.append([pnl19,pnl20])
    pnl21 = Paragraph("<para align=left fontSize=11     lindent=15><b>Expense</b></para>",styles['Normal'])
    data.append([pnl21])
    for i in accountObj.filter(heading = 'expense'):
        val = []
        pnl08 = Paragraph("<para align=left fontSize=11 lindent=30 ><b>%s</b></para>"%(i.title),styles['Normal'])
        pnl081 = Paragraph("<para align=right fontSize=11  >%s</para>"%(i.balance),styles['Normal'])
        totalExpense +=i.balance
        val = [pnl08 , pnl081]
        data.append(val)
    pnl50 = Paragraph("<para align=left fontSize=11    lindent=15><b>Total Expense</b></para>",styles['Normal'])
    pnl51 = Paragraph("<para align=right fontSize=11    >%s</para>"%(totalExpense),styles['Normal'])
    data.append([pnl50,pnl51])
    grossProfit = grossProfit - totalExpense
    pnl52 = Paragraph("<para align=left fontSize=11    ><b>Net Ordinary Income</b></para>",styles['Normal'])
    pnl501 = Paragraph("<para align=right fontSize=11    >%s</para>"%(grossProfit),styles['Normal'])
    data.append([pnl52,pnl501])
    pnl53 = Paragraph("<para align=left fontSize=11    ><b>Other Income/Expense</b></para>",styles['Normal'])
    data.append([pnl53])
    pnl54 = Paragraph("<para align=left fontSize=11    lindent=7><b>Other Income </b> </para>",styles['Normal'])
    data.append([pnl54])
    for i in accountObj.filter(heading = 'otherIncome'):
        val = []
        pnl08 = Paragraph("<para align=left fontSize=11 lindent=30 ><b>%s</b></para>"%(i.title),styles['Normal'])
        pnl081 = Paragraph("<para align=right fontSize=11  >%s</para>"%(i.balance),styles['Normal'])
        totalOtherIncome +=i.balance
        val = [pnl08 , pnl081]
        data.append(val)
    pnl59 = Paragraph("<para align=left fontSize=11    lindent=15>  <b>Total Other Income  </b></para>",styles['Normal'])
    pnl58 = Paragraph("<para align=right fontSize=11    > %s  </para>"%(totalOtherIncome),styles['Normal'])
    data.append([pnl59,pnl58])
    grossProfit = grossProfit - totalOtherIncome
    pnl61 = Paragraph("<para align=left fontSize=11    >  <b>Net Other Income </b> </para>",styles['Normal'])
    pnl62 = Paragraph("<para align=right fontSize=11    >%s  </para>"%(grossProfit),styles['Normal'])
    data.append([pnl61,pnl62])
    pnl63 = Paragraph("<para align=left fontSize=11    ><b>  Net Income  </b></para>",styles['Normal'])
    pnl64 = Paragraph("<para align=right fontSize=11    ><b>%s </b> </para>"%(grossProfit),styles['Normal'])
    data.append([pnl63,pnl64])
    proLossData = Table(data,rowHeights=0.7*cm)
    proLossData.setStyle(TableStyle([('TEXTFONT', (0, 0), (-1, -1), 'Times-Bold'),
    ('TEXTCOLOR',(0,0),(-1,-1),black),
    ('VALIGN',(0,0),(-1,-1),'TOP'),
    ('ALIGN',(0,0),(-1,-1),'RIGHT'),

    ]))
    elements.append(proLossData)
    doc.build(elements,canvasmaker=PageNumCanvas)
class ProfitLossPdfAPI(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny,)
    def get(self,request , format= None):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment;filename="ProfitandLoss.pdf"'
        ProLoss(response , request)
        return response


def BalSheet(response , request):
    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(response,pagesize=letter, topMargin=1*cm,leftMargin=1*cm,rightMargin=1*cm)
    doc.request = request
    elements = []
    divsn = request.user.designation.division
    accountObj = Account.objects.filter(group="balanceSheet", division = divsn)
    currentAssetTotal = 0
    fixedAssetTotal = 0
    otherAssetTotal = 0
    totalAssets = 0
    currentLiabilityTotal = 0
    longTermLiabilityTotal = 0
    ownerEquityTotal = 0
    totalLiablility = 0
    bs01 = Paragraph("<para align=left  fontSize=11    spaceAfter=5>%s</para>"%(request.user.designation.division.name),styles['Heading1'])
    bs02 = Paragraph("<para align=right fontSize=11    spaceAfter=5>Balance Sheet <br/>Date :2020-04-16</para>",styles['Heading1'])

    bs04 = Paragraph("<para align=left fontSize=11><b> Assets</b></para>",styles['Normal'])
    bs05 = Paragraph("<para align=right fontSize=11><b> </b></para>",styles['Normal'])
    bs06 = Paragraph("<para align=left fontSize=11><b> Current Assets</b></para>",styles['Normal'])
    bs10 = Paragraph("<para align=right fontSize=11 ></para>",styles['Normal'])
    data = [[bs01,bs02],[bs10],[bs04,bs05],[bs06]]
    for i in accountObj.filter(heading  = 'currentAsset'):
        currentAssetTotal+=i.balance
        bs07 = Paragraph("<para align=left fontSize=11 lindent=15>%s</para>"%(i.title),styles['Normal'])
        bs08 = Paragraph("<para align=right fontSize=11 >%s</para>"%(i.balance),styles['Normal'])
        data.append([bs07,bs08])
    bs14 = Paragraph("<para align=right fontSize=11 lindent=15>Total current assets</para>",styles['Normal'])
    bs15 = Paragraph("<para align=right fontSize=11 lindent=15>%s</para>"%(currentAssetTotal),styles['Normal'])
    data.append([bs14,bs15])
    bs16 = Paragraph("<para align=left fontSize=11><b> Fixed (Long-Term) Assets</b></para>",styles['Normal'])
    data.append([bs16])
    for i in accountObj.filter(heading  = 'fixedAsset'):
        fixedAssetTotal+=i.balance
        bs07 = Paragraph("<para align=left fontSize=11 lindent=15>%s</para>"%(i.title),styles['Normal'])
        bs08 = Paragraph("<para align=right fontSize=11 >%s</para>"%(i.balance),styles['Normal'])
        data.append([bs07,bs08])
    bs24 = Paragraph("<para align=right fontSize=11 lindent=15>Total Fixed Assets</para>",styles['Normal'])
    bs25 = Paragraph("<para align=right fontSize=11 lindent=15>%s</para>"%(fixedAssetTotal),styles['Normal'])
    data.append([bs24,bs25])
    bs26 = Paragraph("<para align=left fontSize=11 ><b>Other assets</b></para>",styles['Normal'])
    data.append([bs26])
    for i in accountObj.filter(heading  = 'otherAsset'):
        otherAssetTotal+=i.balance
        bs07 = Paragraph("<para align=left fontSize=11 lindent=15>%s</para>"%(i.title),styles['Normal'])
        bs08 = Paragraph("<para align=right fontSize=11 >%s</para>"%(i.balance),styles['Normal'])
        data.append([bs07,bs08])
    bs240 = Paragraph("<para align=right fontSize=11 lindent=15>Total Other Assets</para>",styles['Normal'])
    bs250 = Paragraph("<para align=right fontSize=11 lindent=15>%s</para>"%(fixedAssetTotal),styles['Normal'])
    data.append([bs240,bs250])
    totalAssets = currentAssetTotal + fixedAssetTotal + otherAssetTotal
    bs570 =  Paragraph("<para align=left fontSize=11 ><b>Total Assets</b></para>",styles['Normal'])
    bs580 =  Paragraph("<para align=right fontSize=11 ><b>%s</b></para>"%(totalAssets),styles['Normal'])
    data.append([bs570, bs580])
    bs34 = Paragraph("<para align=left fontSize=11 ><b>Current Liabilities</b></para>",styles['Normal'])
    data.append([bs34])
    for i in accountObj.filter(heading  = 'currentLiability'):
        currentLiabilityTotal+=i.balance
        bs07 = Paragraph("<para align=left fontSize=11 lindent=15>%s</para>"%(i.title),styles['Normal'])
        bs08 = Paragraph("<para align=right fontSize=11 >%s</para>"%(i.balance),styles['Normal'])
        data.append([bs07,bs08])
    bs43 = Paragraph("<para align=right fontSize=11 >Total current liabilities</para>",styles['Normal'])
    bs44 = Paragraph("<para align=right fontSize=11 >%s</para>"%(currentLiabilityTotal),styles['Normal'])
    data.append([bs43,bs44])
    bs45  = Paragraph("<para align=left fontSize=11 ><b>Long-Term  Liabilities</b></para>",styles['Normal'])
    data.append([bs45])
    for i in accountObj.filter(heading  = 'longTermLiability'):
        longTermLiabilityTotal+=i.balance
        bs07 = Paragraph("<para align=left fontSize=11 lindent=15>%s</para>"%(i.title),styles['Normal'])
        bs08 = Paragraph("<para align=right fontSize=11 >%s</para>"%(i.balance),styles['Normal'])
        data.append([bs07,bs08])
    bs49 =  Paragraph("<para align=right fontSize=11 >Total long-term liabilities</para>",styles['Normal'])
    bs50 =  Paragraph("<para align=right fontSize=11 >%s</para>"%(longTermLiabilityTotal),styles['Normal'])
    data.append([bs49,bs50])
    bs51 =  Paragraph("<para align=left fontSize=11 ><b>Owner's Equity</b></para>",styles['Normal'])
    data.append([bs51])
    for i in accountObj.filter(heading  = 'ownerEquity'):
        ownerEquityTotal+=i.balance
        bs07 = Paragraph("<para align=left fontSize=11 lindent=15>%s</para>"%(i.title),styles['Normal'])
        bs08 = Paragraph("<para align=right fontSize=11 >%s</para>"%(i.balance),styles['Normal'])
        data.append([bs07,bs08])
    bs55 =  Paragraph("<para align=right fontSize=11 >Total owner's equity</para>",styles['Normal'])
    bs56 =  Paragraph("<para align=right fontSize=11 >%s</para>"%(ownerEquityTotal),styles['Normal'])
    data.append([bs55,bs56])
    totalLiablility = currentLiabilityTotal + longTermLiabilityTotal + ownerEquityTotal
    bs57 =  Paragraph("<para align=left fontSize=11 ><b>Total Liabilities and Owner's Equity</b></para>",styles['Normal'])
    bs58 =  Paragraph("<para align=right fontSize=11 ><b>%s</b></para>"%(totalLiablility),styles['Normal'])
    data.append([bs57,bs58])
    bs59 =  Paragraph("<para align=left fontSize=11 lindent=15>Retained earnings </para>",styles['Normal'])
    bs60 =  Paragraph("<para align=right fontSize=11 >0</para>",styles['Normal'])
    bs61 =  Paragraph("<para align=right fontSize=11 >0</para>",styles['Normal'])
    balSheetData = Table(data,rowHeights=0.7*cm)
    balSheetData.setStyle(TableStyle([('TEXTFONT', (0, 0), (-1, -1), 'Times-Bold'),
    ('TEXTCOLOR',(0,0),(-1,-1),black),
    ('VALIGN',(0,0),(-1,-1),'TOP'),
    ('ALIGN',(0,0),(-1,-1),'RIGHT'),

    ]))
    elements.append(balSheetData)
    doc.build(elements,canvasmaker=PageNumCanvas)
class BalanceSheetPdfAPI(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny,)
    def get(self,request , format= None):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment;filename="BalanceSheet.pdf"'
        BalSheet(response , request)
        return response


class GSTR3BAPI(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny,)
    def get(self,request , format= None):
        unObj = Unit.objects.get(pk = request.GET['id'])
        print int(request.GET["FY"].split('-')[0])+1
        outwardSupplies = [
        {"heading":'(a) Outward taxable supplies (other than zero reated,nil rated and exempted)','TotalTaxablevalue':'','IntegratedTax':'','CentralTax':'','State/UTTax':"",'Cess':""},
        {"heading":'(b) Outward taxable supplies (zero rated)','TotalTaxablevalue':'','IntegratedTax':'','CentralTax':'','State/UTTax':"",'Cess':""},
        {"heading":'(c) Other outward supplies(Nil rated,exempted)','TotalTaxablevalue':'','IntegratedTax':'','CentralTax':'','State/UTTax':"",'Cess':""},
        {"heading":'(d) Inward supplies (liable to reverse charge)','TotalTaxablevalue':'','IntegratedTax':'','CentralTax':'','State/UTTax':"",'Cess':""},
        {"heading":'(e) Non-GST outward supplies','TotalTaxablevalue':'','IntegratedTax':'','CentralTax':'','State/UTTax':"",'Cess':""}
        ]
        interStateSupplies = [
        {"heading":"Supplies made to Unregistered Persons","PlaceofSupply(State/UT)":"","TotalTaxablevalue":"","AmountofIntegratedTax":""},
        {"heading":"Supplies made to Composition Taxable Persons","PlaceofSupply(State/UT)":"","TotalTaxablevalue":"","AmountofIntegratedTax":""},
        {"heading":" Supplies made to UIN holders","PlaceofSupply(State/UT)":"","TotalTaxablevalue":"","AmountofIntegratedTax":""},

        ]
        itcA = [
            {"Details":"(1) Imports of goods",'IntegratedTax':" ","CentralTax":"","State/UTTax":"","Cess":""},
            {"Details":"(2) Imports of services",'IntegratedTax':" ","CentralTax":"","State/UTTax":"","Cess":""},
            {"Details":"(3) Inward supplies liable to reverse charge (other than 1 & 2 above)",'IntegratedTax':" ","CentralTax":"","State/UTTax":"","Cess":""},
            {"Details":"(4) Inward supplies from ISD",'IntegratedTax':" ","CentralTax":"","State/UTTax":"","Cess":""},
            {"Details":"(5) All other ITC",'IntegratedTax':" ","CentralTax":"","State/UTTax":"","Cess":""}
        ]
        itcB = [
            {"Details":"(1) As per rules 42 &43 of CGST Rules",'IntegratedTax':" ","CentralTax":"","State/UTTax":"","Cess":""},
            {"Details":"(2) Others",'IntegratedTax':" ","CentralTax":"","State/UTTax":"","Cess":""},
        ]
        itcC = [

        ]
        itcD = [

        {"Details":"(1) As per section 17(5)",'IntegratedTax':" ","CentralTax":"","State/UTTax":"","Cess":""},
        {"Details":"(2) Others",'IntegratedTax':" ","CentralTax":"","State/UTTax":"","Cess":""},

        ]
        nonGST = [
        {"Natureofsupplies":"From a supplier under composition scheme,Exempt and Nil rated" ,  "Inter-Statesupplies":"", "Inter-Intra-Statesupplies":""   },
        {"Natureofsupplies":"supply" ,  "Inter-Statesupplies":"", "Inter-Intra-Statesupplies":""   },
        {"Natureofsupplies":"Non GST supply" ,  "Inter-Statesupplies":"", "Inter-Intra-Statesupplies":""   },
        ]
        payTax = [
        {"Description":"IntegratedTax","Taxpayable":"","IntegratedTax":"","CentralTax":"","State/UTTax":"","cess":"","TaxpaidTDS/TCS":"","Tax/cess paid in cash":"","Interest":"","Late Fee":""},
        {"Description":"CentralTax","Taxpayable":"","IntegratedTax":"","CentralTax":"","State/UTTax":"","cess":"","TaxpaidTDS/TCS":"","Tax/cess paid in cash":"","Interest":"","Late Fee":""},
        {"Description":"State/UTTax","Taxpayable":"","IntegratedTax":"","CentralTax":"","State/UTTax":"","cess":"","TaxpaidTDS/TCS":"","Tax/cess paid in cash":"","Interest":"","Late Fee":""},
        {"Description":"Cess","Taxpayable":"","IntegratedTax":"","CentralTax":"","State/UTTax":"","cess":"","TaxpaidTDS/TCS":"","Tax/cess paid in cash":"","Interest":"","Late Fee":""}
        ]
        tdstcs = [
                {"Details":"TDS",'IntegratedTax':" ","CentralTax":"","State/UTTax":""},
                {"Details":"TCS",'IntegratedTax':" ","CentralTax":"","State/UTTax":""}

        ]
        gstrData = {'Point1':outwardSupplies,"Point2":interStateSupplies,"point4":nonGST,"Point5":payTax,'Point6':tdstcs,"point31":itcA,"point32":itcB,"point33":itcC,"point34":itcD,"Divivsion":unObj.division.name,"unitName":unObj.name,"gstin":unObj.gstin,"Month":request.GET['Quarter'],"FinY": str(int(request.GET["FY"].split('-')[0])+1)}
        return Response(gstrData, status = status.HTTP_200_OK)

class GetCashFlowAPI(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny,)
    def get(self,request , format= None):
        divsn = request.user.designation.division
        accountObj = Account.objects.filter(group='cashflow', division = divsn)
        if len(accountObj)<=0:
            dataVal = [{'title':'Net inome','heading':'operating','group':'cashflow','is_lock':False},
            {'title':' Increase in accounts receivable ','heading':'operating','group':'cashflow','is_lock':False},
            {'title':' Decrease in inventory','heading':'operating','group':'cashflow','is_lock':False},
            {'title':'Decrease in accounts payable','heading':'operating','group':'cashflow','is_lock':False},
            {'title':'Capital expenditures','heading':'investing','group':'cashflow','is_lock':False},
            {'title':'Proceeds from sale of property','heading':'investing','group':'cashflow','is_lock':False},
            {'title':'Borrowings of long-term debt','heading':'financing','group':'cashflow','is_lock':False},
            {'title':'Cash dividends','heading':'financing','group':'cashflow','is_lock':False},
            {'title':'Purchase of treasury stock','heading':'financing','group':'cashflow','is_lock':False},
            ]
            for i in dataVal:
                obj = Account.objects.create(**i)
                obj.is_lock = True
                obj.division = divsn
                obj.save()
            accountObj = Account.objects.filter(group='cashflow' , division = divsn)
        operatingObj = accountObj.filter(heading = 'operating')
        investingObj = accountObj.filter(heading = 'investing')
        finanaceObj = accountObj.filter(heading = 'financing')
        operating = AccountLiteSerializer(operatingObj,many=True).data
        investing = AccountLiteSerializer(investingObj,many=True).data
        finance = AccountLiteSerializer(finanaceObj,many=True).data
        data = {'operating':operating,'investing':investing,'finance':finance}
        return Response(data, status = status.HTTP_200_OK)


class GetBalanceSheetAPI(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny,)
    def get(self,request , format= None):
        divsn = request.user.designation.division
        accountObj = Account.objects.filter(group='balanceSheet', division = divsn)
        if len(accountObj)<=0:
            dataVal = [{'title':'Cash','heading':'currentAsset','group':'balanceSheet','is_lock':True,'division':divsn},
            {'title':'Accounts receivable','heading':'currentAsset','group':'balanceSheet','is_lock':True,'division':divsn},
            {'title':'Inventory','heading':'currentAsset','group':'balanceSheet','is_lock':True,'division':divsn},
            {'title':'Prepaid expenses','heading':'currentAsset','group':'balanceSheet','is_lock':True,'division':divsn},
            {'title':'Short-term investments','heading':'currentAsset','group':'balanceSheet','is_lock':True,'division':divsn},
            {'title':'Long-term investments','heading':'fixedAsset','group':'balanceSheet','is_lock':True,'division':divsn},
            {'title':'Property,plant,and equipment','heading':'fixedAsset','group':'balanceSheet','is_lock':True,'division':divsn},
            {'title':'(Less accumulated depreciation)','heading':'fixedAsset','group':'balanceSheet','is_lock':True,'division':divsn},
            {'title':'Intangible assets','heading':'fixedAsset','group':'balanceSheet','is_lock':True,'division':divsn},
            {'title':'Deferred income tax','heading':'otherAsset','group':'balanceSheet','is_lock':True,'division':divsn},
            {'title':'Other','heading':'otherAsset','group':'balanceSheet','is_lock':True,'division':divsn},
            {'title':' Accounts payable','heading':'currentLiability','group':'balanceSheet','is_lock':True,'division':divsn},
            {'title':'Short-term loans','heading':'currentLiability','group':'balanceSheet','is_lock':True},
            {'title':'Income taxes payable','heading':'currentLiability','group':'balanceSheet','is_lock':True,'division':divsn},
            {'title':' Accrued salaries and wages','heading':'currentLiability','group':'balanceSheet','is_lock':True,'division':divsn},
            {'title':' Unearned revenue','heading':'currentLiability','group':'balanceSheet','is_lock':True,'division':divsn},
            {'title':' Current portion of long-term debt','heading':'currentLiability','group':'balanceSheet','is_lock':True,'division':divsn},
            {'title':'Long-term debt','heading':'longTermLiability','group':'balanceSheet','is_lock':True,'division':divsn},
            {'title':'Deferred income tax','heading':'longTermLiability','group':'balanceSheet','is_lock':True,'division':divsn},
            {'title':'Deferred income tax','heading':'longTermLiability','group':'balanceSheet','is_lock':True,'division':divsn},
            {'title':' Owners investment ','heading':'ownerEquity','group':'balanceSheet','is_lock':True,'division':divsn},
            {'title':'Retained earnings','heading':'ownerEquity','group':'balanceSheet','is_lock':True,'division':divsn},
            {'title':'Other','heading':'ownerEquity','group':'balanceSheet','is_lock':True,'division':divsn},
            ]
            for i in dataVal:
                Account.objects.create(**i)
            accountObj = Account.objects.filter(group='balanceSheet' , division = divsn)
        currentAssetObj = accountObj.filter(heading = 'currentAsset')
        currentAssetTotal = currentAssetObj.aggregate(tot=Sum('balance'))
        fixedAssetObj = accountObj.filter(heading = 'fixedAsset')
        fixedAssetTotal = fixedAssetObj.aggregate(tot=Sum('balance'))
        otherAssetObj = accountObj.filter(heading = 'otherAsset')
        otherAssetTotal = otherAssetObj.aggregate(tot=Sum('balance'))
        currentLiabilityObj = accountObj.filter(heading = 'currentLiability')
        currentLiabilityTotal = currentLiabilityObj.aggregate(tot=Sum('balance'))
        longTermLiabilityObj = accountObj.filter(heading = 'longTermLiability')
        longTermLiabilityTotal = longTermLiabilityObj.aggregate(tot=Sum('balance'))
        ownerEquity = accountObj.filter(heading = 'ownerEquity')
        ownerEquityTotal = ownerEquity.aggregate(tot=Sum('balance'))
        currentObj = AccountLiteSerializer(currentAssetObj,many=True).data
        current = {'data':currentObj}
        if currentAssetTotal['tot'] == None:
            current['total'] = 0
        else:
            current['total'] = currentAssetTotal['tot']
        fixedObj = AccountLiteSerializer(fixedAssetObj,many=True).data
        fixed = {'data':fixedObj}
        if fixedAssetTotal['tot'] == None:
            fixed['total'] = 0
        else:
            fixed['total'] = fixedAssetTotal['tot']
        otherObj = AccountLiteSerializer(otherAssetObj,many=True).data
        other = {'data':otherObj}
        if otherAssetTotal['tot'] == None:
            other['total'] = 0
        else:
            other['total'] = otherAssetTotal['tot']
        currentLiabObj = AccountLiteSerializer(currentLiabilityObj,many=True).data
        currentLiability = {'data':currentLiabObj}
        if currentLiabilityTotal['tot'] == None:
            currentLiability['total'] = 0
        else:
            currentLiability['total'] = currentLiabilityTotal['tot']
        longTermLiabObj = AccountLiteSerializer(longTermLiabilityObj,many=True).data
        longTermLiability = {'data':longTermLiabObj}
        if longTermLiabilityTotal['tot'] == None:
            longTermLiability['total'] = 0
        else:
            longTermLiability['total'] = longTermLiabilityTotal['tot']
        ownerEquityObj = AccountLiteSerializer(ownerEquity,many=True).data
        owner = {'data':ownerEquityObj}
        if ownerEquityTotal['tot'] == None:
            owner['total'] = 0
        else:
            owner['total'] = ownerEquityTotal['tot']
        assetTotal = current['total']  + fixed['total']  + other['total']
        liabilityTotal = currentLiability['total'] + longTermLiability['total']  + owner['total']
        data = {'current':current,'fixed':fixed,'other':other,'currentLiability':currentLiability,'longTermLiability':longTermLiability,'owner':owner,'assetTotal':assetTotal,'liabilityTotal':liabilityTotal}
        return Response(data, status = status.HTTP_200_OK)


def gstrPDF(response , request,untObj,fYr,qur):
    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(response,pagesize=letter, topMargin=1*cm,leftMargin=1*cm,rightMargin=1*cm)
    doc.request = request
    elements = []
    bs01 = Paragraph("<para align=Center  fontSize=11    spaceAfter=5>FORM GSTR-3B </para>",styles['Heading1'])
    elements.append(bs01)
    gstR03 = Paragraph("<para align=left fontSize=11><b> Year</b></para>",styles['Normal'])
    gstR04 = Paragraph("<para align=left fontSize=11><b> Month</b></para>",styles['Normal'])
    gstR05 = Paragraph("<para align=left  fontSize=11><b>GSTIN </b></para>",styles['Normal'])
    gstR06 = Paragraph("<para align=left  fontSize=11><b>Legal name of the registered person </b> </para>",styles['Normal'])
    gstR07 = Paragraph("<para align=left  fontSize=11> <b></b></para>",styles['Normal'])
    gstR08 = Paragraph("<para align=left  fontSize=11> </para>",styles['Normal'])
    gstR09 = Paragraph("<para align=left  fontSize=11> </para>",styles['Normal'])
    gstR010 = Paragraph("<para align=left  fontSize=11> </para>",styles['Normal'])
    gstR011 = Paragraph("<para align=left  fontSize=11> </para>",styles['Normal'])
    gstR012 = Paragraph("<para align=left  fontSize=11> </para>",styles['Normal'])
    gstR013 = Paragraph("<para align=left  fontSize=11> </para>",styles['Normal'])
    gstR014 = Paragraph("<para align=left  fontSize=11> </para>",styles['Normal'])
    gstR015 = Paragraph("<para align=left  fontSize=11> </para>",styles['Normal'])
    gstR016 = Paragraph("<para align=left  fontSize=11> </para>",styles['Normal'])
    gstR017 = Paragraph("<para align=left  fontSize=11> </para>",styles['Normal'])
    gstR018 = Paragraph("<para align=left  fontSize=11> </para>",styles['Normal'])
    gstR019 = Paragraph("<para align=left  fontSize=11> </para>",styles['Normal'])
    gstR020 = Paragraph("<para align=left  fontSize=11> </para>",styles['Normal'])
    gstR021 = Paragraph("<para align=left  fontSize=11> </para>",styles['Normal'])
    gstR022 = Paragraph("<para align=left  fontSize=11> </para>",styles['Normal'])
    gstR023 = Paragraph("<para align=left  fontSize=11> 1. </para>",styles['Normal'])
    gstR024 = Paragraph("<para align=left  fontSize=11>2. </para>",styles['Normal'])
    gstRuni = Paragraph("<para align=left  fontSize=11><b> %s  </b></para>"%(untObj.name),styles['Normal'])
    gstR0Yr = Paragraph("<para align=left  fontSize=11><b>Year</b> </para>",styles['Normal'])
    gstR0Mn = Paragraph("<para align=left  fontSize=11><b>Month</b> </para>",styles['Normal'])
    gstR0Yr1 = Paragraph("<para align=left  fontSize=11> </para>",styles['Normal'])
    gstR0Yr2 = Paragraph("<para align=left  fontSize=11> </para>",styles['Normal'])
    gstR0Yr3 = Paragraph("<para align=left  fontSize=11> </para>",styles['Normal'])
    gstR0Yr4 = Paragraph("<para align=left  fontSize=11> </para>",styles['Normal'])
    gstR0Mn4 = Paragraph("<para align=left  fontSize=11><b> %s</b></para>"%(qur),styles['Normal'])

    gstrYr= [gstR0Yr]
    for yr in fYr:
        gstR0Yr1 = Paragraph("<para align=left  fontSize=11><b> %s </b></para>"%(yr),styles['Normal'])
        gstrYr += [gstR0Yr1]
    gstYrMnData = [gstrYr,[gstR0Mn,gstR0Mn4]]
    gstYrMnTab = Table(gstYrMnData,hAlign="RIGHT")
    gstYrMnTab.setStyle(TableStyle([('TEXTFONT', (0, 0), (-1, -1), 'Times-Bold'),
    ('TEXTCOLOR',(0,0),(-1,-1),black),
    ('VALIGN',(0,0),(-1,-1),'TOP'),
    ('ALIGN',(0,0),(-1,-1),'RIGHT'),
    ('INNERGRID',(0,0),(-1,-1),0.1,colors.black),
    ('BOX',(0,0),(-1,-1),0.5,colors.black),
    ('SPAN',(1,1),(4,1)),
    ]))
    gstYrMnTab._argW[0]=2*cm
    gstYrMnTab._argW[1]=0.75*cm
    gstYrMnTab._argW[2]=0.75*cm
    gstYrMnTab._argW[3]=0.75*cm
    gstYrMnTab._argW[4]=0.75*cm
    elements.append(gstYrMnTab)
    elements.append(Spacer(2,0.5*cm))
    allData = [gstR023,gstR05]
    for k in str(untObj.gstin):
        gstR07 = Paragraph("<para align=left  fontSize=11> <b> %s</b></para>"%(k),styles['Normal'])
        allData+=[gstR07]
        print k,"unithasdjfahsdjkfasdfjasdf"
    gstData =  [allData,[gstR024,gstR06,gstRuni]]
    gstTab = Table(gstData,rowHeights=0.7*cm)
    gstTab.setStyle(TableStyle([('TEXTFONT', (0, 0), (-1, -1), 'Times-Bold'),
    ('TEXTCOLOR',(0,0),(-1,-1),black),
    ('VALIGN',(0,0),(-1,-1),'TOP'),
    ('ALIGN',(0,0),(-1,-1),'RIGHT'),
    ('INNERGRID',(0,0),(-1,-1),0.1,colors.black),
    ('BOX',(0,0),(-1,-1),0.5,colors.black),
    ('SPAN',(2,1),(16,1)),
    ]))
    gstTab._argW[0]=0.75*cm
    gstTab._argW[1]=7.5*cm
    gstTab._argW[2]=0.75*cm
    gstTab._argW[3]=0.75*cm
    gstTab._argW[4]=0.75*cm
    gstTab._argW[5]=0.75*cm
    gstTab._argW[6]=0.75*cm
    gstTab._argW[7]=0.75*cm
    gstTab._argW[8]=0.75*cm
    gstTab._argW[9]=0.75*cm
    gstTab._argW[10]=0.75*cm
    gstTab._argW[11]=0.75*cm
    gstTab._argW[12]=0.75*cm
    gstTab._argW[13]=0.75*cm
    gstTab._argW[14]=0.75*cm
    gstTab._argW[15]=0.75*cm
    gstTab._argW[16]=0.75*cm
    elements.append(gstTab)
    gstR025 = Paragraph("<para align=left  fontSize=11 spaceAfter=10 spaceBefore=10><b>3.1 Details of Outward Supplies and inward supplies liable to reverse charge </b> </para>",styles['Normal'])
    elements.append(gstR025)
    gstR026 = Paragraph("<para align=center  fontSize=11>Nature of Supplies </para>",styles['Normal'])
    gstR027 = Paragraph("<para align=center  fontSize=11>Total Taxable Value </para>",styles['Normal'])
    gstR028 = Paragraph("<para align=center  fontSize=11>Integrated Tax </para>",styles['Normal'])
    gstR029 = Paragraph("<para align=center  fontSize=11>Central Tax </para>",styles['Normal'])
    gstR030 = Paragraph("<para align=center  fontSize=11>State/UT Tax </para>",styles['Normal'])
    gstR031 = Paragraph("<para align=center  fontSize=11>Cess </para>",styles['Normal'])
    gstR032 = Paragraph("<para align=center  fontSize=11>1 </para>",styles['Normal'])
    gstR033 = Paragraph("<para align=center  fontSize=11>2 </para>",styles['Normal'])
    gstR034 = Paragraph("<para align=center  fontSize=11>3 </para>",styles['Normal'])
    gstR035 = Paragraph("<para align=center  fontSize=11>4 </para>",styles['Normal'])
    gstR036 = Paragraph("<para align=center  fontSize=11>5 </para>",styles['Normal'])
    gstR037 = Paragraph("<para align=center  fontSize=11>6 </para>",styles['Normal'])
    gstR038 = Paragraph("<para align=center  fontSize=11>(a) Outward taxable supplies (other than zero reated,nil rated and exempted) </para>",styles['Normal'])
    gstR039 = Paragraph("<para align=center  fontSize=11> </para>",styles['Normal'])
    gstR040 = Paragraph("<para align=center  fontSize=11> (b) Outward taxable supplies (zero rated)</para>",styles['Normal'])
    gstR041 = Paragraph("<para align=center  fontSize=11> (c) Other outward supplies(Nil rated,exempted)</para>",styles['Normal'])
    gstR042 = Paragraph("<para align=center  fontSize=11> (d) Inward supplies (liable to reverse charge)</para>",styles['Normal'])
    gstR043 = Paragraph("<para align=center  fontSize=11> (e) Non-GST outward supplies</para>",styles['Normal'])
    outWard = [[gstR026,gstR027,gstR028,gstR029,gstR030,gstR031],[gstR032,gstR033,gstR034,gstR035,gstR036,gstR037],
    [gstR038,gstR039,gstR039,gstR039,gstR039,gstR039],
    [gstR040,gstR039,gstR039,gstR039,gstR039,gstR039],
    [gstR041,gstR039,gstR039,gstR039,gstR039,gstR039],
    [gstR042,gstR039,gstR039,gstR039,gstR039,gstR039],
    [gstR043,gstR039,gstR039,gstR039,gstR039,gstR039],
    ]
    outWardTab = Table(outWard)
    outWardTab.setStyle(TableStyle([('TEXTFONT', (0, 0), (-1, -1), 'Times-Bold'),
    ('TEXTCOLOR',(0,0),(-1,-1),black),
    ('VALIGN',(0,0),(-1,-1),'TOP'),
    ('ALIGN',(0,0),(-1,-1),'RIGHT'),
    ('INNERGRID',(0,0),(-1,-1),0.1,colors.black),
    ('BOX',(0,0),(-1,-1),0.5,colors.black),
    ]))
    outWardTab._argW[0]= 3.75*cm
    outWardTab._argW[1]= 4*cm
    outWardTab._argW[2]= 3.5*cm
    outWardTab._argW[3]= 2.5*cm
    outWardTab._argW[4]= 3.5*cm
    outWardTab._argW[5]= 2*cm
    elements.append(outWardTab)
    gstR044 = Paragraph("<para align=left  fontSize=11 spaceAfter=10 spaceBefore=10><b>3.2 Of the supplies shown in 3.1(a) above,details of inter-State supplies made to unregistered persons, composition taxable persons and UIN holders </b> </para>",styles['Normal'])
    elements.append(gstR044)
    gstR045 = Paragraph("<para align=center  fontSize=11> </para>",styles['Normal'])
    gstR046 = Paragraph("<para align=center  fontSize=11>Place of Supply (State/UT) </para>",styles['Normal'])
    gstR047 = Paragraph("<para align=center  fontSize=11>Total Taxable value </para>",styles['Normal'])
    gstR048 = Paragraph("<para align=center  fontSize=11>Amount of Itengrated Tax </para>",styles['Normal'])
    gstR049 = Paragraph("<para align=center  fontSize=11>1 </para>",styles['Normal'])
    gstR050 = Paragraph("<para align=center  fontSize=11>2 </para>",styles['Normal'])
    gstR051 = Paragraph("<para align=center  fontSize=11>3 </para>",styles['Normal'])
    gstR052 = Paragraph("<para align=center  fontSize=11>4 </para>",styles['Normal'])
    gstR053 = Paragraph("<para align=center  fontSize=11>Supplies made to Unregistered Persons </para>",styles['Normal'])
    gstR054 = Paragraph("<para align=center  fontSize=11>Supplies made to Composition Taxable Persons </para>",styles['Normal'])
    gstR055 = Paragraph("<para align=center  fontSize=11> Supplies made to UIN holders </para>",styles['Normal'])
    gstR056 = Paragraph("<para align=center  fontSize=11>  </para>",styles['Normal'])
    IntrState = [[gstR045,gstR046,gstR047,gstR048],[gstR049,gstR050,gstR051,gstR052],
    [gstR053,gstR056,gstR056,gstR056],
    [gstR054,gstR056,gstR056,gstR056],
    [gstR055,gstR056,gstR056,gstR056],

    ]
    IntrStateTab = Table(IntrState)
    IntrStateTab.setStyle(TableStyle([('TEXTFONT', (0, 0), (-1, -1), 'Times-Bold'),
    ('TEXTCOLOR',(0,0),(-1,-1),black),
    ('VALIGN',(0,0),(-1,-1),'TOP'),
    ('ALIGN',(0,0),(-1,-1),'RIGHT'),
    ('INNERGRID',(0,0),(-1,-1),0.1,colors.black),
    ('BOX',(0,0),(-1,-1),0.5,colors.black),
    ]))

    elements.append(IntrStateTab)
    gstR0Itc = Paragraph("<para align=left  fontSize=11 spaceAfter=10 spaceBefore=10><b>4. Eligible ITC</b></para>",styles['Normal'])
    elements.append(gstR0Itc)
    gstR0Itc1 = Paragraph("<para align=center  fontSize=11>Details</para>",styles['Normal'])
    gstR0Itc2 = Paragraph("<para align=center  fontSize=11>Integrated Tax</para>",styles['Normal'])
    gstR0Itc3 = Paragraph("<para align=center  fontSize=11>Central Tax</para>",styles['Normal'])
    gstR0Itc4 = Paragraph("<para align=center  fontSize=11>State/UT Tax</para>",styles['Normal'])
    gstR0Itc5 = Paragraph("<para align=center  fontSize=11>Cess</para>",styles['Normal'])
    gstR0Itc6 = Paragraph("<para align=center  fontSize=11>1</para>",styles['Normal'])
    gstR0Itc7 = Paragraph("<para align=center  fontSize=11>2</para>",styles['Normal'])
    gstR0Itc8 = Paragraph("<para align=center  fontSize=11>3</para>",styles['Normal'])
    gstR0Itc9 = Paragraph("<para align=center  fontSize=11>4</para>",styles['Normal'])
    gstR0Itc10 = Paragraph("<para align=center  fontSize=11>5</para>",styles['Normal'])
    gstR0Itc11 =  Paragraph("<para align=left  fontSize=11 ><b>(A) ITC Available (whether in full or part)</b></para>",styles['Normal'])
    gstR0Itc12 = Paragraph("<para align=left  fontSize=11 lindent=10>(1) Imports of goods</para>",styles['Normal'])
    gstR0Itc13 = Paragraph("<para align=left  fontSize=11 lindent=10>(2) Imports of services</para>",styles['Normal'])
    gstR0Itc14 = Paragraph("<para align=left  fontSize=11 lindent=10>(3) Inward supplies liable to reverse charge (other than 1 & 2 above)</para>",styles['Normal'])
    gstR0Itc15 = Paragraph("<para align=left  fontSize=11 lindent=10>(4) Inward supplies from ISD</para>",styles['Normal'])
    gstR0Itc16 = Paragraph("<para align=left  fontSize=11 lindent=10>(5) All other ITC</para>",styles['Normal'])
    gstR0Itc17 =  Paragraph("<para align=left  fontSize=11 lindent=10></para>",styles['Normal'])
    gstR0Itc18 =  Paragraph("<para align=left  fontSize=11 ><b>(B) ITC Revesrsed</b></para>",styles['Normal'])
    gstR0Itc19 =  Paragraph("<para align=left  fontSize=11  lindent=10>(1) As per rules 42 &43 of CGST Rules</para>",styles['Normal'])
    gstR0Itc20 =  Paragraph("<para align=left  fontSize=11  lindent=10>(2) Others</para>",styles['Normal'])
    gstR0Itc21 =  Paragraph("<para align=left  fontSize=11 ><b>(C) Net ITC Available (A)-(B)</b></para>",styles['Normal'])
    gstR0Itc22 =  Paragraph("<para align=left  fontSize=11 ><b>(D) Ineligible ITC</b></para>",styles['Normal'])
    gstR0Itc23 =  Paragraph("<para align=left  fontSize=11  lindent=10>(1) As per section 17(5)</para>",styles['Normal'])
    gstR0Itc24 =  Paragraph("<para align=left  fontSize=11  lindent=10>(2) Others</para>",styles['Normal'])


    gstItc = [[gstR0Itc1,gstR0Itc2,gstR0Itc3,gstR0Itc4,gstR0Itc5],
    [gstR0Itc6,gstR0Itc7,gstR0Itc8,gstR0Itc9,gstR0Itc10],
    [gstR0Itc11,gstR0Itc17,gstR0Itc17,gstR0Itc17,gstR0Itc17],
    [gstR0Itc12,gstR0Itc17,gstR0Itc17,gstR0Itc17,gstR0Itc17],
    [gstR0Itc13,gstR0Itc17,gstR0Itc17,gstR0Itc17,gstR0Itc17],
    [gstR0Itc14,gstR0Itc17,gstR0Itc17,gstR0Itc17,gstR0Itc17],
    [gstR0Itc15,gstR0Itc17,gstR0Itc17,gstR0Itc17,gstR0Itc17],
    [gstR0Itc16,gstR0Itc17,gstR0Itc17,gstR0Itc17,gstR0Itc17],
    [gstR0Itc18,gstR0Itc17,gstR0Itc17,gstR0Itc17,gstR0Itc17],
    [gstR0Itc19,gstR0Itc17,gstR0Itc17,gstR0Itc17,gstR0Itc17],
    [gstR0Itc20,gstR0Itc17,gstR0Itc17,gstR0Itc17,gstR0Itc17],
    [gstR0Itc21,gstR0Itc17,gstR0Itc17,gstR0Itc17,gstR0Itc17],
    [gstR0Itc22,gstR0Itc17,gstR0Itc17,gstR0Itc17,gstR0Itc17],
    [gstR0Itc23,gstR0Itc17,gstR0Itc17,gstR0Itc17,gstR0Itc17],
    [gstR0Itc24,gstR0Itc17,gstR0Itc17,gstR0Itc17,gstR0Itc17],


    ]
    gstItcTab = Table(gstItc)
    gstItcTab.setStyle(TableStyle([('TEXTFONT', (0, 0), (-1, -1), 'Times-Bold'),
    ('TEXTCOLOR',(0,0),(-1,-1),black),
    ('VALIGN',(0,0),(-1,-1),'TOP'),
    ('ALIGN',(0,0),(-1,-1),'RIGHT'),
    ('INNERGRID',(0,0),(-1,-1),0.1,colors.black),
    ('BOX',(0,0),(-1,-1),0.5,colors.black),
    ]))
    gstItcTab._argW[0] = 8*cm
    gstItcTab._argW[1] = 3*cm
    gstItcTab._argW[2] = 3*cm
    gstItcTab._argW[3] = 3*cm
    gstItcTab._argW[4] = 2*cm
    elements.append(gstItcTab)

    gstR057 = Paragraph("<para align=left  fontSize=11 spaceAfter=10 spaceBefore=10><b>5. Values of exempt,nil-rated and non-GST inward supplies</b></para>",styles['Normal'])
    elements.append(gstR057)
    gstR058 = Paragraph("<para align=center  fontSize=11 >Nature of supplies</para>",styles['Normal'])
    gstR059 = Paragraph("<para align=center  fontSize=11 >Inter-State supplies</para>",styles['Normal'])
    gstR060 = Paragraph("<para align=center  fontSize=11 >Intra-State supplies</para>",styles['Normal'])
    gstR061 = Paragraph("<para align=center  fontSize=11 >1</para>",styles['Normal'])
    gstR062 = Paragraph("<para align=center  fontSize=11 >2</para>",styles['Normal'])
    gstR063 = Paragraph("<para align=center  fontSize=11 >3</para>",styles['Normal'])
    gstR064 = Paragraph("<para align=center  fontSize=11 >From a supplier under composition scheme,Exempt and Nil rated</para>",styles['Normal'])
    gstR065 = Paragraph("<para align=center  fontSize=11 >supply</para>",styles['Normal'])
    gstR066 = Paragraph("<para align=center  fontSize=11 >Non GST supply</para>",styles['Normal'])
    gstR067 = Paragraph("<para align=center  fontSize=11 ></para>",styles['Normal'])
    nonGST = [[gstR058,gstR059,gstR060],[gstR061,gstR062,gstR063],[gstR064,gstR067,gstR067],[gstR065,gstR067,gstR067],[gstR066,gstR067,gstR067]]
    nonGSTTab =Table(nonGST)
    nonGSTTab.setStyle(TableStyle([('TEXTFONT', (0, 0), (-1, -1), 'Times-Bold'),
    ('TEXTCOLOR',(0,0),(-1,-1),black),
    ('VALIGN',(0,0),(-1,-1),'TOP'),
    ('ALIGN',(0,0),(-1,-1),'RIGHT'),
    ('INNERGRID',(0,0),(-1,-1),0.1,colors.black),
    ('BOX',(0,0),(-1,-1),0.5,colors.black),
    ]))
    elements.append(nonGSTTab)

    gstR085 = Paragraph("<para align=left  fontSize=11 spaceBefore=15 spaceAfter=10><b>6.1 Payment of tax</b></para>",styles['Normal'])
    elements.append(gstR085)
    gstR086 = Paragraph("<para align=center  fontSize=11 >Description</para>",styles['Normal'])
    gstR087 = Paragraph("<para align=center  fontSize=11 >Tax payable</para>",styles['Normal'])
    gstR088 = Paragraph("<para align=center  fontSize=11 >Paid though ITC</para>",styles['Normal'])
    gstR089 = Paragraph("<para align=center  fontSize=11 >Tax paid TDS/TCS</para>",styles['Normal'])
    gstR090 = Paragraph("<para align=center  fontSize=11 >Tax/cess paid in cash</para>",styles['Normal'])
    gstR091 = Paragraph("<para align=center  fontSize=11 >Interest</para>",styles['Normal'])
    gstR092 = Paragraph("<para align=center  fontSize=11 >Late Fee</para>",styles['Normal'])
    gstR093 = Paragraph("<para align=center  fontSize=11 >Integrated Tax</para>",styles['Normal'])
    gstR094 = Paragraph("<para align=center  fontSize=11 >Central Tax</para>",styles['Normal'])
    gstR095 = Paragraph("<para align=center  fontSize=11 >State/UT Tax</para>",styles['Normal'])
    gstR096 = Paragraph("<para align=center  fontSize=11 >Cess</para>",styles['Normal'])
    gstR097 = Paragraph("<para align=center  fontSize=11 >1</para>",styles['Normal'])
    gstR098 = Paragraph("<para align=center  fontSize=11 >2</para>",styles['Normal'])
    gstR099 = Paragraph("<para align=center  fontSize=11 >3</para>",styles['Normal'])
    gstR0100 = Paragraph("<para align=center  fontSize=11 >4</para>",styles['Normal'])
    gstR0101 = Paragraph("<para align=center  fontSize=11 >5</para>",styles['Normal'])
    gstR0102 = Paragraph("<para align=center  fontSize=11 >6</para>",styles['Normal'])
    gstR0103 = Paragraph("<para align=center  fontSize=11 >7</para>",styles['Normal'])
    gstR0104 = Paragraph("<para align=center  fontSize=11 >8</para>",styles['Normal'])
    gstR0105 = Paragraph("<para align=center  fontSize=11 >9</para>",styles['Normal'])
    gstR0106 = Paragraph("<para align=center  fontSize=11 >10</para>",styles['Normal'])
    gstR0107 = Paragraph("<para align=center  fontSize=11 ></para>",styles['Normal'])
    paymTax = [
    [gstR086,gstR087,gstR088,gstR0107,gstR0107,gstR0107,gstR089,gstR090,gstR091,gstR092],
    [gstR0107,gstR0107,gstR093,gstR094,gstR095,gstR096,gstR0107,gstR0107],
    [gstR097,gstR098,gstR099,gstR0100,gstR0101,gstR0102,gstR0103,gstR0104,gstR0105,gstR0106],
    [gstR093,gstR0107,gstR0107,gstR0107,gstR0107,gstR0107,gstR0107,gstR0107,gstR0107,gstR0107],
    [gstR094,gstR0107,gstR0107,gstR0107,gstR0107,gstR0107,gstR0107,gstR0107,gstR0107,gstR0107],
    [gstR095,gstR0107,gstR0107,gstR0107,gstR0107,gstR0107,gstR0107,gstR0107,gstR0107,gstR0107],
    [gstR096,gstR0107,gstR0107,gstR0107,gstR0107,gstR0107,gstR0107,gstR0107,gstR0107,gstR0107],

    ]
    paymTaxTab =Table(paymTax)
    paymTaxTab.setStyle(TableStyle([('TEXTFONT', (0, 0), (-1, -1), 'Times-Bold'),
    ('TEXTCOLOR',(0,0),(-1,-1),black),
    ('VALIGN',(0,0),(-1,-1),'TOP'),
    ('ALIGN',(0,0),(-1,-1),'RIGHT'),
    ('INNERGRID',(0,0),(-1,-1),0.1,colors.black),
    ('BOX',(0,0),(-1,-1),0.5,colors.black),
    ('SPAN',(2,0),(5,0)),
    ]))
    paymTaxTab._argW[0]=3*cm
    paymTaxTab._argW[1]=1.5*cm
    paymTaxTab._argW[2]=1.5*cm
    paymTaxTab._argW[3]=1.5*cm
    paymTaxTab._argW[4]=1.5*cm
    paymTaxTab._argW[5]=1.5*cm
    paymTaxTab._argW[6]=3*cm
    paymTaxTab._argW[7]=2.5*cm
    paymTaxTab._argW[8]=1.5*cm
    paymTaxTab._argW[9]=1.5*cm
    elements.append(paymTaxTab)
    gstR068 = Paragraph("<para align=left  fontSize=11 spaceBefore=15 spaceAfter=10><b>6.2 TDS/TCS Credit</b></para>",styles['Normal'])
    elements.append(gstR068)
    gstR069 = Paragraph("<para align=center  fontSize=11 >Details</para>",styles['Normal'])
    gstR070 = Paragraph("<para align=center  fontSize=11 >Integrated Tax</para>",styles['Normal'])
    gstR071 = Paragraph("<para align=center  fontSize=11 >Central Tax</para>",styles['Normal'])
    gstR072 = Paragraph("<para align=center  fontSize=11 >State/UT Tax</para>",styles['Normal'])
    gstR073 = Paragraph("<para align=center  fontSize=11 >1</para>",styles['Normal'])
    gstR074 = Paragraph("<para align=center  fontSize=11 >2</para>",styles['Normal'])
    gstR075 = Paragraph("<para align=center  fontSize=11 >3</para>",styles['Normal'])
    gstR076 = Paragraph("<para align=center  fontSize=11 >4</para>",styles['Normal'])
    gstR077 = Paragraph("<para align=center  fontSize=11 >TDS</para>",styles['Normal'])
    gstR078 = Paragraph("<para align=center  fontSize=11 ></para>",styles['Normal'])
    gstR079 = Paragraph("<para align=center  fontSize=11 ></para>",styles['Normal'])
    gstR080 = Paragraph("<para align=center  fontSize=11 ></para>",styles['Normal'])
    gstR081 = Paragraph("<para align=center  fontSize=11 >TCS</para>",styles['Normal'])
    gstR082 = Paragraph("<para align=center  fontSize=11 ></para>",styles['Normal'])
    gstR083 = Paragraph("<para align=center  fontSize=11 ></para>",styles['Normal'])
    gstR084 = Paragraph("<para align=center  fontSize=11 ></para>",styles['Normal'])
    TcsTds = [[gstR069,gstR070,gstR071,gstR072],
    [gstR073,gstR074,gstR075,gstR076],
    [gstR077,gstR078,gstR079,gstR080],
    [gstR081,gstR082,gstR083,gstR084],
    ]
    TcsTdsTab =Table(TcsTds)
    TcsTdsTab.setStyle(TableStyle([('TEXTFONT', (0, 0), (-1, -1), 'Times-Bold'),
    ('TEXTCOLOR',(0,0),(-1,-1),black),
    ('VALIGN',(0,0),(-1,-1),'TOP'),
    ('ALIGN',(0,0),(-1,-1),'RIGHT'),
    ('INNERGRID',(0,0),(-1,-1),0.1,colors.black),
    ('BOX',(0,0),(-1,-1),0.5,colors.black),
    ]))
    elements.append(TcsTdsTab)
    doc.build(elements,canvasmaker=PageNumCanvas)

class GSTR3BPdfAPI(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny,)
    def get(self,request , format= None):
        untObj = Unit.objects.get(pk = request.GET['unit'])
        fYr = str(int(request.GET["FinY"].split('-')[0])+1)
        qur = request.GET['Quarter']
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment;filename="GSTR3B_%s.pdf"'%(untObj.gstin)
        gstrPDF(response , request,untObj,fYr,qur)
        return response

class GetVendorDetailsAPI(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny,)
    def get(self,request , format= None):
        vendorPk = request.GET['id']
        vendorObj = VendorProfile.objects.get(pk = vendorPk)
        vendorServiceObj = VendorService.objects.filter(vendorProfile = vendorObj)
        purchaseCount = InvoiceReceived.objects.filter(vendor = vendorObj,status='GRN').count()
        data = {'vendor':VendorProfileLiteSerializer(vendorObj,many=False).data,'totalPurchase':purchaseCount,'services':VendorServiceLiteSerializer(vendorServiceObj,many=True).data}
        return Response(data, status = status.HTTP_200_OK)

def PoPDF(response , request,poObj):
    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(response,pagesize=letter, topMargin=1*cm,leftMargin=1*cm,rightMargin=1*cm)
    doc.request = request
    elements = []
    headerFilePath = globalSettings.BRAND_LOGO_PDF
    drawing = svg2rlg(os.path.join(globalSettings.BASE_DIR,
                                   'static_shared', 'images', headerFilePath))
    sx = sy = 1
    drawing.width, drawing.height = drawing.minWidth() * sx, drawing.height * sy
    drawing.scale(sx, sy)
    elements.append(drawing)
    print  poObj.user
    userObj = User.objects.get(pk = poObj.user.pk)
    po01 = Paragraph("<para align=center  fontSize=11    spaceAfter=15 spaceBefore=20> PURCHASE ORDER </para>",styles['Heading1'])
    poSpac =  Paragraph("<para align=left fontSize=11 spaceAfter=10 spaceBefore=20><b>Products</b></para>",styles['Normal'])
    poNum = Paragraph("<para align=left  fontSize=11    spaceAfter=10 spaceBefore=20>  <b>Order NO: #%d</b></para>"%(poObj.pk),styles['Normal'])
    elements.append(po01)
    elements.append(poNum)
    po017 = Paragraph("<para  fontSize=11 spaceAfter=5><b>Company Details </b></para>",styles['Normal'])
    po018 = Paragraph("<para  fontSize=11 spaceAfter=5><b> Name : </b> %s  </para>"%(poObj.vendor.service.name),styles['Normal'])
    po019 = Paragraph("<para  fontSize=11 spaceAfter=5><b> Telephone : </b> %s  </para>"%(poObj.vendor.service.telephone),styles['Normal'])
    po0mob = Paragraph("<para  fontSize=11 spaceAfter=5><b> Mobile : </b> %s  </para>"%(poObj.vendor.service.mobile),styles['Normal'])
    po025 = Paragraph("<para  fontSize=11 spaceAfter=5><b>Address : </b> %s,%s<br/><br/>%s,%s<br/><br/>%s</para>"%(poObj.address,poObj.city,poObj.state,poObj.country,poObj.pincode),styles['Normal'])
    po020 = Paragraph("<para align=left fontSize=11 spaceAfter=5><b>Contact Person Details </b>  </para>",styles['Normal'])
    po021 = Paragraph("<para align=left fontSize=11 spaceAfter=5><b>Name: </b> %s  </para>"%(poObj.vendor.contactPerson),styles['Normal'])
    po022 = Paragraph("<para align=left fontSize=11 spaceAfter=5><b>Email: </b> %s  </para>"%(poObj.vendor.email),styles['Normal'])
    po023 = Paragraph("<para align=left fontSize=11 spaceAfter=5><b>Mobile: </b> %s  </para>"%(poObj.vendor.mobile),styles['Normal'])
    po024 = Paragraph("<para align=left fontSize=11 spaceAfter=5><b>Address : </b> %s,%s<br/><br/>%s,%s<br/><br/>%s</para>"%(poObj.address,poObj.city,poObj.state,poObj.country,poObj.pincode),styles['Normal'])
    reqPerson = User.objects.get(pk = poObj.requester.pk)
    po026 = Paragraph("<para  fontSize=11 spaceAfter=5 spaceBefore=10><b>Issuer </b>   </para>",styles['Normal'])
    po030 = Paragraph("<para align=left fontSize=11 spaceAfter=5 spaceBefore=10><b>Requester </b>   </para>",styles['Normal'])
    po027 = Paragraph("<para  fontSize=11 spaceAfter=5><b>Name: </b> %s %s </para>"%(userObj.first_name,userObj.last_name),styles['Normal'])
    po028 = Paragraph("<para  fontSize=11 spaceAfter=5><b>Email: </b> %s  </para>"%(userObj.email),styles['Normal'])
    po029 = Paragraph("<para fontSize=11 spaceAfter=5><b>Mobile: </b> %s  </para>"%(userObj.profile.mobile),styles['Normal'])
    po0R01 = Paragraph("<para fontSize=11 spaceAfter=5><b>Name: </b> %s  %s </para>"%(reqPerson.first_name,reqPerson.last_name),styles['Normal'])
    po0R02 = Paragraph("<para fontSize=11 spaceAfter=5><b>Email: </b> %s </para>"%(reqPerson.email),styles['Normal'])
    po0R03 = Paragraph("<para fontSize=11 spaceAfter=5><b>Mobile: </b> %s </para>"%(reqPerson.profile.mobile),styles['Normal'])
    personTab = [[po020],[po021],[po022],[po023],[po024]]
    compData = [[po017],[po018],[po019],[po0mob],[po025]]
    ReqTab = [[po030],[po0R01],[po0R02],[po0R03]]
    issuTab = [[po026],[po027],[po028],[po029]]
    twtoTable =[ [personTab,compData],[ReqTab,issuTab]]
    compTab = Table(twtoTable)
    compTab.setStyle(TableStyle([('TEXTFONT', (0, 0), (-1, -1), 'Times-Bold'),
    ('TEXTCOLOR',(0,0),(-1,-1),black),
    ('VALIGN',(0,0),(-1,-1),'TOP'),
    ('ALIGN',(0,0),(-1,-1),'LEFT'),
    ('LEFTPADDING',(0,0),(-1,-1),0),
    ]))
    elements.append(compTab)
    elements.append(poSpac)

    po010 =  Paragraph("<para align=center fontSize=11 spaceAfter=5><b>S.No</b> </para>",styles['Normal'])
    po011 =  Paragraph("<para align=center fontSize=11 spaceAfter=5><b>Product</b> </para>",styles['Normal'])
    po012 =  Paragraph("<para align=center fontSize=11 spaceAfter=5><b>Price</b> </para>",styles['Normal'])
    po013 =  Paragraph("<para align=center fontSize=11 spaceAfter=5><b>Quantity</b> </para>",styles['Normal'])
    po015 =  Paragraph("<para align=center fontSize=11 spaceAfter=5><b>Total</b> </para>",styles['Normal'])
    po0subTot  =  Paragraph("<para align=center fontSize=11 spaceAfter=5> <b>Sub Total </b></para>",styles['Normal'])
    po0empty  =  Paragraph("<para align=center fontSize=11 spaceAfter=5>  </para>",styles['Normal'])
    PoData = [[po010,po011,po012,po013,po015]]
    PoQty = InvoiceQty.objects.filter(invoice = poObj)
    subTotal = 0
    po0subTotal = ''
    for idx,pQty in enumerate(PoQty):
        pQty.total = pQty.qty * pQty.price
        po010 =  Paragraph("<para align=center fontSize=11 spaceAfter=5> %d </para>"%(idx+1),styles['Normal'])
        po011 =  Paragraph("<para align=center fontSize=11 spaceAfter=5>%s </para>"%(pQty.product),styles['Normal'])
        po012 =  Paragraph("<para align=center fontSize=11 spaceAfter=5>%d </para>"%(pQty.price),styles['Normal'])
        po013 =  Paragraph("<para align=center fontSize=11 spaceAfter=5>%d </para>"%(pQty.qty),styles['Normal'])
        if pQty.total != None:
            po015 =  Paragraph("<para align=center fontSize=11 spaceAfter=5> %s </para>"%(pQty.total),styles['Normal'])
        else:
            po015 =  Paragraph("<para align=center fontSize=11 spaceAfter=5> 0 </para>",styles['Normal'])

        if pQty.total != None:

            subTotal += pQty.total
            po0subTotal  =  Paragraph("<para align=center fontSize=11 spaceAfter=5> %0.2f </para>"%(subTotal),styles['Normal'])

        else:
            subTotal = 0
            po0subTotal  =  Paragraph("<para align=center fontSize=11 spaceAfter=5> %0.2f </para>"%(subTotal),styles['Normal'])
        PoData += [[po010,po011,po012,po013,po015]]
    fullData = PoData+[[po0subTot,po0empty,po0empty,po0empty,po0subTotal]]
    print subTotal,"subTotalsubTotal"
    POTab = Table(fullData)
    POTab.setStyle(TableStyle([('TEXTFONT', (0, 0), (-1, -1), 'Times-Bold'),
    ('TEXTCOLOR',(0,0),(-1,-1),black),
    ('VALIGN',(0,0),(-1,-1),'TOP'),
    ('ALIGN',(0,0),(-1,-1),'RIGHT'),
    ('INNERGRID',(0,0),(-1,-1),0.1,colors.black),
    ('BOX',(0,0),(-1,-1),0.5,colors.black),

    ]))

    elements.append(POTab)
    po0quote  =  Paragraph("<para align=left fontSize=11 spaceAfter=5 spaceBefore=10> <b>Created On: </b>  %s</para>" %(poObj.quoteDate),styles['Normal'])
    elements.append(po0quote)
    po0Delivery = Paragraph("<para align=left fontSize=11 spaceAfter=5> <b>Tentative Delivery Date: </b>  %s</para>"%(poObj.deliveryDate),styles['Normal'])
    elements.append(po0Delivery)
    print poObj.termsandcondition,"poObj.termsandcondition.body"
    terms = []
    if poObj.termsandcondition != None :
        po031 =  Paragraph("<para align=left fontSize=11 spaceAfter=10 spaceBefore=10> <b>Terms and Conditions: </b></para>",styles['Normal'])
        elements.append(po031)
        terms = poObj.termsandcondition.body.split("||")
        for idx,val in enumerate(terms):
            po030 =  Paragraph("<para align=left fontSize=9 spaceAfter=5 spaceBefore=10> %d. %s </para>"%(idx+1,val),styles['Normal'])
            elements.append(po030)
    print terms,'terms'

    doc.build(elements,canvasmaker=PageNumCanvas)
class PurchaseOrderPDFAPI(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny,)
    def get(self,request , format= None):
        POobj = InvoiceReceived.objects.get(pk = request.GET['id'])
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment;filename="Invoice_%d.pdf"'%(POobj.pk)
        PoPDF(response , request,POobj)
        return response

class GetAllCountAPI(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny ,)
    def get(self,request , format= None):
        toReturn = []
        if 'type' in request.GET:
            if request.GET['type'] == 'account':
                divsn = request.user.designation.division
                accountObj = Account.objects.filter(division = divsn)
                account = accountObj.filter(group = None).exclude(personal = True).count()
                val = {'name' : 'Accounts' , 'is_selected' : False, 'count' : account,  'value' : 'account'}
                toReturn.append(val)
                petty = accountObj.filter(group = None, personal = True).count()
                val = {'name' : 'Petty Cash' , 'is_selected' : False, 'count' : petty,  'value' : 'account'}
                toReturn.append(val)
                pal = accountObj.filter(group = 'profit and loss').exclude(personal = True).count()
                val = {'name' : 'Profit and Loss' , 'is_selected' : False, 'count' : pal , 'value' : 'profit and loss'}
                toReturn.append(val)
                cash = accountObj.filter(group = 'cashflow').exclude(personal = True).count()
                val = {'name' : 'Cashflow' , 'is_selected' : False, 'count' : cash , 'value' : 'cashflow'}
                toReturn.append(val)
                balance = accountObj.filter(group = 'balancesheet').exclude(personal = True).count()
                val = {'name' : 'Balance Sheet' , 'is_selected' : False, 'count' : balance , 'value' : 'balancesheet'}
                toReturn.append(val)
            if request.GET['type'] == 'purchasePO':
                purchaseObj = InvoiceReceived.objects.filter(isInvoice = False)
                created = purchaseObj.filter(status = 'created').count()
                val = {'name' : 'Waiting for Approval' , 'is_selected' : False, 'count' : created, 'value' : 'created'}
                toReturn.append(val)
                approved = purchaseObj.filter(status = 'Approved').count()
                val = {'name' : 'Approved' , 'is_selected' : False, 'count' : approved, 'value' : 'Approved'}
                toReturn.append(val)
        return Response({'data':toReturn }, status = status.HTTP_200_OK)

import random, string
class CreateTransactionsAPI(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny ,)
    def post(self,request , format= None):
        trans =  request.data['data']
        date = request.data['date']

        length = 16
        chars = string.digits
        rnd = random.SystemRandom()
        groupId = ''.join(rnd.choice(chars) for i in range(length))

        for data in trans:
            transObj = Transaction.objects.create( narration  = request.data['narration'], user=request.user, dated = date , groupId = groupId )
            accountObj = Account.objects.get(pk = data['account'])
            transObj.debit = data['debit']
            transObj.credit = data['credit']
            if transObj.credit>0:
                transObj.toAcc = accountObj
            if transObj.debit>0:
                transObj.fromAcc = accountObj
            accountObj.balance = accountObj.balance + float(data['credit'])
            accountObj.balance = accountObj.balance - float(data['debit'])
            transObj.balance = accountObj.balance
            transObj.save()
            accountObj.save()
        res = TransactionSerializer(Transaction.objects.filter(groupId = groupId), many = True ).data
        return Response(res , status = status.HTTP_200_OK)


class JournalAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny ,)
    def get(self,request , format= None):
        month = request.GET['month']
        temp_year = request.GET['year']
        if 'January' in month or 'February'  in month or 'March'  in month:
            year = temp_year.split('-')[1]
            if 'Quarter 4' in month:
                startDate = date.today().replace(month=1, day=1, year=int(year))
                endDate = date.today().replace(month=3, day=31, year=int(year))
            else:
                month_no = month_string_to_number(month)
                startDate = date.today().replace(month=int(month_no), day=1, year=int(year))
                endDate = startDate.replace(day=calendar.monthrange(startDate.year, startDate.month)[1])
        elif 'FY' in month:
            year = temp_year.split('-')[0]
            year1 = temp_year.split('-')[1]
            startDate = date.today().replace(month=4, day=1, year=int(year))
            endDate = date.today().replace(month=3, day=31, year=int(year1))
        else:
            year = temp_year.split('-')[0]
            if 'Quarter 1' in month :
                startDate = date.today().replace(month=4, day=1, year=int(year))
                endDate = date.today().replace(month=6, day=31, year=int(year))
            elif 'Quarter 2' in month :
                startDate = date.today().replace(month=7, day=1, year=int(year))
                endDate = date.today().replace(month=9, day=30, year=int(year))
            elif 'Quarter 3' in month :
                startDate = date.today().replace(month=10, day=1, year=int(year))
                endDate = date.today().replace(month=12, day=31, year=int(year))
            else:
                month_no = month_string_to_number(month)
                startDate = date.today().replace(month=int(month_no), day=1, year=int(year))
                endDate = startDate.replace(day=calendar.monthrange(startDate.year, startDate.month)[1])
        ids = Transaction.objects.filter(dated__range = (startDate , endDate)).exclude(externalRecord = True).order_by('dated').values('groupId').distinct()
        result = []
        for id in ids:
            res = TransactionSerializer(Transaction.objects.filter(groupId = id['groupId']).exclude(externalRecord = True), many = True ).data
            result.append(res)


        return Response(result , status = status.HTTP_200_OK)



class GetAllExpensesAPI(APIView):
    permission_classes = (permissions.IsAuthenticated ,)
    def get(self , request , format = None):
        from datetime import date, timedelta
        print 'sssssssssssssssssssssss'
        tosend = []
        today = date.today()
        divsn = self.request.user.designation.division
        purchaseObj = InvoiceReceived.objects.filter( division = divsn)
        expenseObj = ExpenseSheet.objects.filter(stage = 'submitted' , division = divsn)
        totalAmount = 0
        balanceAmount = 0
        paidAmount = 0
        limit = int(self.request.GET['limit'])
        if 'search' in request.GET:
            search = self.request.GET['search']
            if len(search)>0:
                purchaseObj = purchaseObj.filter(Q(companyName__icontains=search) | Q(phone__icontains=search) | Q(email__icontains=search) | Q(personName__icontains=search) | Q(pincode__icontains=search))
                expenseObj = expenseObj.filter(Q(notes__icontains=search) )
        # if request.GET['status'] == 'invoice':
        #     purchaseObj = purchaseObj.all()
        # if request.GET['status'] == 'purchaseOrder':
        #     purchaseObj = purchaseObj.filter(isInvoice = False)
        if request.GET['status'] == 'invoice':
            purchaseObj = purchaseObj[:limit]
            final_data = InvoiceReceivedSerializer(purchaseObj , many=True).data
            final_data.sort(key=lambda item:item['created'], reverse=True)
        if request.GET['status'] == 'expenseSheet':
            expenseObj = expenseObj[:limit]
            final_data = ExpenseSheetSerializer(expenseObj , many=True).data
            final_data.sort(key=lambda item:item['created'], reverse=True)
        if request.GET['status'] == 'all':
            purchaseObj = purchaseObj[:limit/2]
            expenseObj = expenseObj[:limit/2]
            data = InvoiceReceivedSerializer(purchaseObj , many=True).data
            data1 = ExpenseSheetSerializer(expenseObj , many=True).data
            final_data =  data + data1
            final_data.sort(key=lambda item:item['created'], reverse=True)
        tosend = {'data' : final_data , 'totalAmount' : totalAmount , 'balanceAmount' : balanceAmount , 'paidAmount' : paidAmount}
        print tosend,'aaaaaaaaaaaaaa'
        return Response(tosend, status=status.HTTP_200_OK)

class GetExpensesExcelAPI(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny ,)
    def get(self,request , format= None):
        workbook = Workbook()
        Sheet1 = workbook.active
        hdFont = Font(size=12,bold=True)
        alphaChars = list(string.ascii_uppercase)
        Sheet1.title = 'Collections'
        hd = ["Date","Type", 'Name','Amount']
        hdWidth = [40,40,40,30]
        Sheet1.append(hd)
        divsn = self.request.user.designation.division
        purchaseObj = InvoiceReceived.objects.filter( division = divsn )
        expenseObj = ExpenseSheet.objects.filter(stage = 'submitted' , user__designation__division = divsn )
        if 'fromDate' in request.GET and 'toDate' in request.GET:
            fromDate = request.GET['fromDate']
            toDate = request.GET['toDate']
            purchaseObj = purchaseObj.filter(created__range = (fromDate , toDate))
            expenseObj = expenseObj.filter(created__range = (fromDate , toDate))
        data = InvoiceReceivedSerializer(purchaseObj , many=True).data
        data1 = ExpenseSheetSerializer(expenseObj , many=True).data
        final_data =  data + data1
        final_data.sort(key=lambda item:item['created'], reverse=True)
        for i in final_data:
            dated = datetime.strptime(i['created'], "%Y-%m-%dT%H:%M:%S.%fZ").strftime('%Y-%m-%d')
            typ = i['typ']
            if typ == 'EXPENSE SHEET':
                user = User.objects.get(pk = int(i['user']))
                name = user.first_name + ' ' + user.last_name
            else:
                name = i['companyName']
            amount = i['totalAmount']
            data = [dated,typ,name,amount]
            Sheet1.append(data)
        Sheet1.column_dimensions['A'].width = 40
        Sheet1.column_dimensions['B'].width = 40
        Sheet1.column_dimensions['C'].width = 40
        Sheet1.column_dimensions['D'].width = 30
        response = HttpResponse(content=save_virtual_workbook(workbook),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Expenses.xlsx'
        return response

class OuttbondInvoiceAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny ,)
    def post(self,request , format= None):
        if 'invoicePk' in request.data:
            data  = request.data
            data_to_post = {}
            contractObj = Contract.objects.get(pk = int(data['invoicePk']))
            contractQtyObj =  json.loads(contractObj.data)
            data_to_post = {'personName' : contractObj.contact.name , 'phone' : contractObj.contact.mobile , 'email' : contractObj.contact.email , 'address' :  contractObj.contact.street ,'city' :  contractObj.contact.city ,
            'state' :  contractObj.contact.state , 'country' :  contractObj.contact.country , 'pincode' :  contractObj.contact.pincode  , 'contact' : contractObj.contact }
            if 'poNumber' in data:
                data_to_post['poNumber'] = data['poNumber']
            if 'deliveryDate' in data:
                data_to_post['deliveryDate'] = data['deliveryDate']
            if 'payDueDate' in data:
                data_to_post['payDueDate'] = data['payDueDate']
            if 'isInvoice' in data:
                data_to_post['isInvoice'] = data['isInvoice']
            if 'isPerforma' in data:
                data_to_post['isPerforma'] = data['isPerforma']
            if 'serviceFor' in data:
                data_to_post['serviceFor'] = data['serviceFor']
            outBondObj = Sale(**data_to_post)
            outBondObj.user = request.user
            if 'account' in data:
                outBondObj.account = Account.objects.get(pk = int(data['account']))
            if 'termsandcondition' in data:
                outBondObj.termsandcondition = TermsAndConditions.objects.get(pk = int(data['termsandcondition']))
            if 'terms' in data:
                outBondObj.terms = data['terms']
            if contractObj.contact.company:
                outBondObj.name = contractObj.contact.company.name
                outBondObj.gstIn = contractObj.contact.company.tin
            outBondObj.save()
            totalAmount = 0
            for i in contractQtyObj:
                prodData = {}
                if 'rate' in i:
                    prodData['price'] = i['rate']
                if 'desc' in i:
                    prodData['product'] = i['desc']
                if 'quantity' in i:
                    prodData['qty'] =  i['quantity']
                if 'tax' in i:
                    prodData['tax'] = i['tax']
                if 'subtotal' in i:
                    prodData['total']  = i['subtotal']
                if 'taxPer' in i:
                    prodData['taxPer']  = i['taxPer']
                if 'hsn' in i:
                    prodData['hsn'] = i['hsn']
                qtyObj = SalesQty(**prodData)
                qtyObj.outBound= outBondObj
                qtyObj.save()
                totalAmount += float(qtyObj.total)
            outBondObj.total = totalAmount
            outBondObj.balanceAmount = totalAmount
            outBondObj.save()
            contractObj.status = 'approved'
            contractObj.save()
        else:
            data  = request.data
            data_to_post = {}
            if 'name' in data:
                data_to_post['name'] = data['name']
            if 'poNumber' in data:
                data_to_post['poNumber'] = data['poNumber']
            if 'deliveryDate' in data:
                data_to_post['deliveryDate'] = data['deliveryDate']
            if 'payDueDate' in data:
                data_to_post['payDueDate'] = data['payDueDate']
            if 'personName' in data:
                data_to_post['personName'] = data['personName']
            if 'phone' in data:
                data_to_post['phone'] = data['phone']
            if 'email' in data:
                data_to_post['email'] = data['email']
            if 'address' in data:
                data_to_post['address'] = data['address']
            if 'pincode' in data:
                data_to_post['pincode'] = data['pincode']
            if 'city' in data:
                data_to_post['city'] = data['city']
            if 'state' in data:
                data_to_post['state'] = data['state']
            if 'country' in data:
                data_to_post['country'] = data['country']
            if 'total' in data:
                data_to_post['total'] = data['total']
            if 'isInvoice' in data:
                data_to_post['isInvoice'] = data['isInvoice']
            if 'isPerforma' in data:
                data_to_post['isPerforma'] = data['isPerforma']
            if 'balanceAmount' in data:
                data_to_post['balanceAmount'] = data['balanceAmount']
            if 'serviceFor' in data:
                data_to_post['serviceFor'] = data['serviceFor']
            if 'pk' in data:
                outBondObj = Sale.objects.get( pk = int(data['pk']))
                outBondObj.__dict__.update(data_to_post)
            else:
                outBondObj = Sale(**data_to_post)
                outBondObj.user = request.user
                if 'parent' in data:
                    outBondObj.parent = Sale.objects.get( pk = int(data['parent']))
            if 'costcenter' in data:
                outBondObj.costcenter = CostCenter.objects.get(pk = int(data['costcenter']))
            if 'account' in data:
                outBondObj.account = Account.objects.get(pk = int(data['account']))
            if 'gstIn' in data:
                outBondObj.gstIn = data['gstIn']
            if 'contact' in data:
                outBondObj.contact = Contact.objects.get(pk = int(data['contact']))
            if 'termsandcondition' in data:
                outBondObj.termsandcondition = TermsAndConditions.objects.get(pk = int(data['termsandcondition']))
            if 'terms' in data:
                outBondObj.terms = data['terms']
            outBondObj.save()
            if 'products' in data:
                products = data['products']
                for i in products:
                    prodData = {}
                    if 'price' in i and i['price'] is not None:
                        prodData['price'] = i['price']
                    if 'product' in i and i['product'] is not None:
                        prodData['product'] = i['product']
                    if 'total' in i and i['total'] is not None:
                        prodData['total'] = i['total']
                    if 'qty' in i and i['qty'] is not None:
                        prodData['qty'] = i['qty']
                    if 'tax' in i and i['tax'] is not None:
                        prodData['tax'] = i['tax']
                    if 'taxPer' in i and i['taxPer'] is not None:
                        prodData['taxPer']  = i['taxPer']
                    if 'hsn' in i and i['hsn'] is not None:
                        prodData['hsn'] = i['hsn']
                    if 'pk' in i:
                        qtyObj = SalesQty.objects.get(pk = i['pk'])
                        qtyObj.__dict__.update(prodData)
                        qtyObj.save()
                    else:
                        qtyObj = SalesQty(**prodData)
                        qtyObj.outBound= outBondObj
                        if 'id' in i:
                            tourObj = TourPlanStop.objects.get(pk = int(i['id']))
                            tourObj.billed = True
                            tourObj.save()
                        qtyObj.save()
        data = SaleAllSerializer(outBondObj ,  many = False).data
        if outBondObj.parent is not None:
            outBondAllObj = Sale.objects.filter( parent = outBondObj.parent )
            so = outBondObj.parent
            totalUsed = outBondAllObj.aggregate(tot = Sum('total'))
            billed = 0
            if totalUsed['tot'] is not None:
                billed = totalUsed['tot']
                so.balanceAmount = so.total - billed
                so.paidAmount  = billed
                so.save()
        return Response(data,status = status.HTTP_200_OK)
    def get(self,request , format= None):
        if 'id' in request.GET:
            outObj = Sale.objects.get(pk = int(request.GET['id']))
            data = SaleAllSerializer(outObj ,  many = False).data
            children = SaleAllSerializer(Sale.objects.filter(parent = outObj).order_by('-created') ,  many = True).data
            data['children'] = children
        return Response(data ,status = status.HTTP_200_OK)


class PurchaseOrderInvoiceAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny ,)
    def post(self,request , format= None):
        data  = request.data
        data_to_post = {}
        if 'name' in data:
            data_to_post['name'] = data['name']
        if 'poNumber' in data:
            data_to_post['poNumber'] = data['poNumber']
        if 'deliveryDate' in data:
            data_to_post['deliveryDate'] = data['deliveryDate']
        if 'quoteDate' in data:
            data_to_post['quoteDate'] = data['quoteDate']
        if 'paymentDueDate' in data:
            data_to_post['paymentDueDate'] = data['paymentDueDate']
        if 'quoteNumber' in data:
            data_to_post['quoteNumber'] = data['quoteNumber']
        if 'personName' in data:
            data_to_post['personName'] = data['personName']
        if 'phone' in data:
            data_to_post['phone'] = data['phone']
        if 'email' in data:
            data_to_post['email'] = data['email']
        if 'address' in data:
            data_to_post['address'] = data['address']
        if 'pincode' in data:
            data_to_post['pincode'] = data['pincode']
        if 'city' in data:
            data_to_post['city'] = data['city']
        if 'state' in data:
            data_to_post['state'] = data['state']
        if 'country' in data:
            data_to_post['country'] = data['country']
        if 'totalAmount' in data:
            data_to_post['totalAmount'] = data['totalAmount']
        if 'balanceAmount' in data:
            data_to_post['balanceAmount'] = data['balanceAmount']
        if 'paidAmount' in data:
            data_to_post['paidAmount'] = data['paidAmount']
        if 'isInvoice' in data:
            data_to_post['isInvoice'] = data['isInvoice']
        if 'note' in data:
            data_to_post['note'] = data['note']
        if 'pk' in data:
            poinvObj = InvoiceReceived.objects.get( pk = int(data['pk']))
            poinvObj.__dict__.update(data_to_post)
            poinvObj.user = request.user
        else:
            poinvObj = InvoiceReceived(**data_to_post)
            if 'parent' in data:
                poinvObj.parent = InvoiceReceived.objects.get( pk = int(data['parent']))
        if 'vendor' in data:
            poinvObj.vendor = VendorProfile.objects.get(pk = int(data['vendor']))
        if 'costcenter' in data:
            poinvObj.costcenter = CostCenter.objects.get(pk = int(data['costcenter']))
        if 'gstIn' in data:
            poinvObj.gstIn = data['gstIn']
        if 'requester' in data:
            poinvObj.requester = User.objects.get(pk = int(data['requester']))
        if 'bankName' in data:
            poinvObj.bankName = data['bankName']
        if 'accNo' in data:
            poinvObj.accNo = data['accNo']
        if 'ifsc' in data:
            poinvObj.ifsc = data['ifsc']

        poinvObj.save()
        if 'products' in data:
            products = data['products']
            for i in products:
                prodData = {}
                if 'price' in i and i['price'] is not None:
                    prodData['price'] = i['price']
                if 'product' in i and i['product'] is not None:
                    prodData['product'] = i['product']
                if 'total' in i and i['total'] is not None:
                    prodData['total'] = i['total']
                if 'receivedQty' in i and i['receivedQty'] is not None:
                    prodData['receivedQty'] = i['receivedQty']
                if 'tax' in i and i['tax'] is not None:
                    prodData['tax'] = i['tax']
                if 'hsn' in i and i['hsn'] is not None and i['hsn'] !='':
                    prodData['hsn'] = i['hsn']
                if 'productMeta' in i and i['productMeta'] is not None and i['productMeta'] !='':
                    prodData['productMeta'] = ProductMeta.objects.get(pk = int(i['productMeta']['pk']))
                if 'pk' in i:
                    qtyObj = InvoiceQty.objects.get(pk = i['pk'])
                    qtyObj.__dict__.update(prodData)
                    qtyObj.save()
                else:
                    qtyObj = InvoiceQty(**prodData)
                    qtyObj.invoice= poinvObj
                    qtyObj.save()
        data = InvoiceReceivedAllSerializer(poinvObj ,  many = False).data
        return Response(data,status = status.HTTP_200_OK)
    def get(self,request , format= None):
        if 'id' in request.GET:
            purObj = InvoiceReceived.objects.get(pk = int(request.GET['id']))
            data = InvoiceReceivedAllSerializer(purObj ,  many = False).data
            children = InvoiceReceivedAllSerializer(InvoiceReceived.objects.filter(parent = purObj).order_by('-created') ,  many = True).data
            data['children'] = children
        return Response(data ,status = status.HTTP_200_OK)


class ExpenseInvoiceAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny ,)
    def get(self,request , format= None):
        print request.GET['id'],'ioiooio'
        purchaseObj = InvoiceReceived.objects.filter(account = request.GET['id'])
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment;filename="PettyCashAccounts.pdf"'
        purchaserOrderInv(response,purchaseObj,request)
        return response


class UploadTransactionsAPIView(APIView):
    permission_classes = (permissions.IsAuthenticated , isAdmin)
    def post(self, request, format=None):
        wb = load_workbook(filename = BytesIO(request.FILES['file'].read()))
        ws = wb.worksheets[0]
        row_count = ws.max_row+1
        column_count = ws.max_column
        count = 0
        if 'account' in request.data:
            accObj = Account.objects.get(pk = int(request.data['account']))
        for i in range(2, row_count):
            dated = ws['B' + str(i)].value
            narration = ws['C' + str(i)].value
            extRef = ws['D' + str(i)].value
            debit = ws['E' + str(i)].value
            credit = ws['F' + str(i)].value
            balance = ws['G' + str(i)].value
            dates = datetime.strptime(dated, '%d %b %Y').strftime('%Y-%m-%d')
            transObj = Transaction.objects.create(dated = dates , narration = narration , externalRecord = True , externalReferenceID=extRef ,  user=request.user)
            if isinstance(debit, (int, long, float, complex)):
                transObj.debit = float(debit)
                transObj.fromAcc = accObj
                transObj.type = 'debit'
            if isinstance(credit, (int, long, float, complex)):
                transObj.credit = float(credit)
                transObj.toAcc = accObj
                transObj.type = 'credit'
            transObj.balance = float(balance)
            transObj.save()
        return Response({} ,status = status.HTTP_200_OK)

class GetExpensesDataAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    def get(self, request, format=None):
        unclaimed = 0
        claimed = 0
        approved = 0
        expenseObj = ExpenseSheet.objects.filter(user = request.user)
        invObj = Expense.objects.filter(user = request.user)
        unclaimedObj = invObj.filter(sheet__isnull = True)
        unclaimedTot = unclaimedObj.aggregate(tot = Sum('amount'))
        approvedTot = invObj.filter(sheet__stage = 'approved').aggregate(tot = Sum('amount'))
        claimedTot = invObj.filter(sheet__isnull = False).exclude(sheet__stage = 'approved').aggregate(tot = Sum('amount'))
        if unclaimedTot['tot']!=None:
            unclaimed = unclaimedTot['tot']
        if claimedTot['tot']!=None:
            claimed = claimedTot['tot']
        if approvedTot['tot']!=None:
            approved = approvedTot['tot']
        data = {'unclaimed' : unclaimed, 'claimed' : claimed, 'approved' : approved,'unclaimedObj':ExpenseLiteSerializer(unclaimedObj,many=True).data,'expenseObj':ExpenseSheetLiteSerializer(expenseObj,many=True).data}
        return Response(data ,status = status.HTTP_200_OK)

class GetAllTourAPI(APIView):
    renderer_classes = (JSONRenderer,)
    def get(self, request, format=None):
        tourObj = TourPlanStop.objects.filter(billed = False , contact__serviceName = request.GET['serviceFor'], status = 'completed')
        data = []
        for i in tourObj:
            price = 0
            contractprice = 0
            total = 0
            if i.expense is not None or i.contract is not None:
                name = 'Ticket for ' + i.contact.name + ' on ' +str(i.tourplan.date)
                if i.expense is not None:
                    tempprice = i.expense.invoices.all().aggregate(sum=Sum('amount'))
                    price = tempprice['sum'] if tempprice['sum'] else 0
                if i.contract is not None:
                    contractprice = i.contract.grandTotal
                total = price + contractprice
                data.append({'selected':False,'id' : i.pk,'product' : name , 'price' :total , 'qty' : 1, 'total' : total  })
        return Response(data ,status = status.HTTP_200_OK)


from reportlab.platypus import SimpleDocTemplate, Image , Spacer
from reportlab.lib import utils
def addPageNumbertoBrochure(self,doc):
    print doc.user,'lfkdfl'
    if doc.user.designation.division.logo:
        logoImg = str(doc.user.designation.division.logo)
        logoImg = os.path.join(globalSettings.BASE_DIR,'media_root',logoImg )
        limg = utils.ImageReader(logoImg)
        limg.hAlign ="CENTER"
        iw, ih = limg.getSize()
        aspect = ih / float(iw)
        width = 1.25*inch
        he = 1*inch
        self.drawImage(limg,15*mm,10*inch ,width=width, height=(he * aspect),mask='auto')

    if self._pageNumber == 1:
        bgImg1 = svg2rlg(os.path.join(globalSettings.BASE_DIR,'static_shared', 'images', 'CatalogFront.svg'))
        sx=sy=1.05
        bgImg1.width, bgImg1.height = bgImg1.minWidth()*sx, bgImg1.height*sy
        bgImg1.scale(sx,sy)
        renderPDF.draw(bgImg1, self, 0*mm,0*inch)

    p = Paragraph("<para fontSize=9  alignment='center'><b>%s</b><br/></para>"%(doc.user.designation.division.name),styles['Normal'])
    p.wrapOn(self ,75*mm,10.25*inch)
    p.drawOn(self ,75*mm,10.25*inch)


class PageNumCanvas(canvas.Canvas):

    #----------------------------------------------------------------------
    def __init__(self, *args, **kwargs):
        """Constructor"""
        canvas.Canvas.__init__(self, *args, **kwargs)
        self.pages = []
    #----------------------------------------------------------------------
    def showPage(self):
        """
        On a page break, add information to the list
        """
        page_count = len(self.pages)

        self.pages.append(dict(self.__dict__))
        self._startPage()

    #----------------------------------------------------------------------
    def save(self):
        """
        Add the page number to each page (page x of y)
        """
        page_count = len(self.pages)
        for page in self.pages:
            self.__dict__.update(page)
            self.draw_page_number(page_count)
            self.drawLetterHeadFooter()
            canvas.Canvas.showPage(self)

        canvas.Canvas.save(self)


    #----------------------------------------------------------------------
    def draw_page_number(self, page_count):

        """
        Add the page number
        """

        if self._pageNumber == 2:
            bgImg2 = svg2rlg(os.path.join(globalSettings.BASE_DIR,'static_shared', 'images', 'CatalogList.svg'))
            sx=sy=1.05
            bgImg2.width, bgImg2.height = bgImg2.minWidth()*sx, bgImg2.height*sy
            bgImg2.scale(sx,sy)
            renderPDF.draw(bgImg2, self, 0*mm,0*inch)

        if self._pageNumber == 3:
            bgImg2 = svg2rlg(os.path.join(globalSettings.BASE_DIR,'static_shared', 'images', 'CatalogBack.svg'))
            sx=sy=1.05
            bgImg2.width, bgImg2.height = bgImg2.minWidth()*sx, bgImg2.height*sy
            bgImg2.scale(sx,sy)
            renderPDF.draw(bgImg2, self, 0*mm,0*inch)

        p = Paragraph("<para fontSize=8  alignment='right'><b>Page %d of %d</b></para>"%(self._pageNumber,page_count),styles['Normal'])
        p.wrapOn(self , 50*mm , 10*mm)
        p.drawOn(self , 158*mm  , 5*mm)


    def drawLetterHeadFooter(self):
        p = Paragraph("<para fontSize=8  alignment='center'><b>We Thank You for Your Business.</b></para>",styles['Normal'])
        p.wrapOn(self , 50*mm , 10*mm)
        p.drawOn(self , 15*mm  , 5*mm)

#----------------------------------------------------------------------

stylesN = styles['Normal']
stylesH = styles['Heading1']
from reportlab.platypus import SimpleDocTemplate, Image , Spacer
from reportlab.platypus import PageBreak, SimpleDocTemplate, Table, TableStyle
def getFourProductArray(variants):
    imageData = []
    if len(variants) > 0:
        for variant in variants:
            if variant.img1 :
                imgg = str(variant.img1)
                mediaImg = os.path.join(globalSettings.MEDIA_ROOT,imgg)

                URL = globalSettings.SITE_ADDRESS+'/api/finance/makeImageTransparent/'
                fileData = {'file':mediaImg}
                r = requests.post(url = URL, data = fileData)
                data = r.json()
                convertedImg = os.path.join(globalSettings.MEDIA_ROOT,data['url'])

                imgData = open(convertedImg, 'rb')
                img = Image(convertedImg)
                ratio = img.drawHeight / img.drawWidth
                img.drawHeight = 1.25*inch
                img.drawWidth = 1.25*inch
            else:
                mediaImg = os.path.join(globalSettings.BASE_DIR,'static_shared','images', "no_tour_image.jpg")
                imgData = open(mediaImg, 'rb')
                img = Image(mediaImg)
                ratio = img.drawHeight / img.drawWidth
                img.drawHeight = 1.25*inch
                img.drawWidth = 1.25*inch
            productname = Paragraph('<para spaceBefore = 10 >%s </para>'%(variant.name),stylesN)
            proprice =  Paragraph('<para spaceBefore=5>Price : %0.2f  <strike> %0.2f </strike></para>'%(variant.rate,variant.buyingPrice),stylesN)
            moq =  Paragraph('<para spaceBefore = 5>Quantity Added : %s  </para>'%(variant.qtyAdded),stylesN)
            imageData.append([img,productname,proprice,moq])
    return imageData

from reportlab.platypus import BaseDocTemplate, Frame, Paragraph, NextPageTemplate, PageBreak, PageTemplate

def proCatalog(response, request):
    styles = getSampleStyleSheet()
    stylesN = styles['Normal']
    stylesH = styles['Heading1']
    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(response,pagesize=letter, topMargin=2.5*cm,leftMargin=1*cm,rightMargin=1*cm)
    doc.request = request
    elements = []
    products = []
    doc.user = request.user
    doc.story = []

    inventoryObj = Inventory.objects.filter(division = request.user.designation.division.pk).order_by('pk')
    if 'id' in request.GET:
        inventoryObj = inventoryObj.filter(category__pk = int(request.GET['id']))
    n=4
    if len(inventoryObj) >0 :
        x = [inventoryObj[i:i + n] for i in range(0, len(inventoryObj), n)]
        for idx,i in enumerate(x):
            dataSet = getFourProductArray(x[idx])
            products.append(dataSet)
        dataTabl = Table(products,spaceBefore=15,hAlign="LEFT",spaceAfter=30)
        dataTabl.setStyle(TableStyle([('FONTSIZE', (0, 0), (-1, -1), 8),('VALIGN',(0,0),(-1,-1),'TOP'),('AlIGN',(0, 0), (-1, -1),'LEFT'),('TOPPADDING', (0, 0), (-1, -1),30),('BOTTOMPADDING', (0, 0), (-1, -1),10)]))
        elements.append(dataTabl)
        products= []

    data1=[['']]
    t1=Table(data1)
    elements.append(t1)

    elements.append(PageBreak())

    data2=[['']]
    t2=Table(data2)
    elements.append(t2)

    elements.append(PageBreak())

    data2=[['']]
    t2=Table(data2)
    elements.append(t2)

    elements.append(PageBreak())


    doc.build(elements,onFirstPage= addPageNumbertoBrochure,canvasmaker=PageNumCanvas)


class ProductsCatalogAPI(APIView):
    renderer_classes = (JSONRenderer,)
    def get(self, request, format=None):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="Catalog.pdf"'

        proCatalog(response, request)
        f = open(os.path.join(globalSettings.BASE_DIR, 'media_root/Catalog.pdf'), 'wb')
        f.write(response.content)
        f.close()

        return response


class TransparentImageAPI(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny ,)

    def post(self,request , format= None):
        data = request.data

        print data , 'dddddddddddddddd'

        BLUR = 1
        CANNY_THRESH_1 = 10
        CANNY_THRESH_2 = 200
        MASK_DILATE_ITER = 10
        MASK_ERODE_ITER = 10
        MASK_COLOR = (0.0,0.0,0.0) # In BGR format

        img = cv2.imread(data['file'])
        gray = cv2.cvtColor(img,cv2.COLOR_BGR2GRAY)

        edges = cv2.Canny(gray, CANNY_THRESH_1, CANNY_THRESH_2)
        edges = cv2.dilate(edges, None)
        edges = cv2.erode(edges, None)

        contour_info = []
        contours, _ = cv2.findContours(edges, cv2.RETR_LIST, cv2.CHAIN_APPROX_NONE)
        # Previously, for a previous version of cv2, this line was:
        #  contours, _ = cv2.findContours(edges, cv2.RETR_LIST, cv2.CHAIN_APPROX_NONE)
        # Thanks to notes from commenters, I've updated the code but left this note
        for c in contours:
            contour_info.append((
                c,
                cv2.isContourConvex(c),
                cv2.contourArea(c),
            ))
        contour_info = sorted(contour_info, key=lambda c: c[2], reverse=True)
        max_contour = contour_info[0]

        mask = np.zeros(edges.shape)
        cv2.fillConvexPoly(mask, max_contour[0], (255))

        #-- Smooth mask, then blur it --------------------------------------------------------
        mask = cv2.dilate(mask, None, iterations=MASK_DILATE_ITER)
        mask = cv2.erode(mask, None, iterations=MASK_ERODE_ITER)
        mask = cv2.GaussianBlur(mask, (BLUR, BLUR), 0)
        mask_stack = np.dstack([mask]*3)    # Create 3-channel alpha mask

        #-- Blend masked img into MASK_COLOR background --------------------------------------
        mask_stack  = mask_stack.astype('float32') / 255.0          # Use float matrices,
        img         = img.astype('float32') / 255.0                 #  for easy blending
        masked = (mask_stack * img) + ((1-mask_stack)*(MASK_COLOR)) # Blend
        masked = (masked * 255).astype('uint8')                     # Convert back to 8-bit
        tmp = cv2.cvtColor(masked, cv2.COLOR_BGR2GRAY)
        _,alpha = cv2.threshold(tmp,0,255,cv2.THRESH_BINARY)
        b, g, r = cv2.split(masked)
        rgba = [b,g,r, alpha]
        dst = cv2.merge(rgba,4)
        # cv2.imwrite("media_root/finance/inventory/%s" %(data['file']), dst)
        cv2.imwrite(data['file'], dst)
        httpUrl = data['file']

        return Response({'url':httpUrl} ,status = status.HTTP_200_OK)


class DisbursalViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    serializer_class = DisbursalSerializer
    # queryset = Disbursal.objects.all()
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['disbursed']
    def get_queryset(self):
        divsn = self.request.user.designation.division
        return Disbursal.objects.filter(division = divsn)

class DisbursalliteViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    serializer_class = DisbursalLiteSerializer
    # queryset = Disbursal.objects.all()
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['source' , 'sourcePk']
    def get_queryset(self):
        divsn = self.request.user.designation.division
        return Disbursal.objects.filter(division = divsn)


class InvoiceReceivedViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated, )
    serializer_class = InvoiceReceivedSerializer
    # filter_backends = [DjangoFilterBackend]
    # filter_fields = ['title','group','heading','personal']
    def get_queryset(self):
        divsn = self.request.user.designation.division
        return InvoiceReceived.objects.filter(division = divsn)

class InvoiceReceivedAllViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated, )
    serializer_class = InvoiceReceivedAllSerializer
    # filter_backends = [DjangoFilterBackend]
    # filter_fields = ['title','group','heading','personal']
    def get_queryset(self):
        divsn = self.request.user.designation.division
        return InvoiceReceived.objects.filter(division = divsn)

class InvoiceQtyViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated, )
    serializer_class = InvoiceQtySerializer
    queryset = InvoiceQty.objects.all()


class SaveInvoiceReceived(APIView):
    renderer_classes = (JSONRenderer,)
    def post(self, request, format=None):
        data = request.data
        comp = service.objects.get(pk = int(data['companyReference']))
        dataSave = {'companyReference' : comp , 'personName' : data['personName'] , 'address' : data['address'] , 'state' : data['state'] , 'city' : data['city'] ,'country' : data['country'] , 'pincode' : data['pincode'] , 'gstIn' : data['gstIn'] , 'bankName' : data['bankName'],'ifsc' : data['ifsc'],'accNo' : data['accNo'],'note' :data['note'],'personName' : data['personName'] , 'companyName' : data['companyName'],'invNo' : data['invNo'], 'phone' : data['phone']}
        if 'id' in data:
            obj = InvoiceReceived.objects.get(pk = int(data['id']))
            obj.__dict__.update(dataSave)
        else:
            obj = InvoiceReceived.objects.create(**dataSave)
            obj.user = request.user
            obj.division = request.user.designation.division
        if 'account' in data:
            obj.account = Account.objects.get(pk = int(data['account']))
        if 'costcenter' in data:
            obj.costcenter = CostCenter.objects.get(pk = int(data['costcenter']))
        if 'deliveryDate'  in data:
            obj.deliveryDate = data['deliveryDate']
        if 'paymentDueDate'  in data:
            obj.paymentDueDate = data['paymentDueDate']
        total = 0
        for i in data['products']:
            proddataSave = {'product' : i['product'] , 'price' : i['price'] , 'receivedQty' : i['receivedQty'] , 'taxCode' : i['taxCode'] , 'taxPer' : i['taxPer'] , 'tax' : i['tax'] ,'total' : i['total'] , 'invoice' : obj }
            if 'pk' in i:
                prodObj = InvoiceQty.objects.get(pk = int(i['pk']))
                prodObj.__dict__.update(proddataSave)
            else:
                prodObj = InvoiceQty.objects.create(**proddataSave)
            total +=float(i['total'])
        balance = total - obj.paidAmount
        obj.totalAmount = total
        obj.balance = balance
        obj.save()
        toRet = InvoiceReceivedAllSerializer(obj, many=False).data
        return Response(toRet ,status = status.HTTP_200_OK)


class CreateExpenseTransactionAPI(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny ,)
    def post(self,request , format= None):
        data = request.data
        div = request.user.designation.division
        from datetime import date
        today = date.today()
        purchaseObj = InvoiceReceived.objects.get(pk = int(data['purchase']))
        transObj = Disbursal.objects.create(sourcePk = data['purchase'] , amount = data['amount'] , date = today , source = 'INVOICE', division = div)
        transObj.accountNumber = purchaseObj.accNo
        transObj.ifscCode = purchaseObj.ifsc
        transObj.bankName = purchaseObj.bankName
        transObj.narration = 'Expenses Invoice ' + str(round(float(data['amount']))) +' Rs , ' + str(transObj.date.strftime("%B")) + '-' + str(transObj.date.year)
        transObj.save()
        purchaseObj.paidAmount = float(purchaseObj.paidAmount) + float(data['amount'])
        purchaseObj.balanceAmount = float(purchaseObj.totalAmount) - float(purchaseObj.paidAmount)
        purchaseObj.save()
        data = DisbursalLiteSerializer(transObj,many=False).data
        purchaseObjData = InvoiceReceivedAllSerializer(purchaseObj, many=False).data
        return Response({'purchase':purchaseObjData , 'data': data})


class UpdateTotalAPI(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny ,)
    def get(self,request , format= None):
        inv = InvoiceReceived.objects.get(pk = int(request.GET['id']))
        total =inv.parentInvoice.aggregate(tot = Sum('total'))
        totalAmount = 0
        if total['tot'] is not None:
            totalAmount = total['tot']
        inv.totalAmount = totalAmount
        inv.balanceAmount = totalAmount - inv.paidAmount
        inv.save()

        return Response({})
