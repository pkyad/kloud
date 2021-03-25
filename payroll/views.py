# -*- coding: utf-8 -*-
from django.contrib.auth.models import User , Group
from django.shortcuts import render, redirect , get_object_or_404
from django.contrib.auth import authenticate , login , logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.urlresolvers import reverse
from django.template import RequestContext
from django.conf import settings as globalSettings
from django.core.exceptions import ObjectDoesNotExist , SuspiciousOperation
from django.views.decorators.csrf import csrf_exempt, csrf_protect
# Related to the REST Framework
from rest_framework import viewsets , permissions , serializers
from rest_framework.exceptions import *
from url_filter.integrations.drf import DjangoFilterBackend
from .serializers import *
from API.permissions import *
from ERP.models import *
from organization.models import CompanyHolidays
from ERP.views import getApps
from django.db.models import Q, Sum
from django.http import JsonResponse
import random, string
from django.utils import timezone
from rest_framework.views import APIView
# Create your views here.
from reportlab import *
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4, A5
from reportlab.lib.units import cm, mm
from reportlab.lib import colors , utils
from reportlab.platypus import Paragraph, Table, TableStyle, Image, Frame, Spacer, PageBreak, BaseDocTemplate, PageTemplate, SimpleDocTemplate, Flowable
from PIL import Image
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet, TA_CENTER
from reportlab.graphics import barcode , renderPDF
from reportlab.graphics.shapes import *
from reportlab.graphics.barcode.qr import QrCodeWidget
from rest_framework.views import APIView
from rest_framework.renderers import JSONRenderer
from django.http import HttpResponse
import datetime
import json
import pytz
import requests
from django.template.loader import render_to_string, get_template
from django.core.mail import send_mail, EmailMessage
from reportlab.lib.pagesizes import letter,A5,A4,A3
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import ParagraphStyle,getSampleStyleSheet
from reportlab.lib.enums import TA_LEFT, TA_CENTER
from reportlab.lib.colors import *
from reportlab.lib.units import inch, cm
import calendar
# import datetime
from forex_python.converter import CurrencyCodes
from HR.models import Leave,profile
from django.contrib.auth.models import User
from finance.models import ExpenseSheet
from rest_framework.response import Response
from excel_response import ExcelResponse
from django.core.mail import send_mail, EmailMessage
from django.core.mail import send_mass_mail
from email.mime.application import MIMEApplication
from openpyxl import load_workbook,Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill , Font, Alignment
from marketing.models import TourPlan
from dateutil.relativedelta import relativedelta
from itDeclarationCalc import *


class payrollViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    queryset = payroll.objects.all()
    serializer_class = payrollSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['user']

class PayrollLogsViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    serializer_class = PayrollLogsSerializer
    queryset = PayrollLogs.objects.all()


class payslipViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    queryset = Payslip.objects.all()
    serializer_class = payslipSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['month','year','user']

class payrollReportViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    # queryset = PayrollReport.objects.all()
    serializer_class = payrollReportSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['month','year','status','user']
    def get_queryset(self):
        toRet =  PayrollReport.objects.filter(division = self.request.user.designation.division)
        return toRet

class advancesViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    # queryset = Advances.objects.all()
    serializer_class = advancesSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['user','settled']
    def get_queryset(self):
        toRet =  Advances.objects.filter(division = self.request.user.designation.division)
        return toRet



#code for pdf
def payslip(response ,paySlip,userObj,report,month, year, request):
    months = ["","January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
    monthsList = ['', 'Jan' , 'Feb' , 'March' ,'April' , 'May' , 'June' , 'July' , 'Aug' , 'Sept' , 'Oct' , 'Nov' , 'Dec' ]
    now = datetime.datetime.now()
    monthdays=calendar.monthrange(int(year), int(month))[1]
    if month>3:
        financialYear = str(year)+ '-' +str(year+1)
    else:
        financialYear = str(year-1)+ '-' + str(year)
    tourplanObj = TourPlan.objects.filter(user_id= int(userObj.pk),status='approved',date__year=year,date__month = month)
    payslipObj = Payslip.objects.filter(user_id= int(userObj.pk),month =month,year=year)
    dt=str(year)+'-'+str(month).zfill(2)
    currentMonth = monthsList[month]
    mLeavesObj= Leave.objects.filter(Q(fromDate__contains=dt) | Q(toDate__contains=dt),user_id= int(userObj.pk),status='approved')
    extraLeavesObj= Leave.objects.filter(user_id= int(userObj.pk),status='approved',leavesCount__gte=4)
    leaveObj = mLeavesObj | extraLeavesObj
    holidaysList = list(CompanyHolidays.objects.all().values_list('date',flat=True))
    satOff = userObj.payroll.off
    sleave = 0
    aleave = 0
    cleave = 0
    for leave in leaveObj:
        daysDiff = (leave.toDate-leave.fromDate).days
        if daysDiff>0:
            for j in range(daysDiff+1):
                sDate = leave.fromDate+relativedelta(days=j)
                if sDate.month == int(month):
                    wDay = sDate.weekday()
                    if wDay<6 and sDate in holidaysList:
                        if wDay==5 and satOff:
                            continue
                        if leave.category == 'ML':
                            sleave += 1
                        elif leave.category == 'AL':
                            aleave += 1
                        elif leave.category == 'casual':
                            cleave += 1
        else:
            if leave.category == 'ML':
                sleave += leave.leavesCount
            elif leave.category == 'AL':
                aleave += leave.leavesCount
            elif leave.category == 'casual':
                cleave += leave.leavesCount
    totalLeaves = sleave + aleave + cleave
    daysPresent = monthdays - totalLeaves
    earned = 0
    deduct = 0
    totalpay = 0
    print payslipObj,'objjjjjjjjjjjjjjj'
    ta = 0
    da = 0
    for value in tourplanObj:
        ta += value.ta
        da += value.da
        print ta,'---',da
    currencyType='INR'
    s=CurrencyCodes().get_symbol(currencyType) # currencysymbol
    if currencyType == 'INR':
        s='Rs.'
    absent=0
    paidHolidays=0
    balanceCL = 0
    balanceSL = 0
    balanceCO = 0
    try:
        dept_name = userObj.designation.department.dept_name
    except:
        dept_name = ''
    try:
        role = userObj.designation.role.name
    except:
        role = ''
    accountNumber=paySlip.accountNumber
    if userObj.payroll.pfAccNo is not None:
        pfNo = userObj.payroll.pfAccNo
    else:
        pfNo = ''
    if userObj.payroll.esic is not None:
        escisNo = userObj.payroll.esic
    else:
        escisNo = ''
    if userObj.payroll.pan is not None:
        pan = userObj.payroll.pan
    else:
        pan = ''
    if userObj.payroll.basic is not None:
        sbs = userObj.payroll.basic
    else:
        sbs = 0
    (empCode,name,location,department,grade,designation)=(userObj.profile.empID ,userObj.first_name+' '+userObj.last_name,'Bangalore',dept_name,'E.1',role)
    if userObj.payroll.al!=None:
        balanceCL = userObj.payroll.al
    if userObj.payroll.ml!=None:
        balanceSL = userObj.payroll.ml
    if userObj.payroll.adHocLeaves!=None:
        balanceCO = userObj.payroll.adHocLeaves
    (days,ml,al,cl,adHocLeaves)=(daysPresent,sleave,aleave,cleave,0)
    med = 0

    print financialYear, currentMonth
    incomeObj = ITDecaration.objects.filter(user = userObj, year = financialYear, group_name = 'income', month = currentMonth )
    try:
        basic = incomeObj.get( title = 'Basic').amount
    except:
        basic = 0
    try:
        hra = incomeObj.get(title = 'HRA').amount
    except:
        hra = 0
    try:
        special = incomeObj.get(title = 'Special Allownace').amount
    except:
        special = 0
    try:
        lta = incomeObj.get( title = 'LTA').amount
    except:
        lta = 0
    try:
        adHoc = incomeObj.get( title = 'Fixed Variable').amount
    except:
        adHoc = 0
    try:
        bonus =  incomeObj.get( title = 'Statutory Bonus').amount
    except:
        bonus = 0
    try:
        pfAmnt =  incomeObj.get( title = 'PF (Employee)').amount
    except:
        pfAmnt = 0
    try:
        pfAdmin =  incomeObj.get( title = 'PF (Company)').amount
    except:
        pfAdmin = 0
    payObj = Payslip.objects.get(user__id=userObj.pk,month=month,year=year)
    styles = getSampleStyleSheet()
    styledict={'center':ParagraphStyle(name='center', parent=styles['Normal'], alignment=TA_CENTER, fontSize=10)}
    doc = SimpleDocTemplate(response,pagesize=A5, topMargin=1*cm,leftMargin=1*cm, rightMargin=1*cm,
    bottomMargin=1*cm)
    doc.request = request
    # container for the 'Flowable' objects
    elements = []
    if userObj.profile.empType == "full time":
        emptax = "Professional Tax Deduction"
    else:
        emptax = "Tax Deduction"
    user = request.user
    companyName = ''
    companyAddress = ''
    website = ''
    if user.designation is not None:
        if user.designation.unit is not None:
            companyName = user.designation.unit.name
            companyAddress = user.designation.unit.address + ' ' + user.designation.unit.city + ' ' +  user.designation.unit.state + ' ' + user.designation.unit.country + ' - ' +   user.designation.unit.pincode
        if user.designation.division is not None:
            website = user.designation.division.website
    from reportlab.platypus import Image
    imagePath = os.path.join(globalSettings.MEDIA_ROOT , str(user.designation.division.logo))
    try:
        f = open(imagePath, 'rb')
        ima = Image(f)
        ima.drawHeight = 0.5*inch
        ima.drawWidth = 0.5*inch
        ima.hAlign = 'CENTER'
    except:
        ima = ''

    totalEarnings = basic + hra + special + lta  + adHoc + bonus + payObj.reimbursement
    deductions = pfAmnt + payObj.miscellaneous + payObj.tds
    total = totalEarnings - deductions
    a=[ima,Paragraph(companyName +'<br/><br/>',styledict['center']),
       Paragraph(companyAddress+'<br/>',styledict['center']),
        Paragraph('<strong>'+ website +' </strong><br/>' ,styledict['center']),

       Paragraph("<para fontSize=8 alignment='center'><strong>Employee PaySlip For Month Of {0} {1} </strong></para>".format(months[month],year),styles['Normal'])
       ]
    p1=Paragraph("<para fontSize=8><strong>Bank Details : </strong>Salary Has Been Credited To "+str(userObj.payroll.accountNumber)+' '+str(userObj.payroll.bankName)+ "</para>",styles['Normal'])
    data1 = [[a],['']]
    data=[['Emp Code : %s'%(empCode)],['Name : %s'%(name)], ['Location : %s'%(location),'Department :%s'%(department)],['Grade : %s'%(grade),'Designation : %s'%(designation)],['PF No : %s'%(pfNo),'ESIC No : %s'%(escisNo)], ['PAN : %s'%(pan),'Standard Basic Salary : %s %d'%(s,sbs)],['Days Paid : %d'%(daysPresent),'Days Present : %d'%(daysPresent)],['Paid Holidays : %d'%(paidHolidays),'Lwp/Absent : %d'%(absent)],['Sick Leaves : %d'%(ml),'Annual Leaves : %d'%(al)],['Compensatory Leaves : %d'%(cl),'AdHoc Leaves : %d'%(adHocLeaves)],['Balance SL : %d'%(balanceSL),'Balance CL : %d'%(balanceCL)],['Balance CO : %d'%(balanceCO),''],['Earnings Amount '],['Basic Salary + DA' ,s+' '+str(basic)], ['HRA',s+' '+str(hra)],['Special Allowances',s+' '+str(special)],['LTA',s+' '+str(lta)],['Statutory Bonus',s+' '+str(int(bonus))],['Variable Incentives',s+' '+str(payObj.reimbursement)], ['Fixed Variable',s+' '+str(adHoc)],
    ['Total Earnings ',s+' '+str(int(totalEarnings))], ['Deductions Amount'],['Provident Fund',s+' '+str(int(pfAmnt))],['Miscellaneous Deductions',s+' '+str(int(payObj.miscellaneous))],['Advance Deduction',s+' '+str(float(payObj.advanceDeduction))],['Total Deduction',s+' '+str(int(deductions))],['Net Pay',s+' '+str(round(total,2))],[p1]]

    lines=[('LINEBELOW',(0,5),(-1,5),0.5,black),
           # ('LINEBELOW',(0,4),(-1,4),0.5,black),
           # ('LINEBELOW',(0,7),(-1,7),0.5,black),
           ('LINEBELOW',(0,11),(-1,11),0.5,black),
           ('LINEBELOW',(0,12),(-1,12),0.5,black),
           ('LINEBELOW',(0,19),(-1,19),0.5,black),
           ('LINEBELOW',(0,20),(-1,20),0.5,black),
           ('LINEBELOW',(0,21),(-1,21),0.5,black),
           ('LINEBELOW',(0,24),(-1,24),0.5,black),
           ('LINEBELOW',(0,25),(-1,25),0.5,black),
           ('LINEBELOW',(0,26),(-1,26),0.5,black),

           ]
    spans=[('SPAN',(0,0),(-1,1)),('SPAN',(1,2),(-1,2)),('SPAN',(0,-1),(-1,-1)),('SPAN',(0,12),(-1,12)),('SPAN',(0,21),(-1,21))]
    aligns = [('ALIGN',(0,12),(-1,12),'CENTER'),('ALIGN',(0,21),(-1,21),'CENTER')]
    tmain=Table(data1)
    tmain.setStyle(TableStyle([('BOX', (0,0), (-1,-1), 0.75, black)]))
    t1=Table(data , rowHeights=0.50*cm)
    t1.setStyle(TableStyle([('BOX', (0,0), (-1,-1), 0.75, black),('FONTSIZE', (0,0), (-1,-1),8)]+lines+spans+aligns))
    elements.append(tmain)
    elements.append(t1)
    doc.build(elements)


class GetPayslip(APIView):
    def get(self , request , format = None):
        months = ["","January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
        if 'payslip' in request.GET:
            payrol = payroll.objects.get(user = request.GET['payslip'])
            user = User.objects.get(id = request.GET['payslip'])
            report = PayrollReport.objects.get(id = request.GET['report'])
            month = report.month
            year = report.year
        if 'userid' in request.GET:
            user = User.objects.get(id = request.GET['userid'])
            payrol = payroll.objects.get(user = request.GET['userid'])
            tempmonth = request.GET['month']
            month = months.index(tempmonth)
            year = request.GET['year']
            report = []
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = 'attachment;filename="payslip.pdf"'
            payslip(response , payrol , user ,report,month, year, request)

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment;filename="payslipdownload.pdf"'
        payslip(response , payrol , user ,report,month, year, request)
        if 'output' in request.GET:
            filePath = os.path.join(globalSettings.BASE_DIR, 'media_root/payslip%s.pdf' %
            ( datetime.datetime.now(pytz.timezone('Asia/Kolkata')).year))
            f = open(filePath, 'wrb')
            f.write(response.content)
            f.close()
            file_name = 'media/' + filePath.split('/')[-1]
            return Response({'fileUrl' : file_name }, status = status.HTTP_200_OK)
        return response

months = ["","January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
def convertmonth(mon):
    for i in months:
        if mon in months:
            return months.index(mon)


def paysMonthlyslip(response ,paySlip,userObj,month,year,payrollReport, request):
    settingsFields = application.objects.get(name = 'app.clientRelationships').settings.all()
    now = datetime.datetime.now()
    monthint = convertmonth(month)
    monthdays=calendar.monthrange(int(year), int(monthint))[1]
    tourplanObj = TourPlan.objects.filter(user_id= int(userObj.pk),status='approved',date__year=year,date__month=monthint)
    payslipObj = Payslip.objects.filter(user_id= int(userObj.pk),month =monthint,year=year)
    dt=str(year)+'-'+str(monthint).zfill(2)
    mLeavesObj= Leave.objects.filter(Q(fromDate__contains=dt) | Q(toDate__contains=dt),user_id= int(userObj.pk),status='approved')
    extraLeavesObj= Leave.objects.filter(user_id= int(userObj.pk),status='approved',leavesCount__gte=4)
    leaveObj = mLeavesObj | extraLeavesObj
    holidaysList = list(CompanyHolidays.objects.all().values_list('date',flat=True))
    satOff = userObj.payroll.off
    sleave = 0
    aleave = 0
    cleave = 0
    for leave in leaveObj:
        daysDiff = (leave.toDate-leave.fromDate).days
        if daysDiff>0:
            for j in range(daysDiff+1):
                sDate = leave.fromDate+relativedelta(days=j)
                if sDate.month == int(monthint):
                    wDay = sDate.weekday()
                    if wDay<6 and sDate in holidaysList:
                        if wDay==5 and satOff:
                            continue
                        if leave.category == 'ML':
                            sleave += 1
                        elif leave.category == 'AL':
                            aleave += 1
                        elif leave.category == 'casual':
                            cleave += 1
        else:
            if leave.category == 'ML':
                sleave += leave.leavesCount
            elif leave.category == 'AL':
                aleave += leave.leavesCount
            elif leave.category == 'casual':
                cleave += leave.leavesCount
    totalLeaves = sleave + aleave + cleave
    daysPresent = monthdays - totalLeaves
    earned = 0
    deduct = 0
    totalpay = 0
    for pay in payslipObj:
        earned = pay.amount + pay.taDa + pay.adHoc + pay.pfAmnt
        deduct = pay.tds
        totalpay = pay.totalPayable
    ta = 0
    da = 0
    for value in tourplanObj:
        ta += value.ta
        da += value.da
    currencyType='INR'
    s=CurrencyCodes().get_symbol(currencyType) # currencysymbol
    if currencyType == 'INR':
        s='Rs.'
    absent=0
    paidHolidays=0
    balanceSL  = 0
    balanceCL = 0
    balanceCO = 0
    accountNumber=paySlip.accountNumber
    (empCode,name,location,department,grade,designation,pfNo,escisNo,pan,sbs)=(str(paySlip.joiningDate.year)+str(userObj.pk) ,userObj.first_name+' '+userObj.last_name,'Bangalore',userObj.designation.department.dept_name,'E.1',userObj.designation.role.name,userObj.payroll.pfAccNo,userObj.payroll.esic,userObj.payroll.pan,userObj.payroll.basic)
    (days,ml,al,cl,adHocLeaves)=(daysPresent,sleave,aleave,cleave,0)
    if userObj.payroll.al!=None:
        balanceCL = userObj.payroll.al
    if userObj.payroll.ml!=None:
        balanceSL = userObj.payroll.ml
    if userObj.payroll.adHocLeaves!=None:
        balanceCO = userObj.payroll.adHocLeaves
    taxded = deduct + pay.pfAmnt
    (basic,hra,special,lta,adHoc,pf)=(str(round(userObj.payroll.basic/12.0,2)),str(round(userObj.payroll.hra/12.0,2)),str(round(userObj.payroll.special/12.0,2)),str(round(userObj.payroll.lta/12.0,2)),userObj.payroll.adHoc,userObj.payroll.pfAmnt)
    (spf,iol,od)=(userObj.payroll.pfAmnt,0,0)
    totalEarnings,deductions=(0,0)
    med = 0
    if MonthlyPayslip.objects.filter(user__id=userObj.pk,month=monthint,year=year).count()>0:
        payObj = MonthlyPayslip.objects.get(user__id=userObj.pk,month=monthint,year=year)
    else:
        payData = {
        'user': userObj,
        'month': monthint,
        'year': year,
        'basicSalary': basic,
        'hra': hra,
        'conveyance': special,
        'convRemiburse': lta,
        'medRemiburse': med,
        'empPF': pf,
        'otherEarnings': adHoc,
        'ta': ta,
        'da': da,
        'spFund': spf,
        'ptDeduction': deduct,
        'ioLoan': iol,
        'otherDeductions': od,
        'totalEarnings': earned,
        'totalDeduction': taxded,
        'netpay': totalpay,
        }
        payObj = MonthlyPayslip.objects.create(**payData)
    styles = getSampleStyleSheet()
    styledict={'center':ParagraphStyle(name='center', parent=styles['Normal'], alignment=TA_CENTER, fontSize=10)}
    doc = SimpleDocTemplate(response,pagesize=A3, topMargin=1*cm,)
    doc.request = request
    # container for the 'Flowable' objects
    elements = []
    if userObj.profile.empType == "full time":
        emptax = "Professional Tax Deduction"
    else:
        emptax = "Tax Deduction"
    print payrollReport,'---------------------'
    a=[Paragraph("<para fontSize=25 alignment='Left' textColor=#6375d4><strong>CIOC</strong></para>",styles['Normal']),
       Paragraph(str(settingsFields.get(name = 'companyName').value )+'<br/><br/>',styledict['center']),
       Paragraph(str(settingsFields.get(name = 'companyAddress').value)+'<br/>',styledict['center']),
       Paragraph('<strong>www.cioc.co.in </strong><br/>' ,styledict['center']),
       ]
    if payrollReport:
        if payrollReport.status == "approved":
            a.append(Paragraph("<para fontSize=8 alignment='center'><strong>Employee PaySlip For Month Of {0} {1} </strong></para>".format(month,year),styles['Normal']))
        else:
            a.append(Paragraph("<para fontSize=8 alignment='center'><strong>Provisional Employee PaySlip For Month Of {0} {1} </strong></para>".format(month,year),styles['Normal']))
    else:
            a.append(Paragraph("<para fontSize=8 alignment='center'><strong>Employee PaySlip For Month Of {0} {1} </strong></para>".format(month,year),styles['Normal']))

    p1=Paragraph("<para fontSize=8><strong>Bank Details : </strong>Salary Has Been Credited To "+str(userObj.payroll.accountNumber)+' '+str(userObj.payroll.bankName) + "</para>",styles['Normal'])
    print balanceCL,balanceSL,balanceCO,'qqqqqqqqqqqqqqqqqqqqqq'
    data=[[a,'','',''],['','','',''],
          ['Emp Code : %s'%(empCode),'Name : %s'%(name),'',''],
          ['Location : %s'%(location),'Department :%s'%(department),'Grade : %s'%(grade),'Designation : %s'%(designation)],
          ['PF No : %s'%(pfNo),'ESIC No : %s'%(escisNo),'PAN : %s'%(pan),'Standard Basic Salary : %s %d'%(s,sbs)],
          ['Days Paid : %d'%(daysPresent),'Days Present : %d'%(daysPresent),'Paid Holidays : %d'%(paidHolidays),'Lwp/Absent : %d'%(absent)],
          ['Sick Leaves : %d'%(ml),'Annual Leaves : %d'%(al),'Compensatory Leaves : %d'%(cl),'AdHoc Leaves : %d'%(adHocLeaves)],
          ['Balance SL : %d'%(balanceSL),'Balance CL : %d'%(balanceCL),'Balance CO : %d'%(balanceCO),''],
          ['Earnings','Amount','Deductions','Amount'],
          ['Basic Salary' ,s+' '+str(payObj.basicSalary),'Saturatory Provident Fund',s+' '+str(payObj.spFund)],
          ['HRA',s+' '+str(payObj.hra),str(emptax),s+' '+str(payObj.ptDeduction)],
          ['Conveyance',s+' '+str(payObj.conveyance),'Interest On Loan',s+' '+str(payObj.ioLoan)],
          ['Conveyance Reimbursement',s+' '+str(payObj.convRemiburse),'Other Deduction ',s+' '+str(payObj.otherDeductions)],
          ['Medical Reimbursement',s+' '+str(payObj.medRemiburse),'',''],
          ['Employee Contribution towards PF',s+' '+str(payObj.empPF),'',''],
          ['Other Earnings',s+' '+str(payObj.otherEarnings),'',''],
          ['TA',s+' '+str(payObj.ta),'',''],
          ['DA',s+' '+str(payObj.da),'',''],
          ['Total Earnings ',s+' '+str(payObj.totalEarnings),'Total Deduction',s+' '+str(payObj.totalDeduction)],
          ['','','','Net Pay : %s %d'%(s,float(payObj.netpay))],
          [p1,'','',''],
          ]

    lines=[('LINEBELOW',(0,1),(-1,1),0.5,black),
           ('LINEBELOW',(0,4),(-1,4),0.5,black),
           ('LINEBELOW',(0,7),(-1,7),0.5,black),
           ('LINEBELOW',(0,8),(-1,8),0.5,black),
           ('LINEBELOW',(0,17),(-1,17),0.5,black),
           ('LINEBELOW',(0,18),(-1,18),0.5,black),
           ('LINEBELOW',(0,19),(-1,19),0.5,black),
           ('LINEBEFORE',(2,8),(2,19),0.5,black),]
    spans=[('SPAN',(0,0),(-1,1)),('SPAN',(1,2),(-1,2)),('SPAN',(0,-1),(-1,-1))]
    aligns=[('ALIGN',(1,8),(1,18),'RIGHT'),('ALIGN',(-1,8),(-1,19),'RIGHT')]
    rheights=21*[0.3*inch]
    rheights[1]=0.7*inch
    t1=Table(data,rowHeights=rheights)
    t1.setStyle(TableStyle([('BOX', (0,0), (-1,-1), 0.75, black),]+lines+spans+aligns))
    elements.append(t1)
    doc.build(elements)


class GetPayMonthlyslip(APIView):
    def get(self , request , format = None):
        user = User.objects.get(id = request.GET['userid'])
        payrol = payroll.objects.get(user = request.GET['userid'])
        month = request.GET['month']
        year = request.GET['year']
        print request.GET['userid'], request.GET['month'], request.GET['year']
        report = []
        print month,year
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment;filename="payslip.pdf"'
        paysMonthlyslip(response , payrol , user ,month,year,report, request)
        if 'output' in request.GET:
            filePath = os.path.join(globalSettings.BASE_DIR, 'media_root/payslip%s.pdf' %
                                  ( datetime.datetime.now(pytz.timezone('Asia/Kolkata')).year))
            f = open(filePath, 'wrb')
            f.write(response.content)
            f.close()
            print filePath,'aaaaaaaaaaaa'
            file_name = 'media/' + filePath.split('/')[-1]
            return Response({'fileUrl' : file_name }, status = status.HTTP_200_OK)
        return response


#code for excelSheet
class PayslipsReport(APIView):
    permission_classes = (permissions.IsAuthenticated,)
    def get(self , request , format = None):
        objs = Payslip.objects.filter(report_id = request.GET['report'])
        toReturn = []
        for o in objs:
            toReturn.append({"Employee Name" : o.user.first_name + ' ' + o.user.last_name , "payslipID" : o.pk , "totalPayable" : o.totalPayable,"tds" : o.tds , "accountNumber" : o.user.payroll.accountNumber ,"bankName" : o.user.payroll.bankName , "ifscCode" : o.user.payroll.ifscCode, "pan" : o.user.payroll.pan , "PF Account Number" : o.user.payroll.pfAccNo})
        return ExcelResponse(toReturn)

class TDSslipsReport(APIView):
    permission_classes = (permissions.IsAuthenticated,)
    def get(self , request , format = None):
        objs = Payslip.objects.filter(report_id = request.GET['report'])
        toReturn = []
        for o in objs:
            toReturn.append({"Employee Name" : o.user.first_name + ' ' + o.user.last_name , "payslipID" : o.pk,"total" : o.amount , "tds" : o.tds , "totalPayable" : o.totalPayable, "pan" : o.user.payroll.pan })
        return ExcelResponse(toReturn)

class PFslipsReport(APIView):
    permission_classes = (permissions.IsAuthenticated,)
    def get(self , request , format = None):
        hdFont = Font(size=12,bold=True)
        alphaChars = list(string.ascii_uppercase)
        workbook = Workbook()
        Sheet1 = workbook.active
        Sheet1.title = 'Payroll Report'
        hd = ["Employee Name", "payslipID",'PF Account Number','PF Universal Number','PAN','Bank Name' , 'Bank Account Number' , 'Bank IFSC Number' , 'Amount' , 'Bonus' , 'Penalty' , 'PF' , 'TDS' , 'Take Home' , 'PF (Company)' , 'Months CTC']
        # hdWidth = [80,80,80,80,80,80,80,80,80,80,80,80,80,80,80,80]
        Sheet1.append(hd)
        objs = Payslip.objects.filter(report_id = request.GET['report'])
        toReturn = []
        for o in objs:
            data = [o.user.first_name + ' ' + o.user.last_name ,  o.pk ,  o.user.payroll.pfAccNo ,  o.user.payroll.pfUniNo , o.user.payroll.pan ,  o.user.payroll.bankName,  o.user.payroll.accountNumber, o.user.payroll.ifscCode, o.amount , o.reimbursement ,  o.miscellaneous , o.pfAmnt ,  o.tds ,  o.totalPayable , o.pfAdmin ,  o.grandTotal]
            Sheet1.append(data)
            # for idx,k in enumerate(data):
            #     print hdWidth[idx], len(str(k))+5, hdWidth[idx]
            #     if (len(str(k))+5) > hdWidth[idx]:
            #         hdWidth[idx] = len(str(k)) + 5
        response = HttpResponse(content=save_virtual_workbook(workbook),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=payrollReport.xlsx'
        if 'download' in request.GET:
            filePath = os.path.join(globalSettings.BASE_DIR, 'media_root/payrollReport%s.pdf' %
                                  ( request.GET['report']))
            f = open(filePath, 'wrb')
            f.write(response.content)
            f.close()
            file_name = 'media/' + filePath.split('/')[-1]
            return Response({'fileUrl' : file_name }, status = status.HTTP_200_OK)
        return response

class GetReimbursement(APIView):
    renderer_classes = (JSONRenderer,)
    def get(self , request , format = None):
        expenseObj=ExpenseSheet.objects.filter(user_id=request.GET['user'] , approved = 'Yes')
        amt = 0
        for i in expenseObj:
            for j in i.invoices.all():
                if j.approved:
                    amt += j.amount
        tosend = {'amount':amt}
        return JsonResponse(tosend,status = status.HTTP_200_OK)

# ---------------------------------------------------------------
class GetAllMonths(APIView):
    def get(self , request , format = None):
        reportData = PayrollReport.objects.filter(year=int(request.GET['year']), division = request.user.designation.division)
        print reportData
        toReturn = []
        month_lst = [1,2,3,4,5,6,7,8,9,10,11,12]
        returnData =[]
        for o in reportData:
            toReturn.append(o.month)
        for value in month_lst:
            if value not in toReturn:
                returnData.append(value)

        return Response(returnData,status = status.HTTP_200_OK)


class SendPayslipEmailAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    def get(self, request, format=None):
        print request.GET['month'],'GETTTTT'
        # mn = request.GET['month']
        userObj = request.user
        report = PayrollReport.objects.get(pk = request.GET['report'])
        year = report.year
        months = ["","January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
        monthsList = ['', 'Jan' , 'Feb' , 'March' ,'April' , 'May' , 'June' , 'July' , 'Aug' , 'Sept' , 'Oct' , 'Nov' , 'Dec' ]
        tempmonth = request.GET['month']
        monthIndx = months.index(tempmonth)
        for pay in report.payslips.all():
            print "sending email to " , pay.user.username
            user = User.objects.get(id = pay.user.pk)
            payrol = payroll.objects.get(user = pay.user)
            # month = mn
            # year = yr
            response = './static_shared/document/payslip.pdf'
            contactData =[]
            p = profile.objects.get(user = payslip.user.pk )
            contactData.append(p.email)

            email_subject =str('Payslip')
            companyName = ''
            companyAddress = ''
            website = ''
            if userObj.designation is not None:
                if userObj.designation.unit is not None:
                    companyName = userObj.designation.unit.name
                    companyAddress = userObj.designation.unit.address + ' ' + userObj.designation.unit.city + ' ' +  userObj.designation.unit.state + ' ' + userObj.designation.unit.country + ' - ' +   userObj.designation.unit.pincode
                    mobile = userObj.designation.unit.mobile
                if userObj.designation.division is not None:
                    website = userObj.designation.division.website

            msgBody='Please find the attachment'
            ctx = {
                'message': msgBody,
                'linkUrl': website,
                'linkText' : 'View Online',
                'sendersAddress' : companyName +' ' +companyAddress ,
                'sendersPhone' : '841101',
                'linkedinUrl' : 'https://www.linkedin.com/company/13440221/',
                'fbUrl' : 'facebook.com',
                'twitterUrl' : 'twitter.com',
            }
            email_body = get_template('app.clientRelationships.email.html').render(ctx)
            msg = EmailMessage(email_subject, msgBody,  to= contactData)
            payslip(response , payrol , user ,report,monthIndx, year, request)
            # paysMonthlyslip(response  , payrol , user ,month,year,pr, request)
            fp = open('./static_shared/document/payslip.pdf', 'rb')
            att = MIMEApplication(fp.read(), _subtype="pdf")
            fp.close()
            att.add_header('Content-Disposition', 'attachment',
                           filename=fp.name.split('/')[-1])
            msg.attach(att)
            msg.send()
        return Response({},status = status.HTTP_200_OK)


class GetDisbursalSheetAPI(APIView):
    permission_classes = (permissions.IsAuthenticated ,)
    def get(self , request , format = None):
        print 'entered','*******************'
        if 'pkList' in request.GET:
            print 'requesttttttt',request.GET,request.data
            try:
                pkList = list(eval(request.GET['pkList']))
                disbursalList = Disbursal.objects.filter(pk__in=pkList)
            except:
                disbursalList = Disbursal.objects.filter(pk=int(request.GET['pkList']))
            hdFont = Font(size=12,bold=True)
            alphaChars = list(string.ascii_uppercase)
            workbook = Workbook()
            Sheet1 = workbook.active
            Sheet1.title = 'Disbursal Summary'
            hd = ["Date", "Account Number",'Bank Name','Amount','IFSC Code','Narration']
            hdWidth = [30,30,30,30,30,30]
            Sheet1.append(hd)
            print disbursalList.count(),'-----------fdfdff'
            print disbursalList,'-------disb list'
            for i in disbursalList:
                data = [i.date,i.accountNumber,i.bankName,i.amount,i.ifscCode,i.narration]
                Sheet1.append(data)
                for idx,k in enumerate(data):
                    if (len(str(k))+5) > hdWidth[idx]:
                        hdWidth[idx] = len(str(k)) + 5

            for idx,j in enumerate(hd):
                cl = str(alphaChars[idx])+'1'
                Sheet1[cl].fill = PatternFill(start_color="48dce0", end_color="48dce0", fill_type = "solid")
                Sheet1[cl].font = hdFont
                Sheet1.column_dimensions[str(alphaChars[idx])].width = hdWidth[idx]
            response = HttpResponse(content=save_virtual_workbook(workbook),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=DisbursalSummary.xlsx'
            return response
        else:
            return Response({},status = status.HTTP_200_OK)

class sendDisbursalEmailAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    def post(self, request, format=None):
        print 'Sennnnnnddddinggg-------------mailllllllllllll'
        print request.data['pk'],'----reeqqq'
        advance = Advances.objects.get(pk = request.data['pk'])
        id = advance.user.profile.email
        print advance.typ,'---------typeeeeeeeee'
        email_subject =str('Disbursal Info')
        msgBody='Disbursal Information testing.'
        ctx = {
            'fname': advance.user.first_name,
            'lname': advance.user.last_name,
            'amt' : advance.amount,
            'type' : advance.typ,
            'reason': advance.reason,
            'linkUrl': 'cioc.co.in',
            'linkText' : 'View Online',
            'sendersAddress' : '(C) CIOC FMCG Pvt Ltd',
            'sendersPhone' : '841101',
            'linkedinUrl' : 'https://www.linkedin.com/company/13440221/',
            'fbUrl' : 'facebook.com',
            'twitterUrl' : 'twitter.com',
        }
        email_body = get_template('app.payroll.disbursalemail.html').render(ctx)
        msg = EmailMessage(email_subject, email_body,  to= [id])
        msg.content_subtype = 'html'
        msg.send()
        return Response({},status = status.HTTP_200_OK)

class AdvancesDataAPI(APIView):
    permission_classes = (permissions.IsAuthenticated ,)
    def get(self , request , format = None):
        print 'entered','*******************'
        tosend={}
        today = datetime.date.today()
        totalLoanObj = Advances.objects.filter(created__year=today.year,created__month=02,typ='loan',approved=True,settled=False).aggregate(totalLoan = Sum('amount'))
        totalLoanObj = totalLoanObj['totalLoan'] if totalLoanObj['totalLoan'] else 0
        totalAdvanceObj = Advances.objects.filter(created__year=today.year,created__month=02,typ='advance',approved=True,settled=False).aggregate(totalAdvance = Sum('amount'))
        totalAdvanceObj = totalAdvanceObj['totalAdvance'] if totalAdvanceObj['totalAdvance'] else 0
        tosend['loanamt']= totalLoanObj
        tosend['advamt']= totalAdvanceObj
        return Response(tosend, status=status.HTTP_200_OK)

class sendLoanSettlementEmailAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    def post(self, request, format=None):
        print 'Send loan Settlement Emaillllllllll'
        print request.data['pk'],'----reeqqq'
        settingsFields = application.objects.get(name = 'app.clientRelationships').settings.all()
        for value in settingsFields:
            print value
        dates =  request.data['emiDate']
        advance = Advances.objects.get(pk = request.data['pk'])
        id = advance.user.profile.email
        print advance.typ,'---------typeeeeeeeee'
        print dates,'to this-------',id
        email_subject =str('Disbursal Info')
        msgBody='Disbursal Information testing.'
        ctx = {
            'emiTable' : dates,
            'fname': advance.user.first_name,
            'lname': advance.user.last_name,
            'amt' : advance.amount,
            'reason': advance.reason,
            'reqDate': advance.user.created,
            'linkUrl': 'cioc.co.in',
            'linkText' : 'View Online',
            'sendersAddress' : '(C) CIOC FMCG Pvt Ltd',
            'sendersPhone' : '841101',
            'linkedinUrl' : 'https://www.linkedin.com/company/13440221/',
            'fbUrl' : 'facebook.com',
            'twitterUrl' : 'twitter.com',
        }
        email_body = get_template('app.payroll.loanSettlementemail.html').render(ctx)
        msg = EmailMessage(email_subject, email_body,  to= [id])
        msg.content_subtype = 'html'
        msg.send()
        return Response({},status = status.HTTP_200_OK)


class AllPaySlipsAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    def get(self, request, format=None):
        # print Payslip.objects.filter(user = request.user, report__status="paid")
        # payslipObj = Payslip.objects.filter(user = request.user, report__status="paid").order_by('year')
        # first_pay = payslipObj.first()
        # last_pay =  payslipObj.last()
        # data = []
        # for i in range(first_pay.year , last_pay.year+1):
        #     obj = Payslip.objects.filter(year = i, report__status="paid", user = request.user)
        #     if len(obj) > 0:
        #         data.append({'year' : i , 'data' : payslipLiteSerializer(obj, many=True).data })
        data = []
        if 'year' in request.GET:
            obj = Payslip.objects.filter(year = request.GET['year'], report__status="paid", user = request.user)
            data = payslipLiteSerializer(obj, many=True).data
        return Response(data,status = status.HTTP_200_OK)

class ITDeclarationViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    serializer_class = ITDecarationSerializer
    queryset = ITDecaration.objects.all()

def getCurrentYear():
    today = datetime.date.today()
    currentMonth = today.month
    currentYear= today.year
    if currentMonth>3:
        financialYear = str(currentYear)+ '-' +str(currentYear+1)
    else:
        financialYear = str(currentYear-1)+ '-' + str(currentYear)
    return financialYear


class GetITDecarationAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    def get(self, request, format=None):
        if 'user' in request.GET:
            user = User.objects.get(pk = int(request.GET['user']))
        else:
            user = request.user
        financialYear = getCurrentYear()
        if 'currentFinancialYear' in request.GET:
            financialYear = request.GET['currentFinancialYear']
        if 'onlyTotal' in request.GET:
            totalData = CalculateItDeclaration(financialYear, user)
            return Response(totalData,status = status.HTTP_200_OK)
        payroll = user.payroll
        monthsList = ['April' , 'May' , 'June' , 'July' , 'Aug' , 'Sept' , 'Oct' , 'Nov' , 'Dec' , 'Jan' , 'Feb' , 'March']
        allITdecObj = ITDecaration.objects.filter(user = user, year = financialYear)
        # INCOME DATA
        incomeObj = allITdecObj.filter(group_name = 'income')
        if len(incomeObj) == 0:
            basic = 0
            hra = 0
            special = 0
            lta = 0
            fixedVariable = 0
            bonus = 0
            pfAmnt = 0
            pfAdmin = 0
            totalbasicAmount = 0
            totalhraAmount = 0
            totalspecialAmount = 0
            totalltaAmount = 0
            totalfixedAmount = 0
            totalbonusAmount = 0
            totalPfAmount = 0
            totalPfAdminAmount = 0

            totalBasicData = ITDecaration.objects.create(user = user, year = financialYear, group_name = 'income' , title = 'Basic', amount = totalbasicAmount)
            totalhraData = ITDecaration.objects.create(user = user, year = financialYear, group_name = 'income' , title = 'HRA', amount = totalhraAmount)
            totalspecialData = ITDecaration.objects.create(user = user, year = financialYear, group_name = 'income' , title = 'Special Allownace', amount = totalspecialAmount)
            totalltaData = ITDecaration.objects.create(user = user, year = financialYear, group_name = 'income' , title = 'LTA', amount = totalltaAmount)
            totalfixedVariableData = ITDecaration.objects.create(user = user, year = financialYear, group_name = 'income' , title = 'Fixed Variable', amount = totalfixedAmount)
            totalbonusData = ITDecaration.objects.create(user = user, year = financialYear, group_name = 'income' , title = 'Statutory Bonus', amount = totalbonusAmount)
            totalPfData = ITDecaration.objects.create(user = user, year = financialYear, group_name = 'income' , title = 'PF (Employee)', amount = totalPfAmount)
            totalPfAdminData = ITDecaration.objects.create(user = user, year = financialYear, group_name = 'income' , title = 'PF (Company)', amount = totalPfAdminAmount)
            for m in monthsList:
                basicData = ITDecaration.objects.create(user = user, year = financialYear, group_name = 'income' , month = m, title = 'Basic', amount = basic)
                hraData = ITDecaration.objects.create(user = user, year = financialYear, group_name = 'income' , month = m, title = 'HRA', amount = hra)
                specialData = ITDecaration.objects.create(user = user, year = financialYear, group_name = 'income' , month = m, title = 'Special Allownace', amount = special)
                ltaData = ITDecaration.objects.create(user = user, year = financialYear, group_name = 'income' , month = m, title = 'LTA', amount = lta)
                fixedVariableData = ITDecaration.objects.create(user = user, year = financialYear, group_name = 'income' , month = m, title = 'Fixed Variable', amount = fixedVariable)
                bonusData = ITDecaration.objects.create(user = user, year = financialYear, group_name = 'income' , month = m, title = 'Statutory Bonus', amount = bonus)
                pfData = ITDecaration.objects.create(user = user, year = financialYear, group_name = 'income' , month = m, title = 'PF (Employee)', amount = pfAmnt)
                pfAdminData = ITDecaration.objects.create(user = user, year = financialYear, group_name = 'income' , month = m, title = 'PF (Company)', amount = pfAdmin)

        titleUniq =  ['Basic' , 'HRA' , 'Special Allownace' , 'LTA' , 'Fixed Variable' , 'Statutory Bonus' ,'PF (Employee)']
        allData = []
        grandTotal = 0
        incomeObj = ITDecaration.objects.filter(user = user, year = financialYear, group_name = 'income')
        montlyDataall = {}
        for t in titleUniq:
            montlyData = {}
            for m in monthsList:
                val = incomeObj.get(title = t, month = m)
                montlyData[m] = val.amount
            totalData = incomeObj.get(title =  t , month = None)
            montlyData['annualTotal'] = totalData.amount
            grandTotal +=totalData.amount
            allData.append({'title' : t , 'data' : montlyData})
            montlyDataall['annualTotal'] = grandTotal
        montlyDatap = {}
        montlyTot = {}
        for m in monthsList:
            tot = incomeObj.filter(month = m).exclude(title = 'PF (Company)').aggregate(tot = Sum('amount'))
            if tot['tot'] is not None:
                montlyDataall[m] = tot['tot']
            else:
                montlyDataall[m] = 0
            val = incomeObj.get(title = 'PF (Company)', month = m)
            montlyDatap[m] = val.amount
            montlyTot[m] = montlyDataall[m] - montlyDatap[m]
        allData.append({'title' : 'Total' , 'data' : montlyDataall})
        totalData = incomeObj.get(title =  'PF (Company)' , month = None)
        montlyDatap['annualTotal'] = totalData.amount
        allData.append({'title' : 'PF (Company)' , 'data' : montlyDatap})
        montlyTot['annualTotal'] = montlyDataall['annualTotal'] - montlyDatap['annualTotal']
        totalCTC = montlyTot['annualTotal']
        allData.append({'title' : 'Total CTC' , 'data' : montlyTot})
        incomeData = { 'data' : allData }

        # Deductions
        deductionObj = allITdecObj.filter(group_name = 'deductions')
        if len(deductionObj) == 0:
            totalDeduction = 0
            if payroll.pTax is not None:
                totalDeduction = payroll.pTax
                deduction = totalDeduction/12
            deductiontotalData = ITDecaration.objects.create(user = user, year = financialYear, group_name = 'deductions' , title = 'Prof Tax', amount = totalDeduction)
            for m in monthsList:
                deductionDataAdd = ITDecaration.objects.create(user = user, year = financialYear, group_name = 'deductions' , title = 'Prof Tax', amount = deduction, month = m)
            deductionObj = allITdecObj.filter(group_name = 'deductions')
        dedtitleUniq = ['Prof Tax']
        deductionAllData = []
        grandTotalDed = 0
        for d in dedtitleUniq:
            montlyData = {}
            for m in monthsList:
                val = deductionObj.get(title = d, month = m)
                montlyData[m] = val.amount
            totalData = deductionObj.get(title =  d , month = None)
            montlyData['annualTotal'] = totalData.amount
            grandTotalDed +=totalData.amount
            deductionAllData.append({'title' : d , 'data' : montlyData})
        deductionData = { 'data' : deductionAllData }

        # Deduction under chapter VI A
        allIncomes = [
        {'title' : '80C - 5 Years of Fixed Deposit in Scheduled Bank' ,'limit' : 150000},
        {'title' : '80C - Children Tuition Fees' ,'limit' : 150000},
        {'title' : '80CCC - Contribution to Pension Fund' ,'limit' : 150000},
        {'title' : '80C - Deposit in NSC' ,'limit' : 150000},
        {'title' : '80C - Deposit in NSS' ,'limit' : 150000},
        {'title' : '80C - Deposit in Post Office Savings Schemes' ,'limit' : 150000},
        {'title' : '80C - Equity Linked Savings Scheme ( ELSS )' ,'limit' : 150000},
        {'title' : '80C - Interest on NSC Reinvested' ,'limit' : 150000},
        {'title' : '80C - Kisan Vikas Patra (KVP)' ,'limit' : 150000},
        {'title' : '80C - Life Insurance Premium' ,'limit' : 150000},
        {'title' : '80C - Long term Infrastructure Bonds' ,'limit' : 150000},
        {'title' : '80C - Mutual Funds' ,'limit' : 150000},
        {'title' : '80C - NABARD Rural Bonds' ,'limit' : 150000},
        {'title' : '80C - National Pension Scheme' ,'limit' : 150000},
        {'title' : '80C - NHB Scheme' ,'limit' : 150000},
        {'title' : '80C - Post office time deposit for 5 years' ,'limit' : 150000},
        {'title' : '80C - Pradhan Mantri Suraksha Bima Yojana' ,'limit' : 150000},
        {'title' : '80C - Public Provident Fund' ,'limit' : 150000},
        {'title' : '80C - Repayment of Housing loan(Principal amount)' ,'limit' : 150000},
        {'title' : '80C - Stamp duty and Registration charges' ,'limit' : 150000},
        {'title' : '80C - Sukanya Samriddhi Yojana' ,'limit' : 150000},
        {'title' : '80C - Unit Linked Insurance Premium (ULIP)' ,'limit' : 150000},
        {'title' : '80D - Medical Bills - Senior Citizen' ,'limit' : 50000},
        {'title' : '80D - Preventive Health Checkup - Dependant Parents' ,'limit' : 5000},
        {'title' : '80D - Medical Insurance Premium( age <60 years )' ,'limit' : 25000},
        {'title' : '80D - Medical Insurance Premium - Dependant Parents ( age <60 years )' ,'limit' : 25000},
        {'title' : '80D - Preventive Health Check-up' ,'limit' : 5000},



        {'title' : '80EE - Additional Interest on housing loan borrowed as on 1st Apr 2016' ,'limit' : 50000},
        {'title' : '10(13) - Superannuation Exemption' ,'limit' : 150000},
        {'title' : '80CCD1(B) - Contribution to NPS 2015' ,'limit' : 50000},
        {'title' : '80EEA - Additional Interest on Housing loan borrowed as on 1st Apr 2019' ,'limit' : 150000},
        {'title' : '80EEB - Interest on Electric Vehicle borrowed as on 1st Apr 2019' ,'limit' : 150000},
        {'title' : '80TTB - Interest on Deposits in Savings Account, FDs, Post Office And Cooperative Society for Senior Citizen' ,'limit' : 50000},
        {'title' : '80G - Donation - 100% Exemption' ,'limit' : 99999999},
        {'title' : '80G - Donation - 50% Exemption' ,'limit' : 99999999},
        {'title' : '80G - Donation - Children Education' ,'limit' : 99999999},
        {'title' : '80G - Donation - Political Parties' ,'limit' : 99999999},
        {'title' : '80TTA - Interest on Deposits in Savings Account, FDs, Post Office And Cooperative Society' ,'limit' : 10000},
        {'title' : '80E - Interest on Loan of higher Self education' ,'limit' : 99999999},
        {'title' : '80DD - Medical Treatment / Insurance of handicapped Dependant' ,'limit' : 75000},
        {'title' : '80DD - Medical Treatment / Insurance of handicapped Dependant (Severe)' ,'limit' : 125000},
        {'title' : '80DDB - Medical Treatment ( Specified Disease only )' ,'limit' : 40000},
        {'title' : '80DDB - Medical Treatment (Specified Disease only)- Senior Citizen' ,'limit' : 100000},
        {'title' : '80U - Permanent Physical disability (Above 40%)' ,'limit' : 125000},
        {'title' : '80U - Permanent Physical disability (Below 40%)' ,'limit' : 75000},
        {'title' : '80CCG - Rajiv Gandhi Equity Scheme' ,'limit' : 25000},
        ]
        deductionSixAObj = allITdecObj.filter(group_name = 'deductionvia')
        if len(deductionSixAObj)==0:
            for i in allIncomes:
                deductionvia =  ITDecaration.objects.create(user = user, year = financialYear, group_name = 'deductionvia' , title = i['title'], limit = i['limit'])
            deductionSixAObj = allITdecObj.filter(group_name = 'deductionvia')
        dedTot = 0
        deductn = deductionSixAObj.aggregate(tot = Sum('amount'))
        if deductn['tot'] is not None:
            dedTot = deductn['tot']
        deductionSixAData = ITDecarationSerializer(deductionSixAObj, many = True).data

        # OTHER INCOMES
        otherincomesObj = allITdecObj.filter(group_name = 'otherIncomes')
        otherSourcesTot = 0
        totalOther = otherincomesObj.aggregate(tot = Sum('amount'))
        if totalOther['tot'] is not None:
            otherSourcesTot =  totalOther['tot']
        otherIncomesAData = ITDecarationSerializer(otherincomesObj, many = True).data

        # Income / Loss from house property
        housePropTot = 0
        housePropertyObj = allITdecObj.filter(group_name = 'houseProperty')
        houseProp = housePropertyObj.aggregate(tot = Sum('amount') )
        if houseProp['tot'] is not None:
            housePropTot =  houseProp['tot']
        housePropertyData = ITDecarationAllSerializer(housePropertyObj, many=True).data

        #Income From Previous Employer
        prevEmpIncomeList = ['Total Income' , 'Income Tax' , 'Professional Tax' , 'Provident Fund']
        prevEmpObj = allITdecObj.filter(group_name = 'previousEmployer')
        if len(prevEmpObj)==0:
            for e in prevEmpIncomeList:
                prevEmp =  ITDecaration.objects.create(user = user, year = financialYear, group_name = 'previousEmployer' , title = e)
            prevEmpObj = allITdecObj.filter(group_name = 'previousEmployer')
        prevEmpIncTot = prevEmpObj.get(title = 'Total Income').amount
        prevEmpIncTax = prevEmpObj.get(title = 'Income Tax').amount
        prevEmpData = ITDecarationSerializer(prevEmpObj.exclude(Q(title = 'Professional Tax')|Q(title = 'Provident Fund')), many = True).data

        #ANNUAL EXCEMPTION
        excemptionObj = ITDecaration.objects.filter(user = user, year = financialYear, group_name = 'exemptions')
        exctitleUniq = excemptionObj.values_list('title', flat = True).distinct()
        excallData = []
        currentMonth = datetime.date.today().month - 3
        if currentMonth<0:
            currentMonth = 12 + currentMonth
        for t in exctitleUniq:
            excmontlyData = {}
            count = 0
            for m in monthsList:
                count+=1
                val = excemptionObj.get(title = t, month = m)
                excmontlyData[m] = {'amount' : val.amount , 'pk' : val.pk}
                if count<currentMonth:
                    excmontlyData[m]['canEdit'] = False
                else:
                    excmontlyData[m]['canEdit'] = True
            excallData.append({'title' : t , 'data' : excmontlyData})
        excmontlyData ={}
        totalExc = 0
        for j in monthsList:
            totaData = 0
            try:
                totaData = excemptionObj.get(month = j, title = 'House Rent Allowance Section 10 (13A)').amount +   excemptionObj.get(month = j, title = 'Leave Travel Assistance Section 10(5)').amount
                totalExc+=totaData
                excmontlyData[j] = {'amount' : totaData }
            except:
                pass
        excallData.append({'title' : 'Total Exemptions' , 'data' : excmontlyData})
        annualExcemptionData = { 'data' : excallData}
        propertyOwner = allITdecObj.filter(group_name = 'propertyOwnerDetails')
        if propertyOwner:
            propertyOwnerDetails = ITDecarationAllSerializer(propertyOwner.last(), many = False).data
        else:
            propertyOwner =  ITDecaration.objects.create(user = user, year = financialYear, group_name = 'propertyOwnerDetails' , title = 'Rented Property', tenantName = '' , tenantPan = '', address = '')
            propertyOwnerDetails = ITDecarationAllSerializer(propertyOwner, many = False).data
        selfOccupied = ITDecaration.objects.filter(group_name = 'selfOccupied' , user = user,  year = financialYear)
        if selfOccupied:
            selfOccupiedDetails = ITDecarationAllSerializer(selfOccupied.last(), many = False).data
        else:
            selfOccupied =  ITDecaration.objects.create(user = user,  group_name = 'selfOccupied',  year = financialYear , title = 'Housing Loan', tenantName = '' , tenantPan = '', eligibleAmount = 150000)
            selfOccupiedDetails = ITDecarationAllSerializer(selfOccupied, many = False).data



        payroll = {'pk' : user.payroll.pk, 'isOwnHouse' : user.payroll.isOwnHouse , 'isExtraIncome':user.payroll.isExtraIncome,'isRentedHouse' : user.payroll.isRentedHouse}
        totalData = CalculateItDeclaration(financialYear, user)
        return Response({'incomeData' : incomeData , 'deductionData' : deductionData , 'deductionSixAData' : deductionSixAData , 'otherIncomesAData' : otherIncomesAData , 'housePropertyData' : housePropertyData , 'prevEmpData' : prevEmpData  , 'annualExcemptionData' : annualExcemptionData , 'propertyOwnerDetails' : propertyOwnerDetails , 'selfOccupiedDetails' : selfOccupiedDetails , 'totalData' : totalData , 'payroll' : payroll},status = status.HTTP_200_OK)


class GetLimitAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    def get(self, request, format=None):
        if 'user' in data:
            user = User.objects.get(pk = int(request.GET['user']))
        else:
            user = request.user
        payroll = user.payroll
        limitAmount = 0
        if payroll.basic is not None:
            limitAmount = (payroll.basic*18)/100
        return Response({'limitAmount' : limitAmount},status = status.HTTP_200_OK)

from calendar import monthrange
class ReCalculateAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    def get(self, request, format=None):
        if 'user' in request.GET:
            user = User.objects.get(pk = int(request.GET['user']))
        else:
            user = request.user
        today = datetime.date.today()
        # currentMonth = today.month
        financialYear = getCurrentYear()
        payroll = user.payroll
        monthsAllList = ['April' , 'May' , 'June' , 'July' , 'Aug' , 'Sept' , 'Oct' , 'Nov' , 'Dec' , 'Jan' , 'Feb' , 'March']
        if 'dated' in request.GET:
            dated =  datetime.datetime.strptime(request.GET['dated'], '%Y-%m-%d')
            days = dated.strftime("%d")
            currentMonth = dated.month
            month =  dated.strftime("%b")
            year =  dated.year
            range = monthrange(year, currentMonth)
            totalDays = range[1]
            totalDays = float(totalDays)
        range = slice(currentMonth-4 , len(monthsAllList))
        monthsList = monthsAllList[range]
        basic = 0
        hra = 0
        special = 0
        lta = 0
        fixedVariable = 0
        bonus = 0
        totalbasicAmount = 0
        totalhraAmount = 0
        totalspecialAmount = 0
        totalltaAmount = 0
        totalfixedAmount = 0
        totalbonusAmount = 0
        totalPfAmount = 0
        totalPfAdminAmount = 0
        if payroll.basic is not None:
            basic = payroll.basic/12
            # totalbasicAmount = payroll.basic
        if payroll.hra is not None:
            hra = payroll.hra/12
            # totalhraAmount = payroll.hra
        if payroll.special is not None:
            special = payroll.special/12
            # totalspecialAmount = payroll.special
        if payroll.lta is not None:
            lta = payroll.lta/12
            # totalltaAmount = payroll.lta
        if payroll.adHoc is not None:
            fixedVariable = payroll.adHoc/12
            # totalfixedAmount = payroll.adHoc
        if payroll.bonus is not None:
            bonus = payroll.bonus/12
            # totalbonusAmount = payroll.bonus
        if payroll.pfAmnt is not None:
            pfAmnt = payroll.pfAmnt/12
            # totalPfAmount = payroll.pfAmnt
        if payroll.pfAdmin is not None:
            pfAdmin = payroll.pfAdmin/12
            # totalPfAdminAmount = payroll.pfAdmin
        for m in monthsList:

            basicData = ITDecaration.objects.get(user = user, year = financialYear, group_name = 'income' , month = m, title = 'Basic')
            if month == m:
                oldAmnt = basicData.amount/totalDays
                newAmnt = basic/totalDays
                oldDays = float(days) - 1
                newDays = float(totalDays) - float(days) + 1
                oldAmntDays = oldAmnt * float(oldDays)
                newAmntDays = newAmnt * float(newDays)
                basicData.amount =  oldAmntDays + newAmntDays
            else:
                basicData.amount = basic
            basicData.save()
            hraData = ITDecaration.objects.get(user = user, year = financialYear, group_name = 'income' , month = m, title = 'HRA')
            if month == m:
                oldAmnt = hraData.amount/totalDays
                newAmnt = hra/totalDays
                oldDays = float(days) - 1
                newDays = float(totalDays) - float(days) + 1
                oldAmntDays = oldAmnt * float(oldDays)
                newAmntDays = newAmnt * float(newDays)
                hraData.amount =  oldAmntDays + newAmntDays
            else:
                hraData.amount = hra
            hraData.save()
            specialData = ITDecaration.objects.get(user = user, year = financialYear, group_name = 'income' , month = m, title = 'Special Allownace')
            if month == m:
                oldAmnt = specialData.amount/totalDays
                newAmnt = special/totalDays
                oldDays = float(days) - 1
                newDays = float(totalDays) - float(days) + 1
                oldAmntDays = oldAmnt * float(oldDays)
                newAmntDays = newAmnt * float(newDays)
                specialData.amount =  oldAmntDays + newAmntDays
            else:
                specialData.amount = special
            specialData.save()
            ltaData = ITDecaration.objects.get(user = user, year = financialYear, group_name = 'income' , month = m, title = 'LTA')
            if month == m:
                oldAmnt = ltaData.amount/totalDays
                newAmnt = lta/totalDays
                oldDays = float(days) - 1
                newDays = float(totalDays) - float(days) + 1
                oldAmntDays = oldAmnt * float(oldDays)
                newAmntDays = newAmnt * float(newDays)
                ltaData.amount =  oldAmntDays + newAmntDays
            else:
                ltaData.amount = lta
            ltaData.save()
            fixedVariableData = ITDecaration.objects.get(user = user, year = financialYear, group_name = 'income' , month = m, title = 'Fixed Variable')
            if month == m:
                oldAmnt = fixedVariableData.amount/totalDays
                newAmnt = fixedVariable/totalDays
                oldDays = float(days) - 1
                newDays = float(totalDays) - float(days) + 1
                oldAmntDays = oldAmnt * float(oldDays)
                newAmntDays = newAmnt * float(newDays)
                fixedVariableData.amount =  oldAmntDays + newAmntDays
            else:
                fixedVariableData.amount = fixedVariable
            fixedVariableData.save()
            bonusData = ITDecaration.objects.get(user = user, year = financialYear, group_name = 'income' , month = m, title = 'Statutory Bonus')
            if month == m:
                oldAmnt = bonusData.amount/totalDays
                newAmnt = bonus/totalDays
                oldDays = float(days) - 1
                newDays = float(totalDays) - float(days) + 1
                oldAmntDays = oldAmnt * float(oldDays)
                newAmntDays = newAmnt * float(newDays)
                bonusData.amount =  oldAmntDays + newAmntDays
            else:
                bonusData.amount = bonus
            bonusData.save()
            totalPfData = ITDecaration.objects.get(user = user, year = financialYear, group_name = 'income' , month = m,title = 'PF (Employee)')
            if month == m:
                oldAmnt = totalPfData.amount/totalDays
                newAmnt = pfAmnt/totalDays
                oldDays = float(days) - 1
                newDays = float(totalDays) - float(days) + 1
                oldAmntDays = oldAmnt * float(oldDays)
                newAmntDays = newAmnt * float(newDays)
                totalPfData.amount =  oldAmntDays + newAmntDays
            else:
                totalPfData.amount = pfAmnt
            totalPfData.save()
            totalPfAdminData = ITDecaration.objects.get(user = user, year = financialYear, group_name = 'income' ,month = m, title = 'PF (Company)')
            if month == m:
                oldAmnt = totalPfAdminData.amount/totalDays
                newAmnt = pfAdmin/totalDays
                oldDays = float(days) - 1
                newDays = float(totalDays) - float(days) + 1
                oldAmntDays = oldAmnt * float(oldDays)
                newAmntDays = newAmnt * float(newDays)
                totalPfAdminData.amount =  oldAmntDays + newAmntDays
            else:
                totalPfAdminData.amount = pfAdmin
            totalPfAdminData.save()

        basicTot = ITDecaration.objects.filter(user = user, year = financialYear, group_name = 'income' , title = 'Basic', month__in = monthsAllList).aggregate(tot = Sum('amount'))
        if basicTot['tot'] is not None:
            totalbasicAmount = basicTot['tot']
        HRATot = ITDecaration.objects.filter(user = user, year = financialYear, group_name = 'income' , title = 'HRA', month__in = monthsAllList).aggregate(tot = Sum('amount'))
        if HRATot['tot'] is not None:
            totalhraAmount = HRATot['tot']
        specialTot = ITDecaration.objects.filter(user = user, year = financialYear, group_name = 'income' , title = 'Special Allownace', month__in = monthsAllList).aggregate(tot = Sum('amount'))
        if specialTot['tot'] is not None:
            totalspecialAmount = specialTot['tot']
        ltaTot = ITDecaration.objects.filter(user = user, year = financialYear, group_name = 'income' , title = 'LTA', month__in = monthsAllList).aggregate(tot = Sum('amount'))
        if ltaTot['tot'] is not None:
            totalltaAmount = ltaTot['tot']
        fixedTot = ITDecaration.objects.filter(user = user, year = financialYear, group_name = 'income' , title = 'Fixed Variable', month__in = monthsAllList).aggregate(tot = Sum('amount'))
        if fixedTot['tot'] is not None:
            totalfixedAmount = fixedTot['tot']
        bonusTot = ITDecaration.objects.filter(user = user, year = financialYear, group_name = 'income' , title = 'Statutory Bonus', month__in = monthsAllList).aggregate(tot = Sum('amount'))
        if bonusTot['tot'] is not None:
            totalbonusAmount = bonusTot['tot']
        pfTot = ITDecaration.objects.filter(user = user, year = financialYear, group_name = 'income' , title = 'PF (Employee)', month__in = monthsAllList).aggregate(tot = Sum('amount'))
        if pfTot['tot'] is not None:
            totalPfAmount = pfTot['tot']
        pfAdminTot = ITDecaration.objects.filter(user = user, year = financialYear, group_name = 'income' , title = 'PF (Company)', month__in = monthsAllList).aggregate(tot = Sum('amount'))
        if pfAdminTot['tot'] is not None:
            totalPfAdminAmount = pfAdminTot['tot']
        allDec = ITDecaration.objects.filter(user = user, year = financialYear, group_name = 'income' , month__isnull = True)
        totalBasicData, b = allDec.get_or_create(title = 'Basic')
        totalBasicData.amount = totalbasicAmount
        totalBasicData.save()
        totalhraData, h = allDec.get_or_create(title = 'HRA')
        totalhraData.amount = totalhraAmount
        totalhraData.save()
        totalspecialData, s = allDec.get_or_create(title = 'Special Allownace')
        totalspecialData.amount = totalspecialAmount
        totalspecialData.save()
        totalltaData, l = allDec.get_or_create(title = 'LTA')
        totalltaData.amount = totalltaAmount
        totalltaData.save()
        totalfixedVariableData, i = allDec.get_or_create( title = 'Fixed Variable')
        totalfixedVariableData.amount = totalfixedAmount
        totalfixedVariableData.save()
        totalbonusData, n = allDec.get_or_create(title = 'Statutory Bonus')
        totalbonusData.amount = totalbonusAmount
        totalbonusData.save()
        totalPfData, n = allDec.get_or_create(title = 'PF (Employee)')
        totalPfData.amount = totalPfAmount
        totalPfData.save()
        totalPfAdminData, n = allDec.get_or_create(title = 'PF (Company)')
        totalPfAdminData.amount = totalPfAdminAmount
        totalPfAdminData.save()
        titleUniq = ['Basic' , 'HRA' , 'Special Allownace' , 'LTA' , 'Fixed Variable' , 'Statutory Bonus','PF (Employee)']
        allData = []
        grandTotal = 0
        incomeObj = ITDecaration.objects.filter(user = user, year = financialYear, group_name = 'income')
        montlyDataall = {}
        for t in titleUniq:
            montlyData = {}
            for m in monthsAllList:
                val = incomeObj.get(title = t, month = m)
                montlyData[m] = val.amount
            totalData = incomeObj.get(title =  t , month = None)
            montlyData['annualTotal'] = totalData.amount
            grandTotal +=totalData.amount
            allData.append({'title' : t , 'data' : montlyData})
        montlyDataall['annualTotal'] = grandTotal
        incomeData = { 'data' : allData , 'grandTotal' : grandTotal}
        montlyDatap = {}
        montlyTot = {}
        for m in monthsAllList:
            tot = incomeObj.filter(month = m).exclude(title = 'PF (Company)').aggregate(tot = Sum('amount'))
            if tot['tot'] is not None:
                montlyDataall[m] = tot['tot']
            else:
                montlyDataall[m] = 0
            val = incomeObj.get(title = 'PF (Company)', month = m)
            montlyDatap[m] = val.amount
            montlyTot[m] = montlyDataall[m] - montlyDatap[m]
        allData.append({'title' : 'Total' , 'data' : montlyDataall})
        totalData = incomeObj.get(title =  'PF (Company)' , month = None)
        montlyDatap['annualTotal'] = totalData.amount
        allData.append({'title' : 'PF (Company)' , 'data' : montlyDatap})
        montlyTot['annualTotal'] = montlyDataall['annualTotal'] - montlyDatap['annualTotal']
        allData.append({'title' : 'Total CTC' , 'data' : montlyTot})
        incomeData = { 'data' : allData }

        return Response({'incomeData' : incomeData },status = status.HTTP_200_OK)



class AddITDeclarationAPIView(APIView):
        renderer_classes = (JSONRenderer,)
        def post(self, request, format=None):
            data = request.data
            if 'user' in data:
                user = User.objects.get(pk = int(data['user']))
            else:
                user = request.user
            payroll = user.payroll
            year = getCurrentYear()
            monthsList = ['April' , 'May' , 'June' , 'July' , 'Aug' , 'Sept' , 'Oct' , 'Nov' , 'Dec' , 'Jan' , 'Feb' , 'March']
            if 'group_name' in data:
                if data['group_name'] == 'houseProperty' or data['group_name'] == 'propertyOwnerDetails' or data['group_name'] == 'selfOccupied':
                    if 'pk' in data:
                        allITdecObj = ITDecaration.objects.get(pk = int(data['pk']))
                        if 'title' in data:
                            allITdecObj.title = data['title']
                        if 'amount' in data:
                            allITdecObj.amount = data['amount']
                        if 'annualRent' in data:
                            allITdecObj.annualRent = data['annualRent']
                        if 'muncipalTax' in data:
                            allITdecObj.muncipalTax = data['muncipalTax']
                        if 'unrealizedTax' in data:
                            allITdecObj.unrealizedTax = data['unrealizedTax']
                        if 'netAnnualValue' in data:
                            allITdecObj.netAnnualValue = data['netAnnualValue']
                        if 'standardDeduction' in data:
                            allITdecObj.standardDeduction = data['standardDeduction']
                        if 'interestOnLoan' in data:
                            allITdecObj.interestOnLoan = data['interestOnLoan']
                        if 'tenantName' in data:
                            allITdecObj.tenantName = data['tenantName']
                        if 'tenantPan' in data:
                            allITdecObj.tenantPan = data['tenantPan']
                        if 'address' in data:
                            allITdecObj.address = data['address']
                        if 'eligibleAmount' in data:
                            allITdecObj.eligibleAmount = data['eligibleAmount']
                        allITdecObj.save()
                    else:
                        allITdecObj = ITDecaration.objects.create(user = user, year = year, title = data['title'] , amount = data['amount'] , annualRent = data['annualRent'], muncipalTax = data['muncipalTax'], unrealizedTax = data['unrealizedTax'], netAnnualValue = data['netAnnualValue'] , standardDeduction = data['standardDeduction'], interestOnLoan = data['interestOnLoan'], tenantName = data['tenantName'] , tenantPan = data['tenantPan'], group_name = data['group_name'])
                    resData = ITDecarationAllSerializer(allITdecObj, many=False).data
                elif data['group_name'] == 'exemptions':
                    rent = data['rent']
                    travel = 0
                    hra = 0
                    if 'isRentedHouse' in data:
                        payroll = user.payroll
                        payroll.isRentedHouse = data['isRentedHouse']
                        payroll.save()
                    if payroll.hra is not None:
                        hra = payroll.hra
                    basicRent = 0
                    if payroll.basic is not None:
                        basicRent = payroll.basic*0.3
                    monthlyRent = rent
                    monthlyHra = 0
                    montlyBasicRent = 0
                    montlytravel = 0
                    totalExcemption = 0
                    monthlyLta = 0
                    incomeObj = ITDecaration.objects.filter(user = user, year = year, group_name = 'income')
                    currentMonth = datetime.date.today().month - 3
                    if currentMonth<0:
                        currentMonth = 12 + currentMonth
                        # for t in exctitleUniq:
                        #     excmontlyData = {}
                        #     count = 0
                        #     for m in monthsList:
                        #         count+=1
                        #         val = excemptionObj.get(title = t, month = m)
                        #         excmontlyData[m] = {'amount' : val.amount , 'pk' : val.pk}
                        #         if count<currentMonth:
                        #             excmontlyData[m]['canEdit'] = False
                        #         else:
                        #             excmontlyData[m]['canEdit'] = True

                    count = 0
                    for m in monthsList:
                        count+=1
                        monthlyHra =  incomeObj.get(month = m, title = 'HRA').amount
                        monthlyLta =  incomeObj.get(month = m, title = 'LTA').amount
                        # ITDecaration.objects.filter(user = user, year = year, group_name = 'exemptions' , month = m).delete()
                        montlyBasicRent =  incomeObj.get(month = m, title = 'Basic').amount
                        hraData,h = ITDecaration.objects.get_or_create(user = user, year = year, group_name = 'exemptions' , month = m, title = 'HRA')
                        basicRent, b = ITDecaration.objects.get_or_create(user = user, year = year, group_name = 'exemptions' , month = m, title = '40% of Basic + DA')
                        rentData, r = ITDecaration.objects.get_or_create(user = user, year = year, group_name = 'exemptions' , month = m, title = 'Your rent', )
                        houserentData, ho = ITDecaration.objects.get_or_create(user = user, year = year, group_name = 'exemptions' , month = m, title = 'House Rent Allowance Section 10 (13A)')
                        ltaData, l = ITDecaration.objects.get_or_create(user = user, year = year, group_name = 'exemptions' , month = m, title = 'Approved LTA Limit')
                        travelData, t = ITDecaration.objects.get_or_create(user = user, year = year, group_name = 'exemptions' , month = m, title = 'Leave Travel Assistance Section 10(5)')
                        if count>=currentMonth:
                            print monthlyRent,'aaaaaaaaaaaaaaaaaaaaaaa'
                            hraData.amount = monthlyHra
                            hraData.save()
                            basicRent.amount = montlyBasicRent
                            basicRent.save()
                            rentData.amount = monthlyRent
                            rentData.save()
                            montlyhouseRent = min(float(monthlyRent), monthlyHra, montlyBasicRent)
                            totalExcemption = montlyhouseRent + montlytravel
                            houserentData.amount = montlyhouseRent
                            houserentData.save()
                            ltaData.amount = monthlyLta
                            ltaData.save()
                            travelData.amount = montlytravel
                            travelData.save()
                    resData = calculateExcemption(year, user)
            if 'id' in data:
                amount = int(data['amount'])
                print amount
                decObj = ITDecaration.objects.get(pk = int(data['id']))
                month = decObj.month
                decOtherObj = ITDecaration.objects.filter(month = month, user = user, group_name = 'exemptions', year = year)
                hraData = decOtherObj.get(title = 'HRA')
                basicData = decOtherObj.get(title = '40% of Basic + DA')
                rentData = decOtherObj.get(title = 'Your rent')
                houserentData = decOtherObj.get(title = 'House Rent Allowance Section 10 (13A)')
                ltaData = decOtherObj.get(title = 'Approved LTA Limit')
                travelData = decOtherObj.get(title = 'Leave Travel Assistance Section 10(5)')
                decObj.amount = amount
                decObj.save()
                if decObj.title == 'Your rent':
                    houserentData.amount  = min(hraData.amount , basicData.amount , decObj.amount)
                    houserentData.save()
                elif decObj.title == 'Leave Travel Assistance Section 10(5)':
                    travelData.amount  = min(decObj.amount, ltaData.amount)
                    travelData.save()

                resData = calculateExcemption(year, user)
            CalculateItDeclaration(year, user)
            return Response(resData,status = status.HTTP_200_OK)


class SavepayslipAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    def post(self, request, format=None):
        user = request.user
        data = request.data
        user = User.objects.get(pk = int(data['user']))
        report = PayrollReport.objects.get(pk = int(data['report']))
        data = {'month' : data['month'], 'year' : data['year'] , 'user' : user, report: report,reimbursement: data['reimbursement'] , miscellaneous: data['miscellaneous']  }

        return Response(data,status = status.HTTP_200_OK)

class GetAllPayrollAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    def get(self, request, format=None):
        params = request.GET
        divsn = request.user.designation.division
        months = ["","January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
        monthsList = ['', 'Jan' , 'Feb' , 'March' ,'April' , 'May' , 'June' , 'July' , 'Aug' , 'Sept' , 'Oct' , 'Nov' , 'Dec' ]
        tempmonth = params['month']
        monthIndx = months.index(tempmonth)
        month = monthsList[monthIndx]
        tempYear = params['year']
        if monthIndx>3:
            year = tempYear+'-'+str(int(tempYear)+1)
        else:
            year = str(int(tempYear)-1)+'-'+tempYear
        allUsers = User.objects.filter(is_active = True, payrollAuthored__activatePayroll = True, designation__division = divsn)
        toRet = []
        for user in allUsers:
            try:
                itDec = ITDecaration.objects.filter(user = user, year = year, group_name = 'income' , month = month)
                try:
                    basic = itDec.get( title = 'Basic').amount
                except:
                    basic = 0
                try:
                    hra = itDec.get(title = 'HRA').amount
                except:
                    hra = 0
                try:
                    special = itDec.get(title = 'Special Allownace').amount
                except:
                    special = 0
                try:
                    lta = itDec.get( title = 'LTA').amount
                except:
                    lta = 0
                try:
                    adHoc = itDec.get( title = 'Fixed Variable').amount
                except:
                    adHoc = 0
                try:
                    bonus =  itDec.get( title = 'Statutory Bonus').amount
                except:
                    bonus = 0
                try:
                    pfAmnt =  itDec.get( title = 'PF (Employee)').amount
                except:
                    pfAmnt = 0
                try:
                    pfAdmin =  itDec.get( title = 'PF (Company)').amount
                except:
                    pfAdmin = 0
                deduction = 0
                miscellaneous = 0
                reimbursement = 0
                amount = hra + special + lta + basic + adHoc + bonus
                # totalPayable = amount - adHoc
                totalMonths =  ITDecaration.objects.filter(user = user, year = year, group_name = 'income', month__isnull = False, amount__gt=0,  title = 'Basic').count()
                taxAmount = ITDecaration.objects.get(group_name = 'FINAL TOTAL' , title = 'Total Tax',  year = year ,user = user).amount
                totalTaxable = amount  + reimbursement - miscellaneous  - pfAmnt - deduction
                tds = taxAmount/totalMonths
                # if totalTaxable<250000:
                #     tds = 0
                # if totalTaxable>250000 and totalTaxable<=500000:
                #     tds = (totalTaxable*5)/100
                # if totalTaxable>500000 and totalTaxable<=750000:
                #     tds = (totalTaxable*10)/100 + 12500
                # if totalTaxable>750000 and totalInc<=1000000:
                #     tds = (totalTaxable*15)/100 + 37500
                # if totalTaxable>1000000 and totalTaxable<=1250000:
                #     tds = (totalTaxable*20)/100 + 75000
                # if totalTaxable>1250000 and totalTaxable<=1500000:
                #     tds = (totalTaxable*30)/100 + 125000
                # if totalTaxable>1500000:
                #     tds = (totalTaxable*25)/100 + 187500
                totalPayable = totalTaxable - tds
                grandTotal = totalPayable + pfAdmin
                userObj = userSerializer(user, many = False).data
                data = {'user' : userObj, 'basic' : basic , 'hra' : hra, 'special' : special, 'lta' : lta , 'adHoc' : adHoc , 'bonus' : bonus, 'amount' : amount ,'miscellaneous' : miscellaneous , 'reimbursement' : reimbursement , 'pfAmnt' : pfAmnt, 'deduction' : deduction,'totalTaxable' : totalTaxable , 'tds' : tds , 'totalPayable' : totalPayable, 'pfAdmin' : pfAdmin , 'grandTotal' : grandTotal}
                toRet.append(data)
            except:
                pass
        return Response(toRet,status = status.HTTP_200_OK)

class Form16ViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    serializer_class = Form16Serializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['user' , 'year' , 'period']
    def get_queryset(self):
        toRet = Form16.objects.filter(division = self.request.user.designation.division)
        if 'currentUser' in self.request.GET:
            toRet = toRet.filter(user = self.request.user)
        return toRet


class GetPaySlipDetailsAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    def get(self, request, format=None):
        data = request.GET
        user = User.objects.get(pk = int(data['user']))
        months = ["","January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
        monthsList = ['', 'Jan' , 'Feb' , 'March' ,'April' , 'May' , 'June' , 'July' , 'Aug' , 'Sept' , 'Oct' , 'Nov' , 'Dec' ]
        tempmonth = data['month']
        monthIndx = months.index(tempmonth)
        month = monthsList[monthIndx]
        tempYear = data['year']
        if monthIndx>3:
            year = tempYear+'-'+str(int(tempYear)+1)
        else:
            year = str(int(tempYear)-1)+'-'+tempYear
        itDec = ITDecaration.objects.filter(user = user, year = year, group_name = 'income' , month = month)
        try:
            basic = itDec.get( title = 'Basic').amount
        except:
            basic = 0
        try:
            hra = itDec.get(title = 'HRA').amount
        except:
            hra = 0
        try:
            special = itDec.get(title = 'Special Allownace').amount
        except:
            special = 0
        try:
            lta = itDec.get( title = 'LTA').amount
        except:
            lta = 0
        try:
            adHoc = itDec.get( title = 'Fixed Variable').amount
        except:
            adHoc = 0
        try:
            bonus =  itDec.get( title = 'Statutory Bonus').amount
        except:
            bonus = 0
        try:
            pfAmnt =  itDec.get( title = 'PF (Employee)').amount
        except:
            pfAmnt = 0
        try:
            pfAdmin =  itDec.get( title = 'PF (Company)').amount
        except:
            pfAdmin = 0

        toRet = {'basic' : basic , 'hra' : hra, 'special' : special , 'lta' : lta , 'adHoc' : adHoc , 'bonus' : bonus , 'pfAmnt' : pfAmnt , 'pfAdmin' : pfAdmin, 'miscellaneous' : 0 , 'reimbursement' : 0 , 'tds' : 0}
        try:
            paySlip = Payslip.objects.get(month = monthIndx, year = data['year'] , user = user)
            toRet['miscellaneous'] = paySlip.miscellaneous
            toRet['reimbursement'] = paySlip.reimbursement
        except:
            pass
        totalEarnings = basic + hra + special + lta  + adHoc + bonus +  toRet['reimbursement']
        totalMonths =  ITDecaration.objects.filter(user = user, year = year, group_name = 'income', month__isnull = False, amount__gt=0,  title = 'Basic').count()
        taxAmount = ITDecaration.objects.get(group_name = 'FINAL TOTAL' , title = 'Total Tax',  year = year ,user = user).amount
        toRet['tds'] = taxAmount/totalMonths
        try:
            toRet['tds'] = paySlip.tds
        except:
            pass
        advanceDeduction = 0
        advanceDeductiontot = Advances.objects.filter(user = user, settled = False, returnMethod = 'SALARY_ADVANCE').aggregate(tot = Sum('amount'))
        if advanceDeductiontot['tot'] is not None:
            advanceDeduction = advanceDeductiontot['tot']
        deductions = pfAmnt + toRet['miscellaneous'] + toRet['tds'] + advanceDeduction
        total = totalEarnings - deductions
        toRet['advanceDeduction'] = advanceDeduction
        toRet['totalEarnings'] = totalEarnings
        toRet['deductions'] = deductions
        toRet['total'] = total
        dt=str(request.GET['year'])+'-'+str(monthIndx).zfill(2)
        mLeavesObj= Leave.objects.filter(Q(fromDate__contains=dt) | Q(toDate__contains=dt),user_id= int(user.pk),status='approved')
        extraLeavesObj= Leave.objects.filter(user_id= int(user.pk),status='approved',leavesCount__gte=4)
        leaveObj = mLeavesObj | extraLeavesObj
        holidaysList = list(CompanyHolidays.objects.all().values_list('date',flat=True))
        satOff = user.payroll.off
        sleave = 0
        aleave = 0
        cleave = 0
        for leave in leaveObj:
            daysDiff = (leave.toDate-leave.fromDate).days
            if daysDiff>0:
                for j in range(daysDiff+1):
                    sDate = leave.fromDate+relativedelta(days=j)
                    if sDate.month == int(month):
                        wDay = sDate.weekday()
                        if wDay<6 and sDate in holidaysList:
                            if wDay==5 and satOff:
                                continue
                            if leave.category == 'ML':
                                sleave += 1
                            elif leave.category == 'AL':
                                aleave += 1
                            elif leave.category == 'casual':
                                cleave += 1
            else:
                if leave.category == 'ML':
                    sleave += leave.leavesCount
                elif leave.category == 'AL':
                    aleave += leave.leavesCount
                elif leave.category == 'casual':
                    cleave += leave.leavesCount
        toRet['aleave'] = aleave
        toRet['sleave'] = sleave
        toRet['joiningDate'] = user.payroll.joiningDate
        toRet['accountNumber'] = user.payroll.accountNumber
        toRet['ifscCode'] = user.payroll.ifscCode
        toRet['bankName'] = user.payroll.bankName


        return Response(toRet,status = status.HTTP_200_OK)



# api to update leaves on montly bases
class DivideMontlyLeavesAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    def get(self, request, format=None):
        allUsers = User.objects.all()
        for user in allUsers:
            payroll  = user.payroll
            if payroll.al  is not None:
                alLeaves = int(payroll.al)/12
                if payroll.alCurrMonthLeaves is not None:
                    payroll.alCurrMonthLeaves = payroll.alCurrMonthLeaves + int(alLeaves)
                else:
                    payroll.alCurrMonthLeaves = int(alLeaves)
            if payroll.ml  is not None:
                mlLeaves = payroll.ml/12
                if payroll.mlCurrMonthLeaves is not None:
                    payroll.mlCurrMonthLeaves = payroll.mlCurrMonthLeaves + int(mlLeaves)
                else:
                    payroll.mlCurrMonthLeaves = int(mlLeaves)
            payroll.save()
        return Response({},status = status.HTTP_200_OK)


class GetUniqueYears(APIView):
    renderer_classes = (JSONRenderer,)
    def get(self, request, format=None):
        if 'id' in request.GET:
            user = User.objects.filter(pk = int(request.GET['id']))
        else:
            user = request.user
        yearLists = []
        if 'payslip' in request.GET:
            yearLists = PayrollReport.objects.filter(user = user).values_list('year', flat = True).distinct()
        else:
            yearLists = ITDecaration.objects.filter(user = user).values_list('year', flat = True).distinct()

        return Response({'yearLists' : yearLists},status = status.HTTP_200_OK)
