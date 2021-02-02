from django.contrib.auth.models import User , Group
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate , login , logout
from django.contrib.auth.decorators import login_required
from django.core.urlresolvers import reverse
from django.template import RequestContext
from django.conf import settings as globalSettings
# Related to the REST Framework
from rest_framework import viewsets , permissions , serializers
from rest_framework.exceptions import *
from url_filter.integrations.drf import DjangoFilterBackend
from .serializers import *
from .models import *
from API.permissions import *
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from allauth.account.adapter import DefaultAccountAdapter
from rest_framework.views import APIView
from rest_framework.renderers import JSONRenderer
from openpyxl import load_workbook,Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import PatternFill , Font,Alignment
from io import BytesIO,StringIO
import csv
from django.db.models import Case, IntegerField, Sum, When, Count
import json
import operator
from excel_response import ExcelResponse
# from ics import Calendar, Event
from email.MIMEBase import MIMEBase
import datetime
from dateutil import parser
from django.views.decorators.csrf import csrf_exempt
from django.template.loader import render_to_string, get_template
from django.core.mail import send_mail, EmailMessage
from clientRelationships.models import Contact, Contract
from ERP.models import service
from rest_framework import filters
import random, string
from finance.models import Sale,SalesQty
from finance.serializers import SaleLiteSerializer
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from HR.models import *
from organization.serializers import UnitLiteSerializer
from organization.models import Unit
import os,subprocess
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill , Font
from reportlab import *
from reportlab.pdfgen import canvas
from reportlab.lib.units import cm, mm
from reportlab.lib.pagesizes import A4
from reportlab.platypus import Paragraph, Table, TableStyle, Image, Frame, Spacer, PageBreak, BaseDocTemplate, PageTemplate, SimpleDocTemplate, Flowable
from PIL import Image
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet, TA_CENTER
from reportlab.graphics import barcode, renderPDF
from reportlab.graphics.shapes import *
from reportlab.graphics.barcode.qr import QrCodeWidget
import pytz
import requests, urllib
import sys, traceback
from django.template import Context , Template
from performance.models import TimeSheet
from HR.serializers import *
from num2words import num2words
from django.views.decorators.clickjacking import xframe_options_exempt
import googlemaps
from PIM.models import notification
from expenses import *
from ERP.send_push_message import send_push_message

class GetCampaignStatsAPIView(APIView):
    permission_classes = (permissions.IsAuthenticated , isAdmin)
    def get(self, request, format=None):
        campaign = Campaign.objects.get(pk = request.GET['id'],typ="call")
        myList = campaign.callitems.filter(owner = request.user , skipped = False , followUp = False, status = 'persuing')

        if myList.count()>0:
            item = myList.first()
        else:
            item = campaign.callitems.filter(owner__isnull = True , skipped = False , followUp = False , status = 'persuing' ).first()
            if item is None:
                return Response({"status" : "end"  }, status = status.HTTP_200_OK)
            item.owner = request.user
            item.save()

        logs = CampaignLogsSerializer(item.contact.CampaignLogs.all()  , many = True ).data

        d = EmailCampaignItemSerializer(item , many = False).data

        total = campaign.callitems.all().count()
        pending = campaign.callitems.filter(attempted = False).count()

        return Response({'total' : total , 'called' : total-pending , 'pending' :pending  , 'nextnumber' : d  })


def getContactsFromRequest(request):
    cntry = []
    for cntr in request.GET['country'].split(','):
        if cntr != "," or cntr != "":
            cntry.append(cntr)

    srces = []
    for src in request.GET['sources'].split(','):
        if src != "," or src != "":
            srces.append(src)
    conts = Contacts.objects.filter(country__in = cntry , source__in =  srces)
    for filt in request.GET['filters'].split(','):
        if filt == "":
            continue
        if filt == "name":
            conts = conts.exclude(name__isnull=True).exclude(name__exact='')
        if filt == "email":
            conts = conts.exclude(email__isnull=True).exclude(email__exact='')
        if filt == "companyName":
            conts = conts.exclude(companyName__isnull=True).exclude(companyName__exact='')
        if filt == "country":
            conts = conts.exclude(country__isnull=True).exclude(country__exact='')
        if filt == "city":
            conts = conts.exclude(city__isnull=True).exclude(city__exact='')
        if filt == "mobile":
            conts = conts.exclude(mobile__isnull=True).exclude(mobile__exact='')
        if filt == "website":
            conts = conts.exclude(website__isnull=True).exclude(website__exact='')

    return conts

class TagViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated ,)
    serializer_class = TagSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['name',]
    def get_queryset(self):
        request = self.request
        if 'filter' in request.GET:
            conts = getContactsFromRequest(request)
            return Tag.objects.filter(contacts__in  =conts ).distinct()

        if 'fetch' in self.request.GET:
            return Tag.objects.filter(name = self.request.GET['name__icontains'])
        else:
            return Tag.objects.all()
class CampaignViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated ,)
    serializer_class = CampaignSerializer
    queryset = Campaign.objects.all().order_by('-created')
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['name','lead','typ']


class CampaignLogsViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated ,)
    serializer_class = CampaignLogsSerializer
    queryset = CampaignLogs.objects.all()
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['user', 'contact' , 'campaign']

from django.db.models import Q


class EmailCampaignItemViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated ,)
    serializer_class = EmailCampaignItemSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['campaign' , 'contact','level'  ]
    def get_queryset(self):
        print "Campaign item viewset"
        qs = EmailCampaignItem.objects.all()
        if 'email' in self.request.GET:
            qs = qs.filter(contact__email__icontains= self.request.GET['email']  )

        return qs

class CallCampaignItemViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated ,)
    serializer_class = CallCampaignItemSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['campaign', 'status' , 'contact'  ]
    def get_queryset(self):
        print "Campaign item viewset"
        qs = CallCampaignItemViewSet.objects.all()
        if 'email' in self.request.GET:
            qs = qs.filter(contact__email__icontains= self.request.GET['email']  )

        return qs


class TourPlanViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    serializer_class = TourPlanSerializer
    queryset = TourPlan.objects.all().order_by('-id')
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['user','date','status']

class TourPlanStopViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    serializer_class = TourPlanStopSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['tourplan','id','contact','status','serviceName']
    def get_queryset(self):
        divsn = self.request.user.designation.division
        toReturn = TourPlanStop.objects.filter(division = divsn)
        if 'ascending' in self.request.GET:
            toReturn =  toReturn.order_by('status')
        else:
            toReturn =  toReturn.order_by('-id')
        if 'status' in self.request.GET:
            count = 0
            for i in self.request.GET['status'].split(','):
                count+=1
                if 'tourplan__user' in self.request.GET:
                    data = toReturn.filter(status = i,tourplan__user = self.request.GET['tourplan__user'])
                else:
                    data = toReturn.filter(status = i)
                print data
                if count == 1:
                    val = data
                else:
                    val = val | data
                toReturn =  val
            return toReturn
        else:
            return toReturn


class ContactsViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.AllowAny ,)
    serializer_class = ContactsSerializer
    # filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['name','source','email','pinCode','mobile' , 'created' , ]
    search_fields = ('name', 'email', 'pinCode', 'state' , 'website' , 'mobile' ,'notes', 'companyName' , 'directNumber' , 'altNumber' , 'altNumber2')
    def get_queryset(self):
        if not self.request.user.is_authenticated:
            return
        divsn = self.request.user.designation.division
        # toReturn = Contacts.objects.filter(creater__designation__division = divsn)
        toReturn = Contacts.objects.all()
        for i in globalSettings.SOURCE_LIST:
            if i in self.request.GET and int(self.request.GET[i]) == 0:
                toReturn = toReturn.exclude(source=i)
        if 'createddt' in self.request.GET:
            toReturn = toReturn.filter(created__year = 2020 )
            pass
        if 'leads' in self.request.GET:
            toReturn = toReturn.filter(need_service = True)
        if 'source' in self.request.GET:
            toReturn = toReturn.filter(source = self.request.GET['source'])
        if 'updated_search' in self.request.GET:
            dated = self.request.GET['updated_search']
            frmdate = datetime.datetime.strptime(str(dated), '%Y-%m-%d').strftime('%Y-%m-%d %H:%M:%S')
            frmdate = datetime.datetime.strptime(str(frmdate), '%Y-%m-%d %H:%M:%S')
            toDate = frmdate.replace(hour=23, minute=59, second=59, microsecond=999999)
            toReturn = toReturn.filter(updated__range = (frmdate , toDate))
        return toReturn

class GetCountriesApi(APIView):
    permission_classes = (permissions.IsAuthenticated , isAdmin)
    def get(self, request, format=None):
        toRet = Contacts.objects.values('country').annotate(Count('country'))
        return Response(toRet)

class GetSourcesApi(APIView):
    permission_classes = (permissions.IsAuthenticated , isAdmin)
    def get(self, request, format=None):
        return Response(Contacts.objects.values('source').annotate(Count('source')))

class GetContactsCountApi(APIView):
    permission_classes = (permissions.IsAuthenticated , isAdmin)
    def get(self, request, format=None):
        conts  = getContactsFromRequest(request)
        campaignData = list(conts.values('pk'))
        return Response({"total" : conts.count(),"campaignData":campaignData}, status = status.HTTP_200_OK)

class BulkContactsAPIView(APIView):
    permission_classes = (permissions.IsAuthenticated , isAdmin)
    def post(self, request, format=None):
        print 'ttttttttttt',request.FILES['fil'],request.POST['source'],

        fil = StringIO(request.FILES['fil'].read().decode('utf-8'))
        reader = csv.reader(fil, delimiter=':')
        count = 0
        for row in reader:
            dat = row[0].split(',')
            print 'aaaaaaaaaaaaa',dat
            try:
                check = Contacts.objects.get(email=dat[2],source=str(request.POST['source']))
            except:
                check = None
            if check:
                try:
                    check.name = dat[1]
                    check.referenceId = dat[0]
                    if len(dat)>3:
                        check.mobile = dat[3]

                    if len(dat) >4:
                        check.pinCode = dat[4]
                    check.save()
                    if 'tags' in request.POST:
                        for i in request.POST['tags'].split(','):
                            check.tags.add(Tag.objects.get(pk = int(i)))
                except:
                    pass
            else:
                contactData = {"name" : dat[1] ,  "email" : dat[2] ,"referenceId" : dat[0] , "source" : str(request.POST['source'])}
                if len(dat)>3:
                    contactData['mobile'] = dat[3]
                if len(dat) >4:
                    contactData['pinCode'] = dat[4]
                cObj = Contacts.objects.create(**contactData)
                if 'tags' in request.POST:
                    for i in request.POST['tags'].split(','):
                        cObj.tags.add(Tag.objects.get(pk = int(i)))
                count += 1

        return Response({"count" : count}, status = status.HTTP_200_OK)

class ContactsScrapedAPIView(APIView):
    permission_classes = (permissions.AllowAny,)
    def post(self, request, format=None):

        print 'ttttttttttt',request.POST
        count = 0
        check = Contacts.objects.filter(mobile=str(request.POST['mobile']))
        if len(check)>0:
            pass
        else:
            contactData = {"name" : str(request.POST['name']) , "mobile" : str(request.POST['mobile']) ,"source" : str(request.POST['source']) , "pinCode" : str(request.POST['pincode'])}
            print contactData
            cObj = Contacts.objects.create(**contactData)
            tgObj,created = Tag.objects.get_or_create(name=str(request.POST['tag']))
            cObj.tags.add(tgObj)
            count += 1
        print count

        return Response({"count" : count}, status = status.HTTP_200_OK)

class ConvertLeadApi(APIView):
    permission_classes = (permissions.AllowAny,)
    def post(self, request, format=None):
        print 'ttttttttttt',request.data,request.user
        leadObj = Leads.objects.get(pk=int(request.data['leadPk']))
        contactData = {'user':request.user,'name':leadObj.name,'email':leadObj.emailId,'mobile':leadObj.mobileNumber}
        if leadObj.jobLevel:
            contactData['designation'] = leadObj.jobLevel
        if leadObj.company:
            try:
                companyObj = service.objects.filter(name__iexact=leadObj.company)[0]
            except:
                companyObj = service(name=leadObj.company,user=request.user)
                companyObj.save()
            contactData['company'] = companyObj
        contactObj = Contact.objects.create(**contactData)
        leadObj.delete()
        return Response({}, status = status.HTTP_200_OK)

class SourceSuggestAPIView(APIView):
    permission_classes = (permissions.IsAuthenticated,)
    def post(self, request, format=None):

        print 'ttttttttttt',request.data

        if len(request.data['fd']) > 0:
            fromDate = request.data['fd'].split('-')
            toDate = request.data['td'].split('-')
            fd = date(int(fromDate[0]), int(fromDate[1]), int(fromDate[2]))
            td = date(int(toDate[0]), int(toDate[1]), int(toDate[2]))
            print fd,td
            if 'fetch' in request.data:
                duplicates = Contacts.objects.filter(source=str(request.data['source'])).values('source').annotate(pk=models.Max('id'),sourceCount=models.Count(Case(
                When(created__range=(str(fd),str(td)), then=1),
                output_field=IntegerField()
                )))
            else:
                duplicates = Contacts.objects.filter(source__contains=str(request.data['source'])).values('source').annotate(pk=models.Max('id'),sourceCount=models.Count(Case(
                When(created__range=(str(fd),str(td)), then=1),
                output_field=IntegerField()
                )))
            print list(duplicates)

            return Response({'val':list(duplicates)})
        else:
            duplicates = Contacts.objects.filter(source__contains=str(request.data['source'])).values('source').annotate(pk=models.Max('id'),sourceCount=models.Count('source'))
            print list(duplicates)
            return Response({'val':list(duplicates)})

from django.db.models import Max

class GetLastIDApi(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny , )
    def get(self, request, format=None):
        return Response({"res" : Contacts.objects.aggregate(Max('referenceId'))}, status = status.HTTP_200_OK)

class InvitationMailApi(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny , )
    def post(self, request, format=None):
        emailid=[]
        cc = ['ankita.k@cioc.in']
        print  request.data,'aaaaaaaaaa'
        schedule =  Schedule.objects.get(pk = request.data['value'])
        time = schedule.slot.split(' - ')
        strttime = int(time[0])
        endtime = int(time[1])
        newDate =  datetime.datetime.strptime(str(schedule.dated), '%Y-%m-%d').strftime('%Y%m%d %H:%M:%S')
        yourDate = parser.parse(newDate)

        begindate = yourDate.replace(hour=strttime, minute=00)
        enddate = yourDate.replace(hour=endtime, minute=00)
        emailid.append(schedule.emailId)
        email_subject ="Meeting as per your request"
        e = Event()
        e.name = "Meeting"
        e.begin = begindate
        e.end = enddate
        e.organizer = 'ankita.k@cioc.in'
        c.events.add(e)
        c.attendee =  schedule.emailId
        with open('my.ics', 'w') as f:
             f.writelines(c)
        ctx = {
            'dated': schedule.dated,
            'slot' :  schedule.slot,
            'linkUrl': 'cioc.co.in',
            'linkText' : 'View Online',
            'sendersAddress' : '(C) CIOC FMCG Pvt Ltd',
            'sendersPhone' : '841101',
            'linkedinUrl' : 'https://www.linkedin.com/company/13440221/',
            'fbUrl' : 'facebook.com',
            'twitterUrl' : 'twitter.com',
        }

        icspart = MIMEBase('text', 'calendar', **{'method' : 'REQUEST', 'name' : 'my.ics'})
        icspart.set_payload( open("my.ics","rb").read() )
        icspart.add_header('Content-Transfer-Encoding', '8bit')
        icspart.add_header('Content-class', 'urn:content-classes:calendarmessage')
        email_body = get_template('app.homepage.inviteemail.html').render(ctx)
        msg = EmailMessage(email_subject, email_body,  to= emailid, cc= cc )
        msg.attach(icspart)
        msg.content_subtype = 'html'
        msg.send()

        return Response({}, status = status.HTTP_200_OK)

class AddCampaign(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny , )
    def post(self, request, format=None):
        campaignObj = Campaign.objects.create(name = request.data['name'],typ =request.data['typ'], status = request.data['status']  )
        campaignObj.save()
        # if 'contacts' in request.data:
        #     campaignObj.participants.clear()
        #     for i in request.data['contacts']:
        #         campaignObj.participants.add(User.objects.get(pk = int(i)))
        if 'msgBody' in request.data:
            campaignObj.msgBody = request.data['msgBody']
        if 'emailSubject' in request.data:
            campaignObj.emailSubject = request.data['emailSubject']
        if 'emailBody' in request.data:
            campaignObj.emailBody = request.data['emailBody']
        if 'emailTemplate' in request.data:
            campaignObj.emailTemplate = request.data['emailTemplate']
        if 'limitPerDay' in request.data:
            campaignObj.limitPerDay = request.data['limitPerDay']
        if 'directions' in request.data:
            campaignObj.directions = request.data['directions']
        if 'team' in request.data:
            campaignObj.team = Team.objects.get(pk = request.data['team'])
        if 'lead' in request.data:
            campaignObj.lead = User.objects.get(pk = request.data['lead'])
        campaignObj.save()
        if len(request.data['contacts'])>0:
            tags = Tag.objects.filter(pk__in = request.data['tags'])
            print "tags" , tags
            if tags.count()>0:
                contcts = Contacts.objects.filter(pk__in = request.data['contacts'] , tags__in = tags ).distinct()
            else:
                contcts = Contacts.objects.filter(pk__in = request.data['contacts'] ).distinct()

            print "contcts" , contcts

            for k in contcts:
                if request.data['typ'] =='email':
                    campaigndata = EmailCampaignItem.objects.create(contact = k ,campaign = campaignObj)
                    campaigndata.save()
                if request.data['typ']=='call':
                    campaigndata = CallCampaignItem.objects.create(contact = k ,campaign = campaignObj)
                    campaigndata.save()


        return Response({}, status = status.HTTP_200_OK)

class GetNextEntryApiView(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny , )
    def get(self, request, format=None):
        if 'level' in request.GET:
            item = EmailCampaignItem.objects.filter(  status = 'persuing', campaign__lead = request.user , level = request.GET['level']).first()
            if item is None:
                return Response({"status" : "end"  }, status = status.HTTP_200_OK)
            item.owner = request.user
            item.save()

            logs = CampaignLogsSerializer(item.contact.CampaignLogs.all()  , many = True ).data

            d = EmailCampaignItemSerializer(item , many = False).data
            return Response({"contact" : d , "logs" : logs, "status" : "ok"  }, status = status.HTTP_200_OK)
        else:
            myList = Campaign.objects.get(pk = request.GET['id']).callitems.filter(owner = request.user , skipped = False , followUp = False, status = 'persuing')

            if myList.count()>0:
                item = myList.first()
            else:
                item = Campaign.objects.get(pk = request.GET['id']).callitems.filter(owner__isnull = True , skipped = False , followUp = False , status = 'persuing' ).first()
                if item is None:
                    return Response({"status" : "end"  }, status = status.HTTP_200_OK)
                item.owner = request.user
                item.save()

            logs = CampaignLogsSerializer(item.contact.CampaignLogs.all()  , many = True ).data

            d = CallCampaignItemSerializer(item , many = False).data
            return Response({"contact" : d , "logs" : logs, "status" : "ok"  }, status = status.HTTP_200_OK)

def cleanName(name):
    if name is None:
        return
    for salut in ['Miss ' , 'Dr ' , 'Mr ' , 'Mrs ', 'Ms ' ]:
        if name.startswith(salut):
            return name.replace(salut , '')
    return name


class ContactsBulkUploadAPIView(APIView):
    permission_classes = (permissions.IsAuthenticated , isAdmin)
    def post(self, request, format=None):
        excelFile = request.FILES['exFile']
        wb = load_workbook(filename = BytesIO(excelFile.read()) ,  read_only=True)
        count = 0
        contacts = []
        isSales = False
        isHigh = False
        if 'dataType' in request.data:
            if request.data['dataType'] == 'Sales':
                isSales = True
        if 'priority' in request.data:
            if request.data['priority'] == 'urgent':
                isHigh = True
        for ws in wb.worksheets:
            wsTitle = ws.title
            for i in range(2,ws.max_row+1):
                name = cleanName(ws['A' + str(i)].value)
                # print name

                try:
                    email = ws['B' + str(i)].value
                except:
                    email = None

                try:
                    mobile = ws['C' + str(i)].value
                except:
                    mobile = None

                try:
                    pincode = ws['D' + str(i)].value
                except:
                    pincode = None

                try:
                    address = ws['E' + str(i)].value
                except:
                    address = None
                try:
                    city = ws['F' + str(i)].value
                except:
                    city = None
                try:
                    state = ws['G' + str(i)].value
                except:
                    state = None
                try:
                    country = ws['H' + str(i)].value
                except:
                    country = None

                try:
                    areaCode = ws['I' + str(i)].value
                except:
                    areaCode = None
                count+=1
                print email,'aaaaaaaaaaaaa'
                contactObj , n = Contacts.objects.get_or_create( mobile = mobile)
                contactObj.name = name
                contactObj.email = email
                contactObj.pinCode = pincode
                contactObj.addrs = address
                contactObj.city = city
                contactObj.state = state
                contactObj.country = country
                if 'source' in request.data:
                    contactObj.source = request.data['source']
                contactObj.isSales = isSales
                contactObj.save()
                print contactObj.pk
                contacts.append(contactObj)
        if len(contacts)>0:
            contactList = ContactsLiteSerializer(contacts,many = True).data
            if 'generate' in request.data:
                userObj =  request.user
                teamObj = Team.objects.filter(manager = userObj)
                name = request.data['source'] + ' - ' + userObj.first_name
                campaignObj = Campaign.objects.create(lead = userObj, typ='call' , name = name, isSales = isSales,isHigh = isHigh)
                if teamObj.count()>0:
                    campaignObj.team = teamObj.first()
                    campaignObj.save()
                for i in contacts:
                    campItemObj = CampaignItem.objects.create(contact = i,campaign =  campaignObj)

        else:
            contactList = []
        return Response({"count" : count , 'contactList' : contactList}, status = status.HTTP_200_OK)

class CreatePlanAPI(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny ,)
    def post(self,request , format= None):
        divsn =  request.user.designation.division
        try:
            dated = datetime.datetime.strptime(request.data['date'], "%Y-%m-%dT%H:%M:%S.%fZ").date()
        except:
            dated = request.data['date']
        planObj , a = TourPlan.objects.get_or_create(date = dated, user = User.objects.get(pk = request.data['user']))
        if 'pk' in request.data:
            planStop = TourPlanStop.objects.get(pk = int(request.data['pk']))
            planStop.visitType = request.data['visitType']
            planStop.timeslot = request.data['timeslot']
            planStop.tourplan = planObj
            planStop.status = 'assigned'
            planStop.save()
            if 'customer' in request.data:
                customer = request.data['customer']
                custObj = Contacts.objects.get(pk = int(customer['pk']))
                custObj.name = customer['name']
                custObj.email = customer['email']
                custObj.mobile = customer['mobile']
                custObj.source = customer['source']
                custObj.addrs = customer['addrs']
                custObj.city = customer['city']
                custObj.state = customer['state']
                custObj.country = customer['country']
                custObj.pinCode = customer['pinCode']
                custObj.lat = customer['lat']
                custObj.lng = customer['lng']
                custObj.about = customer['about']
                custObj.save()
        else:
            custObj = Contacts.objects.get(pk = request.data['contact'])
            planStop = TourPlanStop.objects.create(contact = custObj, tourplan = planObj, visitType = request.data['visitType'], timeslot = request.data['timeslot'], division = divsn)
            if 'serviceName' in request.data:
                planStop.serviceName = request.data['serviceName']
            planStop.save()
            custObj.need_service = False
            custObj.save()
        txt = 'Technician ' + planStop.tourplan.user.first_name + ' ' + planStop.tourplan.user.last_name + 'is been assigned'
        logObj = Log.objects.create(txt = txt, type ='assigned' , plan = planStop )
        logObj.save()

        if planStop.tourplan.user.profile.expoPushToken!=None and len(planStop.tourplan.user.profile.expoPushToken) > 0:
            try:
                send_push_message(str(planStop.tourplan.user.profile.expoPushToken) , 'One task was assigned '  , {'id' : planStop.pk  }  )
            except:
                pass

        if 'firstComment' in  request.data:
            logObj = Log.objects.create(txt = request.data['firstComment']  , type ='text' , plan = planStop )
            logObj.save()
        contactObj = planStop.contact
        if contactObj.contact_ref == None:
            newContact = Contact.objects.create(mobile = contactObj.mobile, user=request.user)
        else:
            newContact = Contact.objects.get(pk = contactObj.contact_ref.pk)
        newContact.name = contactObj.name
        newContact.email = contactObj.email
        newContact.mobile = contactObj.mobile
        newContact.street = contactObj.addrs
        newContact.city = contactObj.city
        newContact.state = contactObj.state
        newContact.country = contactObj.country
        newContact.pincode = contactObj.pinCode
        newContact.save()
        contactObj.contact_ref = newContact
        contactObj.save()

        if 'pk' in request.data:
            pass
        else:
            message = 'New service request received. Customer : {0} Time : {1} Address : {2}'.format(planStop.contact.name, planStop.timeslot, planStop.contact.addrs)
            number = str(planStop.tourplan.user.profile.mobile)
            try:
                globalSettings.SEND_WHATSAPP_MSG(number,message)
            except:
                pass

        return Response({},status = status.HTTP_200_OK)

from marketing.models import Contacts
def generateOTPCode():
    length = 4
    chars = string.digits
    rnd = random.SystemRandom()
    return ''.join(rnd.choice(chars) for i in range(length))

class RequestOtpAPI(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny ,)
    def post(self,request , format= None):
        print request.data,'aaaaaaa'
        if request.data['typ'] == 'startOtp':
            tourObj = TourPlanStop.objects.get(pk = request.data['id'])
            otp = generateOTPCode()
            tourObj.startOTP = otp
            tourObj.save()
            msg = 'Your otp is ' + otp
            print msg
            txt = 'OTP is been sent to the user to start the job'
            logObj = Log.objects.create(txt = txt, type ='text' , plan = tourObj )
            logObj.save()
            globalSettings.SEND_SMS(tourObj.contact.mobile, msg)
        if request.data['typ'] == 'endOtp':
            tourObj = TourPlanStop.objects.get(pk = request.data['id'])
            otp = generateOTPCode()
            tourObj.completeOTP = otp
            tourObj.save()
            txt = 'OTP is been sent to the user to stop the job'
            logObj = Log.objects.create(txt = txt, type ='text' , plan = tourObj )
            logObj.save()
            msg = 'Your otp is ' + otp
            globalSettings.SEND_SMS(tourObj.contact.mobile, msg)
        if request.data['typ'] == 'postponedOtp':
            tourObj = TourPlanStop.objects.get(pk = request.data['id'])
            otp = generateOTPCode()
            print otp,'aaaaaaaaaaaaa'
            tourObj.cancelOTP = otp
            tourObj.save()
            txt = 'OTP is been sent to the user to postponed the job'
            logObj = Log.objects.create(txt = txt, type ='text' , plan = tourObj )
            logObj.save()
            msg = 'Your otp is ' + otp
            globalSettings.SEND_SMS(tourObj.contact.mobile, msg)
        return Response({},status = status.HTTP_200_OK)



class VerifyOtpAPI(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny ,)
    def post(self,request , format= None):
        from datetime import datetime
        statusVal = 'Un-Verified'
        tourObj = TourPlanStop.objects.get(pk = request.data['id'])
        if 'typ' in request.data:
            if request.data['typ'] == 'verifystartOtp':
                tourObj.status = 'ongoing'
                tourObj.save()
                statusVal = 'Verified'
                msg = 'Your job has been started'
                txt = 'Your job has been started'
                logObj = Log.objects.create(txt = txt, type ='text' , plan = tourObj )
                logObj.save()
                globalSettings.SEND_SMS(tourObj.contact.mobile, msg)
            if request.data['typ'] == 'verifyendOtp':
                tourObj.status = 'completed'
                tourObj.save()
                contactObj = tourObj.contact
                if contactObj.contact_ref is None:
                    newContact = Contact.objects.create(name = contactObj.name , email = contactObj.email , mobile = mobile, street =  contactObj.addrs , city =  contactObj.city , state =  contactObj.state , country = contactObj.country, pincode =  contactObj.pinCode )
                    contactObj.contact_ref = newContact
                contactObj.save()
                statusVal = 'Verified'
                txt = 'Your job has been ended'
                logObj = Log.objects.create(txt = txt, type ='text' , plan = tourObj )
                logObj.save()
                msg = 'Your job has been ended'
                globalSettings.SEND_SMS(tourObj.contact.mobile, msg)
            if request.data['typ'] == 'verifypostpondedOtp':
                tourObj.status = 'postponed'
                tourObj.is_postponded = True
                tourObj.postponded_date =  request.data['postponded_date']
                tourObj.timeslot = request.data['timeslot']
                if 'general_comment' in request.data:
                    tourObj.general_comment = request.data['general_comment']
                    logObj = Log.objects.create(txt = request.data['general_comment'], type ='text' , plan = tourObj )
                    logObj.save()
                tourObj.save()
                statusVal = 'Verified'
                txt = 'Your job has been postponed'
                logObj = Log.objects.create(txt = txt, type ='text' , plan = tourObj )
                logObj.save()
                msg = 'Your job has been postponed'
                globalSettings.SEND_SMS(tourObj.contact.mobile, msg)
        if 'techStatus' in request.data:
            tourObj.techStatus = request.data['techStatus']
            tourObj.save()
            statusVal = tourObj.techStatus
        return Response({"status" : statusVal  }, status = status.HTTP_200_OK)



class CreateInvoiceAPI(APIView): # from the field staff mobile app
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny ,)
    def post(self,request , format= None):
        from dateutil.relativedelta import relativedelta
        # if 'delete' in request.data:
        divsn =  request.user.designation.division
        print request.data,'aaaaaaaaaaaaaaaaaaaaaaa', request.data['tourplan']
        products = request.data['products']
        price = 0
        totalTax = 0
        print request.user
        contact = Contacts.objects.get(pk = request.data['contact'])
        tourplan = TourPlanStop.objects.get(pk = request.data['tourplan'])
        invoicePk = None
        try:
            invoicePk = tourplan.invoice.pk
        except:
            pass

        print invoicePk, 'invoicePk'
        dated = datetime.date.today()
        data = {'personName':contact.name,'phone' : contact.mobile , 'email': contact.email ,'address': contact.addrs , 'pincode': contact.pinCode  ,'state': contact.state ,'city':  contact.city,'country': contact.country , 'user' : request.user, 'recDate': dated }
        if invoicePk is not None:
            invObj = OutBoundInvoice.objects.get(pk=invoicePk)
            invObj.personName = data['personName']
            invObj.phone = data['phone']
            invObj.email = data['email']
            invObj.address = data['address']
            invObj.pincode = data['pincode']
            invObj.state = data['state']
            invObj.city = data['city']
            invObj.country = data['country']
            invObj.user = data['user']
            invObj.recDate = data['recDate']
            invObj.poNumber = invObj.pk
            invObj.save()

        else:
            invObj = OutBoundInvoice(**data)
            invObj.poNumber = invObj.pk
            invObj.save()

        for i in products:
            assetObj = Asset.objects.get(pk = i['asset'])
            checkinObj = Checkin.objects.filter(asset = assetObj, checkout = False, returned = False, to = request.user)
            allOutBondObj = OutBoundInvoiceQty.objects.filter(asset = i['asset'], addon = i['addon']  ,outBound = invObj).count()
            if allOutBondObj< int(i['quantity']):
                quantity = int(i['quantity']) - allOutBondObj
                if 'amc' in request.data:
                    for checkin in checkinObj:
                        print checkin.price
                    if request.data['amc'] == 'applicable':
                        if assetObj.primary:
                            for j in range(int(quantity)):
                                taxRate = 0
                                if assetObj.taxPercentage is not None:
                                    taxRate = assetObj.taxPercentage
                                addonObj = AssetAddon.objects.get(pk = int(i['addon']))
                                taxPrice = (addonObj.rate*taxRate)/100
                                subtotal = addonObj.rate + taxPrice
                                if 'addons' in i:
                                    today = datetime.date.today()
                                    addonObj = AssetAddon.objects.get(pk = int(i['addons']))
                                    period = addonObj.period
                                    if addonObj.typ == 'monthly':
                                        service = 12*int(period)
                                        monthslimit = 1
                                    if addonObj.typ == 'quarterly':
                                        service = 4*int(period)
                                        monthslimit = 3
                                    if addonObj.typ == 'half-yearly':
                                        service = 2*int(period)
                                        monthslimit = 6
                                    if addonObj.typ == 'yearly':
                                        service = 1*int(period)
                                        monthslimit = 12
                                    for c in range(0 , service):
                                        if c == 0:
                                            tourDate =  today
                                        else:
                                            val = monthslimit * c
                                            tourDate =  today + relativedelta(months=val)
                                        comments =  assetObj.name + '||'+addonObj.label
                                        tourObj, a = TourPlan.objects.get_or_create(date = tourDate, user = request.user)
                                        stopObj = TourPlanStop.objects.create(contact = contact , tourplan = tourObj ,  assignedBy = request.user, status="upcoming" , timeslot = '9 AM' , comments = comments, division = divsn)
                                outData = {'product' : assetObj.name , 'qty' : 1,'price' : checkinObj[j].price ,'tax':taxRate, 'total' : subtotal , 'outBound' : invObj, 'asset' : assetObj.pk, 'addon' : i['addon']}

                                outObj = OutBoundInvoiceQty(**outData)
                                outObj.save()

                                try:
                                    check = checkinObj[j]
                                    check.checkout = True
                                    check.name = contact.name
                                    check.email = contact.email
                                    check.phone = contact.mobile
                                    check.address = contact.addrs
                                    check.usedIn = outObj
                                    check.save()
                                except:
                                    print "Trying to add a non assigned asset in the invoice"
                                    pass
                        else:
                            assetObj = Asset.objects.get(pk = i['asset'])
                            checkinObj = Checkin.objects.filter(asset = assetObj, checkout = False, returned = False, to = request.user)
                            for j in range(int(quantity)):
                                outData = {'product' : assetObj.name , 'qty' : 1,'price' : 0 ,'total' : 0 , 'outBound' : invObj, 'asset' : assetObj.pk}
                                outObj = OutBoundInvoiceQty(**outData)
                                outObj.save()
                                try:
                                    check = checkinObj[j]
                                    check.checkout = True
                                    check.name = contact.name
                                    check.email = contact.email
                                    check.phone = contact.mobile
                                    check.address = contact.addrs
                                    check.usedIn = outObj
                                    check.save()
                                except:
                                    print "Trying to add a non assigned asset in the invoice"
                                    pass

                    elif request.data['amc'] == 'notapplicable':
                        print i['asset']
                        assetObj = Asset.objects.get(pk = i['asset'])
                        checkinObj = Checkin.objects.filter(asset = assetObj, checkout = False, returned = False, to = request.user)
                        for j in range(int(quantity)):
                            taxRate = 0
                            if assetObj.taxPercentage is not None:
                                taxRate = assetObj.taxPercentage
                            taxPrice = (checkinObj[j].price*taxRate)/100
                            subtotal = checkinObj[j].price + taxPrice
                            outData = {'product' : assetObj.name , 'qty' : 1,'price' : checkinObj[j].price , 'total' :subtotal  ,'tax':taxRate, 'outBound' : invObj, 'asset' : assetObj.pk}
                            outObj = OutBoundInvoiceQty(**outData)
                            outObj.save()
                            try:
                                check = checkinObj[j]
                                check.checkout = True
                                check.name = contact.name
                                check.email = contact.email
                                check.phone = contact.mobile
                                check.address = contact.addrs
                                check.usedIn = outObj
                                check.save()
                            except:
                                print "Trying to add a non assigned asset in the invoice"
                                pass


        price = 0
        totalTax = 0
        for inv in invObj.outBoundQty.all():
            price+=inv.total
        invObj.total = price
        invObj.save()
        tourplan.invoice = invObj
        tourplan.save()
        data = OutBoundInvoiceLiteSerializer(invObj, many=False).data
        msg = 'Your invoice has been initiated with an amount of Rs ' + str(price)
        return Response({'invoice':data }, status = status.HTTP_200_OK)

class DeleteInvoiceAPI(APIView): # from the field staff mobile app
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny ,)
    def post(self,request , format= None):

        if 'tourpk' in request.data:
            try:
                invoiceObj = TourPlanStop.objects.get(pk = request.data['tourpk']).invoice
            except:
                pass
        if 'invpk' in request.data:
            invoiceObj = OutBoundInvoice.objects.get(pk=int(request.data['invpk']))
        outBondObj = OutBoundInvoiceQty.objects.filter(asset = request.data['asset'], addon = request.data['addon']  ,outBound = invoiceObj).delete()
        price = 0
        totalTax = 0
        for inv in invoiceObj.outBoundQty.all():
            totalTax+=inv.tax
            price+=inv.total
        invoiceObj.total = price
        invoiceObj.totalGST = totalTax
        invoiceObj.save()
        data = OutBoundInvoiceLiteSerializer(invoiceObj, many=False).data
        return Response({'invoice':data }, status = status.HTTP_200_OK)

class getVisitsAPI(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny ,)
    def get(self,request , format= None):
        user = request.user
        newData = []
        ongoingData = []
        followUpData = []
        pending = []
        from datetime import date
        today = date.today()
        if 'id' in request.GET:
            newObj = TourPlanStop.objects.get(pk = int(request.GET['id']))
            data = TourPlanStopLiteSerializer(newObj, many = False).data
            return Response({'data' : data }, status = status.HTTP_200_OK)
        if 'typ' in request.GET:
            newObj = TourPlanStop.objects.filter(status = 'assigned' , tourplan__user = user).exclude(is_postponded = True)
            ongoingObj = TourPlanStop.objects.filter(status = 'ongoing' , tourplan__user = user).exclude(is_postponded = True)
            completedObj = TourPlanStop.objects.filter(status = 'completed' , tourplan__user = user).exclude(is_postponded = True)
            if request.GET['typ'] == 'CustomPeriod':
                if 'date1' in request.GET:
                    date1 = request.GET['date1']
                if 'date2' in request.GET:
                    date2 = request.GET['date2']
                if date2<date1:
                    fromDate = date2
                    toDate = date1
                else:
                    fromDate = date1
                    toDate = date2
                newObj = newObj.filter(tourplan__date__range = (fromDate , toDate))
                ongoingObj = ongoingObj.filter(tourplan__date__range = (fromDate , toDate))
                completedObj = completedObj.filter(tourplan__date__range = (fromDate , toDate))
            elif request.GET['typ'] == 'Today':
                newObj = newObj.filter(tourplan__date = today)
                ongoingObj = ongoingObj.filter(tourplan__date = today)
                completedObj = completedObj.filter(tourplan__date = today)
                print newObj , ongoingObj
            elif request.GET['typ'] == 'Yesterday':
                yesterday = today - timedelta(days = 1)
                newObj = newObj.filter(tourplan__date = yesterday)
                ongoingObj = ongoingObj.filter(tourplan__date = yesterday)
                completedObj = completedObj.filter(tourplan__date = yesterday)
            elif request.GET['typ'] == 'ThisWeek':
                noDays =  int(today.strftime('%w'))
                fromDate = today - timedelta(days=int(noDays))
                toDate = fromDate + timedelta(days=6)
                newObj = newObj.filter(tourplan__date__range = (fromDate , toDate))
                ongoingObj = ongoingObj.filter(tourplan__date__range = (fromDate , toDate))
                completedObj = completedObj.filter(tourplan__date__range = (fromDate , toDate))
            elif request.GET['typ'] == 'LastWeek':
                noDays =  int(today.strftime('%w'))
                sundayDate = today - timedelta(days=int(noDays))
                fromDate = sundayDate - timedelta(days=7)
                toDate = sundayDate - timedelta(days=1)
                newObj = newObj.filter(tourplan__date__range = (fromDate , toDate))
                ongoingObj = ongoingObj.filter(tourplan__date__range = (fromDate , toDate))
                completedObj = completedObj.filter(tourplan__date__range = (fromDate , toDate))
            elif request.GET['typ'] == 'ThisMonth':
                fromDate = today.replace(day=1)
                toDate = today + relativedelta(day=31)
                newObj = newObj.filter(tourplan__date__range = (fromDate , toDate))
                ongoingObj = ongoingObj.filter(tourplan__date__range = (fromDate , toDate))
                completedObj = completedObj.filter(tourplan__date__range = (fromDate , toDate))
            elif request.GET['typ'] == 'LastMonth':
                toDate = today.replace(day=1) - datetime.timedelta(days=1)
                fromDate = toDate.replace(day=1)
                newObj = newObj.filter(tourplan__date__range = (fromDate , toDate))
                ongoingObj = ongoingObj.filter(tourplan__date__range = (fromDate , toDate))
                completedObj = completedObj.filter(tourplan__date__range = (fromDate , toDate))
            completed = TourPlanStopLiteSerializer(completedObj, many = True).data
            newData = TourPlanStopLiteSerializer(newObj, many = True).data
            ongoingData = TourPlanStopLiteSerializer(ongoingObj, many = True).data
        pending = TourPlanStopLiteSerializer(TourPlanStop.objects.filter(status = 'assigned' , tourplan__user = user,tourplan__date__lt = today).exclude(is_postponded = True), many = True).data

        followUpObj1 = TourPlanStop.objects.filter(tourplan__user = user,tourplan__date__gt = today,status = 'assigned')
        followUpObj2 =  TourPlanStop.objects.filter(tourplan__user = user, postponded_date__gt = today , is_postponded = True)
        followUpObj = followUpObj1 | followUpObj2
        followUpData = TourPlanStopLiteSerializer(followUpObj, many = True).data
        data = {'new' : newData , 'ongoing' : ongoingData , 'followup' : followUpData , 'pending':pending , 'completed' :completed}
        return Response({'data' : data }, status = status.HTTP_200_OK)


from pydub import AudioSegment
class addAudioAPI(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny ,)
    def post(self,request , format= None):
        user = request.user
        if 'visit' in request.data:
            visit_pk = request.data['visit']
            visitObj = TourPlanStop.objects.get(pk = visit_pk)
            type = 'Recording'
        else:
            print request.FILES['audioFile'].name
            filename = request.FILES['audioFile'].name
            type = 'Recording'
            tempno = filename.split('[+')[1]
            phone = tempno.split(']')[0]
            visitObj = TourPlanStop.objects.filter(contact__mobile = phone).last()
        file_Data = request.FILES['audioFile']
        file_name = file_Data.name
        logObj = Log.objects.create( type = type , plan = visitObj , file = file_Data)
        logObj.save()
        return Response({}, status = status.HTTP_200_OK)


class addCallAudioAPI(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny ,)
    def post(self,request , format= None):
        user = request.user
        if 'visit' in request.data:
            visit_pk = request.data['visit']
            visitObj = TourPlanStop.objects.get(pk = visit_pk)
        else:
            print request.FILES['audioFile'].name
            filename = request.FILES['audioFile'].name
            if (filename.find('INC') != -1):
                type = 'Incoming'
            elif (filename.find('OUT') != -1):
                type = 'Outgoing'
            else:
                type = 'Missed'
            print filename,'filenamefilename'

            mobileNumber = filename.split(']')[1]
            mobileNumber = mobileNumber.split('_[')[1]
            if len(mobileNumber)>=13:
                mobileNumber = mobileNumber.split('+91')[1]
            elif len(mobileNumber)>10:
                mobileNumber = mobileNumber.split('91')[1]
        visitObj = TourPlanStop.objects.filter(contact__mobile = mobileNumber).last()
        file_Data = request.FILES['audioFile']
        file_name = file_Data.name
        logObj = Log.objects.create( type = type , plan = visitObj , file = file_Data)
        logObj.save()
        basePath = globalSettings.MEDIA_ROOT
        print basePath, logObj.file, 'file name'
        wav = str(basePath)+'/'+str(logObj.file)
        print wav, 'wav path'
        cmd = 'lame --preset insane %s' % wav
        subprocess.call(cmd, shell=True)
        print logObj.file
        logObj.file = str(logObj.file).split('.')[0]+'.mp3'
        logObj.save()
        os.remove(wav)

        return Response({}, status = status.HTTP_200_OK)

from pydub import AudioSegment
class checkAssetAPI(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny ,)
    def get(self,request , format= None):
        if 'serialNo' in request.GET:
            checkinObj = Checkin.objects.filter(serialNo = request.GET['serialNo']).first()
            checkinCount = Checkin.objects.filter(to = None , returned = False , usedIn = None  ).count()
            try:
                msg = 'Successfull'
                data = {'count' : checkinCount, 'asset':  checkinObj.asset.pk , 'assetName':  checkinObj.asset.name ,'price':checkinObj.asset.price,'is_charges':checkinObj.asset.is_charges}
            except:
                msg ='Unsuccessfull'
                data = {'count' : 0}
        if 'asset' in request.GET:
            data = []
            checkinObj = Assets.objects.filter(name__icontains = request.GET['asset'])
            for i in checkinObj:
                checkinCount = 0
                checkinCount = i.checkins.filter(to = None , returned = False , usedIn = None).count()
                objData = {'count' : checkinCount, 'asset':  i.pk , 'assetName':  i.name ,'price':i.price,'is_charges':i.is_charges}
                data.append(objData)
            msg = 'Successfully'
        return Response({'msg':msg , 'data' : data}, status = status.HTTP_200_OK)
    def post(self,request , format= None):
        if request.user.is_staff == False:
            return Response({ 'error' : 'Not a permitted User'}, status = status.HTTP_200_OK)
        try:
            serialNos = request.data['serialNos']
            serialNos = serialNos.split(' , ')
        except:
            serialNos = [request.data['serialNos']]
        data = []
        user = request.data['user']
        UserObj = User.objects.get(pk = user)
        for value in serialNos:
            chekinObj = Checkin.objects.filter(serialNo = value, returned = False , usedIn = None , to = None).first()
            try:
                from datetime import datetime
                chekinObj.to = UserObj
                chekinObj.assignedBy = request.user
                chekinObj.assignedOn = datetime.now()
                chekinObj.save()
                print chekinObj.assignedBy.pk, chekinObj.to.pk, chekinObj.assignedOn
                val = {'serialNo' : chekinObj.serialNo , 'pk': chekinObj.pk , 'assetPK' : chekinObj.asset.pk ,'msg' : 'Successfull'}
                data.append(val)
            except:
                val = {'serialNo' : value ,'msg' : 'Already Used' }
                data.append(val)
        print 'heeeeeeeeerrrrrrrrrrrrrrrrrrrrrrr',data
        return Response({ 'data' : data}, status = status.HTTP_200_OK)

class searchChekinAPI(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny ,)
    def get(self,request , format= None):
        if 'user' in request.GET:
            userObj = User.objects.get(pk = request.GET['user'])
            chwckinO = Checkin.objects.filter(to = userObj, checkout = True)
            chekinObj = chwckinO.values_list('price','asset__id').distinct()
            data = []
            for i in chekinObj:
                count = 1
                countData = chwckinO.filter(asset__id = i[1] , price = i[0])
                count = countData.count()
                val = {'pk' : i[1] , 'price' : i[0] , 'count' : count , 'name' : countData[0].asset.name}
                data.append(val)
            print data

        return Response(data, status = status.HTTP_200_OK)


class DashBoardAPI(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny ,)
    def get(self,request , format= None):
        from datetime import date
        today_date = date.today()
        completedCount = 0
        visitCount = 0
        invoiceCount = 0
        first_day_of_month = today_date.replace(day=1)
        first = first_day_of_month - timedelta(days=1)
        last = first_day_of_month + relativedelta(months=1)
        user = request.user
        desg = designation.objects.filter(user = user).first()
        unitObj = UnitLiteSerializer(Unit.objects.get(pk = desg.unit.pk), many = False).data
        userObj = userSearchSerializer(user, many = False).data
        visitObj = TourPlanStop.objects.filter(tourplan__user = user , tourplan__date__gt = first,  tourplan__date__lt = last)
        visitCount = visitObj.count()
        allotedCount = visitObj.filter(status = 'assigned').count()
        completedCount = visitObj.filter(status = 'completed').count()
        invoiceTotal = OutBoundInvoiceQty.objects.filter(outBound__user=user , created__gt = first,  created__lt = last).aggregate(sum=Sum('total'))['sum']
        if invoiceTotal == None:
            invoiceTotal = 0
        invoiceCount = OutBoundInvoice.objects.filter(user=user , created__gt = first,  created__lt = last ).count()
        try:
            visitefficiency = (completedCount / visitCount) * 100
        except:
            visitefficiency = 0
        try:
            invoiceEff = (invoiceCount / visitCount) * 100
        except:
            invoiceEff = 0
        data = {'userObj' : userObj, 'visitCount' : visitCount, 'completedCount' : completedCount , 'invoiceTotal' : invoiceTotal , 'allotedCount' : allotedCount, 'invoiceTotal' : invoiceTotal , 'invoiceCount' : invoiceCount , 'visitefficiency' : visitefficiency , 'invoiceEff' : invoiceEff , 'unitObj' : unitObj}
        return Response({'data' : data}, status = status.HTTP_200_OK)

class AdminDashBoardAPI(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny,)
    def get(self,request , format= None):
        from datetime import date
        invoiceTotal = 0
        totalDays = request.GET['days']
        today_date = date.today()
        if totalDays == '7' :
            from_date = today_date - timedelta(days=7)
            to_date = today_date + timedelta(days=1)
        if totalDays == '15' :
            from_date = today_date - timedelta(days=15)
            to_date = today_date + timedelta(days=1)
        if totalDays == '3' :
            from_date_val = today_date - timedelta(days=1)
            from_date = from_date_val - relativedelta(months=3)
            to_date = today_date + timedelta(days=1)
        visitObj = TourPlanStop.objects.filter(tourplan__date__gt = from_date , tourplan__date__lt = to_date)
        visitCount = visitObj.count()
        newCount = visitObj.filter(status = 'assigned').exclude(is_postponded = True).count()
        postponedCount = visitObj.filter(is_postponded = True).count()
        completedCount = visitObj.filter(status = 'completed').exclude(is_postponded = True).count()
        cancelledCount = visitObj.filter(status = 'cancelled').exclude(is_postponded = True).count()
        ongoingCount = visitObj.filter(status = 'ongoing').exclude(is_postponded = True).count()
        invoiceTotal = OutBoundInvoiceQty.objects.filter(created__gt = from_date , created__lt = to_date).aggregate(sum=Sum('total'))['sum']
        if invoiceTotal == None:
            invoiceTotal = 0
        customerCount = Contacts.objects.filter(created__gt = from_date , created__lt = to_date).count()
        data = { 'visitCount' : visitCount, 'completedCount' : completedCount,'newCount': newCount, 'postponedCount' : postponedCount , 'invoiceTotal' : invoiceTotal ,'cancelledCount' : cancelledCount , 'customerCount': customerCount , 'ongoingCount':ongoingCount}
        return Response({'data' : data}, status = status.HTTP_200_OK)


class GetAreaCodeAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny,)
    def get(self,request , format= None):
        areaCodes = list(Contacts.objects.all().exclude(areaCode = None).values('areaCode').distinct())
        return Response({'areaCodes' : areaCodes}, status = status.HTTP_200_OK)



class getAllLog(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny,)
    def get(self,request , format= None):
        contactObj = Contacts.objects.get(pk = int(request.GET['contact']))
        if request.user.is_superuser:
            logObj = Log.objects.filter(plan__contact = contactObj).order_by('-created')
        else:
            logObj = Log.objects.filter(type='text',plan__contact = contactObj)
        data = LogSerializer(logObj,many = True).data
        return Response({'data' : data} ,status = status.HTTP_200_OK)


class getLocation(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny,)
    def get(self,request , format= None):
        from geopy.geocoders import Nominatim
        geolocator = Nominatim()
        if 'address' in request.GET:
            location = geolocator.geocode(request.GET['address'])
            if location == None:
                return Response({'msg' : 'Location Not Found'} ,status = status.HTTP_200_OK)
            print((location.latitude, location.longitude))
            return Response({'lat' : location.latitude,'lon':location.longitude} ,status = status.HTTP_200_OK)
        else:
            geolocator = Nominatim()
            cord = str(request.GET['lat'])+','+str(request.GET['lon'])
            location = geolocator.reverse(cord)
            if location == None:
                return Response({'msg' : 'Location Not Found'} ,status = status.HTTP_200_OK)
            print(location.address)
            print((location.latitude, location.longitude))
            print(location.raw)
            address=json.loads(json.dumps(location.raw))
            return Response({'address' : address} ,status = status.HTTP_200_OK)

import pandas
class BulkEmailAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny,)

    def get(self, request, format=None):

        testmailSent = False

        for campaign in Campaign.objects.filter(typ = 'email'):
            items = campaign.items.all().filter(emailed = False)[:campaign.limitPerDay]
            print "items count : " , items.count()
            for i in items:
                print "ID : " , i.pk
                campaign = i.campaign
                try:
                    name = i.contact.name.split(' ')[0]
                except:
                    name = ''
                ctx = {
                    'linkUrl': 'cioc.co.in',
                    'linkText' : 'View Online',
                    'sendersAddress' : '(C) CIOC FMCG Pvt Ltd',
                    'sendersPhone' : '841101',
                    'linkedinUrl' : 'https://www.linkedin.com/company/13440221/',
                    'fbUrl' : 'facebook.com',
                    'twitterUrl' : 'twitter.com',
                    'name' : name.lower() ,
                    'tracking_id':i.pk,
                    'senderName' : 'Pradeep'
                }


                try:
                    email_subject = Template(campaign.emailSubject).render(Context(ctx))

                    if campaign.emailTemplate is None:
                        email_body = Template(campaign.emailBody).render(Context(ctx))
                        body_row = campaign.emailBody
                    else:
                        email_body = get_template(campaign.emailTemplate).render(ctx)
                        body_row = campaign.emailTemplate

                    to = i.contact.email
                    to = 'info@cioc.in'
                    # to = 'pkyisky@gmail.com '
                    msg = EmailMessage(email_subject, email_body, 'Pradeep Yadav <pradeep@epsilonai.com>',  to= [to] , cc= [] )
                    msg.content_subtype = 'html'
                    msg.send()

                    if not testmailSent:
                        msg = EmailMessage(email_subject, email_body, 'Pradeep Yadav <pradeep@epsilonai.com>',  to= ['info@cioc.in'] , cc= [] )
                        msg.content_subtype = 'html'
                        msg.send()
                        testmailSent = True


                    i.emailed = True
                    if int(i.attempt) == 0:
                        i.subject1 = campaign.emailSubject
                        i.body1 = body_row
                    elif int(i.attempt) == 1:
                        i.subject2 = campaign.emailSubject
                        i.body2 = body_row
                    elif int(i.attempt) == 2:
                        i.subject3 = campaign.emailSubject
                        i.body3 = body_row
                    elif int(i.attempt) == 3:
                        i.subject4 = campaign.emailSubject
                        i.body4 = body_row
                    elif int(i.attempt) == 4:
                        i.subject5 = campaign.emailSubject
                        i.body5 = body_row

                    i.attempt += 1
                    i.save()


                except:
                    traceback.print_exc(file=sys.stdout)

        return Response({}, status = status.HTTP_200_OK)

class SendTestMailAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny,)

    def get(self, request, format=None):
        campaign = Campaign.objects.get(pk = request.GET['id'])
        ctx = {
            'linkUrl': 'cioc.co.in',
            'linkText' : 'View Online',
            'sendersAddress' : '(C) CIOC FMCG Pvt Ltd',
            'sendersPhone' : '841101',
            'linkedinUrl' : 'https://www.linkedin.com/company/13440221/',
            'fbUrl' : 'facebook.com',
            'twitterUrl' : 'twitter.com',
            'name' : 'Pradeep' ,
            'tracking_id': 0,
            'senderName' : 'Pradeep'
        }


        email_subject = Template(campaign.emailSubject).render(Context(ctx))

        if campaign.emailTemplate is None:
            email_body = Template(campaign.emailBody).render(Context(ctx))
        else:
            email_body = get_template(campaign.emailTemplate).render(ctx)

        to = 'info@cioc.in'
        # to = 'pkyisky@gmail.com '
        msg = EmailMessage(email_subject, email_body, 'Pradeep Yadav <pradeep@epsilonai.com>',  to= [to] , cc= [] )

        msg.content_subtype = 'html'
        msg.send()

        return Response({}, status = status.HTTP_200_OK)




import os.path
class emailTracks(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny,)
    def get(self, request, *args, **kwargs):
        requestPath = str(request.META['PATH_INFO']).replace('/', '')
        imgName = requestPath.split('_')
        trackingPK = int(imgName[1].split('.')[0])
        ext = imgName[1].split('.')[1]
        image_data = open(os.path.join(globalSettings.BASE_DIR, 'static_shared', 'images', imgName[0] + '.' + ext), 'rb').read()
        try:
            res = requests.get('https://erp.cioc.in/updateEmailOpen?id='+ trackingPK)
            print res.text
        except Exception as e:
            pass
        return HttpResponse(image_data, content_type="image/jpg")

class OpenEmailUpdate(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny,)
    def get(self, request, *args, **kwargs):
        trackingObj = CampaignItem.objects.get(pk=request.GET['id'])
        trackingObj.opened = True
        trackingObj.openedTime = datetime.datetime.now(pytz.timezone('Asia/Kolkata'))
        trackingObj.save()
        print "Updating email tracking : " , trackingObj
        return Response({}, status = status.HTTP_200_OK)

class getCampaignDetailsAPI(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny,)
    def get(self, request, format=None):
        numOfCampaigns = []
        numOfRead = []
        numOfOpens = []
        numOfSent = []
        campaignCount = 0
        mailSent = 0
        mailReadCount = 0
        mailOpensCount = 0

        campaignObj = CampaignTracker.objects.all()
        for i in campaignObj:
            if i.campaignId not in numOfCampaigns:
                numOfCampaigns.append(i.campaignId)
                campaignCount = len(numOfCampaigns)

            if i.sent == bool(True):
                numOfSent.append(i.sent)
                mailSent = len(numOfSent)

            if i.read == bool(True):
                numOfRead.append(i.read)
                mailReadCount = len(numOfRead)

            if i.open == bool(True):
                numOfOpens.append(i.open)
                mailOpensCount = len(numOfOpens)
        toSend = {'mailSent': mailSent, 'campaignCount': campaignCount, 'mailReadCount': mailReadCount, 'mailOpensCount': mailOpensCount}
        return Response(toSend, status = status.HTTP_200_OK)


import re
emailRegex = '^\w+([\.-]?\w+)*@\w+([\.-]?\w+)*(\.\w{2,3})+$'
def checkEmail(email):
    if(re.search(emailRegex,email)):
        return True
    else:
        return False

from os import walk
class GetEmailTemplatesAPI(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny,)
    def post(self, request, format=None):
        emailList = request.FILES['dataFile']
        template = str(request.data['template'])+'html'
        excel_data = pandas.read_excel(emailList)
        mainDataFrame = excel_data[['Name','Email','Subject']].to_dict(orient='records')

        if len(excel_data.columns) > 3:
            extraCtxDataFrame = excel_data.drop(['Name', 'Email','Subject'], axis=1)
            extraCtxDataFrame = extraCtxDataFrame.to_dict(orient='records')

        counter = 0
        for i in range(len(mainDataFrame)):
            email = mainDataFrame[i]["Email"]
            name = mainDataFrame[i]["Name"]
            subject = mainDataFrame[i]["Subject"]
            counter += 1
            if checkEmail(email):
                if len(str(name))>0:
                    if len(str(subject))>0:
                        ctx = 0

            if len(excel_data.columns) > 3:
                if checkEmail(email):
                    if len(str(name))>0:
                        if len(str(subject))>0:
                            ctx = extraCtxDataFrame[i]

            emailTrackObj = CampaignTracker.objects.create(to = email,subject=subject,name=name,context=ctx,template=template)

        return Response({'counter':counter}, status = status.HTTP_200_OK)

    def get(self, request, format=None):
        path = globalSettings.BASE_DIR+'/marketing/mailTemplates'
        templates = []
        for (dirpath, dirnames, filenames) in walk(path):
            for a in filenames:
                templates.append(a)
        return Response(templates, status = status.HTTP_200_OK)


class UploadTemplatesAPI(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny,)
    def post(self, request, format=None):
        uploadedTemplate = request.FILES['templateFile']
        saveFile = os.path.join(globalSettings.BASE_DIR, 'marketing', 'mailTemplates', uploadedTemplate.name)
        efile = open(saveFile, "w")
        return Response({}, status = status.HTTP_200_OK)

class FollowUpCampaignAPI(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny,)
    def get(self, request, format=None):
        campaignObj = CampaignTracker.objects.all()
        sentData = campaignObj.filter(sent = True)
        print len(sentData), 'len(sentData)'
        unsentData = campaignObj.filter(sent = False)
        print len(unsentData), 'len(unsentData)'
        for i in unsentData:
            contactData = []
            contactData.append(i.to)
            imageUrl = globalSettings.SITE_ADDRESS+'/api/marketing/images/logo_'+str(i.pk)+'.jpg'

            try:
                cc = []
                email_subject ='Test Email'

                msgBody= ''

                ctx = {
                    'message': msgBody,
                    'linkUrl': 'cioc.co.in',
                    'linkText' : 'View Online',
                    'sendersAddress' : '(C) CIOC FMCG Pvt Ltd',
                    'sendersPhone' : '841101',
                    'linkedinUrl' : 'https://www.linkedin.com/company/13440221/',
                    'fbUrl' : 'facebook.com',
                    'twitterUrl' : 'twitter.com',
                    'recieverName' : i.to,
                    'imageUrl':imageUrl
                }

                email_body = get_template(i.template).render(ctx)
                msg = EmailMessage(email_subject, email_body,  to= contactData, cc= cc )

                msg.content_subtype = 'html'
                msg.send()
                i.sent = True
                i.save()
            except:
                pass
        return Response({}, status = status.HTTP_200_OK)



class CampaignTrackerViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated ,)
    serializer_class = CampaignTrackerSerializer
    queryset = CampaignTracker.objects.all()
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['campaignId']



class GetTechStatusAPI(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny ,)
    def get(self,request , format= None):
        toReturn = []
        val = {'name' : 'assigned' , 'is_selected' : False, 'count' : 0}
        toReturn.append(val)
        val = {'name' : 'ongoing' , 'is_selected' : False, 'count' : 0}
        toReturn.append(val)
        val = {'name' : 'postponed' , 'is_selected' : False, 'count' : 0}
        toReturn.append(val)
        val = {'name' : 'completed' , 'is_selected' : False, 'count' : 0}
        toReturn.append(val)
        val = {'name' : 'cancelled' , 'is_selected' : False, 'count' : 0}
        toReturn.append(val)
        return Response({'data':toReturn }, status = status.HTTP_200_OK)

class GetTechincianInfoAPI(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny ,)
    def get(self,request , format= None):
        workbook = Workbook()
        Sheet1 = workbook.active
        hdFont = Font(size=12,bold=True)
        alphaChars = list(string.ascii_uppercase)
        Sheet1.title = 'Collections'
        if request.user.is_superuser:
            hd = ["InvoiceID","Dated", 'Timeslot','Customer Name','Customer Mobile No','Customer Email','Technician','Status','Total']
            hdWidth = [10,10,10,30,30,30,30,15,20]
            Sheet1.append(hd)
        else:
            hd = ["InvoiceID","Dated", 'Timeslot','Customer Name','Customer Mobile No','Customer Email','Technician','Status']
            hdWidth = [10,10,10,30,30,30,30,15]
            Sheet1.append(hd)
        toPlan  = TourPlanStop.objects.all()
        if 'technician' in request.GET:
            tourplan = toPlan.filter(tourplan__user__pk = request.GET['technician'],status=request.GET['status'])
        else:
            tourplan = toPlan.filter(status=request.GET['status'])
        total = 0
        invTot = 0
        cashtotal = 0
        chequetotal = 0
        noncashtotal = 0
        cardtotal = 0
        for i in tourplan:
            if request.user.is_superuser:
                if i.contract != None:
                    total += i.contract.grandTotal
                    invTot = i.contract.grandTotal
            onDat = i.tourplan.date

            data = [i.pk,onDat,i.timeslot,i.contact.name,i.contact.mobile,i.contact.email,i.tourplan.user.first_name+' '+i.tourplan.user.last_name,i.status,invTot]
            Sheet1.append(data)
        if request.user.is_superuser:
            Sheet1.append(['','','','','','','','',''])
            Sheet1.append(['','','','','','','','Sum',total])
            Sheet1.append(['','','','','','','','',''])
            Sheet1.append(['','','','','','','','',''])
            Sheet1.append(['','','','','','','','',''])

        Sheet1.column_dimensions['A'].width = 20
        Sheet1.column_dimensions['B'].width = 20
        Sheet1.column_dimensions['C'].width = 15
        Sheet1.column_dimensions['D'].width = 30
        Sheet1.column_dimensions['E'].width = 20
        Sheet1.column_dimensions['F'].width = 30
        Sheet1.column_dimensions['G'].width = 30
        Sheet1.column_dimensions['H'].width = 15
        response = HttpResponse(content=save_virtual_workbook(workbook),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Collection.xlsx'
        return response




import json
class GetAllTechincianInfoAPI(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny ,)
    def get(self,request , format= None):
        workbook = Workbook()
        Sheet1 = workbook.active
        hdFont = Font(size=12,bold=True)
        alphaChars = list(string.ascii_uppercase)
        Sheet1.title = 'Collections'
        if request.user.is_superuser:
            hd = ["InvoiceID","Dated", 'Timeslot','Customer Name','Customer Mobile No','Customer Email','Technician','Status','Total']
            hdWidth = [10,10,10,30,30,30,30,15,20]
            Sheet1.append(hd)
        else:
            hd = ["InvoiceID","Dated", 'Timeslot','Customer Name','Customer Mobile No','Customer Email','Technician','Status']
            hdWidth = [10,10,10,30,30,30,30,15]
            Sheet1.append(hd)
        total = 0
        invTot = 0
        cashtotal = 0
        chequetotal = 0
        noncashtotal = 0
        cardtotal = 0
        invTot = 0
        tourplan = []
        onDat =[]
        toPlan = TourPlanStop.objects.all()
        if 'technician' in request.GET :
            toPlans = toPlan.filter(tourplan__user__pk = request.GET['technician'])
        else:
            toPlans = toPlan
        if 'status' in request.GET:
            allStat = request.GET['status'].split(',')
            for k in allStat:
                tourplan = toPlans.filter(status=str(k))
                for i in tourplan:
                    if request.user.is_superuser:
                        if i.contract != None:
                            invTot = i.contract.grandTotal
                        total += invTot
                        onDat = i.tourplan.date
                        data = [i.pk,onDat,i.timeslot,i.contact.name,i.contact.mobile,i.contact.email,i.tourplan.user.first_name+' '+i.tourplan.user.last_name,i.status,invTot]
                        Sheet1.append(data)
                    else:
                        onDat = i.tourplan.date
                        data = [i.pk,onDat,i.timeslot,i.contact.name,i.contact.mobile,i.contact.email,i.tourplan.user.first_name+' '+i.tourplan.user.last_name,i.status]
                        Sheet1.append(data)
        else:
            allStat = toPlans
            for i in allStat:
                if request.user.is_superuser:
                    if i.contract != None:
                        invTot = i.contract.grandTotal
                    total += invTot
                    onDat = i.tourplan.date
                    data = [i.pk,onDat,i.timeslot,i.contact.name,i.contact.mobile,i.contact.email,i.tourplan.user.first_name+' '+i.tourplan.user.last_name,i.status,invTot]
                    Sheet1.append(data)
                else:
                    onDat = i.tourplan.date
                    data = [i.pk,onDat,i.timeslot,i.contact.name,i.contact.mobile,i.contact.email,i.tourplan.user.first_name+' '+i.tourplan.user.last_name,i.status]
                    Sheet1.append(data)

        Sheet1.column_dimensions['A'].width = 20
        Sheet1.column_dimensions['B'].width = 20
        Sheet1.column_dimensions['C'].width = 15
        Sheet1.column_dimensions['D'].width = 30
        Sheet1.column_dimensions['E'].width = 20
        Sheet1.column_dimensions['F'].width = 30
        Sheet1.column_dimensions['G'].width = 30
        Sheet1.column_dimensions['H'].width = 15
        if request.user.is_superuser:
            Sheet1.append(['','','','','','','','',''])
            Sheet1.append(['','','','','','','','Sum',total])
            Sheet1.append(['','','','','','','','',''])
            Sheet1.append(['','','','','','','','',''])
            Sheet1.append(['','','','','','','','',''])
        response = HttpResponse(content=save_virtual_workbook(workbook),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=TechnicianDetails.xlsx'
        return response


class homePageAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny ,)
    def get(self,request , format= None):
        from datetime import date
        mypath = os.path.join(globalSettings.BASE_DIR, 'static_shared', 'images', 'Banners')
        print mypath
        f = []
        for (dirpath, dirnames, filenames) in walk(mypath):
            print filenames
            for a in filenames:
                filename = 'static/images/Banners/' + a
                f.append(filename)
        user = request.user
        today = date.today()
        toPlans = TourPlanStop.objects.filter(tourplan__user = user , tourplan__date = today)
        backlogPlans = TourPlanStop.objects.filter(tourplan__user = user , tourplan__date__lt = today , status = 'assigned').count()
        totalTarget = toPlans.count()
        pendingTarget = toPlans.filter(tourplan__user = user,status = 'assigned',tourplan__date = today).exclude(is_postponded = True).count()
        try:
            pendingTargetPer =  ((pendingTarget * 100)/ totalTarget)
        except:
            pendingTargetPer =  0
        completedTargetPer = 100 - pendingTargetPer
        data = {'totalTarget' : totalTarget , 'pendingTarget' : pendingTarget , 'pendingTargetPer' : pendingTargetPer , 'completedTargetPer' : completedTargetPer , 'iamges' : f,'backlogPlans':backlogPlans}
        return Response(data, status = status.HTTP_200_OK)

themeColor = colors.HexColor('#227daa')
styles = getSampleStyleSheet()
styleN = styles['Normal']
styleH = styles['Heading1']
class GetPendingTasksAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny ,)
    def get(self,request , format= None):
        from datetime import date
        user = request.user
        today = date.today()
        toPlans = TourPlanStop.objects.filter(tourplan__user = user , tourplan__date__lt = today)
        data = TourPlanStopPendingSerializer(toPlans , many = True).data
        return Response(data, status = status.HTTP_200_OK)


class FullPageImage(Flowable):
    def __init__(self, img):
        Flowable.__init__(self)
        self.image = img

    def draw(self):
        img = utils.ImageReader(self.image)

        iw, ih = img.getSize()
        aspect = ih / float(iw)
        width, self.height = PAGE_SIZE
        width -= 3.5 * cm
        self.canv.drawImage(os.path.join(BASE_DIR, self.image), -1 * MARGIN_SIZE +
                            1.5 * cm, -1 * self.height + 5 * cm, width, aspect * width)


class expanseReportHead(Flowable):

    def __init__(self, request, contract):
        Flowable.__init__(self)
        self.req = request
        self.contract = contract
    #----------------------------------------------------------------------

    def draw(self):
        """
        draw the floable
        """
        print self.contract.status
        now = datetime.datetime.now(pytz.timezone('Asia/Kolkata'))
        print self.contract.status

        docTitle = 'INVOICE'

        id = self.contract.pk
        docID = '%s%s%s' % (id, now.year, self.contract.pk)
        utc_time = self.contract.created
        tz = pytz.timezone( 'Asia/Kolkata')
        utc_time =utc_time.replace(tzinfo=pytz.UTC) #replace method
        indian_time=utc_time.astimezone(tz)        #astimezone method
        datecreated = str(indian_time.strftime("%d-%B-%Y - %H:%M %S"))
        try:
            passKey = '%s%s' % (str(self.req.user.date_joined.year),self.req.user.pk)  # also the user ID
            pSrc = '''
            <font size=9>
            <strong>On:</strong> %s<br/>
            <strong>#:</strong> %s<br/><br/>
            </font>
            ''' % ( datecreated, docID)
        except:
            pSrc = '''
            <strong>On:</strong> %s<br/>
            <strong>Document ID:</strong> %s<br/><br/>
            </font>
            ''' % ( datecreated, docID)

        story = []
        head = Paragraph(pSrc, styleN)
        head.wrapOn(self.canv, 200 * mm, 50 * mm)
        head.drawOn(self.canv, 0 * mm, -10 * mm)


def addPageNumber(canvas, doc):
    """
    Add the page number
    """
    print doc.contract
    now = datetime.datetime.now(pytz.timezone('Asia/Kolkata'))

    id = doc.contract.pk
    docID = '%s%s%s' % (id, now.year, doc.contract.pk)
    try:
        passKey = '%s%s' % (str(doc.request.user.date_joined.year),doc.request.user.pk)  # also the user ID
    except:
        passKey = docID
    d = Drawing(60, 60)
    renderPDF.draw(d, canvas, 180 * mm, 270 * mm)
    pass


class PageNumCanvas(canvas.Canvas):

    #----------------------------------------------------------------------
    def __init__(self, *args, **kwargs):
        """Constructor"""
        canvas.Canvas.__init__(self, *args, **kwargs)
        print "dir : ---------", dir(self)
        self.pages = []

    #----------------------------------------------------------------------
    def showPage(self):
        """
        On a page break, add information to the list
        """
        self.pages.append(dict(self.__dict__))
        self._startPage()

    #----------------------------------------------------------------------
    def save(self):
        """
        Add the page number to each page (page x of y)
        """
        page_count = len(self.pages)

        for page in self.pages:
            self.__dict__.update(page)
            self.drawLetterHeadFooter()
            canvas.Canvas.showPage(self)
        canvas.Canvas.save(self)

    #----------------------------------------------------------------------
    def draw_page_number(self, page_count):
        """
        Add the page number
        """

        text = "<font size='8'>Page #%s of %s</font>" % (
            self._pageNumber, page_count)
        p = Paragraph(text, styleN)
        p.wrapOn(self, 50 * mm, 10 * mm)
        p.drawOn(self, 100 * mm, 10 * mm)

    def drawLetterHeadFooter(self):
        settingsFields = application.objects.get(name='app.CRM').settings.all()
        self.setStrokeColor(themeColor)
        self.setFillColor(themeColor)
        self.rect(0, 0, 1500, 70, fill=True)
        compNameStyle = styleN.clone('footerCompanyName')
        compNameStyle.textColor = 'white'

        p = Paragraph(settingsFields.get(
            name='companyName').value, compNameStyle)
        p.wrapOn(self, 50 * mm, 10 * mm)
        p.drawOn(self, 85 * mm, 18 * mm)

        p1 = Paragraph(settingsFields.get(
            name='companyAddress').value, compNameStyle)
        p1.wrapOn(self, 200 * mm, 10 * mm)
        p1.drawOn(self, 15 * mm, 10 * mm)

        p2 = Paragraph(settingsFields.get(
            name='contactDetails').value, compNameStyle)
        p2.wrapOn(self, 200 * mm, 10 * mm)
        p2.drawOn(self, 40 * mm, 4 * mm)

        from svglib.svglib import svg2rlg


        headerFilePath = globalSettings.INVOICE_HEADER


        drawing = svg2rlg(os.path.join(globalSettings.BASE_DIR,
                                       'static_shared', 'images', headerFilePath))
        sx = sy = 0.5
        drawing.width, drawing.height = drawing.minWidth() * sx, drawing.height * sy
        drawing.scale(sx, sy)
        renderPDF.draw(drawing, self, 1 * mm, self._pagesize[1] - 25 * mm)

from num2words import num2words


def genInvoice(response, outbond, request):
        MARGIN_SIZE = 8 * mm
        PAGE_SIZE = A4

        pdf_doc = SimpleDocTemplate(response, pagesize=PAGE_SIZE,
                                    leftMargin=MARGIN_SIZE, rightMargin=MARGIN_SIZE,
                                    topMargin=4 * MARGIN_SIZE, bottomMargin=3 * MARGIN_SIZE)
        pdf_doc.contract = outbond
        pdf_doc.request = request
        styles = getSampleStyleSheet()
        styleN = styles['Normal']
        styleH = styles['Heading1']
        tableHeaderStyle = styles['Normal'].clone('tableHeaderStyle')
        tableHeaderStyle.textColor = colors.white
        tableHeaderStyle.fontSize = 7
        tableBodyStyle = styles['Normal'].clone('tableBodyStyle')
        story = []
        summryParaSrc = """

        <font size='9'>
        %s<br/>
        %s<br/>
        %s<br/>
         %s , %s<br/><br/><br/>
         <strong>Kind Attention:</strong>%s <br/>
        </font>
        """ % ( outbond.phone , outbond.email , outbond.address , outbond.pincode , outbond.country, outbond.personName )

        story.append(Paragraph(summryParaSrc, styleN))
        pHeadDetails = Paragraph('<strong>Details of Products & Services</strong>', tableHeaderStyle)
        pHeadPrice = Paragraph('<strong>Price</strong>', tableHeaderStyle)
        pHeadQty = Paragraph('<strong>Qty</strong>', tableHeaderStyle)
        pHeadTax = Paragraph('<strong>Tax Rate </strong>', tableHeaderStyle)
        pHeadTotal = Paragraph('<strong>Total</strong>', tableHeaderStyle)
        data = [[pHeadDetails, pHeadPrice,pHeadQty, pHeadTax, pHeadTotal]]
        for i in outbond.outBoundQty.all():
            pBodyDetails = Paragraph(str(i.product), tableBodyStyle)
            pBodyPrice =  Paragraph(str(i.price), tableBodyStyle)
            pBodyQty =  Paragraph(str(i.qty), tableBodyStyle)
            pBodytax =  Paragraph(str(i.tax), tableBodyStyle)
            pBodytotal =  Paragraph(str(i.total), tableBodyStyle)
            data.append([pBodyDetails, pBodyPrice,pBodyQty, pBodytax, pBodytotal])

        grandTotal = Paragraph(str(outbond.total), tableBodyStyle)
        data.append(['', '','', '', grandTotal])

        t = Table(data)
        ts = TableStyle([('ALIGN', (1, 1), (-3, -3), 'RIGHT'),
                         ('VALIGN', (0, 1), (-1, -3), 'TOP'),
                         ('VALIGN', (0, -2), (-1, -2), 'TOP'),
                         ('VALIGN', (0, -1), (-1, -1), 'TOP'),
                         # ('SPAN', (-3, -1), (-2, -1)),
                         # ('SPAN', (-2, -1), (-1, -1)),
                         ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                         ('BACKGROUND', (0, 0), (-1, 0), themeColor),
                         ('LINEABOVE', (0, 0), (-1, 0), 0.25, themeColor),
                         ('LINEABOVE', (0, 1), (-1, 1), 0.25, themeColor),
                         ('BACKGROUND', (-2, -1), (-1, -1), themeColor),
                         ('LINEABOVE', (0, -1), (-1, -1), 0.25, colors.gray),
                     ])
        t.setStyle(ts)
        t._argW[0] = 8 * cm
        t._argW[1] = 3 * cm
        t._argW[2] = 3 * cm
        t._argW[3] = 3 * cm
        t._argW[4] = 3 * cm
        story.append(t)
        pdf_doc.build(story, onFirstPage=addPageNumber,
                      onLaterPages=addPageNumber, canvasmaker=PageNumCanvas)

class GenerateInvoiceAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny ,)
    def get(self,request , format= None):
        if 'invoice' not in request.GET:
            return Response(status=status.HTTP_400_BAD_REQUEST)
        response = HttpResponse(content_type='application/pdf')
        o = OutBoundInvoice.objects.get(id=request.GET['invoice'])
        response.aggrement = o
        response['Content-Disposition'] = 'attachment; filename="invoice_%s.pdf"' % (o.pk)
        genInvoice(response, o, request)
        return response


class SendMailPDFAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny ,)
    def post(self,request , format= None):
        contactData=[]
        print int(request.data['invoice']),'int(request.data'
        o = OutBoundInvoice.objects.get(id=int(request.data['invoice']))
        print o,'invoice'
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment;filename="Invoicedownload.pdf"'
        genInvoice(response, o, request)
        f = open(os.path.join(globalSettings.BASE_DIR, 'media_root/Invoicedownload.pdf'), 'wb')
        f.write(response.content)
        f.close()
        email_subject = "Invoice"
        msgBody = "Hi "  + o.personName + ",\n\n\t\t Find the attachment for the Invoice Details"
        email = request.data['email']

        contactData.append(str(email))
        msg = EmailMessage(email_subject, msgBody,  to= contactData )

        a = str(f).split('media_root/')[1]
        b = str(a).split(', mode')[0]
        c = str(b).split("'")[0]
        msg.attach_file(os.path.join(globalSettings.MEDIA_ROOT,str(c)))
        msg.send()
        return Response({'send':True}, status = status.HTTP_200_OK)

class GetTodayInvoiceIView(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny ,)
    def get(self,request , format= None):
        from datetime import date
        user = request.user
        today = date.today()
        if 'typ' in request.GET:
            tourObj = TourPlanStop.objects.filter(tourplan__user = user , status = "completed" ).order_by('-pk')
            if request.GET['typ'] == 'CustomPeriod':
                if 'date1' in request.GET:
                    date1 = request.GET['date1']
                if 'date2' in request.GET:
                    date2 = request.GET['date2']
                if date2<date1:
                    fromDate = date2
                    toDate = date1
                else:
                    fromDate = date1
                    toDate = date2
                tourObj = tourObj.filter(tourplan__date__range = (fromDate , toDate))
            elif request.GET['typ'] == 'Today':
                tourObj = tourObj.filter(tourplan__date = today)
            elif request.GET['typ'] == 'Yesterday':
                yesterday = today - timedelta(days = 1)
                tourObj = tourObj.filter(tourplan__date = yesterday)
            elif request.GET['typ'] == 'ThisWeek':
                noDays =  int(today.strftime('%w'))
                fromDate = today - timedelta(days=int(noDays))
                toDate = fromDate + timedelta(days=6)
                tourObj = tourObj.filter(tourplan__date__range = (fromDate , toDate))
            elif request.GET['typ'] == 'LastWeek':
                noDays =  int(today.strftime('%w'))
                sundayDate = today - timedelta(days=int(noDays))
                fromDate = sundayDate - timedelta(days=7)
                toDate = sundayDate - timedelta(days=1)
                tourObj = tourObj.filter(tourplan__date__range = (fromDate , toDate))
            elif request.GET['typ'] == 'ThisMonth':
                fromDate = today.replace(day=1)
                toDate = today + relativedelta(day=31)
                tourObj = tourObj.filter(tourplan__date__range = (fromDate , toDate))
            elif request.GET['typ'] == 'LastMonth':
                toDate = today.replace(day=1) - datetime.timedelta(days=1)
                fromDate = toDate.replace(day=1)
                tourObj = tourObj.filter(tourplan__date__range = (fromDate , toDate))
        toRet = TourPlanStopLiteSerializer(tourObj, many=True).data
        return Response(toRet,status = status.HTTP_200_OK)


sid         = 'cioc51'
key         = 'c9c3b834403869391f582cc288ef9f1d4aaf06a23f25620f'
token       = '56849aaa4e89cc4dce38b0bc6c0c70dfccf249646a635800'

class FetchRecordingsView(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny ,)
    def get(self,request , format= None):

        for log in Log.objects.filter(callUrl__isnull = False , fileUrl__isnull = True  ):

            res = requests.get('https://'+ key +':'+  token +'@api.exotel.in'+ log.callUrl )
            data = res.json()
            print data
            log.fileUrl = data['Call']['RecordingUrl']
            if log.fileUrl is not None:
                log.save()

        return Response({},status = status.HTTP_200_OK)


class CallCustomerView(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny ,)
    def get(self,request , format= None):

        plan = TourPlanStop.objects.get(id = request.GET['id'])

        data = {
          'From': request.user.profile.mobile ,
          'To': plan.contact.mobile ,
          'CallerId': '02248966707'
        }

        res = requests.post('https://'+ key +':'+  token +'@api.exotel.in/v1/Accounts/'+ sid +'/Calls/connect.json', data=data)

        resData = res.json()
        print resData

        log = Log(txt = 'Outgoing from ' + request.user.first_name , plan = plan , type = 'routedCall' , callUrl = resData['Call']['Uri'] )

        log.save()

        return Response(resData ,status = status.HTTP_200_OK)


class SendCallReportAPI(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny ,)
    def post(self,request , format= None):
        logObj = Log.objects.get(pk = request.data['log'])
        from time import gmtime, strftime
        creteDat = logObj.created.strftime("%a, %d %b %Y %H:%M")
        contactData = []
        email_subject = "Report"
        msgBody =  "\n\n\t"+"Time stamp : "+creteDat+"\n\n\t"+"Caller : "+ logObj.plan.tourplan.user.first_name+' '+logObj.plan.tourplan.user.last_name+"(%s)"%logObj.plan.tourplan.user.pk+"\n\n\t"+"Customer : "+logObj.plan.contact.name+"(%s)"%logObj.plan.contact.mobile+"\n\n\t"+"Comment : "+  request.data['text']+",\n\n\t"+"Recording : "+logObj.fileUrl
        email = globalSettings.G_ADMIN
        contactData.append(str(email))
        msg = EmailMessage(email_subject, msgBody,  to= contactData )
        msg.send()
        return Response({'send':True} ,status = status.HTTP_200_OK)

class GetEmailCampaignStatsView(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny ,)
    def get(self,request , format= None):
        qs = EmailCampaignItem.objects.filter(campaign = Campaign.objects.get(pk = request.GET['id']))

        total = qs.count()
        sent = qs.filter(emailed = True).count()
        opned = qs.filter(opened = True).count()
        clicked = 0
        if total > 0:
            sentPercent = (sent/float(total))*100
        else:
            sentPercent = 0


        if sent == 0:
            return Response({'total': total , 'sent' : sent , 'sentPercent':sentPercent , 'opened' : opned , 'openedPercent' : 0, 'clicked' : clicked , 'clickedPercent' : 0 } ,status = status.HTTP_200_OK)
        else:
            return Response({'total': total , 'sent' : sent , 'sentPercent':sentPercent , 'opened' : opned , 'openedPercent' : (opned/float(sent))*100 , 'clicked' : clicked , 'clickedPercent' : (clicked/float(sent))*100 } ,status = status.HTTP_200_OK)

class CorrectContactDetailsView(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny ,)
    def get(self,request , format= None):
        mailDomains = [
            '@googlemail',
            '@gmail',
            '@yahoo',
            '@outlook',
            '@hotmail',
            '@me.com',
        ]
        contacts = Contacts.objects.filter( website__isnull = True , email__isnull = False)
        for mailDomain in mailDomains:
            contacts = contacts.filter( ~Q(email__icontains= mailDomain)  )

        print contacts.count()

        count = 0
        for contact in contacts:
            count += 1
            contact.website = contact.email.split('@')[1]
            contact.save()
            print contact.email , contact.email.split('@')[1]
        return Response({} ,status = status.HTTP_200_OK)

styleN = styles['Normal']
styleH = styles['Heading1']
tableHeaderStyle = styles['Normal'].clone('tableHeaderStyle')
tableHeaderStyle.textColor = colors.black
tableHeaderStyle.fontSize = 7
class ServicerecieptCanvas(canvas.Canvas):

    #----------------------------------------------------------------------
    def __init__(self, *args, **kwargs):
        """Constructor"""
        canvas.Canvas.__init__(self, *args, **kwargs)
        print "dir : ---------", dir(self)
        self.pages = []

    #----------------------------------------------------------------------
    def showPage(self):
        """
        On a page break, add information to the list
        """
        self.pages.append(dict(self.__dict__))
        self._startPage()

    #----------------------------------------------------------------------
    def save(self):
        """
        Add the page number to each page (page x of y)
        """
        page_count = len(self.pages)

        for page in self.pages:
            self.__dict__.update(page)
            self.draw_page_number(page_count)
            self.drawLetterHeadFooter()
            canvas.Canvas.showPage(self)

        canvas.Canvas.save(self)

    #----------------------------------------------------------------------
    def draw_page_number(self, page_count):
        """
        Add the page number
        """

        text = "<font size='8'>Page #%s of %s</font>" % (
            self._pageNumber, page_count)
        p = Paragraph(text, tableHeaderStyle)
        p.wrapOn(self, 50 * mm, 10 * mm)
        p.drawOn(self, 100 * mm, 20 * mm)

    def drawLetterHeadFooter(self):
        try:
            company = servicecontactObj.creater.designation.division.name
        except:
            company = ''
        text = "<para size=11 align=center>{0}<br/></para>".format(company)
        p = Paragraph(text, tableHeaderStyle)
        p.wrapOn(self, 200 * mm, 10 * mm)
        p.drawOn(self, 10 * mm, 16 * mm)
        try:
            address =  servicecontactObj.creater.designation.unit.address
        except:
            address = ''
        try:
            city = servicecontactObj.creater.designation.unit.city
        except:
            city = ''
        try:
            state = servicecontactObj.creater.designation.unit.state
        except:
            state = ''
        try:
            country = servicecontactObj.creater.designation.unit.country
        except:
            country = ''
        try:
            pincode = servicecontactObj.creater.designation.unit.pincode
        except:
            pincode = ''
        text2 = "<para align=center size=9><strong>{0} {1} {2} {3} {4}</strong></para>".format(address, city, pincode, state, country)

        p1 = Paragraph(text2, tableHeaderStyle)
        p1.wrapOn(self, 220 * mm, 0 * mm)
        p1.drawOn(self, 0 * mm,  10* mm)
        try:
            site =  servicecontactObj.creater.designation.division.website
        except:
            site = ''
        text3 = "<para align=center size=10>{0}</para>".format(site)

        p2 = Paragraph(text3, tableHeaderStyle)
        p2.wrapOn(self, 50 * mm, 0 * mm)
        p2.drawOn(self, 12 * mm,  4* mm)
        try:
            email =  servicecontactObj.creater.designation.unit.email
        except:
            email = ''
        text4 = "<para align=center size=10>{0}</para>".format(email)

        p3 = Paragraph(text4, tableHeaderStyle)
        p3.wrapOn(self, 120 * mm, 0 * mm)
        p3.drawOn(self, 50 * mm,  4* mm)
        try:
            mobile =  servicecontactObj.creater.designation.unit.mobile
        except:
            mobile = ''
        text5 = "<para align=center size=10>{0}</para>".format(mobile)

        p3 = Paragraph(text5, tableHeaderStyle)
        p3.wrapOn(self, 250 * mm, 0 * mm)
        p3.drawOn(self, 50 * mm,  4* mm)
        from svglib.svglib import svg2rlg
        headerFilePath = globalSettings.INVOICE_HEADER
        drawing = svg2rlg(os.path.join(globalSettings.BASE_DIR,
                                       'static_shared', 'images', headerFilePath))
        sx = sy = 0.5
        drawing.width, drawing.height = drawing.minWidth() * sx, drawing.height * sy
        drawing.scale(sx, sy)
        renderPDF.draw(drawing, self, 1 * mm, self._pagesize[1] - 25 * mm)

def genServiceReciept(response,request, contactObj):
        MARGIN_SIZE = 8 * mm
        PAGE_SIZE = A4

        pdf_doc = SimpleDocTemplate(response, pagesize=PAGE_SIZE,
                                    leftMargin=MARGIN_SIZE, rightMargin=MARGIN_SIZE,
                                    topMargin=4 * MARGIN_SIZE, bottomMargin=3 * MARGIN_SIZE)
        global servicecontactObj
        servicecontactObj = contactObj
        pdf_doc.request = request
        pdf_doc.contact  = contactObj
        styles = getSampleStyleSheet()
        styleN = styles['Normal']
        styleH = styles['Heading1']
        tableHeaderStyle = styles['Normal'].clone('tableHeaderStyle')
        tableHeaderStyle.textColor = colors.black
        tableHeaderStyle.fontSize = 7
        tableBodyStyle = styles['Normal'].clone('tableBodyStyle')
        story = []
        material_charges_wordData =  num2words(round(contactObj.material_charges), lang='en_IN')
        summryParaSrc = Paragraph("""

        <para size='15' >

         <strong>Material Inward Note ( for Service)</strong> <br/>
        </para>
        """,tableHeaderStyle)

        today = datetime.date.today()
        recievedOn = Paragraph('<para><strong>Recieved On : </strong> {0}</para>'.format(today), tableHeaderStyle)
        reciept = Paragraph('<para><strong>Reciept # : </strong>{0}</para>'.format(contactObj.pk), tableHeaderStyle)
        recieved = Paragraph('<para><strong>Recieved by : </strong>{0} {1}</para>'.format(contactObj.creater.first_name , contactObj.creater.last_name), tableHeaderStyle)

        customer =  Paragraph('<para><strong>Customer Name: </strong>{0}</para>'.format(contactObj.name), tableHeaderStyle)
        contact = Paragraph('<para><strong>Contact Number : </strong>{0}</para>'.format(contactObj.mobile), tableHeaderStyle)
        company = Paragraph('<para><strong>Company Name : </strong>{0}</para>'.format(contactObj.companyName), tableHeaderStyle)
        address = Paragraph('<para>{0} </para>'.format(contactObj.addrs), tableHeaderStyle)
        if contactObj.city==None:
            city = ''
        else:
            city = contactObj.city
        address1 = Paragraph('<para>{0}</para>'.format(city), tableHeaderStyle)
        if contactObj.state==None:
            state = ''
        else:
            state = contactObj.state
        if contactObj.country==None:
            country = ''
        else:
            country = contactObj.country
        if contactObj.areaCode==None:
            areaCode = ''
        else:
            areaCode = contactObj.areaCode
        address2 = Paragraph('<para>{0},{1},{2}</para>'.format(state, areaCode, country ), tableHeaderStyle)
        details = Paragraph('<para><strong>Details of items recieved</strong></para>', tableHeaderStyle)
        warranty = Paragraph('<para><strong>Warranty Status</strong></para>', tableHeaderStyle)
        issues = Paragraph('<para><strong>Issues Reported</strong></para>', tableHeaderStyle)



        engineerNotes = Paragraph('''<para><strong>Engineer's Notes</strong></para>''', tableHeaderStyle)
        estimatedRepair = Paragraph('<para ><strong>Estimated Repair Charges(in words):</strong></para>', tableHeaderStyle)
        estimatedRepairDetail = Paragraph('<para>{0} only</para>'.format(material_charges_wordData), tableHeaderStyle)
        customerSign = Paragraph('<para><strong>Customer Signature</strong></para>', tableHeaderStyle)
        termsandcndtns = Paragraph('<para><strong>Terms and Conditions :</strong></para>', tableHeaderStyle)
        data = [[recievedOn],[reciept],[recieved],[customer, address],[contact,address1],[company,address2]]
        matData = [[summryParaSrc]]
        mat = Table(matData,rowHeights=2*cm)
        mats = TableStyle([

                         ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                         ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),

                     ])

        mat.setStyle(mats)
        story.append(mat)
        estimatedTab = [[estimatedRepair,''],[estimatedRepairDetail,'']]
        signTab = [[customerSign]]

        termsConditionsTab = [[termsandcndtns]]
        count = 0
        for t in contactObj.material_terms.body.split('||'):
            count+=1
            termsandcndtns1 = Paragraph('<para><strong>{0}.</strong> {1}</para>'.format(count, t), tableHeaderStyle)
            termsConditionsTab.append([termsandcndtns1])
        ter = Table(termsConditionsTab)
        terts = TableStyle([

                         ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                         ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),

                     ])

        ter.setStyle(terts)

        s = Table(signTab,rowHeights=2*cm,colWidths=10*cm,spaceAfter=0.5*cm)
        sts = TableStyle([

                         ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                         ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                          ('BOX', (0, 0), (-1, -1), 0.25,colors.gray),

                     ])

        s.setStyle(sts)
        s.hAlign = 'LEFT'
        e = Table(estimatedTab,spaceAfter=0.5*cm)
        ets = TableStyle([

                         ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                         ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),

                     ])

        e.setStyle(ets)

        t = Table(data,rowHeights=5*mm)
        ts = TableStyle([

                         ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                         ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),

                     ])
        t.setStyle(ts)

        story.append(t)
        story.append(Spacer(2.5, 0.5 * cm))
        detailsData = [[details,warranty,issues]]
        itemsData = eval(contactObj.material_items)
        for i in itemsData:
            detail = Paragraph('<para>{0} </para>'.format(i['item']), tableHeaderStyle)
            warranty1 = Paragraph('<para>{0}</para>'.format(i['warranty']), tableHeaderStyle)
            issues1 = Paragraph('<para>{0}</para>'.format(i['issue']), tableHeaderStyle)
            val = [detail,warranty1,issues1]
            detailsData.append(val)
        d = Table(detailsData)
        dts = TableStyle([

                         ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                         ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                         ('BOX', (0, 0), (-1, -1), 0.25,colors.gray),
                         ('LINEBELOW', (0, 0), (-1, -1), 0.25,colors.gray),

                     ])
        d.setStyle(dts)
        story.append(d)

        story.append(Spacer(2.5, 0.5 * cm))
        noteData = [[engineerNotes]]
        n = Table(noteData,rowHeights=2*cm,spaceBefore=0.5*cm,spaceAfter=0.5*cm)
        nts = TableStyle([

                         ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                         ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                         ('BOX', (0, 0), (-1, -1), 0.25,colors.gray),


                     ])
        n.setStyle(nts)
        story.append(n)
        story.append(e)
        story.append(s)
        story.append(ter)
        pdf_doc.build(story,  canvasmaker=ServicerecieptCanvas)
class ServiceRecieptAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny ,)
    def get(self,request , format= None):
        response = HttpResponse(content_type='application/pdf')
        contactObj = Contacts.objects.get(pk = int(request.GET['id']))
        response['Content-Disposition'] = 'attachment; filename="servicereciept.pdf"'
        genServiceReciept(response, request, contactObj)
        return response

class GetOnlyContactAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny ,)
    def get(self,request , format= None):
        contObj = Contacts.objects.filter(contact_ref = True)
        today = datetime.date.today()
        if 'typ' in request.GET:
            if request.GET['typ'] == 'CustomPeriod':
                if 'date1' in request.GET and 'date2' in request.GET:
                    date1 = request.GET['date1']
                    date2 = request.GET['date2']
                    if date2<date1:
                        fromDate = date2
                        toDate = date1
                    else:
                        fromDate = date1
                        toDate = date2
                    contObj = contObj.filter(created__range = (fromDate , toDate))
            elif request.GET['typ'] == 'Today':
                contObj = contObj.filter(created = today)
            elif request.GET['typ'] == 'Yesterday':
                yesterday = today - timedelta(days = 1)
                contObj = contObj.filter(created = yesterday)
            elif request.GET['typ'] == 'ThisWeek':
                noDays =  int(today.strftime('%w'))
                fromDate = today - timedelta(days=int(noDays))
                toDate = fromDate + timedelta(days=6)
                contObj = contObj.filter(created__range = (fromDate , toDate))
            elif request.GET['typ'] == 'LastWeek':
                noDays =  int(today.strftime('%w'))
                sundayDate = today - timedelta(days=int(noDays))
                fromDate = sundayDate - timedelta(days=7)
                toDate = sundayDate - timedelta(days=1)
                contObj = contObj.filter(created__range = (fromDate , toDate))
            elif request.GET['typ'] == 'ThisMonth':
                fromDate = today.replace(day=1)
                toDate = today + relativedelta(day=31)
                contObj = contObj.filter(created__range = (fromDate , toDate))
            elif request.GET['typ'] == 'LastMonth':
                toDate = today.replace(day=1) - datetime.timedelta(days=1)
                fromDate = toDate.replace(day=1)
                contObj = contObj.filter(created__range = (fromDate , toDate))
        if 'search' in request.GET:
            search = request.GET['search']
            contObj = contObj.filter(Q(name__icontains = search) | Q(email__icontains = search) |  Q(mobile__icontains = search) )
        data = ContactsSerializer(contObj , many =True).data
        return Response({'data' : data} ,status = status.HTTP_200_OK)

class SaveNextStopAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    def post(self , request , format = None):
        data = request.POST
        if 'pk' in data:
            contactsObj = Contacts.objects.get(pk = int(data['pk']))
            prof =  request.user.profile
            prof.contact = contactsObj.pk
            prof.nextlat = contactsObj.lat
            prof.nextlng = contactsObj.lon
        return Response({}, status = status.HTTP_200_OK)


def convert(seconds):
    min, sec = divmod(seconds, 60)
    hour, min = divmod(min, 60)
    return "%d:%02d" % (hour, min)

class GetVisitDetailsAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny ,)
    def get(self , request , format = None):
        data = request.GET
        import datetime

        if 'user' in data:
            userObj = User.objects.get(pk = int(data['user']))
            dated = data['selectedDate']
            timesheetObj = None
            try:
                timesheetObj = TimeSheet.objects.get(date = dated , user = userObj)
                total_distance = round(timesheetObj.distanceTravelled/1000, 2)
            except:
                total_distance = 0
            tourTrackerObj = UserTourTracker.objects.filter(user = userObj , created__contains = dated)
            tourTrackerList = UserTourTrackerSerializer(tourTrackerObj , many = True).data
            min = 0
            max = tourTrackerObj.count()
            options = []
            options = list(range(0, len(tourTrackerList)))
            tourValues = []
            tourObj = TourPlanStop.objects.filter(tourplan__user =  userObj, tourplan__date = dated)
            for t in tourObj:
                totalDist = 0
                total_time = 0
                driving_time = 0
                duration = {}
                distance = {}
                lastIndex = len(tourTrackerList)
                if lastIndex>0:
                    origins = str(tourTrackerList[lastIndex-1]['lat']) +', '+ str(tourTrackerList[lastIndex-1]['lng'])
                    destinations = str(t.contact.lat) +', '+ str(t.contact.lng)
                    url = "https://maps.googleapis.com/maps/api/distancematrix/json?origins={0}&destinations={1}&mode=driving&language=en-EN&sensor=false&key={2}".format(str(origins),str(destinations),str(globalSettings.MAP_API_KEY))
                    result= json.load(urllib.urlopen(url))
                    if result['status'] == 'OK':
                        if result['rows'][0]['elements'][0]['status'] == 'OK':
                            output = result['rows'][0]['elements'][0]
                            duration = output['duration']
                            distance = output['distance']
                    if t.techStatus == 'in-transit':
                        totalDist =  result['rows'][0]['elements'][0]['distance']['value']
                        total_time = result['rows'][0]['elements'][0]['duration']['value']
                    tourValues.append({'customer' : t.contact.name , 'lat' : t.contact.lat , 'lng' : t.contact.lng, 'expectedDuration': duration, 'expectedDistance': distance,'techStatus':t.techStatus , 'city' : t.contact.city , 'state' : t.contact.state , 'country' : t.contact.country , 'addrs' : t.contact.addrs, 'pinCode' : t.contact.pinCode,'technician': t.tourplan.user.first_name , 'totalDist' : round((totalDist/1000),2) , 'total_time' : convert(total_time) })
            data = {'battery' : userObj.profile.battery_level , 'total_distance' : total_distance, 'tourTrackerList' : tourTrackerList , 'min' : min , 'max' : max, 'options': options , 'tourValues': tourValues}
        return Response({'data' : data}, status = status.HTTP_200_OK)

class CallCurrentCustomerView(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny ,)
    def get(self,request , format= None):

        data = {
          'From': request.user.profile.mobile ,
          'To': request.GET['mobile'] ,
          'CallerId': '02248966707'
        }

        res = requests.post('https://'+ key +':'+  token +'@api.exotel.in/v1/Accounts/'+ sid +'/Calls/connect.json', data=data)
        resData = res.json()
        print resData
        log = Log(txt = 'Outgoing from ' + request.user.first_name  , type = 'routedCall' , callUrl = resData['Call']['Uri'] )
        log.save()
        return Response(resData ,status = status.HTTP_200_OK)

class FetchSIPCalls(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny ,)
    def get(self,request , format= None):
        res = requests.get(globalSettings.CALL_RECORDING_PATH , params = request.GET )

        return Response(res.json() ,status = status.HTTP_200_OK)

class UpdatePlayTime(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny ,)
    def get(self,request , format= None):
        pk = request.GET['pk']
        playTime = request.GET['playTime']
        dataToSend = {'playTime': playTime}
        res = requests.patch(globalSettings.CALL_RECORDING_PATH+ pk+'/', data = dataToSend )

        return Response(res.json() ,status = status.HTTP_200_OK)


class GetCountAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny ,)
    def get(self,request , format= None):
        try:
            dated = datetime.datetime.strptime(request.GET['date'], "%Y-%m-%dT%H:%M:%S.%fZ").date()
        except:
            dated = request.GET['date']
        if 'user' in request.GET:
            user = User.objects.get(pk = int(request.GET['user']))
        tourPlanObj = 0
        tourPlanObj = TourPlanStop.objects.filter(tourplan__user = user , tourplan__date = dated).count()

        return Response(tourPlanObj ,status = status.HTTP_200_OK)

import wave
class GSMCallRecordingAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny ,)
    def post(self,request , format= None):

        dirname = globalSettings.MEDIA_ROOT
        if 'inAudioFile' in request.data:
            completeName = os.path.join(dirname, request.data['inAudioFileName'])
            file1 = wave.open(completeName, "wb")
            file1.writeframesraw(request.data['inAudioFile'])
            file1.close()
        if 'outAudioFile' in request.data:
            completeNameout = os.path.join(dirname, request.data['outAudioFileName'])
            fileout = wave.open(completeNameout, "wb")
            fileout.writeframesraw(request.data['outAudioFile'])
            fileout.close()
        return Response({} ,status = status.HTTP_200_OK)



class GetCampDetailsAPIView(APIView):
    permission_classes = (permissions.AllowAny,)
    def get(self, request, format=None):
        if 'id' in request.GET:
            campObj = Campaign.objects.get(pk = int(request.GET['id']))
            campItemObj = Campaign.objects.filter(parent = campObj)
            camp = CampaignSerializer(campItemObj, many=True).data
            print
            data = []
            for i in camp:
                campItemsObj = CampaignItem.objects.filter(campaign__id = int(i['pk']))
                i['campCount'] = campItemsObj.count()
                contObj = campItemsObj.filter(status = 'converted')
                i['converted'] = contObj.count()
                userObj = User.objects.get(pk = int(i['lead']))
                i['user'] = userObj.first_name + ' ' + userObj.last_name
                data.append(i)
            return Response(data,status = status.HTTP_200_OK)

class DownloadDailyCallReportAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    def get(self , request , format = None):
        if 'user' in request.GET:
            user = User.objects.get(pk = int(request.GET['user']))
            from datetime import date
            today = date.today()
            print today, 'todayDate'
            workbook = Workbook()
            hdFont = Font(size=12,bold=True)
            vertical_align = Alignment(wrap_text=True, horizontal='center',vertical='center')
            right_align = Alignment(wrap_text=True, horizontal='right')
            alphaChars = list(string.ascii_uppercase)

            excelName = "Daily_Call_Report_%s_%s"%(today, user.first_name)

            count = 1
            sheet = workbook.active
            hd = ['S.N.','Date','Data Sample Name', 'Contact Name', 'Contact Number','Status','attempted','skipped','Followup', 'Level']
            sheet.append(hd)

            frmdate = datetime.datetime.strptime(str(today), '%Y-%m-%d').strftime('%Y-%m-%d %H:%M:%S')
            frmdate = datetime.datetime.strptime(str(frmdate), '%Y-%m-%d %H:%M:%S')
            toDate = frmdate.replace(hour=23, minute=59, second=59, microsecond=999999)

            campaignItemObjs = CampaignItem.objects.filter(campaign__lead = user, updated__range = (frmdate , toDate), attempted = True)
            count = 1
            for campaignItem in campaignItemObjs:
                tz_IN = pytz.timezone('Asia/Kolkata')
                number = campaignItem.contact.mobile
                length = len(campaignItem.contact.mobile)
                followSchedule = ''
                if campaignItem.followUpDate is not None:
                    followUpDateIndia = campaignItem.followUpDate.astimezone(tz_IN)
                    followUpDate = followUpDateIndia.date()
                    time = followUpDateIndia.time()
                    followSchedule = str(followUpDate) + ' , '+ str(time).split('.')[0][:-3]
                newNumber ="X" * (length-2)+campaignItem.contact.mobile[-2:]
                sheet.append([count,campaignItem.updated.date(),campaignItem.campaign.name,campaignItem.contact.name,newNumber,campaignItem.status,campaignItem.attempted,campaignItem.skipped,followSchedule,campaignItem.level])
                count +=1

            sheet.column_dimensions['A'].width = 20
            sheet.column_dimensions['B'].width = 20
            sheet.column_dimensions['C'].width= 20
            sheet.column_dimensions['D'].width = 20
            sheet.column_dimensions['E'].width = 20
            sheet.column_dimensions['F'].width = 20
            sheet.column_dimensions['G'].width = 20
            sheet.column_dimensions['H'].width = 20
            sheet.column_dimensions['I'].width = 20
            sheet.column_dimensions['J'].width = 20
            sheet.column_dimensions['K'].width = 20
            sheet.column_dimensions['L'].width = 20
            response = HttpResponse(content=save_virtual_workbook(workbook),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment;filename="%s.xlsx"'%(excelName)
            return response
        else:
            return None

class SendNotfcnAPIView(APIView):
    permission_classes = (permissions.AllowAny,)
    def post(self, request, format=None):
        if 'comment' in request.data:
            msg = request.data['comment']
            userObj = request.user
            notification.objects.create(user = userObj , message = msg)
            team = userObj.designation.team
            if team is not None:
                teamLeadObj = team.manager
                notification.objects.create(user = teamLeadObj , message = msg)
            adminObj = User.objects.filter(is_superuser = True)
            for i in adminObj:
                notification.objects.create(user = i , message = msg)
            contactObj = Contacts.objects.get(pk = int(request.data['contact']))
            contactObj.qcReviewed = True
            contactObj.qcApproved = False
            contactObj.save()
            return Response(status = status.HTTP_200_OK)


class createExpenseSheetAPIView(APIView):
    permission_classes = (permissions.AllowAny,)
    def post(self, request, format=None):
        data = request.data
        tourplanstopObj = TourPlanStop.objects.get(pk = int(data['tourPlanPk']))
        notes = request.user.first_name + ' - Expenses'
        expenseObj = ExpenseSheet.objects.create(notes = notes , user =request.user)
        tourplanstopObj.expense = expenseObj
        tourplanstopObj.save()
        obj = TourPlanStopSerializer(tourplanstopObj, many = False).data
        return Response(obj,status = status.HTTP_200_OK)


class DownloadExpensesAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny ,)
    def get(self,request , format= None):
        if 'id' not in request.GET:
            return Response(status=status.HTTP_400_BAD_REQUEST)
        response = HttpResponse(content_type='application/pdf')
        tourplanObj = TourPlanStop.objects.get(id=request.GET['id'])
        o = tourplanObj.expense
        response.aggrement = o
        response['Content-Disposition'] = 'attachment; filename="expenses_%s.pdf"' % (o.pk)
        genExpense(response,tourplanObj, o, request)
        return response
