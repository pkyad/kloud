from django.contrib.auth.models import User , Group
from django.shortcuts import render, redirect , get_object_or_404
from django.contrib.auth import authenticate , login , logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.urlresolvers import reverse
from django.template import RequestContext
from django.conf import settings as globalSettings
from django.core.exceptions import ObjectDoesNotExist , SuspiciousOperation
from django.views.decorators.csrf import csrf_exempt, csrf_protect
# Related to the REST Framework
from rest_framework import viewsets , permissions , serializers
from rest_framework.exceptions import *
from url_filter.integrations.drf import DjangoFilterBackend
from .serializers import *
from API.permissions import *
from ERP.models import *
from organization.models import CompanyHolidays
from clientRelationships.models import Contact, CustomerSession
from ERP.views import getApps
from django.db.models import Q
from django.http import JsonResponse
import random, string
from django.utils import timezone
from rest_framework.views import APIView
from datetime import date,timedelta,datetime
from dateutil.relativedelta import relativedelta
import calendar
from rest_framework.response import Response
from rest_framework.renderers import JSONRenderer
import json
from django.core.mail import send_mail , EmailMessage
from finance.models import Sale
from payroll.models import Payslip, payroll
from django.template.loader import render_to_string, get_template
from openpyxl import load_workbook
from io import BytesIO,StringIO
from performance.models import TimeSheet
from HR.models import *
from ERP.send_email import send_email
from ERP.views import CreateUnit
from payroll.serializers import payrollSerializer
import django
import requests

class userProfileViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    serializer_class = userProfileSerializer
    queryset = profile.objects.all()

class userProfileAdminModeViewSet(viewsets.ModelViewSet):
    permission_classes = (isAdmin ,)
    serializer_class = userProfileAdminModeSerializer
    queryset = profile.objects.all()
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['user','location_track','job_type']

class userDesignationViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    # queryset = designation.objects.all()
    serializer_class = userDesignationSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['user','division','unit','role','team']
    def get_queryset(self):
        toReturn = designation.objects.all()
        if 'search' in self.request.GET:
            val = self.request.GET['search']
            toReturn = designation.objects.filter(Q(user__first_name__icontains = val) | Q(user__email__icontains = val) | Q(user__profile__mobile__icontains = val)| Q(unit__name__icontains = val,unit__city__icontains=val)|Q(department__dept_name__icontains = val)|Q(role__name__icontains = val))
            return toReturn
        return toReturn
class userAdminViewSet(viewsets.ModelViewSet):
    permission_classes = (isAdmin ,)
    queryset = User.objects.all()
    serializer_class = userAdminSerializer

class userCreateViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    queryset = User.objects.all()
    serializer_class = userAdminSerializer

class DocumentViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    # queryset = Documents.objects.all()
    serializer_class = DocumentSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['name']
    def get_queryset(self):
        # if self.request.user.is_superuser:
        divsn = self.request.user.designation.division
        rt = Document.objects.filter( division = divsn)
        if 'search' in self.request.GET:
            rt = rt.filter(Q(name__icontains = self.request.GET['search'] ) | Q(description__icontains = self.request.GET['search'] ))
        return rt

def usersAccessToModify(root):
    if root:
        return User.objects.all().order_by('-date_joined')
    else:
        return User.objects.filter(is_superuser = False).order_by('-date_joined')


class UserViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated ,)
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['username','email','is_staff','is_active','designation','profile','first_name','last_name']
    search_fields = ('email','designation','profile','first_name','last_name')
    serializer_class = userSerializer
    def get_queryset(self):
        user = self.request.user
        divsn = user.designation.division
        profileObj = user.profile
        # if profileObj.sipUserName is None:
        #     try:
        #         idVal = 100 + int(user.pk)
        #         URL = 'https://'+globalSettings.SIP_WSS_SERVER +"/createAnEndpoint/?exten="+str(idVal)+"&username="+profileObj.mobile+str(divsn.pk)
        #         r = requests.get(url = URL)
        #         data = r.json()
        #         print data, 'datatata'
        #         profileObj.sipUserName = data['auths'][0]['username']
        #         profileObj.sipPassword = data['auths'][0]['password']
        #         profileObj.sipExtension = data['exten']
        #         profileObj.save()
        #
        #     except:
        #         print 'issue in createAnEndpoint'
        #         pass

        if 'reportingTo' in self.request.GET:
            userPks = list(designation.objects.filter(reportingTo = int(self.request.GET['reportingTo']) , division = divsn ).values_list('user__pk',flat=True))
            print userPks,'userpkkkkkkkkkkk'
            if len(userPks)>0:
                toRet = User.objects.filter(pk__in=userPks)
            else:
                toRet = User.objects.none()
            return toRet
        if 'registeredBy' in self.request.GET:
            user = self.request.user
            divisionObj = user.designation.division
            userData = User.objects.filter(username__icontains = str(self.request.GET['registeredBy']), designation__division = divisionObj)
            return  userData

        if 'unit' in self.request.GET:
            userPks = list(designation.objects.filter(division = divsn).values_list('user__pk',flat=True))
            print userPks,'userpkkkkkkkkkkk'
            if len(userPks)>0:
                toRet = User.objects.filter(pk__in=userPks)
            else:
                toRet = User.objects.none()

            return toRet
        if 'division' in self.request.GET:
            toRet = User.objects.filter(designation__division = divsn)
            return toRet
        if 'mode' in self.request.GET:
            if self.request.GET['mode']=="mySelf":
                if self.request.user.is_authenticated:
                    return User.objects.filter(username = self.request.user.username)
                else:
                    raise PermissionDenied()
            else :
                return usersAccessToModify(self.request.user.is_superuser)
        else:
            return usersAccessToModify(self.request.user.is_superuser)


class UserSearchViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated ,)
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['username','is_staff' , 'first_name' , 'last_name']
    serializer_class = userSearchSerializer
    queryset = User.objects.all()
    def get_queryset(self):
        divsn = self.request.user.designation.division
        if 'mode' in self.request.GET:
            if self.request.GET['mode']=="mySelf":
                if self.request.user.is_authenticated:
                    return User.objects.filter(username = self.request.user.username)
                else:
                    raise PermissionDenied()
            else :
                return User.objects.all().order_by('-date_joined')
        elif 'areaCode' in  self.request.GET:
            print 'herrrrrrrrrr'
            return User.objects.filter(designation__unit__areaCode = self.request.GET['areaCode'])
        elif 'staff' in  self.request.GET:
            print self.request.GET['staff']
            print User.objects.filter(is_staff = self.request.GET['staff'])
            return User.objects.filter(is_staff = self.request.GET['staff'])
        else:
            userObj = User.objects.filter(designation__division = divsn).order_by('-date_joined')
            if 'search' in self.request.GET:
                userObj = userObj.filter(Q(first_name__icontains = self.request.GET['search']) | Q(last_name__icontains = self.request.GET['search']))
            return userObj


class GroupViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    queryset = Group.objects.all()
    serializer_class = groupSerializer




class leaveViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    serializer_class = leaveSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['user' , 'status','fromDate']
    def get_queryset(self):
        import datetime
        print self.request.user.pk,'aaaaaaaaaaa'
        if 'mode' in self.request.GET:
            dated = datetime.date.today() - relativedelta(months=3)
            return Leave.objects.filter(user =  self.request.user,created__gt =  dated)
        if 'leaves' in self.request.GET:
            return Leave.objects.filter(user = self.request.user)
        if self.request.user.is_superuser:
            return Leave.objects.all()
        desigs = self.request.user.managing.all()
        reportees = []
        for d in desigs:
            reportees.append(d.user)
        return Leave.objects.filter(user__in = reportees )

class ExitManagementViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    # queryset = ExitManagement.objects.all()
    serializer_class = ExitManagementSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['is_exited' ]
    def get_queryset(self):
        toRet = ExitManagement.objects.all()
        if 'search' in self.request.GET:
            toRet = toRet.filter(Q(user__first_name__icontains = self.request.GET['search']) | Q(user__last_name__icontains = self.request.GET['search'])| Q(user__email__icontains = self.request.GET['search'])| Q(user__profile__mobile__icontains = self.request.GET['search']))
        return toRet


class AppraisalViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    serializer_class = AppraisalSerializer
    # queryset = Appraisal.objects.all()
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['user','status']
    def get_queryset(self):
        print self.request.GET,'ssssssssss'
        if 'getMyHRapprovals' in self.request.GET:
            return Appraisal.objects.filter(Q(createdUser=self.request.user) | Q(hr=self.request.user))
        if 'getMyManagerapprovals' in self.request.GET:
            return Appraisal.objects.filter(Q(manager=self.request.user) | Q(superManager=self.request.user))
        return Appraisal.objects.all()


class TeamViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    serializer_class = TeamSerializer
    queryset = Team.objects.all()
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['title']


class TeamAllViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    serializer_class = TeamAllSerializer
    queryset = Team.objects.all()
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['title']

# class GetAllTeam(APIView):
#     permission_classes = (permissions.AllowAny,)
#     def get(self, request, format=None):
#         teamObj = Team.objects.all()
#         data = TeamSerializer(teamObj, many=True).data
#         return Response({}, status = status.HTTP_200_OK)



class LeavesCalAPI(APIView):
    def get(self , request , format = None):
        payrollObj = payroll.objects.get(user = self.request.user.pk)
        print payrollObj,payrollObj.off
        fromDate = self.request.GET['fromDate'].split('-')
        toDate = self.request.GET['toDate'].split('-')
        fd = date(int(fromDate[0]), int(fromDate[1]), int(fromDate[2]))
        td = date(int(toDate[0]), int(toDate[1]), int(toDate[2]))
        fromDate = fd + relativedelta(days=1)
        toDate = td + relativedelta(days=1)
        chObj = CompanyHolidays.objects.filter(date__range=(str(fromDate),str(toDate)))
        print fromDate,toDate
        total = (toDate-fromDate).days + 1
        if toDate<fromDate:
            total = 0
        holidays = []
        sundays = []
        saturdays = []
        leaves = 0

        if total > 0:
            daysList = [fromDate + relativedelta(days=i) for i in range(total)]
            print daysList
            for i in daysList:
                print i,i.weekday()
                if i.weekday() < 5:
                    print 'holidays',holidays
                    for j in chObj:
                        if j.date == i:
                            holidays.append({'date':i,'name':j.name})
                elif payrollObj.off and i.weekday() == 5:
                    print 'saturday'
                    saturdays.append(i)
                elif i.weekday() == 6:
                    print 'sunday'
                    sundays.append(i)
            leaves = total - (len(holidays) + len(sundays) + len(saturdays))
        toSend = {'total':total,'holidays':holidays,'sundays':sundays,'saturdays':saturdays,'leaves':leaves,'fromDate':fromDate,'toDate':toDate}

        return Response({'data':toSend}, status = status.HTTP_200_OK)

# class ProfileOrgChartsViewSet(viewsets.ModelViewSet):
#     permission_classes = (permissions.IsAuthenticated,)
#     queryset = designation.objects.all()
#     serializer_class = ProfileOrgChartsSerializer

def findChild(d, pk = None):
    toReturn = []
    sameLevel = False
    for des in  d.user.managing.all():
        try:
            dp = des.user.profile.displayPicture.url
            if dp == None:
                dp = '/static/images/userIcon.png'
        except:
            dp = '/static/images/userIcon.png'

        if des.role:
            role = des.role.name
        else:
            role = ''

        if str(des.user.pk) == pk:
            for tr in toReturn:
                tr['className'] = 'rd-dept'

            clsName = 'middle-level'
            sameLevel = True
        else:
            clsName = 'product-dept'
            if sameLevel:
                clsName = 'rd-dept'

        print des.user , clsName

        toReturn.append({
            "id" : des.user.pk,
            "name" : des.user.first_name + ' ' +  des.user.last_name,
            "dp" : dp,
            "children" : findChild(des),
            "role" : role,
            "className" :  clsName
        })

    return toReturn



class OrgChartAPI(APIView):
    def get(self , request , format = None):
        d = User.objects.get(pk = request.GET['user']).designation
        print d.role,d.reportingTo
        if d.reportingTo is not None:
            d = d.reportingTo.designation
        try:
            dp = d.user.profile.displayPicture.url
            if dp == None:
                dp = '/static/images/userIcon.png'

        except:
            dp = '/static/images/userIcon.png'

        if d.role:
            role = d.role.name
        else:
            role = ''


        if str(d.user.pk) == request.GET['user']:
            clsName = 'middle-level'
        else:
            clsName = 'product-dept'


        toReturn = {
            "id" : d.user.pk,
            "name" : d.user.first_name + ' ' +  d.user.last_name,
            "dp" : dp,
            "children" : findChild(d , pk = request.GET['user']),
            "role" : role,
            "className" :  clsName
        }

        return Response(toReturn )

class sendDeboardingEmailAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    def post(self, request, format=None):
        print 'Sennding-----mailll-------manager'
        date = request.data['startDate']
        emp =  User.objects.get(pk = request.data['userPk'])
        emaiIDs = []
        manager = profile.objects.get(user__pk = request.data['manPk'] )
        sManager = profile.objects.get(user__pk = request.data['sManPk'])
        emaiIDs.append(manager.email)
        emaiIDs.append(sManager.email)
        empMail_subject = str('Deboarding Initiated')
        email_subject =str('Deboard Approval')
        # msgBody='Deboarding Approval Request.'
        ctx = {
            'fname': emp.first_name,
            'lname': emp.last_name,
            'initiatedOn':date,
            'linkUrl': 'cioc.co.in',
            'linkText' : 'View Online',
            'sendersAddress' : '(C) CIOC FMCG Pvt Ltd',
            'sendersPhone' : '841101',
            'linkedinUrl' : 'https://www.linkedin.com/company/13440221/',
            'fbUrl' : 'facebook.com',
            'twitterUrl' : 'twitter.com',
        }
        email_body = get_template('app.HR.deboardEmail.html').render(ctx)
        msg = EmailMessage(email_subject, email_body,  to=emaiIDs)
        msg.content_subtype = 'html'
        msg.send()
        email_emp = get_template('app.HR.deboardEmailtoEmployee.html').render(ctx)
        empMsg = EmailMessage(empMail_subject, email_emp,  to=[emp.profile.email])
        empMsg.content_subtype = 'html'
        empMsg.send()
        return Response({},status = status.HTTP_200_OK)

class UserBulkUploadAPIView(APIView):
    permission_classes = (permissions.IsAuthenticated , isAdmin)
    def post(self, request, format=None):


        excelFile = request.FILES['exFile']
        wb = load_workbook(filename = BytesIO(excelFile.read()) ,  read_only=True)
        count = 1
        for ws in wb.worksheets:
            wsTitle = ws.title
            for i in range(2,ws.max_row+1):
                print ws['A' + str(i)].value,'ppppp'
                try:
                    first_name = ws['A' + str(i)].value
                except:
                    first_name = ''

                try:
                    last_name = ws['B' + str(i)].value
                    if len(last_name) == None:
                        last_name = ''
                except:
                    last_name = ''

                try:
                    email = ws['C' + str(i)].value
                except:
                    email = None

                try:
                    phone = ws['D' + str(i)].value
                except:
                    phone = None
                try:
                    unit = ws['E' + str(i)].value
                except:
                    unit = None

                try:
                    department = ws['F' + str(i)].value
                except:
                    department = None
                try:
                    role = ws['G' + str(i)].value
                except:
                    role = None
                try:
                    ReportingTo = ws['H' + str(i)].value
                except:
                    ReportingTo = None



                username = phone

                if username is not None:
                    userObj , n = User.objects.get_or_create( username = username)
                    userObj.email = email
                    userObj.first_name = first_name
                    userObj.last_name = last_name
                    userObj.save()
                    profile = userObj.profile
                    profile.mobile = phone
                    profile.email = email
                    profile.save()

                    division = Division.objects.get(pk = request.data['division'] )

                    designation = userObj.designation
                    designation.division = division
                    if unit is not None:
                        print unit, 'unit'
                        unitObjs = Unit.objects.filter(name__icontains=str(unit))
                        if unitObjs.count()>0:
                            unit = unitObjs[0]
                            designation.unit = unit
                    if department is not None:
                        departmentObjs = Departments.objects.filter(dept_name__icontains=str(department))
                        if departmentObjs.count()>0:
                            department = departmentObjs[0]
                            designation.department = department
                    if role is not None:
                        roleObjs = Role.objects.filter(name__icontains=str(role))
                        if roleObjs.count()>0:
                            role = roleObjs[0]
                            designation.role = role
                    if ReportingTo is not None:
                        reportingTo = User.objects.filter(profile__mobile__icontains=str(ReportingTo))
                        if reportingTo.count()>0:
                            reportingTo = reportingTo[0]
                            designation.reportingTo = reportingTo


                    designation.save()
                    count+=1

                    # if areaCode != None:
                    #     try:
                    #         print "in get"
                    #         unitObj = Unit.objects.get(areaCode = areaCode)
                    #         print "unitObj : " , unitObj
                    #         designation = userObj.designation
                    #         designation.unit = unitObj
                    #         designation.division = division
                    #         designation.save()
                    #     except:
                    #         print "in create"
                    #         unitObj = Unit.objects.create(areaCode = areaCode , division = division)
                    #         unitObj.save()
                    #         unitObj.name = areaCode
                    #         unitObj.pincode = areaCode
                    #         unitObj.save()
                    #         designation = userObj.designation
                    #         designation.division = division
                    #         designation.unit = unitObj
                    #         designation.save()

            return Response({"count" : count}, status = status.HTTP_200_OK)


from geopy import distance
class SaveLocationAPIView(APIView):
    permission_classes = (permissions.IsAuthenticated ,)
    def post(self, request, format=None):
        try:
            data = request.body
        except Exception as e:
            data = request.data
        import datetime
        print data, 'data'
        today = datetime.date.today()
        user = request.user
        profObj = user.profile
        timeObj, t = TimeSheet.objects.get_or_create(date = today , user = profObj.user)
        if profObj.lat == 0 and  profObj.lon == 0:
            profObj.lat = data['lat']
            profObj.lon = data['lon']
        else:
            oldLat = profObj.lat
            oldLan = profObj.lon
            newLat = data['lat']
            newLon = data['lon']
            profObj.lat = data['lat']
            profObj.lon = data['lon']
            new_dimensions = (newLat, newLon)
            old_dimensions = (oldLat, oldLan)
            dist =  (distance.distance(new_dimensions, old_dimensions).miles)
            timeObj.distanceTravelled = float(timeObj.distanceTravelled) + float(dist)
            timeObj.save()
        trackerObj = UserTourTracker.objects.filter(user = request.user ,  created__contains = today)
        print trackerObj , len(trackerObj)
        dataObj = {'user' : user , 'lat' : data['lat'] , 'lng' : data['lon']}
        if len(trackerObj)>0:
            now = datetime.datetime.now()
            lastTracker = trackerObj.last()
            totalPrevMin = float(lastTracker.created.hour*60) + lastTracker.created.minute
            totalNewMin = float(now.hour*60) + now.minute
            difference = float(totalNewMin) - float(totalPrevMin)
            if difference>0:
                traker = UserTourTracker.objects.create(**dataObj)
        else:
            traker = UserTourTracker.objects.create(**dataObj)

        if 'battery_level' in data:
            if data['battery_level']!=None:
                profObj.battery_level = data['battery_level']
        profObj.save()
        return Response({}, status = status.HTTP_200_OK)


class GetUsersAPIView(APIView):
    permission_classes = (permissions.IsAuthenticated , isAdmin)
    def get(self, request, format=None):
        data = request.data
        user = request.user
        if permission.objects.filter(app__name = 'app.geoLocation', user = user).count()>0 or user.is_superuser == True:
            profObj = profile.objects.filter()
            print 'ssssssssssssssssssssss'
            data  = userProfileViewSerializer(profObj,many=True).data
            return Response({'data' : data}, status = status.HTTP_200_OK)
        else:
            return Response({'error':'Permission Denied'}, status = status.HTTP_200_OK)


class fetchAllDetailsAPIView(APIView):
    permission_classes = (permissions.IsAuthenticated , isAdmin)
    def get(self, request, format=None):
        user=request.user
        userObj = {'name' : user.first_name + ' '+user.last_name, 'pk' : user.pk,'email':user.email,'phone':user.profile.mobile, 'is_staff' : user.is_staff, 'prefix':user.profile.prefix,'profile_pk':user.profile.pk,'dp':'','role_name':'','reportingTo_name':'','hra':'','special':'','lta':'','basic':'','adHoc':'','policyNumber':'','provider':'','amount':'','PFUan':'','al':'','ml':'','adHocLeaves':'','joiningDate':'','accountNumber':'','ifscCode':'','bankName':'','hr_name':'','hr_email':'','hr_mobile':'','hr_dp':'','reportingTo_dp':''}
        if user.profile.displayPicture!=None:
            userObj['dp'] =  str(user.profile.displayPicture)
        if user.designation!=None:
            userObj['designation_pk'] = user.designation.pk
            if user.designation.role!=None:
                userObj['role_name'] = user.designation.role.name
                userObj['role_pk'] = user.designation.role.pk
            if user.designation.team!=None:
                userObj['reportingTo_pk'] = user.designation.team.manager.pk
                userObj['reportingTo_name'] = user.designation.team.manager.first_name + ' ' + user.designation.team.manager.last_name

                userObj['reportingTo_email'] = user.designation.team.manager.email
                userObj['reportingTo_mobile'] = user.designation.team.manager.profile.mobile
                if user.designation.team.manager.profile.displayPicture!=None:
                    userObj['reportingTo_dp'] = str(user.designation.team.manager.profile.displayPicture)
            if user.designation.hrApprover!=None:
                userObj['hr_pk'] = user.designation.hrApprover.pk
                userObj['hr_name'] = user.designation.hrApprover.first_name + ' ' + user.designation.hrApprover.last_name
                userObj['hr_email'] = user.designation.hrApprover.email
                userObj['hr_mobile'] = user.designation.hrApprover.profile.mobile
                if user.designation.hrApprover.profile.displayPicture!=None:
                    userObj['hr_dp'] =str(user.designation.hrApprover.profile.displayPicture)
        if user.payroll!=None:
            print user.payroll.pk
            userObj['payroll_pk'] = user.payroll.pk
            userObj['hra'] = user.payroll.hra
            userObj['special'] = user.payroll.special
            userObj['lta'] = user.payroll.lta
            userObj['basic'] = user.payroll.basic
            userObj['adHoc'] = user.payroll.adHoc
            userObj['policyNumber'] = user.payroll.policyNumber
            userObj['provider'] = user.payroll.provider
            userObj['amount'] = user.payroll.amount
            userObj['PFUan'] = user.payroll.PFUan
            userObj['al'] = user.payroll.al
            userObj['ml'] = user.payroll.ml
            userObj['adHocLeaves'] = user.payroll.adHocLeaves
            userObj['joiningDate'] = user.payroll.joiningDate
            userObj['accountNumber'] = user.payroll.accountNumber
            userObj['ifscCode'] = user.payroll.ifscCode
            userObj['bankName'] = user.payroll.bankName
        return Response(userObj, status = status.HTTP_200_OK)

def myProfile(request):
    u = request.user
    p = u.profile
    return JsonResponse({'name' : u.first_name +  ' ' + u.last_name , 'dp' : None , 'dashboardEnabled' : p.isDashboard , 'isManager' : p.isManager, 'is_staff' : u.is_staff})

class GetMyAppsView(APIView):
    permission_classes = (permissions.IsAuthenticated , isAdmin)
    def get(self, request, format=None):
        u=request.user
        adminApps = ['app.organization' , 'sudo.configure', 'app.employees' ]
        prfl = u.profile
        # if prfl.apps is not None:
        #     print "returning from the profile"
        #     return Response(json.loads(u.profile.apps), status = status.HTTP_200_OK)


        if 'mode' in request.GET and request.GET['mode']== 'recent':
            apps = u.menuapps.all()[:4]

        if 'mode' in request.GET and request.GET['mode']== 'others':
            apps = u.menuapps.all()

        appsJson = []

        if 'displayName__icontains' in request.GET:
            apps = apps.filter(app__displayName__icontains = request.GET['displayName__icontains'] )

        if 'inMenu' in request.GET:
            apps = apps.filter(app__inMenu = True)

        for userapp in apps:
            app = userapp.app
            state = None
            if app.name.startswith('app.'):
                if app.url == None:
                    state = app.name.replace('app.' , app.module + ".")
                displayName = app.displayName
            elif app.name.startswith('sudo.'):
                if app.url == None:
                    state = app.name.replace('sudo.' , app.module + ".")
                displayName = app.displayName
            else:
                continue
            appsJson.append({"name" : app.name , "pk" : app.pk , "icon" : app.icon , "state" : state, "displayName" : displayName , 'url' : app.url , 'appStoreUrl'  : app.appStoreUrl, 'playStoreUrl' : app.playStoreUrl , 'index' : userapp.index })


        # prfl.apps = json.dumps({"apps" : appsJson , "settings" : u.is_staff })
        # prfl.save()
        # print "returning from the database fetch"

        return Response({"apps" : appsJson , "settings" : u.is_staff }, status = status.HTTP_200_OK)

from rest_framework.request import Request
from rest_framework.test import APIRequestFactory
from ERP.serializers import UserAppsLiteSerializer
class UpdatepayrollDesignationMasterAccountAPI(APIView):
    # permission_classes = (permissions.IsAuthenticated)
    def post(self, request, format=None):
        if 'pk' in request.data:
            profObj = profile.objects.get(pk = int(request.data['pk']))
            if 'displayPicture' in request.data:
                profObj.displayPicture = request.data['displayPicture']
            if 'TNCandBond' in request.data:
                profObj.TNCandBond = request.data['TNCandBond']
            if 'resume' in request.data:
                profObj.resume = request.data['resume']
            if 'certificates' in request.data:
                profObj.certificates = request.data['certificates']
            if 'transcripts' in request.data:
                profObj.transcripts = request.data['transcripts']
            if 'otherDocs' in request.data:
                profObj.otherDocs = request.data['otherDocs']
            if 'IDPhoto' in request.data:
                profObj.IDPhoto = request.data['IDPhoto']
            profObj.save()
            return JsonResponse({}, status = status.HTTP_200_OK,safe=False)
        if 'basic' in request.data:
            basicObj = profile.objects.get(pk = int(request.data['basic']['pk']))
            basicObj.empID = request.data['basic']['empID']
            basicObj.empType = request.data['basic']['empType']
            basicObj.prefix = request.data['basic']['prefix']
            basicObj.gender = request.data['basic']['gender']
            basicObj.permanentAddressStreet = request.data['basic']['permanentAddressStreet']
            basicObj.permanentAddressCity = request.data['basic']['permanentAddressCity']
            basicObj.permanentAddressPin = request.data['basic']['permanentAddressPin']
            basicObj.permanentAddressState = request.data['basic']['permanentAddressState']
            basicObj.permanentAddressCountry = request.data['basic']['permanentAddressCountry']
            basicObj.sameAsLocal = request.data['basic']['sameAsLocal']
            basicObj.localAddressStreet = request.data['basic']['localAddressStreet']
            basicObj.localAddressCity = request.data['basic']['localAddressCity']
            basicObj.localAddressPin = request.data['basic']['localAddressPin']
            basicObj.localAddressState = request.data['basic']['localAddressState']
            basicObj.localAddressCountry = request.data['basic']['localAddressCountry']
            basicObj.emergency = request.data['basic']['emergency']
            basicObj.bloodGroup = request.data['basic']['bloodGroup']
            basicObj.website = request.data['basic']['website']
            basicObj.almaMater = request.data['basic']['almaMater']
            basicObj.pgUniversity = request.data['basic']['pgUniversity']
            basicObj.docUniversity = request.data['basic']['docUniversity']
            basicObj.fathersName = request.data['basic']['fathersName']
            basicObj.mothersName = request.data['basic']['mothersName']
            basicObj.wifesName = request.data['basic']['wifesName']
            basicObj.wifesName = request.data['basic']['wifesName']
            basicObj.childCSV = request.data['basic']['childCSV']
            basicObj.note1 = request.data['basic']['note1']
            basicObj.note2 = request.data['basic']['note2']
            basicObj.note3 = request.data['basic']['note3']
            if 'married' in request.data['basic']:
                basicObj.married = request.data['basic']['married']
            if 'mobile' in request.data['basic']:
                basicObj.mobile = request.data['basic']['mobile']
            if 'email' in request.data['basic']:
                basicObj.email = request.data['basic']['email']
            if 'dateOfBirth' in request.data['basic']:
                basicObj.dateOfBirth = request.data['basic']['dateOfBirth']
            if 'job_type' in request.data['basic']:
                basicObj.job_type = request.data['basic']['job_type']
            basicObj.save()
        if 'designation' in request.data:
            desigObj = designation.objects.get(pk = request.data['designation']['pk'])
            if 'reportingTo' in request.data['designation']:
                desigObj.reportingTo = User.objects.get(pk=int(request.data['designation']['reportingTo']))
            if 'hrApprover' in request.data['designation']:
                desigObj.hrApprover = User.objects.get(pk = request.data['designation']['hrApprover'])
            if 'unit' in request.data['designation']:
                desigObj.unit = Unit.objects.get(pk=  request.data['designation']['unit'])
                desigObj.division = desigObj.unit.division
            if 'department' in request.data['designation']:
                desigObj.department = Departments.objects.get(pk=  request.data['designation']['department'])
            if 'role' in request.data['designation']:
                desigObj.role = Role.objects.get(pk=  request.data['designation']['role'])
            if 'team' in request.data['designation']:
                desigObj.team = Team.objects.get(pk=  request.data['designation']['team'])
            desigObj.save()
        if 'payroll' in request.data :
            print request.data['payroll']
            payObj = payroll.objects.get(pk = request.data['payroll']['pk'] )
            payObj.activatePayroll = request.data['payroll']['activatePayroll']
            payObj.accountNumber = request.data['payroll']['accountNumber']
            payObj.adHocLeaves = request.data['payroll']['adHocLeaves']
            payObj.ctc = request.data['payroll']['ctc']
            payObj.al = request.data['payroll']['al']
            payObj.bankName = request.data['payroll']['bankName']
            payObj.esic = request.data['payroll']['esic']
            payObj.ifscCode = request.data['payroll']['ifscCode']
            payObj.ml = request.data['payroll']['ml']
            payObj.off = request.data['payroll']['off']
            payObj.pan = request.data['payroll']['pan']
            payObj.pfAccNo = request.data['payroll']['pfAccNo']
            payObj.pfUniNo = request.data['payroll']['pfUniNo']
            payObj.hra = request.data['payroll']['hra']
            payObj.special = request.data['payroll']['special']
            payObj.lta = request.data['payroll']['lta']
            payObj.basic = request.data['payroll']['basic']
            payObj.taxSlab = request.data['payroll']['taxSlab']
            payObj.adHoc = request.data['payroll']['adHoc']
            payObj.esicAmount = request.data['payroll']['esicAmount']
            payObj.pTax = request.data['payroll']['pTax']
            payObj.bonus = request.data['payroll']['bonus']
            payObj.pfAmnt = request.data['payroll']['pfAmnt']
            payObj.esicAdmin = request.data['payroll']['esicAdmin']
            payObj.pfAdmin = request.data['payroll']['pfAdmin']
            if 'joiningDate' in request.data['payroll']:
                payObj.joiningDate = request.data['payroll']['joiningDate']
            if 'lastWorkingDate' in request.data['payroll']:
                payObj.lastWorkingDate = request.data['payroll']['lastWorkingDate']
            print request.data['payroll']['PFUan'],'aaaaaaaaaa'
            if 'PFUan' in request.data['payroll']:
                payObj.PFUan = request.data['payroll']['PFUan']
            payObj.policyNumber = request.data['payroll']['policyNumber']
            payObj.provider = request.data['payroll']['provider']
            payObj.amount = request.data['payroll']['amount']
            payObj.noticePeriodRecovery = request.data['payroll']['noticePeriodRecovery']
            payObj.notice = request.data['payroll']['notice']
            payObj.probation = request.data['payroll']['probation']
            payObj.probationNotice = request.data['payroll']['probationNotice']
            payObj.save()

        if 'master' in request.data:
            masterUser = User.objects.get(pk = request.data['master']['pk'])
            masterUser.email = request.data['master']['email']
            masterUser.first_name = request.data['master']['first_name']
            masterUser.last_name = request.data['master']['last_name']
            masterUser.is_active = request.data['master']['is_active']
            masterUser.is_staff = request.data['master']['is_staff']
            masterUser.username = request.data['master']['username']
            if 'password' in request.data['master']:
                masterUser.set_password(request.data['master']['password'])
            masterUser.save()


        serializer_context = {
            'request': Request(request),
        }
        print userAdminSerializer(masterUser,context=serializer_context).data,"0923902039209"
        return JsonResponse({'designation':userDesignationSerializer(desigObj,many=False).data,'payroll':payrollSerializer(payObj,many=False).data,'master':userAdminSerializer(masterUser,context=serializer_context).data, 'simpleMode' : request.user.designation.division.simpleMode}, status = status.HTTP_200_OK,safe=False)
    def get(self, request, format=None):
        masterUser = User.objects.get(pk = int(request.GET['id']))
        desigObj = masterUser.designation
        profObj = masterUser.profile
        payObj = masterUser.payroll
        serializer_context = {
            'request': Request(request),
        }

        # if profObj.sipUserName is None:
        #     try:

        #         idVal = 100 + int(profObj.user.pk)
        #         URL = globalSettings.EXTERNAL_SITE +"/createAnEndpoint/?exten="+str(idVal)+"&username="+profObj.user.username
        #         r = requests.get(url = URL)
        #         data = r.json()
        #         print data, 'datatata'
        #         profObj.sipUserName = data['auths'][0]['username']
        #         profObj.sipPassword = data['auths'][0]['password']
        #         profObj.sipExtension = data['exten']
        #         profObj.save()
        #     except:
        #         print 'issue in createAnEndpoint'
        #         pass

        div = request.user.designation.division

        return JsonResponse({
            'designation':userDesignationSerializer(desigObj,many=False).data,
            'apps':UserAppsLiteSerializer(masterUser.menuapps.all(),many=True).data,
            'payroll':payrollSerializer(payObj,many=False).data,
            'profile':userProfileSerializer(profObj,many=False).data,
            'master':userAdminSerializer(masterUser,context=serializer_context).data,
            'basic' : userAllProfileSerializer(profObj , many=False).data,
            'simpleMode' : div.simpleMode ,
            'telephony' : div.telephony,
            'canChangeStaffStatus' : masterUser.pk != request.user.pk
            }, status = status.HTTP_200_OK,safe=False)

@csrf_exempt
def TokenLogin(request):
    key = request.GET['token']
    account = profile.objects.get(linkToken = key)
    user = User.objects.get(pk = account.user.pk)
    user.backend = 'django.contrib.auth.backends.ModelBackend'
    login(request , user)
    if request.method == 'GET':
        return redirect(reverse(globalSettings.LOGIN_REDIRECT))
    else:
        csrf_token =django.middleware.csrf.get_token(request)
        unit = ''
        designation = ''
        department = ''
        role = ''
        if user.designation.unit is not None:
            unit = user.designation.unit.pk
        if user.designation.division is not None:
            designation = user.designation.division.pk
        try:
            if user.designation.department is not None:
                department = user.designation.department.pk
        except:
            department = ''
        if user.designation.role is not None:
            role = user.designation.role.pk
        return JsonResponse({'csrf_token':csrf_token , "pk" : user.pk ,'first_name':user.first_name,'last_name' : user.last_name , 'unit' : unit , 'designation' : designation , 'department' : department,'role' : role} , status = 200)

def randomPassword():
    length = 12
    chars = string.digits
    rnd = random.SystemRandom()
    return ''.join(rnd.choice(chars) for i in range(length))

class GeneratePassKey(APIView):
    renderer_classes = (JSONRenderer,)
    def get(self , request , format = None):
        if 'profile' in request.GET:
            prof = profile.objects.get(pk = request.GET['profile'])
            if prof.linkToken == None or 'generate' in request.GET:
                token = randomPassword()
                prof.linkToken = token
                prof.save()

            data = {'passkey' : prof.linkToken, 'siteAddress' : globalSettings.SITE_ADDRESS}
        return Response(data, status = status.HTTP_200_OK)


def sendWelcomeKite(request):
    toReturn = {}
    toReturn['message'] = 'There is some issue with email client.'
    toReturn['messageType'] = 'warning'
    toReturn['success'] = True

    if 'user' in request.GET:
        profileObj = profile.objects.get(user__pk = request.GET['user'])
        if profileObj.linkToken is None:
            token = randomPassword()
            profileObj.linkToken = token
            profileObj.save()
        print profileObj.linkToken, 'profileObj.linkToken'


        if profileObj.user.email is not None:
            ctx = userSearchSerializer(profileObj.user, many = False).data
            ctx['loginUrl'] = globalSettings.SITE_ADDRESS + '/welcome/'+profileObj.linkToken
            email_body = get_template('app.HR.welcomeKit.html').render(ctx)
            email_subject = 'Welcome Kit'
            email_to=[]
            email_to.append(str(profileObj.user.email))
            email_cc = []
            email_bcc = []
            print 'sending email now...'
            msg = EmailMessage(email_subject, email_body, to=email_to ,  )
            msg.content_subtype = 'html'
            msg.send()
            print 'Sent'

            toReturn['message'] = 'SuccessFully send welcome kit'
            toReturn['messageType'] = 'success'
            toReturn['success'] = True

    return JsonResponse(toReturn, status = status.HTTP_200_OK)


class AddTeamMemberAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    def post(self, request, format=None):
        userObj = User.objects.get(pk = int(request.data['user']))
        if 'typ' in request.data:
            designation = userObj.designation
            designation.team = None
            designation.save()
            return Response({},status=status.HTTP_200_OK)
        teamObj = Team.objects.get(pk = int(request.data['team']))
        designation = userObj.designation
        designation.team = teamObj
        designation.save()
        data = userDesignationLiteSerializer(designation,many=False).data
        return Response(data,status=status.HTTP_200_OK)



class GetWelcomeDetailsAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny,)
    def post(self, request, format=None):
        data = request.data
        user = request.user
        profile = user.profile
        payroll = user.payroll
        if 'submit' in data:
            profile.onboarding = True
            payroll.activatePayroll = True
            profile.save()
            payroll.save()
        else:
            if 'permanentAddressStreet' in data:
                profile.permanentAddressStreet = data['permanentAddressStreet']
            if 'permanentAddressPin' in data:
                profile.permanentAddressPin = data['permanentAddressPin']
            if 'permanentAddressCity' in data:
                profile.permanentAddressCity = data['permanentAddressCity']
            if 'permanentAddressState' in data:
                profile.permanentAddressState = data['permanentAddressState']
            if 'permanentAddressCountry' in data:
                profile.permanentAddressCountry = data['permanentAddressCountry']
            if 'localAddressStreet' in data:
                profile.localAddressStreet = data['localAddressStreet']
            if 'localAddressPin' in data:
                profile.localAddressPin = data['localAddressPin']
            if 'localAddressCity' in data:
                profile.localAddressCity = data['localAddressCity']
            if 'localAddressState' in data:
                profile.localAddressState = data['localAddressState']
            if 'localAddressCountry' in data:
                profile.localAddressCountry = data['localAddressCountry']
            if 'pan' in data:
                payroll.pan = data['pan']
            if 'dateOfBirth' in data:
                profile.dateOfBirth = data['dateOfBirth']
            if 'PFUan' in data:
                payroll.PFUan = data['PFUan']
            if 'gender' in data:
                profile.gender = data['gender']
            if 'bloodGroup' in data:
                profile.bloodGroup = data['bloodGroup']
            if 'married' in data:
                profile.married = data['married']
            if 'note1' in data:
                profile.note1 = data['note1']
            if 'note2' in data:
                profile.note2 = data['note2']
            if 'accountNumber' in data:
                payroll.accountNumber = data['accountNumber']
            if 'ifscCode' in data:
                payroll.ifscCode = data['ifscCode']
            if 'bankName' in data:
                payroll.bankName = data['bankName']
            if 'anivarsary' in data:
                profile.anivarsary = data['anivarsary']
            if 'adhar' in data:
                profile.adhar = data['adhar']
            emergency = ''
            if 'emergencyName' in data:
                emergency = data['emergencyName']
            if 'emergencyNumber' in data:
                emergency = emergency +'::'+data['emergencyNumber']
            profile.emergency = emergency
            if 'displayPicture' in data:
                profile.displayPicture = data['displayPicture']
            profile.save()
            payroll.save()
        return Response({},status=status.HTTP_200_OK)
    def get(self, request, format=None):
        user = request.user
        data = {
        'first_name' : user.first_name,
        'last_name' : user.last_name,
        'pk':user.pk,
        # 'profilePk' : user.profile.pk,
        'permanentAddressStreet' : user.profile.permanentAddressStreet,
        'permanentAddressPin' : user.profile.permanentAddressPin,
        'localAddressStreet' : user.profile.localAddressStreet,
        'localAddressPin' : user.profile.localAddressPin,
        'localAddressCity' : user.profile.localAddressCity,
        'localAddressCountry' : user.profile.localAddressCountry,
        'localAddressState' : user.profile.localAddressState,
        'permanentAddressCity' : user.profile.permanentAddressCity,
        'permanentAddressState' : user.profile.permanentAddressState,
        'permanentAddressCountry' : user.profile.permanentAddressCountry,
        'pan' : user.payroll.pan,
        'dateOfBirth' : user.profile.dateOfBirth,
        'PFUan' : user.payroll.PFUan,
        'adhar' : user.profile.adhar,
        'gender' : user.profile.gender,
        'bloodGroup' : user.profile.bloodGroup,
        'married' : user.profile.married,
        'note1' : user.profile.note1,
        'note2' : user.profile.note2,
        'accountNumber' : user.payroll.accountNumber,
        'reaccountNumber' : user.payroll.accountNumber,
        'ifscCode' : user.payroll.ifscCode,
        'bankName' : user.payroll.bankName,
        'anivarsary' : user.profile.anivarsary,
        'adhar' : user.profile.adhar,
        'emergencyName' :'',
        'emergencyNumber' :''
        # 'displayPicture' : user.profile.displayPicture
        }
        if user.profile.displayPicture:
            data['displayPicture'] = globalSettings.SITE_ADDRESS + '/media/'+str(user.profile.displayPicture)
        else:
            data['displayPicture'] = ''
        if user.profile.emergency and len(user.profile.emergency)>0:
            data['emergencyName'] = user.profile.emergency.split('::')[0]
            data['emergencyNumber'] = user.profile.emergency.split('::')[1]
        return Response(data,status=status.HTTP_200_OK)

def resetAppsInProfile(user):
    p = user.profile
    p.apps = None
    print "reseting the apps in the profile" , p.pk , p.user.pk
    p.save()

class AppInstallerView(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.IsAuthenticated,)
    def post(self, request, format=None):

        app = InstalledApp.objects.get(pk = request.data['app']).app
        ua = UserApp(app = app , user = designation.objects.get(pk = request.data['designation']).user )
        ua.save()

        return Response({} , status = status.HTTP_200_OK)

    def get(self ,request ,format = None):
        installations = request.user.designation.division.installations.filter(~Q(app__name__in =  ['app.dashboard' , 'app.messenger'] ))
        if 'search' in request.GET:
            installations = installations.filter(app__displayName__icontains = request.GET['search'])
        data = InstalledAppSerializer(installations , many = True).data
        return Response(data , status = status.HTTP_200_OK)

    def delete(self , request, format = None):
        app = UserApp.objects.get(pk = request.GET['app']).delete()
        return Response({} , status = status.HTTP_200_OK)


class  UpdatePushTokenView(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.IsAuthenticated,)
    def post(self, request, format=None):
        if 'userID' in request.data:
            prof = User.objects.get(pk = int(request.data['userID'])).profile
            if 'expoPushToken' in request.data:
                prof.expoPushToken = request.data['expoPushToken']
                prof.save()
                return Response( status = status.HTTP_200_OK)


class RegNewUserView(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny,)
    def post(self, request, format=None):
        data = request.data
        if 'pk' in data:
            userObj = User.objects.get(pk = int(data['pk']))
            if 'name' in data:
                userObj.first_name = data['name']
                userObj.save()
            if 'company' in data:
                divObj = Division.objects.create(name = data['company'])
                des = userObj.designation
                des.division = divObj
                val = {'name':userObj.first_name, 'pk': userObj.pk, 'userNumber' : '' , 'email' : '', 'division_pk' : des.division.pk}
                res = CreateUnit(val)
                unit = Unit.objects.get(pk = int(res['unit']))
                des.unit = unit
                print res
                des.save()

        return Response({} , status = status.HTTP_200_OK)
