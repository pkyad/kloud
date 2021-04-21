from django.contrib.auth.models import User, Group
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.core.urlresolvers import reverse
from django.template import RequestContext
from django.conf import settings as globalSettings
# Related to the REST Framework
from rest_framework import viewsets, permissions, serializers
from rest_framework.exceptions import *
from url_filter.integrations.drf import DjangoFilterBackend
from .serializers import *
from API.permissions import *
from django.db.models import Q, F
from django.http import HttpResponse
from allauth.account.adapter import DefaultAccountAdapter
from rest_framework.views import APIView
from rest_framework.renderers import JSONRenderer
from ERP.models import service, appSettingsField,address
from HR.models import profile
# Create your views here.
from reportlab import *
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm, mm
from reportlab.lib import colors, utils
from reportlab.platypus import Paragraph, Table, TableStyle, Image, Frame, Spacer, PageBreak, BaseDocTemplate, PageTemplate, SimpleDocTemplate, Flowable
from PIL import Image
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet, TA_CENTER
from reportlab.graphics import barcode, renderPDF
from reportlab.graphics.shapes import *
from reportlab.graphics.barcode.qr import QrCodeWidget
import datetime
import calendar as pythonCal
import json
import pytz
import requests
from django.template.loader import render_to_string, get_template
from django.core.mail import send_mail, EmailMessage
from django.db.models import Sum, Count
from dateutil.relativedelta import relativedelta
from PIM.models import calendar
from excel_response import ExcelResponse
from django.db.models.functions import Concat
from django.db.models import Value
import xlsxwriter
from clientRelationships.serializers import ContractSerializer,ContactSerializer,DealLiteSerializer
from projects.models import project,Issues, ProjectObjective
from projects.serializers import IssueSerializer
from finance.models import Sale,SalesQty, TermsAndConditions
from finance.serializers import SaleSerializer,SalesQtySerializer
from mail.models import mailAttachment
from ERP.models import GenericPincode
from rest_framework import filters
from django.http import JsonResponse
from svglib.svglib import svg2rlg
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from aggrement import *
# from Invoice1 import *
# from Invoice2 import *
import random, string
from openpyxl.writer.excel import save_virtual_workbook
import basehash
hash_fn = basehash.base36()





class TeamMembersStatsView(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny,)
    def get(self, request, format=None):
        max = None
        toReturn = []
        divsn = self.request.user.designation.division
        for u in User.objects.filter(designation__division = divsn).annotate(total_contracts = Count('contracts') ).order_by('-total_contracts')[:10]:
            if max is None:
                max = u.total_contracts
            if u.profile.displayPicture:
                dp = u.profile.displayPicture.url
            else:
                dp = None
            if max == 0:
                max = 1
            toReturn.append( {"name" : u.first_name , "count" : u.total_contracts , "percent" : int(u.total_contracts*100/max) , "id" : u.pk , "dp" : dp } )
        return Response(toReturn, status=status.HTTP_200_OK)



class DownloadInvoice(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny,)
    def get(self, request, format=None):
        if 'contract' not in request.GET:
            return Response(status=status.HTTP_400_BAD_REQUEST)
        response = HttpResponse(content_type='application/pdf')
        if 'user' in request.GET:
            user = User.objects.get(pk = int(request.GET['user']))
        else:
            user =  request.user
        o = Contract.objects.get(id=request.GET['contract'] , division = user.designation.division)
        response.contract = o
        response.division = user.designation.division
        response.unit = user.designation.unit
        if o.termsAndCondition is not None:
            if o.termsAndCondition.version == 'V2':
                from Invoice2 import *
            elif o.termsAndCondition.version == 'V3':
                from Invoice3 import *
            else:
                from Invoice1 import *
        else:
            # try:
            crmObj = CRMTermsAndConditions.objects.filter(division = user.designation.division).first()
            if crmObj.version == 'V2':
                from Invoice2 import *
            elif crmObj.version == 'V3':
                from Invoice3 import *
            else:
                from Invoice1 import *
            # except:
            #     pass
        response['Content-Disposition'] = 'attachment; filename="CR_%s%s_%s_%s.pdf"' % (o.status,
            o.pk, datetime.datetime.now(pytz.timezone('Asia/Kolkata')).year, o.pk)
        genInvoice(response, o, request)
        f = open(os.path.join(globalSettings.BASE_DIR, 'media_root/CR_%s%s_%s.pdf' %
                              (o.pk, datetime.datetime.now(pytz.timezone('Asia/Kolkata')).year, o.pk)), 'wb')
        f.write(response.content)
        f.close()
        if 'saveOnly' in request.GET:
            return Response(status=status.HTTP_200_OK)
        if 'output' in request.GET:
            filePath = os.path.join(globalSettings.BASE_DIR, 'media_root/CR_%s%s_%s_%s.pdf' %
                                 (o.status,o.pk, datetime.datetime.now(pytz.timezone('Asia/Kolkata')).year, o.pk))
            f = open(filePath, 'wrb')
            f.write(response.content)
            f.close()
            file_name = 'media/' + filePath.split('/')[-1]
            return Response({'fileUrl' : file_name }, status = status.HTTP_200_OK)
        return response



class ContactLiteViewSet(viewsets.ModelViewSet):
    permission_classes = (isOwner, )
    serializer_class = ContactLiteSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['name', 'company', 'mobile', 'email','typ']

    def get_queryset(self):
        return Contact.objects.all()

class ServiceAllViewSet(viewsets.ModelViewSet):
    permission_classes = (isOwner, )
    serializer_class = serviceAllSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['name', 'company', 'mobile', 'email','typ']
    def get_queryset(self):
        divsn = self.request.user.designation.division
        return service.objects.filter(division = divsn)

class ContactViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.AllowAny, )
    serializer_class = ContactSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filter_fields = ['name','company' , 'email' , 'mobile' , 'emailSecondary' , 'mobileSecondary','updated','typ']
    search_fields = ('name', 'email', 'company__name', 'mobile','typ')
    def get_queryset(self):
        divsn = self.request.user.designation.division
        toReturn = Contact.objects.filter(user__designation__division = divsn).order_by('-created')
        if 'lastUpdate' in self.request.GET:
            lastUpdate = datetime.datetime.strptime(self.request.GET['lastUpdate'], '%Y-%m-%dT%H:%M:%S.%fZ').strftime('%Y-%m-%dT%H:%M:%S.%fZ')
            toReturn = toReturn.filter(updated__gt = lastUpdate)
        if 'search' in self.request.GET:
            val = self.request.GET['search']
            toReturn = toReturn.filter(Q(name__icontains =val)|Q(email__icontains = val)|Q(company__name__icontains = val)|Q(mobile__icontains = val))
        if 'orderBy' in self.request.GET:
            toReturn = toReturn.order_by('name')
        return toReturn


class ContactAllViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.AllowAny, )
    serializer_class = ContactAllSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filter_fields = ['name','company' , 'email' , 'mobile' , 'emailSecondary' , 'mobileSecondary']
    search_fields = ('name', 'email', 'company__name', 'mobile')
    def get_queryset(self):
        user = self.request.user
        toReturn =   Contact.objects.filter(user__designation__division = user.designation.division).order_by('-created')
        if 'search' in self.request.GET:
            val = self.request.GET['search']
            toReturn =   toReturn.filter(Q(name__icontains =val)|Q(email__icontains = val)|Q(company__name__icontains = val)|Q(mobile__icontains = val))
        if 'products' in self.request.GET:
            toReturn = toReturn.annotate(number_of_entries=Count('productContact'))
            toReturn = toReturn.filter(number_of_entries__gt = 0)
        return toReturn

class DealLiteViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticatedOrReadOnly, readOnly, )
    serializer_class = DealLiteSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['name', 'result', 'company']

    def get_queryset(self):
        return Deal.objects.filter(active=True)


class DealViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated, )
    serializer_class = DealSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['name', 'state', 'result', 'company']

    def get_queryset(self):

        if 'created' in self.request.GET and 'won' in self.request.GET and 'lost' in self.request.GET:
            qs1, qs2, qs3 = Deal.objects.none(), Deal.objects.none(), Deal.objects.none()
            if self.request.GET['created'] == '1':
                qs1 = Deal.objects.filter(state='created')
            if self.request.GET['won'] == '1':
                qs2 = Deal.objects.filter(result='won')
            if self.request.GET['lost'] == '1':
                qs2 = Deal.objects.filter(result='lost')

            toReturn = qs1 | qs2 | qs3
        else:
            toReturn = Deal.objects.all()

        toReturn = toReturn.filter(active=True)
        if 'company__contains' in self.request.GET:
            comName = self.request.GET['company__contains']
            toReturn = toReturn.filter(
                company__in=service.objects.filter(name__icontains=comName))

        if 'created' in self.request.GET and self.request.GET['created'] == 'false':
            toReturn = toReturn.exclude(state='created')

        if 'board' in self.request.GET:
            toReturn = toReturn.filter(result='na')

        return toReturn


class FilesViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated, )
    serializer_class = FilesSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['title', 'description']
    def get_queryset(self):
        divsn = self.request.user.designation.division
        return  Files.objects.filter(user__designation__division = divsn)


class EmailTemplateViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated, )
    serializer_class = EmailTemplateSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['title']
    def get_queryset(self):
        divsn = self.request.user.designation.division
        emailObj =  EmailTemplate.objects.filter(creator = True)
        emailObj = emailObj.filter(user = self.request.user)
        emailData = EmailTemplate.objects.filter(creator = False , user__designation__division = divsn)
        return emailData | emailObj


class ContractViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated, )
    serializer_class = ContractSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['status', 'value','id','contact','user' , 'deal']
    def get_queryset(self):
        divisn =  self.request.user.designation.division
        if 'search' in self.request.GET:
            try:
                toRet = Contract.objects.filter(value__gte=int(self.request.GET['search']), division = divisn)
                return toRet
            except:
                return Contract.objects.all(division = divisn).order_by('-created')
        if 'filters' in self.request.GET:
            prts =  self.request.GET['filters'].split(',')
            statss = []
            if prts[0] == 'true':
                statss.append('quoted')
            if prts[1] == 'true':
                statss.append('billed')
            if prts[2] == 'true':
                statss.append('dueElapsed')
            return Contract.objects.filter(status__in =  statss, division = divisn).order_by('-created')
        else:
            return Contract.objects.filter(division = divisn).order_by('-created')


class ActivityViewSet(viewsets.ModelViewSet):
    permission_classes = (isOwner, )
    serializer_class = ActivitySerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['contact', 'deal', 'notes', 'data', 'typ','user']

    def get_queryset(self):
        return Activity.objects.order_by('-created')



class CRMTermsAndConditionsViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,  )
    serializer_class = CRMTermsAndConditionsSerializer
    def get_queryset(self):
        divsn = self.request.user.designation.division
        return CRMTermsAndConditions.objects.filter(division = divsn)

class LegalAgreementViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,  )
    serializer_class = LegalAgreementSerializer
    def get_queryset(self):
        return LegalAgreement.objects.all()


class LegalAgreementTermsViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,  )
    serializer_class = LegalAgreementTermsSerializer
    filter_fields = ['parent' ]
    def get_queryset(self):
        return LegalAgreementTerms.objects.all()



from email.mime.application import MIMEApplication

class SendNotificationAPIView(APIView):
    renderer_classes = (JSONRenderer,)

    def post(self, request, format=None):
        print request.data
        toEmail = []
        cc = []

        if request.data['sendEmail']:
            for c in request.data['contacts']:
                em = Contact.objects.get(pk=c).email
                if em is not None:
                    toEmail.append(em)
            print toEmail

            for c in request.data['internal']:
                p = profile.objects.get(user_id=c)
                u = User.objects.get(pk=c)
                cc.append(u.email)
            print cc

        toSMS = []
        if request.data['sendSMS']:
            for c in request.data['contacts']:
                mob = Contact.objects.get(pk=c).mobile
                if mob is not None:
                    toSMS.append(mob)
            print toSMS

        c = Contract.objects.get(pk=request.data['contract'], division = request.user.designation.division)
        docID = '%s%s%s' % (c.deal.pk, c.billedDate.year, c.pk)
        value = c.grandTotal

        typ = request.data['type']
        print "will send invoice generated mail to ", toEmail, cc, toSMS
        print docID, value
        if typ == 'invoiceGenerated':
            email_subject = 'Invoice %s generated' % (docID)
            heading = 'Invoice Generated'
            msgBody = ['We are pleased to share invoice number <strong>%s</strong> for the amount of INR <strong>%s</strong>.' %
                       (docID, value), 'The due date to make payment is <strong>%s</strong>.' % (c.dueDate), 'In case you have any query please contact us.']
            smsBody = 'Invoice %s generated for the amount of INR %s. Due date is %s. Please check youe email for more information.' % (
                docID, value, c.dueDate)
        elif typ == 'dueDateReminder':
            email_subject = 'Payment reminder for invoice %s' % (docID)
            heading = 'Payment reminder'
            msgBody = ['We are sorry but invoice number <strong>%s</strong> for the amount of INR <strong>%s</strong> is still unpaid.' %
                       (docID, value), 'The due date to make the payment is <strong>%s</strong>. Please make payment at the earliest to avoid late payment fee.' % (c.dueDate), 'In case you have any query please contact us.']
            smsBody = 'REMINDER : Invoice no. %s is sill unpaid. Due date is %s. Please ignore if paid.' % (
                docID, c.dueDate)
        elif typ == 'dueDateElapsed':
            email_subject = 'Payment overdue for invoice number %s' % (docID)
            heading = 'Due date missed'
            msgBody = ['We are pleased to share the updated copy of invoice number <strong>%s</strong> for the amount of INR <strong>%s</strong> including the late payment fees.' %
                       (docID, value), 'The payment is now due <strong>Immediately</strong>.', 'In case you have any query please contact us.']
            smsBody = 'ALERT : Invoice no. %s updated to include late payment fee. Please pay immediately. Check your email for more info.' % (
                docID)

        ctx = {
            'heading': heading,
            'recieverName': Contact.objects.get(pk=request.data['contacts'][0]).name,
            'message': msgBody,
            'linkUrl': 'cioc.co.in',
            'linkText': 'View Online',
            'sendersAddress': '(C) CIOC FMCG Pvt Ltd',
            'sendersPhone': '841101',
            'linkedinUrl': 'https://www.linkedin.com/company/13440221/',
            'fbUrl': 'facebook.com',
            'twitterUrl': 'twitter.com',
        }

        email_body = get_template(
            'app.clientRelationships.email.html').render(ctx)
        msg = EmailMessage(email_subject, email_body, to=toEmail,
                           cc=cc, from_email='do_not_reply@cioc.co.in')
        msg.content_subtype = 'html'

        if typ != 'dueDateReminder':
            fp = open('./media_root/clientRelationships/doc%s%s_%s.pdf' %
                      (c.deal.pk, c.pk, c.status), 'rb')
            att = MIMEApplication(fp.read(), _subtype="pdf")
            fp.close()
            att.add_header('Content-Disposition', 'attachment',
                           filename=fp.name.split('/')[-1])
            msg.attach(att)

        msg.send()

        for n in toSMS:
            url = globalSettings.SMS_API_PREFIX + \
                'number=%s&message=%s' % (n, smsBody)
            requests.get(url)

        return Response(status=status.HTTP_200_OK)

def QuoteDouwnload(data):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "UID Wise"
    heading = ['#','Date','Customer', 'Customer Mobile', 'Customer Email', 'Customer Address', 'Company' , 'Value' , 'Sender']
    heading_font = Font(bold=True, size=14)
    ws1.append(heading)
    for idx,cell in enumerate(ws1["1:1"]):
        cell.font = heading_font
        ws1.column_dimensions[get_column_letter(idx+1)].width = 25 if idx<3 else 50
    for row in data:
        data =[str(row['pk']),str(row['created'].strftime("%b %d, %Y")),]
        if row['contact__pk']:
            data.append(str(row['contact__name']))
            data.append(str(row['contact__mobile']))
            data.append(str(row['contact__email']))
        else:
            data.append(' ')
            data.append(' ')
            data.append(' ')


        address = ''
        if row['contact__street']:
            address+=row['contact__street']+' '
        if row['contact__city']:
            address+=row['contact__city']+' '
        if row['contact__state']:
            address+=row['contact__state']+' '
        if row['contact__country']:
            address+=row['contact__country']+' '
        if row['contact__pincode']:
            address+=row['contact__pincode']+' '
        data.append(address)
        if row['contact__company__name']:
            data.append(str(row['contact__company__name']))
        else:
            data.append(' ')
        data.append(str(row['value']))
        data.append(str(row['user__first_name'])+ ' ' +str(row['user__last_name']) )
        ws1.append(data)
    return wb


class ClientHomeCalAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    def get(self, request, format=None):
        today = datetime.date.today()
        divsn = self.request.user.designation.division
        approvedAmount = list(Contract.objects.filter( division = divsn,
            status='approved').values('status').annotate(total=Sum('value')))
        if len(approvedAmount) > 0:
            approvedAmount = approvedAmount[0]['total']
        else:
            approvedAmount = 0
        billedAmount = list(Contract.objects.filter(
            status='billed', dueDate__lt=today).values('status').annotate(total=Sum('value')))
        if len(billedAmount) > 0:
            billedAmount = billedAmount[0]['total']
        else:
            billedAmount = 0
        monthdays = pythonCal.monthrange(today.year, today.month)
        FstDate = datetime.date(today.year, today.month, 1)
        lstDate = datetime.date(today.year, today.month, monthdays[1])
        receivedAmount = list(Contract.objects.filter(status='received', recievedDate__gte=FstDate,
                                                      recievedDate__lte=lstDate).values('status').annotate(total=Sum('value')))
        if len(receivedAmount) > 0:
            receivedAmount = receivedAmount[0]['total']
        else:
            receivedAmount = 0
        target = 0
        complete = 0
        period = 'yearly'
        calLi = list(Deal.objects.filter(user__designation__division = divsn, result='na').values(
            'state').annotate(sum_val=Sum('value'), count_val=Count('state')))
        sumLi = [0, 0, 0, 0, 0, 0]
        countLi = [0, 0, 0, 0, 0, 0]
        for i in calLi:
            if i['state'] == 'contacted':
                sumLi[0] = i['sum_val']
                countLi[0] = i['count_val']
            elif i['state'] == 'demo':
                sumLi[1] = i['sum_val']
                countLi[1] = i['count_val']
            elif i['state'] == 'requirements':
                sumLi[2] = i['sum_val']
                countLi[2] = i['count_val']
            elif i['state'] == 'proposal':
                sumLi[3] = i['sum_val']
                countLi[3] = i['count_val']
            elif i['state'] == 'negotiation':
                sumLi[4] = i['sum_val']
                countLi[4] = i['count_val']
            elif i['state'] == 'conclusion':
                sumLi[5] = i['sum_val']
                countLi[5] = i['count_val']
        currencyObj = appSettingsField.objects.filter(name='currency', app=69)
        if len(currencyObj) > 0:
            currencyTyp = currencyObj[0].value
        else:
            currencyTyp = ''
        contact_count = Contact.objects.filter(division = divsn, created__range = (FstDate,lstDate)).count()
        contract_count = Contract.objects.filter(division = divsn, created__range = (FstDate,lstDate)).count()
        date1 = request.GET['frm'] +' 00:00:00.243860'
        date2 = request.GET['to'] +' 23:59:59.9999'
        # quoted = Contract.objects.filter(division = divsn, created__range = (request.GET['frm'], request.GET['to']))
        quoted = Contract.objects.filter(division = divsn,  created__range = (date1, date2))
        if 'typ' in request.GET:
            quoted = quoted.filter(status = request.GET['typ'])
        if 'tandc' in request.GET:
            quoted = quoted.filter(termsAndCondition__pk = int(request.GET['tandc']))
        if 'customer' in request.GET:
            quoted = quoted.filter(contact__id = request.GET['customer'])
        if request.user.is_staff:
            quotedQuote = quoted.filter(division = divsn).order_by('-created')
        else:
            quotedQuote = quoted.filter(user = request.user).order_by('-created')
        quotedQuote = quotedQuote.values('pk', 'data', 'value','created','updated','status','deal__name','deal__pk','deal__company__name','deal__company__pk','contact__name','contact__pk','contact__company__name','contact__company__pk','contact__dp','user__pk','user__first_name','user__last_name','dueDate','termsAndCondition__heading','termsAndCondition__canSupplyOrder' , 'termsAndCondition__canInvoice','contact__mobile','contact__email','contact__street','contact__city','contact__state','contact__country','contact__pincode')
        billedQuote = Contract.objects.filter(status = 'billed',user = request.user).order_by('-value').values('pk', 'data', 'value','created','updated','status','deal__name','deal__pk','contact__name','contact__pk','deal__company__name','deal__company__pk','contact__name','contact__pk','contact__company__name','contact__dp','user__pk','user__first_name','user__last_name','dueDate','termsAndCondition__canSupplyOrder' , 'termsAndCondition__canInvoice','contact__mobile','contact__email','contact__street','contact__city','contact__state','contact__country','contact__pincode')
        dueElapsedQuote = Contract.objects.filter(status = 'dueElapsed',user = request.user).values('pk', 'data', 'value','created','updated','status','deal__name','deal__pk','contact__name','contact__pk','deal__company__name','deal__company__pk','contact__name','contact__pk','contact__company__name','contact__dp','user__pk','user__first_name','user__last_name','dueDate','termsAndCondition__canSupplyOrder' , 'termsAndCondition__canInvoice','contact__mobile','contact__email','contact__street','contact__city','contact__state','contact__country','contact__pincode')

        if 'download' in request.GET:
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=excelReport.xlsx'
            excelData = QuoteDouwnload(quotedQuote)
            excelData.save(response)
            return response

        toSend = {'sumLi': sumLi, 'countLi': countLi, 'approvedAmount': approvedAmount, 'billedAmount': billedAmount, 'receivedAmount': receivedAmount, 'currencyTyp': currencyTyp, 'target': target, 'complete': complete, 'period': period,'contact_count':contact_count,'opportunities_count':0,'contract_count':contract_count,'quotedQuote':quotedQuote,'billedQuote':billedQuote,'dueElapsedQuote':dueElapsedQuote}

        return Response(toSend, status=status.HTTP_200_OK)

class GetUserTargetView(APIView):
    renderer_classes = (JSONRenderer,)
    def get(self, request, format=None):
        userPk = request.GET['user']
        target = 0
        complete = 0
        period = ''
        kraObj = KRA.objects.filter(
            responsibility__title='CRM.SalesTarget', user=userPk)
        if len(kraObj) > 0:
            target = kraObj[0].target
            period = kraObj[0].period
            created = str(kraObj[0].created).split(' ')[0].split('-')
            print created
            print "period : " , period
            if period == 'daily':
                userTarget = list(Deal.objects.filter(result='won', user=userPk, closeDate=today).values(
                    'result').annotate(sum_val=Sum('value')))
            else:
                if period == 'yearly':
                    fDate = datetime.date(int(created[0]), 1, 1)
                    lDate = datetime.date(int(created[0]), 12, 31)
                elif period == 'monthly':
                    mr = pythonCal.monthrange(int(created[0]), int(created[1]))
                    fDate = datetime.date(int(created[0]), int(created[1]), 1)
                    lDate = datetime.date(
                        int(created[0]), int(created[1]), mr[1])
                elif period == 'quaterly':
                    if int(created[1]) <= 3:
                        fDate = datetime.date(int(created[0]), 1, 1)
                        lDate = datetime.date(int(created[0]), 3, 31)
                    elif int(created[1]) > 3 and int(created[1]) <= 6:
                        fDate = datetime.date(int(created[0]), 4, 1)
                        lDate = datetime.date(int(created[0]), 6, 30)
                    elif int(created[1]) > 6 and int(created[1]) <= 9:
                        fDate = datetime.date(int(created[0]), 7, 1)
                        lDate = datetime.date(int(created[0]), 9, 30)
                    elif int(created[1]) > 9 and int(created[1]) <= 12:
                        fDate = datetime.date(int(created[0]), 10, 1)
                        lDate = datetime.date(int(created[0]), 12, 31)
                elif period == 'weekly':
                    dt = datetime.date(int(created[0]), int(
                        created[1]), int(created[2]))
                    fDate = dt - relativedelta(days=dt.weekday())
                    lDate = fDate + relativedelta(days=6)

                userTarget = list(Deal.objects.filter(result='won', user=userPk, closeDate__range=(
                    fDate, lDate)).values('result').annotate(sum_val=Sum('value')))
            if len(userTarget) > 0:
                complete = userTarget[0]['sum_val']
        return Response({'target': target, 'complete': complete, 'period': period}, status=status.HTTP_200_OK)

class ReportHomeCalAPIView(APIView):
    renderer_classes = (JSONRenderer,)

    def get(self, request, format=None):
        today = datetime.date.today()
        minTime = datetime.time.min
        maxTime = datetime.time.max
        toSend = []
        if len(request.GET['typ']) > 0:
            fd = request.GET['fdate'].split('-')
            td = request.GET['tdate'].split('-')
            fd = datetime.date(int(fd[0]), int(fd[1]), int(fd[2]))
            td = datetime.date(int(td[0]), int(td[1]), int(td[2]))
            fd = datetime.datetime.combine(fd, minTime)
            td = datetime.datetime.combine(td, maxTime)
            if 'download' in request.GET:
                if len(request.GET['usr']) < 3:
                    usr = []
                else:
                    usr = []
                    a = str(request.GET['usr'][1:-1]).split(',')
                    for i in a:
                        usr.append(int(i))
            else:
                if request.GET['usr'] == '':
                    usr = []
                else:
                    ur = str(request.GET['usr']).split(',')
                    usr = []
                    for i in ur:
                        usr.append(int(i))

            if request.GET['typ'] == 'call':
                print '************ call'
                if len(usr) == 0:
                    toSend = list(Activity.objects.filter(typ='call', created__range=(fd, td)).values('created', 'data',
                        'contact__name', 'contact__mobile','user', userName=Concat('user__first_name', Value(' '), 'user__last_name')))
                else:
                    print list(Activity.objects.filter(typ='call', created__range=(fd, td), user__in=usr).values('data'))


                    toSend = list(Activity.objects.filter(typ='call', created__range=(fd, td), user__in=usr).values('created', 'data',
                        'contact__name', 'contact__mobile','user', userName=Concat('user__first_name', Value(' '), 'user__last_name')))

            elif request.GET['typ'] == 'contacts':
                print '************ contacts'

                if len(usr) == 0:
                    toSend = list(Contact.objects.filter(created__range=(fd, td)).values('name', 'email', 'mobile', 'designation', 'company__name', 'user', 'user__first_name', 'user__last_name' ))
                else:
                    toSend = list(Contact.objects.filter(created__range=(fd, td), user__in=usr).values('name', 'email', 'mobile', 'designation', 'company__name', 'user', 'user__first_name', 'user__last_name'))

            elif request.GET['typ'] == 'leads':
                print '************ leads'
                if len(usr) == 0:
                    toSend = list(Deal.objects.filter(created__range=(fd, td)).values('name', 'state', 'result', 'value', 'company__name', 'user', userName=Concat('user__first_name', Value(' '), 'user__last_name')))
                else:
                    toSend = list(Deal.objects.filter(created__range=(fd, td), user__in=usr).values('name', 'state', 'result', 'value',
                    'company__name','user', userName=Concat('user__first_name', Value(' '), 'user__last_name')))

            elif request.GET['typ'] == 'conversion':
                print '************ conversion'
                if len(usr) == 0:
                    toSend = list(Contract.objects.filter(status='approved', approvedDate__range=(fd, td)).values('status', 'value', 'grandTotal', 'deal__name', 'deal__company__name','contact__name','contact__company__name','user', userName=Concat('user__first_name', Value(' '), 'user__last_name')))

                else:
                    toSend = list(Contract.objects.filter(status='approved', approvedDate__range=(fd, td), user__in=usr).values('status', 'value', 'grandTotal','deal__name', 'deal__company__name','contact__name','contact__company__name', 'user', userName=Concat('user__first_name', Value(' '), 'user__last_name')))

            elif request.GET['typ'] == 'pipeline':
                print '************ pipeline'
                if len(usr) == 0:
                    toSend = list(Deal.objects.filter(result='na', created__range=(fd, td)).values(
                        'state').annotate(totalValue=Sum('value'), count=Count('state')))
                else:
                    toSend = list(Deal.objects.filter(result='na', created__range=(fd, td), user__in=usr).values(
                        'state').annotate(totalValue=Sum('value'), count=Count('state')))
                if 'download' not in request.GET:
                    sumLi = [0, 0, 0, 0, 0, 0]
                    countLi = [0, 0, 0, 0, 0, 0]
                    for i in toSend:
                        if i['state'] == 'contacted':
                            sumLi[0] = i['totalValue']
                            countLi[0] = i['count']
                        elif i['state'] == 'demo':
                            sumLi[1] = i['totalValue']
                            countLi[1] = i['count']
                        elif i['state'] == 'requirements':
                            sumLi[2] = i['totalValue']
                            countLi[2] = i['count']
                        elif i['state'] == 'proposal':
                            sumLi[3] = i['totalValue']
                            countLi[3] = i['count']
                        elif i['state'] == 'negotiation':
                            sumLi[4] = i['totalValue']
                            countLi[4] = i['count']
                        elif i['state'] == 'conclusion':
                            sumLi[5] = i['totalValue']
                            countLi[5] = i['count']
                        i['sumLi'] = sumLi
                        i['countLi'] = countLi
                    print sumLi, countLi

            elif request.GET['typ'] == 'salesInflow':
                print '************ salesInflow'
                if len(usr) == 0:
                    toSend = list(Contract.objects.filter(status='received', recievedDate__range=(fd, td)).values('pk', 'deal__name', 'contact__name','deal__company__name','deal__company__mobile','contact__company__name','contact__company__mobile','user').annotate(grandTotal=Sum('grandTotal'), dealCount=Count('deal', distinct=True)))
                else:
                    toSend = list(Contract.objects.filter(status='received', recievedDate__range=(fd, td), user__in=usr).values('pk').annotate(grandTotal=Sum('grandTotal'), dealCount=Count('deal', distinct=True)))

                if 'download' not in request.GET:
                    for i in toSend:
                        i['comapnyName'] = i['deal__company__name']
                        i['details'] = []
                        if len(usr) == 0:
                            dealsList = list(Contract.objects.filter(status='received', recievedDate__range=(
                                fd, td), deal__company__name=i['deal__company__name']).values_list('deal__name', flat=True).distinct())
                        else:
                            dealsList = list(Contract.objects.filter(status='received', recievedDate__range=(
                                fd, td), user__in=usr, deal__company__name=i['deal__company__name']).values_list('deal__name', flat=True).distinct())
                        for j in dealsList:
                            d = {'dealName': j}
                            if len(usr) == 0:
                                deals = list(Contract.objects.filter(status='received', recievedDate__range=(fd, td), deal__company__name=i['deal__company__name'], deal__name=j).values(
                                    'grandTotal', 'dueDate', 'pk', 'deal__contacts', 'deal__contacts__name', 'deal__contacts__email', 'deal__contacts__mobile', userName=Concat('user__first_name', Value(' '), 'user__last_name')))
                            else:
                                deals = list(Contract.objects.filter(status='received', recievedDate__range=(fd, td), user__in=usr, deal__company__name=i['deal__company__name'], deal__name=j).values(
                                    'grandTotal', 'dueDate', 'pk', 'deal__contacts', 'deal__contacts__name', 'deal__contacts__email', 'deal__contacts__mobile', userName=Concat('user__first_name', Value(' '), 'user__last_name')))
                            d['data'] = deals
                            i['details'].append(d)

            elif request.GET['typ'] == 'nonPerforming':
                print '************ nonPerforming'
                if len(usr) == 0:
                    toSend = list(Contract.objects.filter(status='billed', dueDate__lt=today).values('deal__company__name', 'user').annotate(grandTotal=Sum('grandTotal'), dealCount=Count('deal', distinct=True)))

                else:
                    toSend = list(Contract.objects.filter(status='billed', dueDate__lt=today, user__in=usr).values('deal__company__name', 'user').annotate(grandTotal=Sum('grandTotal'), dealCount=Count('deal', distinct=True)))
                print toSend

                if 'download' not in request.GET:
                    for i in toSend:
                        i['comapnyName'] = i['deal__company__name']
                        i['details'] = []
                        if len(usr) == 0:
                            dealsList = list(Contract.objects.filter(status='billed', dueDate__lt=today,
                                                                     deal__company__name=i['deal__company__name']).values_list('deal__name', flat=True).distinct())
                        else:
                            dealsList = list(Contract.objects.filter(status='billed', dueDate__lt=today, user__in=usr,
                                                                     deal__company__name=i['deal__company__name']).values_list('deal__name', flat=True).distinct())
                        print dealsList
                        for j in dealsList:
                            print j
                            d = {'dealName': j}
                            if len(usr) == 0:
                                deals = list(Contract.objects.filter(status='billed', dueDate__lt=today, deal__company__name=i['dealCompany'], deal__name=j).values(
                                    'grandTotal', 'dueDate', 'pk', 'deal__contacts', 'deal__contacts__name', 'deal__contacts__email', 'deal__contacts__mobile', userName=Concat('user__first_name', Value(' '), 'user__last_name')))
                            else:
                                deals = list(Contract.objects.filter(status='billed', dueDate__lt=today, user__in=usr, deal__company__name=i['dealCompany'], deal__name=j).values(
                                    'grandTotal', 'dueDate', 'pk', 'deal__contacts', 'deal__contacts__name', 'deal__contacts__email', 'deal__contacts__mobile', userName=Concat('user__first_name', Value(' '), 'user__last_name')))
                            d['data'] = deals
                            i['details'].append(d)

            print toSend
            if 'download' in request.GET and len(toSend) > 0:
                return ExcelResponse(toSend)

        return Response(toSend,status=status.HTTP_200_OK)

class SendEmailAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    def post(self, request, format=None):
        cc = []
        contactData=[]
        for c in request.data['contact']:
            u = Contact.objects.get(pk = c)
            contactData.append(u.email)

        for c in request.data['cc']:
            u = User.objects.get(pk = c)
            cc.append(str(u.email))
        print cc

        email_subject =request.data['emailSubject']

        msgBody= request.data['emailbody']

        ctx = {
            'message': msgBody,
            'linkUrl': 'cioc.co.in',
            'linkText' : 'View Online',
            'sendersAddress' : '(C) CIOC FMCG Pvt Ltd',
            'sendersPhone' : '841101',
            'linkedinUrl' : 'https://www.linkedin.com/company/13440221/',
            'fbUrl' : 'facebook.com',
            'twitterUrl' : 'twitter.com',
        }

        email_body = get_template('app.clientRelationships.email.html').render(ctx)
        msg = EmailMessage(email_subject, msgBody,  to= contactData, cc= cc )
        msg.content_subtype = 'html'
        msg.send()

        return Response({}, status = status.HTTP_200_OK)


class CustomerAccessAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny,)
    def get(self, request, format=None):
        print '****** entered', request.GET
        toSend = {}
        custObj = CustomerSession.objects.get(sessionId=str(request.GET['sessionId']))
        outboundObj = OutBoundInvoice.objects.filter(email=custObj.email)
        outBoundData = OutBoundInvoiceSerializer(outboundObj,many=True).data
        for i in outBoundData:
            qtyObj = OutBoundInvoiceQty.objects.filter(outBound=int(i['pk']))
            a = OutBoundInvoiceQtySerializer(qtyObj,many=True).data
            i['qtyData'] = a
        toSend['invData'] = outBoundData

        if request.GET['custTyp'] == 'client':
            if request.GET['typ'] == 'getAllData':
                try:
                    projectsList = list(project.objects.filter(company__id=int(custObj.company.pk)).values('pk','title'))
                    toSend['companyProjects'] = projectsList
                except:
                    toSend['companyProjects'] = []
                bugsData = []
                for i in projectsList:
                    bugsObjs = Issues.objects.filter(project__id=int(i['pk']))
                    bugDt = IssueSerializer(bugsObjs,many=True).data
                    bugsData.append({'name':i['title'],'data':bugDt})
                toSend['bugsData'] = bugsData
                if custObj.contact:
                    contactObj = Contact.objects.get(pk=custObj.contact.pk)
                    contactdata = ContactSerializer(contactObj).data
                    if not contactdata['dp']:
                        contactdata['dp'] = '/static/images/userIcon.png'
                    toSend['contactData'] = contactdata
                if custObj.company:
                    dealsObj = Deal.objects.filter(company=custObj.company)
                    dealData = DealLiteSerializer(dealsObj,many=True).data
                    for i in dealData:
                        contractObj = Contract.objects.filter(deal=int(i['pk']))
                        a = ContractSerializer(contractObj,many=True).data
                        i['contractData'] = a
                    toSend['dealsData'] = dealData
        return Response(toSend, status=status.HTTP_200_OK)
    def post(self, request, format=None):
        if request.GET['typ'] == 'bug':
            projObj = project.objects.get(pk=int(request.data['project']))
            dt = datetime.datetime.strptime(str(request.data['tentresdt']),'%Y-%m-%d').date()
            data = {'title':str(request.data['title']),'description':str(request.data['description']),'priority':str(request.data['priority']),'project':projObj,'tentresdt':dt}
            if 'fil' in request.FILES:
                data['file'] = request.FILES['fil']
            bugObj = Issues.objects.create(**data)

            email_body = ''' <div class="col-md-12" style="font-size:15px">
            <strong style="font-size:25px">{0}</strong><br>
            <strong>Priority : </strong>{1}<br>
            <strong>Tentative Date : </strong>{2}<br>
            <strong>Title : </strong>{3}<br>
            <strong>Description : </strong>{4}
            </div>'''.format(bugObj.project.title,bugObj.priority.upper(),str(bugObj.tentresdt),bugObj.title,bugObj.description)
            usersEmails = list(User.objects.filter(is_superuser=True).values_list('email',flat=True))
            msg = EmailMessage('Fix The Bug', email_body, to=usersEmails)
            msg.content_subtype = 'html'
            try:
                msg.attach_file(bugObj.file.path)
            except:
                pass
            msg.send()
            print 'email send successfully'
        return Response({}, status = status.HTTP_200_OK)


import pandas as pd
import re
import sys, traceback

class BulkContactsAPIView(APIView):
    permission_classes = (permissions.IsAuthenticated , isAdmin)
    def post(self, request, format=None):
        df = pd.read_excel(request.FILES['xl'])
        fieldsData =df.T.to_dict().values()
        count = 0
        for i in fieldsData:
            mob1 = str(i['Mobile'])

            if "," in mob1:
                mob1prts = mob1.split(',')
            elif "/" in mob1:
                mob1prts = mob1.split("/")
            else:
                mob1prts = [mob1]
            mob1Arr  = []
            for mob1 in mob1prts:
                mob1 = mob1.replace('-' , "").replace("+91" , "")
                mob1 = ' '.join(mob1.split())
                mob1 = re.sub("\D", "", mob1)
                mob1Arr.append(mob1)

            mob2 = i['Mobile2']

            street = i['Street']

            print "street : " , street

            if str(i['Name']) is None or str(i['Name']) is "nan":
                continue

            count+=1
            try:
                contactObj , created = Contact.objects.get_or_create(name = str(i['Name']),user = request.user)
            except:
                continue
            contactObj.email = i['Email']
            contactObj.emailSecondary = i['Email2']

            if len(mob1Arr)==1:
                contactObj.mobile = mob1Arr[0]
            if len(mob1Arr)>=2:
                contactObj.mobileSecondary = mob1Arr[1]

            pincode = str(i['Pincode'])
            if pincode == 'nan':
                try:
                    pincode = re.findall(r'(\d{6})', str(street))[0]
                    street = street.replace(str(pincode),'')
                except:
                    pass

            try:
                pincodeDetails = GenericPincode.objects.get(pincode = str(pincode))
                city = pincodeDetails.city
                state =  pincodeDetails.state
                country = pincodeDetails.country
            except:
                city = ''
                state =  ''
                country = ''
            try:
                companyObj = service.objects.get(name = str(i['Company Name']))
                if companyObj.address == None:
                    addrObj = address.objects.create(street = street,city = city,state = state,country = country)
                    if pincode != "nan" and pincode is not None:
                        addrObj.pincode = pincode
                    try:
                        addrObj.save()
                        companyObj.address = addrObj
                    except:
                        pass
                    companyObj.save()
                else:
                    companyObj.address.street = street
                    companyObj.address.pincode = pincode
                    companyObj.address.city = city
                    companyObj.address.state = state
                    companyObj.address.country = country
                    companyObj.save()
            except:

                traceback.print_exc(file=sys.stdout)

                try:
                    companyObj, created = service.objects.get_or_create(name = str(i['Company Name']),user = request.user)
                    addrObj = address(street = street, pincode = pincode,city = city,state = state,country = country)
                    addrObj.save()
                    companyObj.address = addrObj
                except:
                    pass
                companyObj.save()

            contactObj.company = companyObj
            if i['Gender'] == 'M':
                contactObj.male = True
            elif i['Gender'] == "F":
                contactObj.male = False
            else:
                contactObj.male = True

            contactObj.save()
        print count
        return Response({"count" : count}, status = status.HTTP_200_OK)

def createOpportunity(contObj,newOpp,user):
    if str(newOpp['Company Name'])=='nan':
        try:
            companyObj = service.objects.get(name=str(newOpp['Company Name']))
        except:
            companyObj = service.objects.create(name=str(newOpp['Company Name']),user=user, division = request.user.designation.division)
    else:
        try:
            companyObj = service.objects.get(name='Unknown')
        except:
            companyObj = service.objects.create(name='Unknown',user=user, division = request.user.designation.division)
    oppObj= Deal.objects.create(name = str(newOpp['Name']),user=user,company=companyObj)
    oppObj.value = newOpp['Value']
    oppObj.state = newOpp['State']
    oppObj.requirements = '<p>'+ newOpp['Requirements'] + '</p>'
    oppObj.save()
    return oppObj


class BulkOpportunityAPIView(APIView):
    permission_classes = (permissions.IsAuthenticated , isAdmin)
    def post(self, request, format=None):
        df = pd.read_excel(request.FILES['xl'])
        fieldsData =df.T.to_dict().values()
        count = 0
        unaddedCount = 0
        for i in fieldsData:
            print str(i['Contact Name'])
            if str(i['Contact Name']) == 'nan' and str(i['Contact Mobile']) == 'nan':
                unaddedCount +=1
            elif str(i['Contact Name']) == 'nan' and str(i['Contact Mobile']) != 'nan':
                try:
                    contObj = Contact.objects.get(mobile=i['Contact Mobile'])
                    createOpportunity(contObj,i,request.user)
                    count+=1
                except:
                    unaddedCount +=1
            else:
                contObj = Contact.objects.create(name=i['Contact Name'],mobile=i['Contact Mobile'],user=request.user)
                createOpportunity(contObj,i,request.user)
                count+=1

        return Response({"count" : count,"unaddedCount":unaddedCount}, status = status.HTTP_200_OK)

from django.core.files import File
class sendEmailAttachment(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny,)
    def get(self, request, format=None):
        if 'contract' not in request.GET:
            return Response(status=status.HTTP_400_BAD_REQUEST)
        response = HttpResponse(content_type='application/pdf')
        o = Contract.objects.get(id=request.GET['contract'])
        response.contract = o
        response.division = request.user.designation.division
        response.unit = request.user.designation.unit
        response['Content-Disposition'] = 'attachment; filename="CR_invoice%s_%s_%s.pdf"' % (
            o.pk, datetime.datetime.now(pytz.timezone('Asia/Kolkata')).year, o.pk)
        if o.termsAndCondition is not None:
            if o.termsAndCondition.version == 'V2':
                from Invoice2 import *
            elif o.termsAndCondition.version == 'V3':
                from Invoice3 import *
            else:
                from Invoice1 import *
        else:
            # try:
            crmObj = CRMTermsAndConditions.objects.filter(division = request.user.designation.division).first()
            if crmObj.version == 'V2':
                from Invoice2 import *
            elif crmObj.version == 'V3':
                from Invoice3 import *
            else:
                from Invoice1 import *
        genInvoice(response, o, request)
        filePath = os.path.join(globalSettings.BASE_DIR, 'media_root/CR_invoice%s%s_%s.pdf' %
                              (o.pk, datetime.datetime.now(pytz.timezone('Asia/Kolkata')).year, o.pk))
        f = open(filePath, 'wrb')
        f.write(response.content)
        f.close()
        f = open(filePath, 'rb')
        mailObj = mailAttachment(user = request.user,contractid = o.pk)
        mailObj.save()
        django_file = File(f)
        mailObj.attachment.save(f.name.split('/')[-1], django_file, save=True)
        mailObj.save()

        m={'user':mailObj.user.pk,'pk':mailObj.pk,'contractid':mailObj.contractid, "filename" : filePath.split('/')[-1]}

        return Response(m, status = status.HTTP_200_OK)


from reportlab.platypus import SimpleDocTemplate, Image
def genDeal(response, deal, request):

    MARGIN_SIZE = 8 * mm
    PAGE_SIZE = A4

    pdf_doc = SimpleDocTemplate(response, pagesize=PAGE_SIZE,
                                leftMargin=MARGIN_SIZE, rightMargin=MARGIN_SIZE,
                                topMargin=1 * MARGIN_SIZE, bottomMargin=3 * MARGIN_SIZE)


    pdf_doc.deal = deal
    pdf_doc.request = request

    tableHeaderStyle = styles['Normal'].clone('tableHeaderStyle')
    tableHeaderStyle.textColor = colors.white
    tableHeaderStyle.fontSize = 7

    dealReq = DealRequirement.objects.filter(parent = deal.pk)
    dealPayment = DealPaymentsTerm.objects.filter(parent = deal.pk)
    story = []
    headerFilePath = globalSettings.BRAND_LOGO_PDF
    drawing = svg2rlg(os.path.join(globalSettings.BASE_DIR,
                                   'static_shared', 'images', headerFilePath))
    sx = sy = 0.5
    drawing.width, drawing.height = drawing.minWidth() * sx, drawing.height * sy
    drawing.scale(sx, sy)
    story.append(drawing)
    title_style = styles['Heading1']
    title_style.alignment = 1
    title = Paragraph("Project Requirement", title_style)
    story.append(title)
    story.append(Spacer(2.5, 0.5 * cm))
    summryParaSrc = """
    <font size='9'>
    <strong>Client Name :</strong> %s<br/>
    <strong>Project Name :</strong> %s<br/>
    <strong>Closing Date :</strong> %s<br/>
    <strong>Delivery Time :</strong> %s days<br/>
    <strong>Specification :</strong><br/>
    %s <br/>
    </font>
    """ % (deal.company.name, deal.name , deal.closeDate.date() , deal.deliveryTime , deal.requirements )
    story.append(Paragraph(summryParaSrc, styleN))
    story.append(Spacer(2.5, 0.5 * cm))
    title1 = Paragraph("Annexure 1 : ", styleN)
    story.append(title1)
    story.append(Spacer(2.5, 0.5 * cm))
    data = [['Sl. No', 'Title', 'Details']]
    index = 0
    for i in dealReq:
        index+=1
        data.append([index, Paragraph(i.title, styleN), Paragraph(i.details, styleN)])
    t = Table(data)
    data1 = [['Sl. No', 'Offset', 'Percent' ,'MileStone']]
    indexp = 0
    for j in dealPayment:
        indexp+=1
        data1.append([indexp, Paragraph(j.offset, styleN), j.percent+' %', Paragraph(j.milestone, styleN)])
    t1 = Table(data1)
    ts = TableStyle([
     ('VALIGN', (0, 1), (-1, -3), 'TOP'),
     ('VALIGN', (0, -2), (-1, -2), 'TOP'),
     ('VALIGN', (0, -1), (-1, -1), 'TOP'),
     ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
     ('TEXTCOLOR', (1, -1), (-1, -1), colors.black),
     ('BACKGROUND', (0, 0), (-1, 0), themeColor),
     ])
    t.setStyle(ts)
    t1.setStyle(ts)
    t._argW[0] = 2.5 * cm
    t._argW[1] = 7.4 * cm
    t._argW[2] = 9.5 * cm
    t1._argW[0] = 2 * cm
    t1._argW[1] = 7.5 * cm
    t1._argW[2] = 2.5 * cm
    t1._argW[3] = 7.5 * cm
    story.append(t)
    story.append(Spacer(2.5, 0.5 * cm))
    title2 = Paragraph("Payment : ", styleN)
    story.append(title2)
    story.append(Spacer(2.5, 0.5 * cm))
    story.append(t1)
    pdf_doc.build(story)


class DownloadOpportunityAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny,)
    def get(self, request, format=None):
        if 'deal' not in request.GET:
            return Response(status=status.HTTP_400_BAD_REQUEST)
        response = HttpResponse(content_type='application/pdf')
        o = Deal.objects.get(id=request.GET['deal'])
        response.deal = o
        response['Content-Disposition'] = 'attachment; filename="deal_%s.pdf"' % (o.pk)
        genDeal(response, o, request)
        f = open(os.path.join(globalSettings.BASE_DIR, 'media_root/deal_%s.pdf' %
                              (o.pk)), 'wb')
        f.write(response.content)
        f.close()
        if 'saveOnly' in request.GET:
            return Response(status=status.HTTP_200_OK)
        return response


class ConvertAsProject(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny,)
    def post(self, request, format=None):
        data = request.data
        dealObj = Deal.objects.get(pk = data['deal'])
        projectObj = project.objects.create(title = data['title'],dueDate = data['dueDate'],description = data['description'], company = dealObj.company,user = User.objects.get(pk=data['user']) )
        projectObj.save
        reqObj = DealRequirement.objects.filter(parent = dealObj.pk)
        if len(reqObj)>0:
            for j in reqObj:
                projectObjevtives = ProjectObjective.objects.create(details = j.title, user = User.objects.get(pk=data['user']), parent =  projectObj)
                projectObjevtives.save()
        if len(data['team'])>0:
            for i in data['team']:
                projectObj.team.add(User.objects.get(pk = i))
            projectObj.save
        dealObj.result = 'won'
        dealObj.save()
        return Response(status=status.HTTP_200_OK)


import json
class CreateLegalAgreementTerms(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny,)
    def post(self, request, format=None):
        data =  request.data
        legalObj , created = LegalAgreement.objects.get_or_create(deal = Deal.objects.get(pk = data['deal']))
        if created:
            config =  json.loads(open(os.path.join(globalSettings.BASE_DIR,'clientRelationships','contractTemplate.json')).read())
            for i in config:
                termsObj = LegalAgreementTerms.objects.create(content=i['content'],typ=i['typ'],parent=legalObj)
        return JsonResponse({'legalPk' : legalObj.pk},status =200 )


class PageNumCanvasLegal(canvas.Canvas):

    #----------------------------------------------------------------------
    def __init__(self, *args, **kwargs):
        """Constructor"""
        canvas.Canvas.__init__(self, *args, **kwargs)
        print "dir : ---------",  args[0]
        self.deal =  args[0]
        self.pages = []

    #----------------------------------------------------------------------
    def showPage(self):
        """
        On a page break, add information to the list
        """
        self.pages.append(dict(self.__dict__))
        self._startPage()

    #----------------------------------------------------------------------
    def save(self):
        """
        Add the page number to each page (page x of y)
        """
        page_count = len(self.pages)

        for page in self.pages:
            self.__dict__.update(page)
            self.draw_page_number(page_count)
            self.drawLetterHeadFooter()
            canvas.Canvas.showPage(self)

        canvas.Canvas.save(self)

    #----------------------------------------------------------------------
    def draw_page_number(self, page_count):
        """
        Add the page number
        """

        text = "<font size='8'>Page #%s of %s</font>" % (
            self._pageNumber, page_count)
        p = Paragraph(text, styleN)
        p.wrapOn(self, 50 * mm, 10 * mm)
        p.drawOn(self, 180 * mm, 290 * mm)

    def drawLetterHeadFooter(self):
        settingsFields = application.objects.get(name='app.CRM').settings.all()
        self.setStrokeColor(themeColor)
        self.setFillColor(themeColor)
        self.rect(0, 0, 1500, 70, fill=True)
        compNameStyle = styleN.clone('footerCompanyName')
        compNameStyle.textColor = 'white'

        p = Paragraph(settingsFields.get(
            name='companyName').value, compNameStyle)
        p.wrapOn(self, 50 * mm, 10 * mm)
        p.drawOn(self, 85 * mm, 18 * mm)

        p1 = Paragraph(settingsFields.get(
            name='companyAddress').value, compNameStyle)
        p1.wrapOn(self, 200 * mm, 10 * mm)
        p1.drawOn(self, 15 * mm, 10 * mm)

        p2 = Paragraph(settingsFields.get(
            name='contactDetails').value, compNameStyle)
        p2.wrapOn(self, 200 * mm, 10 * mm)
        p2.drawOn(self, 40 * mm, 4 * mm)


def genAggrement(response, aggrement, request):

    MARGIN_SIZE = 8 * mm
    PAGE_SIZE = A4

    pdf_doc = SimpleDocTemplate(response, pagesize=PAGE_SIZE,
                                leftMargin=MARGIN_SIZE, rightMargin=MARGIN_SIZE,
                                topMargin=1 * MARGIN_SIZE, bottomMargin=3 * MARGIN_SIZE)


    pdf_doc.aggrement = aggrement
    pdf_doc.request = request

    tableHeaderStyle = styles['Normal'].clone('tableHeaderStyle')
    tableHeaderStyle.textColor = colors.white
    tableHeaderStyle.fontSize = 7

    req = LegalAgreementTerms.objects.filter(parent = aggrement.pk)
    customerName = aggrement.deal.company.name
    splitted = customerName.split()
    first = splitted[0]
    customerName = customerName + " ('" + first + "') "
    customerNameShort = first
    startDate = aggrement.effectiveDate.date()
    endDate = aggrement.effectiveDateEnd.date()
    story = []
    headerFilePath = globalSettings.BRAND_LOGO_PDF
    drawing = svg2rlg(os.path.join(globalSettings.BASE_DIR,
                                   'static_shared', 'images', headerFilePath))
    sx = sy = 0.5
    drawing.width, drawing.height = drawing.minWidth() * sx, drawing.height * sy
    drawing.scale(sx, sy)
    story.append(drawing)
    story.append(Spacer(2.5, 0.5 * cm))
    title_style = styles['Heading2']
    title_style.alignment = 1
    title = Paragraph(aggrement.title, title_style)
    story.append(title)
    story.append(Spacer(2.5, 0.5 * cm))
    count = 0
    custCount = 0

    for i in req:
        story.append(Spacer(2.5, 0.5 * cm))
        if custCount == 0:
            customerName = customerName
        else:
            customerName = customerNameShort
        content = i.content.replace("{{customerName}}", customerName)
        custCount+=1
        content = content.replace("{{startDate}}", str(startDate))
        content = content.replace("{{endDate}}", str(endDate))
        if i.typ == 'point':
            count += 1
            textData = '<b>' + str(count) + '</b>' + '. ' +content
            data =  Paragraph( textData, styleN)
        elif i.typ == 'para':
            data =  Paragraph(content, styleN)
        else:
            data =  Paragraph(content, styleN)
        story.append(data)
    story.append(PageBreak())
    dealData = "<b>In witness thereof the parties hereto have caused this Agreement to be executed by their respective duly authorized representatives</b>"
    dealDataT = Paragraph(dealData, styleN)
    story.append(dealDataT)
    story.append(Spacer(2.5, 0.5 * cm))
    dealDataF = Paragraph('For CIOC:', styleN)
    story.append(dealDataF)
    story.append(Spacer(2.5, 0.5 * cm))
    dealDataS = Paragraph('Signature:', styleN)
    story.append(dealDataS)
    story.append(Spacer(2.5, 0.5 * cm))
    summryParaSrc = """
    Printed Name :<br/>
    Title :<br/>
    Phone :<br/>
    Email :<br/>
    Address :<br/>
    Date :<br/>
    """
    dealDataD = Paragraph(summryParaSrc, styleN)
    story.append(dealDataD)
    story.append(Spacer(2.5, 0.5 * cm))
    dealDataC = Paragraph('For Customer:', styleN)
    story.append(dealDataC)
    story.append(Spacer(2.5, 0.5 * cm))
    story.append(dealDataS)
    story.append(Spacer(2.5, 0.5 * cm))
    dealDataD = Paragraph(summryParaSrc, styleN)
    story.append(dealDataD)
    pdf_doc.build(story,canvasmaker=PageNumCanvasLegal)


class DownloadLegalAggrementAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny,)
    def get(self, request, format=None):
        if 'aggrement' not in request.GET:
            return Response(status=status.HTTP_400_BAD_REQUEST)
        response = HttpResponse(content_type='application/pdf')
        o = LegalAgreement.objects.get(id=request.GET['aggrement'])
        response.aggrement = o
        response['Content-Disposition'] = 'attachment; filename="aggrement_%s.pdf"' % (o.pk)
        genAggrement(response, o, request)
        f = open(os.path.join(globalSettings.BASE_DIR, 'media_root/aggrement_%s.pdf' %
                              (o.pk)), 'wb')
        f.write(response.content)
        f.close()
        if 'saveOnly' in request.GET:
            return Response(status=status.HTTP_200_OK)
        return response


class AddContactAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny,)
    def post(self, request, format=None):
        if 'secretKey' in request.data:
            if request.data['secretKey'] == globalSettings.SECRET_KEY:
                contactObj = Contact.create(name = request.data['name'],mobile = request.data['mobile'],email = request.data['email'])
                contactObj.save()
                if 'company' in request.data:
                    companyObj = service.objects.get_or_create(name = request.data['company'])
                    contactObj.company =  companyObj
                    if 'pincode' in request.data:
                        try:
                            pincodeObj = GenericPincode.objects.get(pincode = request.data['pincode'])
                            AddrObj = address.objects.create(pincode = pincodeObj.pincode, city = pincodeObj.city , state = pincodeObj.state )
                            AddrObj.save()
                            if 'street'in request.data:
                                AddrObj.street = request.data['street']
                                AddrObj.save()
                            contactObj.address = AddrObj
                        except:
                            pass
                elif 'pincode' in request.data:
                    try:
                        pincodeObj = GenericPincode.objects.get(pincode = request.data['pincode'])
                        contactObj.pincode =  pincodeObj.pincode
                        contactObj.state =  pincodeObj.state
                        contactObj.city =  pincodeObj.city
                        contactObj.country =  pincodeObj.country
                        if 'street'in request.data:
                            contactObj.street =request.data['street']
                    except:
                        pass
                if 'notes' in request.data:
                    contactObj.notes =  request.data['notes']
                if 'emailSecondary' in request.data:
                    contactObj.emailSecondary =  request.data['emailSecondary']
                if 'mobileSecondary' in request.data:
                    contactObj.mobileSecondary =  request.data['mobileSecondary']
                if 'designation' in request.data:
                    contactObj.designation =  request.data['designation']
                if 'linkedin' in request.data:
                    contactObj.linkedin =  request.data['linkedin']
                if 'facebook' in request.data:
                    contactObj.facebook =  request.data['facebook']
                if 'male' in request.data:
                    contactObj.male =  request.data['male']
                contactObj.save()
                toReturn = {'name' : contactObj.name , 'mobile' : contactObj.mobile ,  'email' : contactObj.email}
                return Response(toReturn,status=status.HTTP_200_OK)
            else:
                return Response(status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response(status=status.HTTP_400_BAD_REQUEST)

class AddOpportunityAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny,)
    def post(self, request, format=None):
        if 'secretKey' in request.data:
            if request.data['secretKey'] == globalSettings.SECRET_KEY:
                dealObj = Deal.objects.create(name=request.data['name'],state = 'created')
                dealObj.save()
                if 'company' in request.data:
                    companyObj = service.objects.get_or_create(name = request.data['company'])
                    dealObj.company =  companyObj
                if 'requirements'in request.data:
                    dealObj.requirements =  request.data['requirements']
                if 'value'in request.data:
                    dealObj.value =  request.data['value']
                if 'currency'in request.data:
                    dealObj.currency =  request.data['currency']
                if 'currency'in request.data:
                    dealObj.currency =  request.data['currency']
                if 'closeDate'in request.data:
                    dealObj.closeDate =  request.data['closeDate']
                if 'contacts' in request.data:
                    for i in request.data['contacts']:
                        dealObj.contacts.add(Contact.objects.get(pk = i))
                        dealObj.save()
                if 'internalUsers' in request.data:
                    for j in request.data['internalUsers']:
                        dealObj.internalUsers.add(User.objects.get(pk = j))
                        dealObj.save()
                dealObj.save()
                toReturn = {'name' : contactObj.name , 'mobile' : contactObj.mobile ,  'email' : contactObj.email}
                return Response(toReturn,status=status.HTTP_200_OK)
            else:
                return Response(status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response(status=status.HTTP_400_BAD_REQUEST)

class ChartDataView(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny,)
    def get(self, request, format=None):

        labelsArr = []
        today = datetime.datetime.now()
        dataContactArr = []
        dataContractsArr = []
        divsn = self.request.user.designation.division
        for j in range(7):
            i = 6-j
            dt =  today- datetime.timedelta(days=i)
            labelsArr.append(dt.strftime("%b %d"))
            dataContactArr.append(Contact.objects.filter(created__range =[ dt- datetime.timedelta(days=1) , dt] , user__designation__division = divsn).count())
            dataContractsArr.append( Contract.objects.filter(created__range =[ dt- datetime.timedelta(days=1) , dt] , user__designation__division = divsn ).count())

        toReturn = {
            "data" : [dataContactArr , dataContractsArr],
            "labels" : labelsArr,
            "series" : ["Quotations" , "Contacts"]
        }


        return Response( toReturn , status=status.HTTP_200_OK)

class InvoiceSettingsView(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny,)
    def get(self, request, format=None):
        logo = None
        if request.user.designation.division.logo:
            logo = request.user.designation.division.logo.url


        toReturn = {
            "headerImg" : '/media/' +str(request.user.designation.division.logo),
            "name" : request.user.designation.division.name,
            "website" : request.user.designation.division.website,
            "logo" : logo,
            "address" : request.user.designation.unit.address,
            "city" : request.user.designation.unit.city,
            "state" : request.user.designation.unit.state,
            "pincode" : request.user.designation.unit.pincode,
            "country" : request.user.designation.unit.country,
            "color" : globalSettings.INVOICE_THEME_COLOR_V2,
            "type" : globalSettings.INVOICE_TYPE,
            "email" : request.user.designation.unit.email,
            "mobile" : request.user.designation.unit.mobile,
            "gst":request.user.designation.unit.gstin,
            "pan" : request.user.designation.division.pan
        }

        return Response( toReturn , status=status.HTTP_200_OK)

# class CreateContactView(APIView):
#     renderer_classes = (JSONRenderer,)
#     permission_classes = (permissions.AllowAny,)
#     def post(self, request, format=None):
#         data = request.data
#         div = request.user.designation.division
#
#         try:
#             contactObj, created  = Contact.objects.get_or_create(mobile = data['mobile'], division = div)
#             if created:
#                 contactObj.user = request.user
#
#         except:
#                 contactObj  = Contact.objects.filter(mobile = data['mobile'], division = div).first()
#         contactObj.name = data['name']
#         # if 'pk' in data:
#         #     contactObj = Contact.objects.get(pk = int(data['pk']))
#         #     if 'name' in data:
#         #         contactObj.name = data['name']
#         #     if 'mobile' in data:
#         #         contactObj.mobile = data['mobile']
#         # else:
#         #     contactObj = Contact.objects.create(name = data['name'] , mobile = data['mobile'] , user = request.user)
#         if 'isGst' in data:
#             contactObj.isGst = data['isGst']
#         if 'designation' in data:
#             contactObj.designation = data['designation']
#         if 'email' in data:
#             contactObj.email = data['email']
#         if 'companypk' in data and 'company' in data:
#             companyObj = service.objects.get(pk = int(data['companypk']))
#             contactObj.company = companyObj
#             if companyObj.address == None:
#                 if 'street' in data:
#                     addressObj = address.objects.create(street = data['street'])
#                     companyObj.address = addressObj
#                     companyObj.save()
#             else:
#                 addressObj = companyObj.address
#                 if 'street' in data:
#                     addressObj.street = data['street']
#                     addressObj.save()
#         elif 'company' in data:
#             companyObj = service.objects.create(name = data['company'], user = request.user , division = request.user.designation.division)
#             contactObj.company = companyObj
#             if 'street' in data:
#                 addressObj = address.objects.create(street = data['street'])
#                 companyObj.address = addressObj
#                 companyObj.save()
#         if 'company' in data:
#             if 'gstin' in data:
#                 companyObj.tin = data['gstin']
#             if 'city' in data:
#                 addressObj.city = data['city']
#             if 'state' in data:
#                 addressObj.state = data['state']
#             if 'country' in data:
#                 addressObj.country = data['country']
#             if 'pincode' in data:
#                 addressObj.pincode = data['pincode']
#             addressObj.save()
#             companyObj.save()
#
#         if 'company' not in data:
#             contactObj.company = None
#
#         if 'street' in data:
#             contactObj.street = data['street']
#         if 'city' in data:
#             contactObj.city = data['city']
#         if 'state' in data:
#             contactObj.state = data['state']
#         if 'country' in data:
#             contactObj.country = data['country']
#         if 'pincode' in data:
#             contactObj.pincode = data['pincode']
#
#         contactObj.save()
#         data = ContactSerializer(contactObj , many = False).data
#         return Response( data , status=status.HTTP_200_OK)


class CreateContactView(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny,)
    def post(self, request, format=None):
        data = request.data
        print data,'aaaaaaaaaaaaaaa'
        # print data['gstin'],'aaaaaaaaaaaaaaaa'
        div = request.user.designation.division
        if 'companypk' in data and 'company' in data:
            companyObj = service.objects.get(pk = int(data['companypk']))
            if companyObj.address == None:
                if 'street' in data:
                    addressObj = address.objects.create(street = data['street'])
                    companyObj.address = addressObj
                    companyObj.save()
            else:
                addressObj = companyObj.address
                if 'street' in data:
                    addressObj.street = data['street']
                    addressObj.save()
        elif 'company' in data:
            companyObj = service.objects.create(name = data['company'], user = request.user , division = request.user.designation.division)
            # contactObj.company = companyObj
            if 'street' in data:
                addressObj = address.objects.create(street = data['street'])
                companyObj.address = addressObj
                companyObj.save()
        try:
            contactObj, created  = Contact.objects.get_or_create(mobile = data['mobile'], division = div)
            if created:
                contactObj.user = request.user
        except:
                contactObj  = Contact.objects.filter(mobile = data['mobile'], division = div).first()
        contactObj.name = data['name']
        try:
            if companyObj:
                contactObj.company = companyObj
        except:
            pass
        if 'isGst' in data:
            contactObj.isGst = data['isGst']
        if 'designation' in data:
            contactObj.designation = data['designation']
        if 'email' in data:
            contactObj.email = data['email']
        if 'company' in data:
            if 'gstin' in data:
                companyObj.tin = data['gstin']
            if 'city' in data:
                addressObj.city = data['city']
            if 'state' in data:
                addressObj.state = data['state']
            if 'country' in data:
                addressObj.country = data['country']
            if 'pincode' in data:
                addressObj.pincode = data['pincode']
            addressObj.save()
            companyObj.save()
            # companyObj.save()

        # if 'company' not in data:
        #     contactObj.company = None

        if 'street' in data:
            contactObj.street = data['street']
        if 'city' in data:
            contactObj.city = data['city']
        if 'state' in data:
            contactObj.state = data['state']
        if 'country' in data:
            contactObj.country = data['country']
        if 'pincode' in data:
            contactObj.pincode = data['pincode']

        contactObj.save()
        data = ContactSerializer(contactObj , many = False).data
        return Response( data , status=status.HTTP_200_OK)


class GetBoardOptionsView(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny,)
    def get(self, request, format=None):
        if 'deal' in request.GET:
            data = globalSettings.DEAL_BOARD
        if 'allDeal' in request.GET:
            data = globalSettings.DEAL_BOARD_OPTIONS
        if 'jobs' in request.GET:
            data = globalSettings.RECRUITMENT_OPTIONS
        return Response( data , status=status.HTTP_200_OK)

class ContractTrackerViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticatedOrReadOnly, )
    serializer_class = ContractTrackerSerializer
    queryset = ContractTracker.objects.all()
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['contract']

class SaveDealView(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny,)
    def post(self, request, format=None):
        data = request.data
        if 'companyPk' in data:
            companyObj = service.objects.get(pk = int(data['companyPk']))
        else:
            companyObj = service.objects.create(name = data['company'] , user = request.user, division = request.user.designation.division)
        dealObj = Deal.objects.create(name = data['name'] ,  user = request.user, company = companyObj)
        if 'contactPk' in data:
            contactObj = Contact.objects.get(pk = int(data['contactPk']))
        else:
            contactObj = Contact.objects.create(name = data['contact'],email = data['email'] , mobile = data['mobile'] , user = request.user)
        dealObj.contacts.add(contactObj)
        dealObj.save()
        return Response( status=status.HTTP_200_OK)


class DownloadAggrement(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny, )
    renderer_classes = (JSONRenderer,)
    def get(self, request, format=None):
        contactObj = Contact.objects.get(pk = request.GET['contact'])
        productObj =  []
        filename ='amcAggrement.pdf'
        response = HttpResponse(content_type='amcAggrement/pdf')
        response.division = request.user.designation.division
        response.unit = request.user.designation.unit
        response['Content-Disposition'] = 'attachment; filename="amcAggrement.pdf"'
        genAMCAggrement(response, contactObj, productObj, request)

        return response

class FixDivisionView(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny,)
    def get(self, request, format=None):
        success = 0
        failure = 0
        for c in Contract.objects.all():
            try:
                c.division = c.user.designation.division
                c.save()
                success += 1
            except:
                failure += 1

        return Response({"failure" : failure , "success" : success}, status=status.HTTP_200_OK)


class AddProductView(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny,)
    def post(self, request, format=None):
        data = request.data
        # {u'startDate': u'2021-01-05', u'serialNo': u'3423432', u'period': u'Quaterly', u'totalServices': 1, u'contact': u'1', u'address': u'', u'seperateAddress': False, u'productName': u'asdsa', u'pincode': u''}
        # print data,'aaaaaaaaaaa'
        # if 'id' in data:
        #     pass
        # else:
        if 'serialNo' in data:
            serialNo = data['serialNo']
        else:
            serialNo = ''
        if 'notes' in data:
            notes = data['notes']
        else:
            notes = ''
        contactObj = Contact.objects.get(pk = int(data['contact']))
        toSave = {'contact' : contactObj, 'period': data['period'] , 'totalServices' : data['totalServices'] , 'startDate' : data['startDate'] , 'notes' : notes , 'serialNo' : serialNo, 'productName' : data['productName']}


        if data['seperateAddress'] == True:
            address = data['address']
            city = data['city']
            state = data['state']
            pincode = data['pincode']
            country = data['country']
        else:
            if contactObj.street is not None:
                address = contactObj.street
            else:
                address = ''
            if contactObj.city is not None:
                city = contactObj.city
            else:
                city = ''
            if contactObj.state is not None:
                state = contactObj.state
            else:
                state = ''
            if contactObj.pincode is not None:
                pincode = contactObj.pincode
            else:
                pincode = ''
            if contactObj.country is not None:
                country = contactObj.country
            else:
                country = ''
        toSave['installationAddress'] = address
        toSave['city'] = city
        toSave['state'] = state
        toSave['pincode'] = pincode
        toSave['country'] = country
        amc = RegisteredProducts.objects.create(**toSave)

        if  data['period'] == 'Yearly':
            months = 12
        elif data['period'] == 'Quaterly':
            months = 3
        elif data['period'] == 'Half Yearly':
            months = 6
        nextDate =  datetime.datetime.strptime(data['startDate'], '%Y-%m-%d')
        for i in range(0,int(data['totalServices'])):
            division = request.user.designation.division
            ticketData = {'referenceContact' : contactObj , 'name' : contactObj.name , 'phone' : contactObj.mobile , 'email'  : contactObj.email , 'productName' : data['productName']  ,'notes' : notes , 'productSerial' : serialNo , 'address' : address , 'pincode' : pincode , 'city' : city, 'state' : state , 'country' : country , 'referenceAMC' : amc , 'division' : division, 'status' : 'upcoming','preferredDate': nextDate , 'requireOnSiteVisit' : True}
            nextDate = nextDate+ relativedelta(months=+months)
            ticket = ServiceTicket.objects.create(**ticketData)
        toRet = RegisteredProductsSerializer(amc, many = False).data
        return Response(toRet, status=status.HTTP_200_OK)


class RegisteredProductsViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticatedOrReadOnly, )
    serializer_class = RegisteredProductsSerializer
    queryset = RegisteredProducts.objects.all()
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['contact']


class ServiceTicketViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticatedOrReadOnly, )
    serializer_class = ServiceTicketSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['status','engineer']
    # search_fields = ('name', 'email', 'phone')
    def get_queryset(self):
        division = self.request.user.designation.division
        toRet = ServiceTicket.objects.filter(division = division).order_by('-created')
        if 'search' in self.request.GET:
            toRet = toRet.filter(Q(name__icontains = self.request.GET['search']) | Q(phone__icontains = self.request.GET['search']) )
        return toRet

    # filter_backends = [DjangoFilterBackend]
    # filter_fields = ['contact']

class ConfigureTermsAndConditionsViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,  )
    serializer_class = ConfigureTermsAndConditionsSerializer
    def get_queryset(self):
        divsn = self.request.user.designation.division
        return ConfigureTermsAndConditions.objects.filter(division = divsn)


class DownloadAllVisitsAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny ,)
    def get(self,request , format= None):
        workbook = Workbook()
        divsn = self.request.user.designation.division
        obj = ServiceTicket.objects.filter(division = divsn).order_by('-created')
        Sheet1 = workbook.active
        hdFont = Font(size=12,bold=True)
        alphaChars = list(string.ascii_uppercase)
        Sheet1.title = 'Assigned'
        hd1 = [ 'ID' 'Name' , 'Phone' , 'Email' , 'Product' , 'Serial No.' , ' Preferred Date ' ,' Prefered Time Slot ' , 'Technician' ]
        hdWidth = [10,10,10,30,30,30,30,15]
        Sheet1.append(hd1)
        data = []
        for i in obj.filter(status = 'assigned'):
            if i.preferredDate is not None:
                date = i.preferredDate
            else:
                date = ''
            if i.preferredTimeSlot is not None:
                timeslot = i.preferredTimeSlot
            else:
                timeslot = ''
            if i.engineer is not None:
                engineer =  i.engineer.first_name+' ' +i.engineer.last_name
            else:
                engineer = ''
            data = [i.name, i.email, i.phone, i.productName , i.productSerial , date, timeslot , engineer]
            Sheet1.append(data)
        Sheet2 = workbook.create_sheet('Ongoing')
        Sheet2.append(hd1)
        data = []
        for i in obj.filter(status = 'ongoing'):
            if i.preferredDate is not None:
                date = i.preferredDate
            else:
                date = ''
            if i.preferredTimeSlot is not None:
                timeslot = i.preferredTimeSlot
            else:
                timeslot = ''
            if i.engineer is not None:
                engineer = i.engineer.first_name+' ' +i.engineer.last_name
            else:
                engineer = ''
            data = [i.name, i.email, i.phone, i.productName , i.productSerial , date, timeslot , engineer]
            Sheet2.append(data)
        Sheet3 = workbook.create_sheet('Completed')
        Sheet3.append(hd1)
        data = []
        for i in obj.filter(status = 'completed'):
            if i.preferredDate is not None:
                date = i.preferredDate
            else:
                date = ''
            if i.preferredTimeSlot is not None:
                timeslot = i.preferredTimeSlot
            else:
                timeslot = ''
            if i.engineer is not None:
                engineer =  i.engineer.first_name+' ' +i.engineer.last_name
            else:
                engineer = ''
            data = [i.name, i.email, i.phone, i.productName , i.productSerial , date, timeslot , engineer]
            Sheet3.append(data)
        Sheet4 = workbook.create_sheet('Postponed')
        Sheet4.append(hd1)
        data = []
        for i in obj.filter(status = 'postponed'):
            if i.preferredDate is not None:
                date = i.preferredDate
            else:
                date = ''
            if i.preferredTimeSlot is not None:
                timeslot = i.preferredTimeSlot
            else:
                timeslot = ''
            if i.engineer is not None:
                engineer = i.engineer.first_name+' ' +i.engineer.last_name
            else:
                engineer = ''
            data = [i.name, i.email, i.phone, i.productName , i.productSerial , date, timeslot , engineer]
            Sheet4.append(data)
        Sheet5 = workbook.create_sheet('Cancelled')
        Sheet5.append(hd1)
        data = []
        for i in obj.filter(status = 'cancelled'):
            if i.preferredDate is not None:
                date = i.preferredDate
            else:
                date = ''
            if i.preferredTimeSlot is not None:
                timeslot = i.preferredTimeSlot
            else:
                timeslot = ''
            if i.engineer is not None:
                engineer = i.engineer.first_name+' ' +i.engineer.last_name
            else:
                engineer = ''
            data = [i.name, i.email, i.phone, i.productName , i.productSerial , date, timeslot , engineer]
            Sheet5.append(data)

        Sheet6 = workbook.create_sheet('Upcoming')
        Sheet6.append(hd1)
        data = []
        for i in obj.filter(status = 'upcoming'):
            if i.preferredDate is not None:
                date = i.preferredDate
            else:
                date = ''
            if i.preferredTimeSlot is not None:
                timeslot = i.preferredTimeSlot
            else:
                timeslot = ''
            if i.engineer is not None:
                engineer = i.engineer.first_name+' ' +i.engineer.last_name
            else:
                engineer = ''
            data = [i.name, i.email, i.phone, i.productName , i.productSerial , date, timeslot , engineer]
            Sheet6.append(data)

        response = HttpResponse(content=save_virtual_workbook(workbook),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Collection.xlsx'
        return response






class MaterialIssuedNoteAPIView(APIView):
    def get(self , request , format = None):
        data = request.GET
        from materialInward import *
        ticket = ServiceTicket.objects.get(pk = int(data['id']))
        response = HttpResponse(content_type='application/pdf')
        response.division = request.user.designation.division
        response.unit = request.user.designation.unit
        response['Content-Disposition'] = 'attachment;filename="Quotationdownload.pdf"'
        genMaterialIssueNote(response , ticket,request)
        return response


class OrderAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny,)
    def post(self, request, format=None):
        data = request.data
        toRet = {}
        amount = 0
        contactObj =  Contact.objects.get(pk = int(data['id']))
        divisionId = hash_fn.unhash(data['division'])
        division = Division.objects.get(pk = int(divisionId))
        cartObj = Cart.objects.filter(contact = contactObj, division = division )
        costcenterObj  = division.divisionCostCenter.all().first()
        accountObj = division.divisionAccount.filter(personal = False, heading = 'income').first()
        data = []
        saleData = {'division' : division , 'contact' : contactObj, 'personName' : contactObj.name , 'phone' : contactObj.mobile , 'address' :  contactObj.street , 'pincode' : contactObj.pincode, 'state' : contactObj.state, 'city' : contactObj.city , 'country' : contactObj.country, 'costcenter' : costcenterObj, 'account' : accountObj , 'sameasbilling' : True , 'billingAddress' : contactObj.street , 'billingPincode' : contactObj.pincode , 'billingState' : contactObj.state , 'billingCity' : contactObj.city , 'billingCountry' : contactObj.country}

        saleObj = Sale(**saleData)
        saleObj.save()
        try:
            terms = TermsAndConditions.objects.all().first()
            saleObj.termsandcondition = terms
            saleObj.terms = terms.body
            saleObj.save()
        except:
            pass
        try:
            userObj = User.objects.filter(designation__division = division, is_staff = True)
            saleObj.user = userObj.first()
            saleObj.save()
        except:
            pass
        for i in cartObj:
            if i.addon is not None:
                i.addon = json.loads(i.addon)
                price = float(i.price) + float(i.addon['price'])
                total = price * i.qty
            else:
                price = float(i.price)
                total = i.total
            saleQtyObj = {'outBound' : saleObj , 'product' : i.product.name , 'qty' : i.qty , 'price' : price , 'taxPer' : i.product.taxRate , 'total' : total , 'division' : division,'hsn':i.product.taxCode}
            saleqtyObj = SalesQty(**saleQtyObj)
            saleqtyObj.save()
            amount+=total
            i.delete()
        saleObj.total = amount
        saleObj.balanceAmount = amount
        saleObj.save()
            # val = {"currency":"INR","desc":i.product.name,"quantity":i.qty,"rate":i.price,"saleType":"Product","total":i.total,"type":"onetime","extraFieldOne":"","extraFieldTwo":"","subtotal":i.total,"totalTax":0,"tax":0,"taxCode":0}
        toRet = {'id' : saleObj.pk }
        return Response(toRet, status=status.HTTP_200_OK)
