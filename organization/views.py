# -*- coding: utf-8 -*-
from __future__ import unicode_literals

from django.shortcuts import render
from rest_framework import viewsets , permissions , serializers
from url_filter.integrations.drf import DjangoFilterBackend
from .serializers import *
from API.permissions import *
from .models import *
from django.db.models import Q
from rest_framework.views import APIView
from openpyxl import load_workbook
from io import BytesIO,StringIO
from openpyxl import Workbook
from django.http import HttpResponse
from openpyxl.writer.excel import save_virtual_workbook
from rest_framework.renderers import JSONRenderer
import os
import json

from ERP.models import service, address, Division
from website.models import *

from clientRelationships.models import CRMTermsAndConditions, Contact, Contract

from finance.models import Inventory
from HR.serializers import userSearchSerializer
from ERP.serializers import UserAppSerializer
from rest_framework.request import Request
from rest_framework.test import APIRequestFactory

# Create your views here.


class DivisionViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated, )
    serializer_class = DivisionSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['name']
    def get_queryset(self):
        queryset = Division.objects.all().order_by('-pk')
        if self.request.user.is_superuser:
            if 'division' in self.request.GET:
                queryset =  queryset.filter(name__icontains = str(self.request.GET['division']))
                return queryset
            return queryset
        else:
            queryset = Division.objects.filter(pk = self.request.user.designation.division.pk)
            # queryset = Division.objects.all()
            return queryset

class DivisionLiteViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated, )
    serializer_class = DivisionLiteSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['name']
    def get_queryset(self):
        queryset = Division.objects.all().order_by('-pk')
        return queryset



class UnitLiteViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated, readOnly)

    serializer_class = UnitLiteSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['name' , 'division','warehouse']
    def get_queryset(self):
        if 'warehouse' in self.request.GET:
            return Unit.objects.filter(warehouse = True)
        return self.request.user.designation.division.units.all()

class UnitFullViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated, readOnly)
    # queryset = Unit.objects.all()
    serializer_class = UnitFullSerializer
    def get_queryset(self):
        return self.request.user.designation.division.units.all()

class UnitSuperliteViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated, readOnly)
    serializer_class = UnitSuperLiteSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['name' , 'division']
    queryset = Unit.objects.all()


class UnitViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated, )
    serializer_class = UnitSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['name','division']
    def get_queryset(self):
        return self.request.user.designation.division.units.all()

class FirstLevelUnitViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated, )
    serializer_class = UnitLiteSerializer
    queryset = Unit.objects.filter(parent=None)
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['division']


class RoleViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated, )
    serializer_class = RoleSuperLiteSerializer
    queryset = Role.objects.all()
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['name', 'id']
    def get_queryset(self):
        return self.request.user.designation.division.rolesdivisions.all()

class UpdatePermissionAPI(APIView):
    def post(self, request, format = None):
        app = application.objects.get(pk = request.data['app'])
        role = Role.objects.get(pk = request.data['role'])

        if request.data['value']:
            role.permissions.add(app)
        else:
            role.permissions.remove(app)

        return Response({}, status = status.HTTP_200_OK)

class UpdateReportsPermissionAPI(APIView):
    def post(self, request, format = None):
        report = HomeChart.objects.get(pk = request.data['report'])
        role = Role.objects.get(pk = request.data['role'])

        if request.data['value']:
            role.reports.add(report)
        else:
            role.reports.remove(report)

        return Response({}, status = status.HTTP_200_OK)


class UpdateChartsPermissionAPI(APIView):
    def post(self, request, format = None):
        chart = HomeChart.objects.get(pk = request.data['chart'])
        role = Role.objects.get(pk = request.data['role'])

        if request.data['value']:
            role.charts.add(chart)
        else:
            role.charts.remove(chart)

        return Response({}, status = status.HTTP_200_OK)


class UnitsBulkUploadAPIView(APIView):
    permission_classes = (permissions.IsAuthenticated , isAdmin)
    def post(self, request, format=None):
        excelFile = request.FILES['exFile']
        wb = load_workbook(filename = BytesIO(excelFile.read()) ,  read_only=True)
        count = 0
        for ws in wb.worksheets:
            wsTitle = ws.title
            for i in range(2,ws.max_row+1):
                try:
                    name = ws['A' + str(i)].value
                except:
                    name = None

                try:
                    address = ws['B' + str(i)].value
                except:
                    address = None

                try:
                    pincode = ws['C' + str(i)].value
                except:
                    pincode = None

                try:
                    l1 = ws['D' + str(i)].value
                except:
                    l1 = None

                try:
                    l2 = ws['E' + str(i)].value
                except:
                    l2 = None
                try:
                    mobile = ws['F' + str(i)].value
                except:
                    mobile = None
                try:
                    telephone = ws['G' + str(i)].value
                except:
                    telephone = None
                try:
                    tax = ws['H' + str(i)].value
                except:
                    tax = None
                try:
                    areaCode = ws['I' + str(i)].value
                except:
                    areaCode = None
                count+=1
                unitObj , n = Unit.objects.get_or_create( areaCode =  areaCode)
                print unitObj
                unitObj.address = address
                unitObj.l1 = l1
                unitObj.l2 = l2
                unitObj.mobile = mobile
                unitObj.telephone = telephone
                unitObj.tax = tax
                unitObj.areaCode = areaCode
                unitObj.name = name
                unitObj.pincode = pincode
                unitObj.save()

        return Response({"count" : count}, status = status.HTTP_200_OK)


class HomeChartViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    queryset = HomeChart.objects.all()
    serializer_class = HomeChartSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['name', 'type', 'division']

class DashboardDataAPI(APIView):
    def get(self, request, format = None):
        toReturn = []
        for chart in HomeChart.objects.filter(type = "Chart" , enabled = True):
            error = None
            try:
                exec(chart.query)
            except:
                error = "Error while running the query"
            toReturn.append({"name" : chart.name , "chartType" : chart.chartType , "configuration" : {} , "query" : chart.query , "enabled" : chart.enabled , "data" : data , "id" : chart.id , "error" : error})

        return Response(toReturn, status = status.HTTP_200_OK)

class DownloadChart(APIView):
    def get(self, request, format = None):
        chart = HomeChart.objects.get(id=int(request.GET['id']))
        exec(chart.query)
        workbook = Workbook()
        Sheet1 = workbook.active
        dArr = []
        dArr.append(data['labels'])
        dArr.append(data['values'])

        if chart.chartType == 'Pie' or chart.chartType == 'Doughnut':
            for i in dArr:
                Sheet1.append(i)
        elif chart.chartType == 'Bar' or chart.chartType == 'Line':
            Sheet1.append(dArr[0])
            for i in dArr[1]:
                Sheet1.append(i)

        response = HttpResponse(content=save_virtual_workbook(workbook),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=chart.xlsx'
        return response

class GetMyDivision(APIView):
    def get(self, request, format = None):
        return Response(DivisionSerializer(request.user.designation.division).data , status = status.HTTP_200_OK)


class InstalledAppViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    queryset = InstalledApp.objects.all()
    serializer_class = InstalledAppSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = [ 'parent','app','addedBy']

class CompanyHolidayViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.AllowAny,)
    serializer_class = CompanyHolidaySerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = [ 'name','date','typ']
    def get_queryset(self):
        divsn = self.request.user.designation.division
        toReturn = CompanyHolidays.objects.filter(division = divsn ) | CompanyHolidays.objects.filter(division = None)
        if 'search' in self.request.GET:
            val = self.request.GET['search']
            toReturn = toReturn.filter(Q(typ__icontains = val)| Q(name__icontains = val))
            return toReturn
        return toReturn.distinct()


class Getheaderandfooter(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny, )
    def post(self, request, format=None):
        userDivision  = request.user.designation.division.pk
        data = request.data
        print hash_fn.hash(userDivision),'439802849083993'
        d = Division.objects.get(pk = userDivision )
        d.apiKey = hash_fn.hash(userDivision)
        if 'defaultTitle' in data:
            d.defaultTitle = data['defaultTitle']
        if 'defaultDescription' in data:
            d.defaultDescription = data['defaultDescription']
        if 'enableChatbot' in data:
            if data['enableChatbot'] =='true':
                d.enableChatbot = True
            else:
                d.enableChatbot = False

        if 'defaultOgImage' in request.FILES:
            d.defaultOgImage = request.FILES['defaultOgImage']
        d.save()
        ogImage = os.path.join(globalSettings.MEDIA_ROOT, str(d.defaultOgImage))
        im = Image.open(ogImage)
        width, height = im.size
        d.defaultOgWidth = width
        d.defaultOgHeight = height
        if 'footerTemplate' in data :
            try:
                uielementObj =  UIelementTemplate.objects.get(pk=data['footerTemplate'])
                d.footerTemplate =  uielementObj.template
                d.footerData =  uielementObj.defaultData
                d.footerCss =  uielementObj.css
                d.footerTemplate = d.footerTemplate.replace('$data','data.footerData')
            except:
                pass
        if 'headerTemplate' in data :
            try:
                uielementObj =  UIelementTemplate.objects.get(pk=data['headerTemplate'])
                d.headerTemplate =  uielementObj.template
                d.headerData =  uielementObj.defaultData
                d.headerCss =  uielementObj.css
                d.headerTemplate = d.headerTemplate.replace('$data','data.headerData')
            except:
                pass

        # if d.defaultOgImage != None:
        #     ogImage = os.path.join(globalSettings.MEDIA_ROOT , str(d.defaultOgImage))
        #     im = Image.open(ogImage)
        #     width, height = im.size
        #     d.defaultOgWidth = width
        #     d.defaultOgHeight = height
        d.save()
        serializer_context = {
            'request': Request(request),
        }
        print serializer_context['request'],"serializer_context"
        divsionObj = DivisionSerializer(d,context=serializer_context,many=False).data
        return Response({'data':divsionObj}, status=status.HTTP_200_OK)




class DataMigrationsAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny, )
    def get(self, request, format=None):
        divisionPK = 4

        divisionObj = Division.objects.get(pk = divisionPK)
        unitObj = divisionObj.units.last()

        # #creating users
        userData = None
        userDataPath = os.path.join(globalSettings.BASE_DIR, 'static_shared','sdpl_user.json')
        with open(userDataPath) as json_file:
            userData = json.load(json_file)
        #     for item in userData:
        #         userObj, created = User.objects.get_or_create(username = item['username'])
        #         userObj.first_name = item['first_name']
        #         userObj.last_name = item['last_name']
        #         userObj.email = item['email']
        #         userObj.is_active = item['is_active']
        #         userObj.is_staff = item['is_staff']
        #         userObj.set_password('essgi')
        #         userObj.save()
        #
        #         profileObj = userObj.profile
        #         profileObj.married = item['profile']['married']
        #         profileObj.empType = item['profile']['empType']
        #         profileObj.isManager = item['profile']['isManager']
        #         profileObj.mobile = item['profile']['mobile']
        #         profileObj.permanentAddressState = item['profile']['permanentAddressState']
        #         profileObj.localAddressStreet = item['profile']['localAddressStreet']
        #         profileObj.permanentAddressCountry = item['profile']['permanentAddressCountry']
        #         profileObj.permanentAddressPin = item['profile']['permanentAddressPin']
        #         profileObj.permanentAddressCity = item['profile']['permanentAddressCity']
        #         profileObj.localAddressCountry = item['profile']['localAddressCountry']
        #         profileObj.save()
        #
        #         payrollObj = userObj.payroll
        #         payrollObj.pfAccNo = item['payroll']['pfAccNo']
        #         payrollObj.off = item['payroll']['off']
        #         payrollObj.alHold = item['payroll']['alHold']
        #         payrollObj.save()
        #
        #         designationObj = userObj.designation
        #         designationObj.division = divisionObj
        #         designationObj.unit = unitObj
        #         designationObj.save()
        #
        #         print 'user', userObj.first_name, userObj.pk



        # getting company creation
        companyPath = os.path.join(globalSettings.BASE_DIR, 'static_shared','sdpl_company.json')
        with open(companyPath) as json_file:
            comapnyData = json.load(json_file)
            for item in comapnyData:

                try:
                    if item['user'] is not None:
                        userObj = User.objects.get(username = item['user']['username'])
                except:
                    print item['user'], 'user'
                    userObjData = [x for x in userData if x['pk'] == item['user']]
                    userObj = User.objects.get(username = userObjData[0]['username'])
                else:
                    userObj = User.objects.get(email = 'sachin@superiorsystems.in')


                objs = service.objects.filter(name = item['name'],user = userObj)
                obj = None
                if len(objs)>0:
                    obj = objs[0]
                if len(objs) == 0:
                    print item['name'], 'name '
                    if item['name'] != 'cioc':
                        obj = service.objects.create(name = item['name'],user = userObj)

                if obj is not None:
                    obj.contactPerson = item['contactPerson']
                    obj.mobile = item['mobile']

                    if item['address'] is not None:
                        addressObj = address.objects.create(city = item['address']['city'], country = item['address']['country'], lon = item['address']['lon'] , pincode = item['address']['pincode'], state = item['address']['state'],street = item['address']['street'],lat = item['address']['lat'])

                        obj.address = addressObj

                    obj.save()
                    print 'company', obj.name, obj.pk

        # #creating crm term and condition details
        termsAndConditionPath = os.path.join(globalSettings.BASE_DIR, 'static_shared','sdpl_crmtermsAndConditions.json')
        with open(termsAndConditionPath) as json_file:
            termsAndConditionData = json.load(json_file)
            for item in termsAndConditionData:
                obj, created = CRMTermsAndConditions.objects.get_or_create(heading = item['heading'], version = 'V2', division = divisionObj,themeColor = '#BB2724')
                obj.body = item['body']
                obj.default = item['default']
                # obj.extraFieldOne = item['extraFieldOne']
                # obj.extraFieldTwo = item['extraFieldTwo']
                # obj.isGst = item['isGst']

                obj.save()
                print 'termsAndCondition', obj.heading, obj.pk
        #
        #
        # #creating inventory data
        # inventoryPath = os.path.join(globalSettings.BASE_DIR, 'static_shared','sdpl_inventory.json')
        # with open(inventoryPath) as json_file:
        #     inventoryData = json.load(json_file)
        #     for item in inventoryData:
        #         obj = Inventory.objects.create(name = item['name'], division = divisionObj)
        #         obj.value = item['value']
        #         obj.refurnishedAdded = item['refurnishedAdded']
        #         obj.sellable = item['sellable']
        #         obj.rate = item['rate']
        #         obj.qtyAdded = item['qtyAdded']
        #         obj.totalRef = item['totalRef']
        #         obj.total = item['total']
        #         obj.richtxtDesc = item['richtxtDesc']
        #         obj.description = item['description']
        #
        #         if item['productMeta'] is not None:
        #             productMetaObj ,created = ProductMeta.objects.get_or_create(code = item['productMeta']['code'],taxRate = item['productMeta']['taxRate'])
        #             productMetaObj.typ  = item['productMeta']['typ']
        #             productMetaObj.description  = item['productMeta']['description']
        #             obj.productMeta = productMetaObj
        #
        #         obj.save()
        #         print 'inventory', obj.name, obj.pk

        # #creating contacts data
        contactPath = os.path.join(globalSettings.BASE_DIR, 'static_shared','sdpl_contacts.json')
        with open(contactPath) as json_file:
            contactData = json.load(json_file)
            for item in contactData:
                print item['mobile']

                try:
                    if item['user'] is not None:
                        userObj = User.objects.get(username = item['user']['username'])
                except:
                    print item['user'], 'user'
                    userObjData = [x for x in userData if x['pk'] == item['user']]
                    userObj = User.objects.get(username = userObjData[0]['username'])
                else:
                    userObj = User.objects.get(email = 'sachin@superiorsystems.in')


                obj = Contact.objects.create(name = item['name'], division = divisionObj, user = userObj)
                if item['company'] is not None:
                    companyObjs = service.objects.filter(name = item['company']['name'])
                    if len(companyObjs)>0:
                        obj.company = companyObjs[0]
                obj.email = item['email']
                obj.emailSecondary = item['emailSecondary']
                obj.mobileSecondary = item['mobileSecondary']
                obj.designation = item['designation']
                obj.notes = item['notes']
                obj.linkedin = item['linkedin']
                obj.facebook = item['facebook']
                obj.male = item['male']
                if item['street'] is not None and len(item['street'])<=100:
                    obj.street = item['street']
                obj.male = item['male']
                obj.city = item['city']
                obj.pincode = item['pincode']
                obj.country = item['country']
                obj.isGst = item['isGst']
                if item['mobile'] is not None and len(item['mobile'])<=12:
                    obj.mobile = item['mobile']
                obj.save()
                print 'contact ',obj.name , obj.mobile
        #
        # #creating contracts
        contractPath = os.path.join(globalSettings.BASE_DIR, 'static_shared','sdpl_contracts.json')
        with open(contractPath) as json_file:
            contractData = json.load(json_file)
            for item in contractData:
                contactObj = None

                if item['contact'] is not None:
                    contactObjs = Contact.objects.filter(email = item['contact']['email'], division = divisionObj, mobile = item['contact']['mobile'])

                    if len(contactObjs)>0:
                        contactObj = contactObjs[0]
                if contactObj is not None:
                    try:
                        if item['user'] is not None:
                            userObj = User.objects.get(username = item['user']['username'])
                    except:
                        print item['user'], 'user'
                        userObjData = [x for x in userData if x['pk'] == item['user']]
                        userObj = User.objects.get(username = userObjData[0]['username'])
                    else:
                        userObj = User.objects.get(email = 'sachin@superiorsystems.in')

                    obj = Contract.objects.create(data = item['data'], contact = contactObj, user = userObj)
                    obj.grandTotal = item['grandTotal']
                    obj.dueDate = item['dueDate']
                    obj.created = item['created']
                    obj.updated = item['updated']
                    obj.archivedDate = item['archivedDate']
                    obj.status = item['status']
                    obj.totalTax = item['totalTax']
                    obj.value = item['value']
                    obj.recievedDate = item['recievedDate']
                    obj.heading = item['subject']
                    obj.read = item['read']

                    if item['termsAndCondition'] is not None:
                        conditionObj = CRMTermsAndConditions.objects.filter(heading = item['termsAndCondition']['heading'], division = divisionObj)
                        if len(conditionObj)>0:
                            obj.termsAndCondition = conditionObj[0]

                    obj.save()
                    print 'contract', obj.pk
        return Response({'status':'ok'}, status=status.HTTP_200_OK)

import basehash
hash_fn = basehash.base36()
from chatbot.uipathHelper import uiPathToken

class SaveSettingsAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes=(permissions.AllowAny,)
    def get(self , request , format = None):
        profile = request.user.designation.division

        if request.GET['type'] == 'uipath':
            return Response({'email' : profile.uipathUsername, 'tenant' : profile.uipathTenent ,'url' :  profile.uipathUrl , 'password' : profile.uipathPass , 'uipathOrgId' : profile.uipathOrgId },status = status.HTTP_200_OK)
        elif request.GET['type'] == 'web':
            return Response({'url' : profile.website ,  'chatterLink' : hash_fn.hash(profile.pk) , 'host' : globalSettings.SITE_ADDRESS  }, status = status.HTTP_200_OK)
        elif request.GET['type'] == 'messenger':
            return Response({'access_token' : profile.access_token,'pageID' : profile.pageID}, status = status.HTTP_200_OK)
        elif request.GET['type'] == 'whatsapp':
            return Response({'twillioAccountSID' : profile.twillioAccountSID,'trillioAuthToken' : profile.trillioAuthToken,'whatsappNumber' : profile.whatsappNumber}, status = status.HTTP_200_OK)
        elif request.GET['type'] == 'buttonSetting':
            return Response({'ios_sdk_enabled' : profile.ios_sdk_enabled,'react_sdk_enabled' : profile.react_sdk_enabled,'rest_sdk_enabled' : profile.rest_sdk_enabled,'android_sdk_enabled' : profile.android_sdk_enabled}, status = status.HTTP_200_OK)
        elif request.GET['type'] == 'uiSettings':
            if profile.dp:
                print dir( profile.dp )
                print profile.dp.url
                dp = profile.dp.url
            else:
                dp = None

            return Response({
                "supportBubbleColor"  : profile.supportBubbleColor,
                "fontColor"  : profile.fontColor,
                "iconColor"  : profile.iconColor,
                "windowColor"  : profile.windowColor,
                "firstMessage"  : profile.firstMessage,
                "chatIconPosition"  : profile.chatIconPosition,
                "chatIconType"  : profile.chatIconType,
                "mascotName"  : profile.mascotName,
                "dp"  : dp,
            }, status = status.HTTP_200_OK)
        elif request.GET['type'] == 'whatsappTest':
            return Response({ 'whatsapp_test_number' : profile.whatsapp_test_number   }, status = status.HTTP_200_OK)


    def post(self , request , format = None):
        custProfile = request.user.designation.division

        print dir(request)
        if request.data['type'] == 'uipath':
            d = request.data['data']
            token = uiPathToken( d['email'] , d['password'] , d['tenant'] ,  d['url'])
            custProfile.uipathUrl = d['url']
            custProfile.uipathPass = d['password']
            custProfile.uipathUsername = d['email']
            custProfile.uipathTenent = d['tenant']
            custProfile.uipathOrgId = d['uipathOrgId']
        elif request.data['type'] == 'messenger':
            d = request.data['data']
            custProfile.access_token = d['access_token']
            custProfile.pageID = d['pageID']
        elif request.data['type'] == 'whatsapp':
            d = request.data['data']
            custProfile.twillioAccountSID = d['twillioAccountSID']
            custProfile.trillioAuthToken = d['trillioAuthToken']
            custProfile.whatsappNumber = d['whatsappNumber']
        elif request.data['type'] == 'web':
            custProfile.website = request.data['data']['url']
        elif request.data['type'] == 'ios':
            custProfile.ios_sdk_enabled = request.data['ios_sdk_enabled']
        elif request.data['type'] == 'react':
            custProfile.react_sdk_enabled = request.data['react_sdk_enabled']
        elif request.data['type'] == 'rest':
            custProfile.rest_sdk_enabled = request.data['rest_sdk_enabled']
        elif request.data['type'] == 'android':
            custProfile.android_sdk_enabled = request.data['android_sdk_enabled']
        elif request.data['type'] == 'whatsappTest':
            custProfile.whatsapp_test_number = request.data['whatsapp_test_number']
        elif request.data['type'] == 'ui':
            custProfile.supportBubbleColor = request.data["supportBubbleColor"]
            custProfile.fontColor = request.data["fontColor"]
            custProfile.iconColor = request.data["iconColor"]
            custProfile.windowColor = request.data["windowColor"]
            custProfile.firstMessage = request.data["firstMessage"]
            custProfile.chatIconPosition = request.data["chatIconPosition"]
            custProfile.chatIconType = request.data["chatIconType"]
            custProfile.mascotName = request.data["mascotName"]
            if 'dp' in request.FILES :
                custProfile.dp = request.FILES["dp"]

        custProfile.save()
        return Response({},status = status.HTTP_200_OK)

    def delete(self , request , format = None):
        custProfile = request.user.designation.division
        custProfile.uipathUrl = None
        custProfile.uipathPass = None
        custProfile.uipathUsername = None
        custProfile.uipathTenent = None
        custProfile.save()
        return Response({},status = status.HTTP_200_OK)

class UnInstallApp(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny, )
    def get(self, request, format=None):
        params = request.GET
        div = request.user.designation.division
        app = InstalledApp.objects.get(pk =int(params['id']))
        mainApp = app.app
        if mainApp.name == 'app.organization':
            div.simpleMode = True
            div.save()
        app.delete()
        usersAppObj = UserApp.objects.filter(user__designation__division = div, app = mainApp)
        usersAppObj.delete()
        # designation = request.user.designation
        # division = designation.division
        # try:
        #     if division.simpleMode == True:
        #         designation.apps.remove(app)
        #         designation.save()
        # except:
        #     pass
        return Response(status=status.HTTP_200_OK)


class InstallUserApp(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.AllowAny, )
    def post(self, request, format=None):
        params = request.data
        if 'type' in params:
            user = request.user
            div = user.designation.division
            appDetails = application.objects.get(pk = params['app'])
            if appDetails.name == 'app.organization':
                div.simpleMode = False
                div.save()
            app, created = InstalledApp.objects.get_or_create(parent = div , app = appDetails, priceAsAdded = params['priceAsAdded'] , addedBy= request.user)
            app.save()
            data = {}
            if appDetails.inMenu == True:
                ua = UserApp(user = user , app = application.objects.get(pk = params['app'])  )
                ua.save()
            # userObj = userSearchSerializer(user, many = False).data
            # userAppObj = UserAppsSerializerr(user, many = False).data
            # data = {'userObj' : userObj, 'apps' : apps}
                data = UserAppSerializer(ua,many=False).data
            return Response({'data':data},status=status.HTTP_200_OK)
        ua = UserApp(user = User.objects.get(pk = int(params['user'])) , app = application.objects.get(pk = int(params['app']))  )
        ua.save()
        data = UserAppSerializer(ua,many=False).data
        return Response({'data':data},status=status.HTTP_200_OK)
