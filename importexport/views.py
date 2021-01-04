from django.contrib.auth.models import User , Group
from django.shortcuts import render, redirect , get_object_or_404
from django.contrib.auth import authenticate , login , logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.urlresolvers import reverse
from django.template import RequestContext
from django.conf import settings as globalSettings
from django.core.exceptions import ObjectDoesNotExist , SuspiciousOperation
from django.views.decorators.csrf import csrf_exempt, csrf_protect
from django.utils.encoding import smart_str
# Related to the REST Framework
from rest_framework import viewsets , permissions , serializers
from rest_framework.exceptions import *
from url_filter.integrations.drf import DjangoFilterBackend
from .models import *
from .serializers import *
from API.permissions import *
from django.db.models import Q,Count ,F,Sum
from django.http import JsonResponse
from rest_framework.renderers import JSONRenderer
import random, string , locale
from django.utils import timezone
from rest_framework.views import APIView
from datetime import date,timedelta
from dateutil.relativedelta import relativedelta
import calendar
from rest_framework.response import Response
from django.contrib.auth.models import User, Group
from allauth.socialaccount.models import *
from rest_framework import filters
from openpyxl import load_workbook,Workbook
# from xlsxwriter import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import Font,Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
from django.db.models.functions import Extract , ExtractDay, ExtractMonth, ExtractYear
from django.db.models import Sum , Avg
import json
import os
from django.db.models import CharField,FloatField, Value , Func
import csv
import pandas as pd
from django.http import HttpResponse
from reportlab import *
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.pagesizes import letter, landscape , LETTER
from reportlab.lib.units import cm, mm
from reportlab.lib import colors , utils
from reportlab.platypus import Paragraph, Table, TableStyle, Image, Frame, Spacer, PageBreak, BaseDocTemplate, PageTemplate, SimpleDocTemplate, Flowable,ListItem,ListFlowable,NextPageTemplate
from PIL import Image
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet, TA_CENTER
from reportlab.graphics import barcode , renderPDF
from reportlab.graphics.shapes import *
from reportlab.graphics.barcode.qr import QrCodeWidget
from reportlab.lib.pagesizes import letter,A5,A4,A3,A2
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import ParagraphStyle,getSampleStyleSheet
from reportlab.lib.enums import TA_LEFT, TA_CENTER
from reportlab.lib.colors import *
from reportlab.lib.units import inch, cm
from django.template.loader import render_to_string, get_template
from django.core.mail import send_mail, EmailMessage
from excel_response import ExcelResponse
from num2words import num2words
import re


class ProductsViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.AllowAny , )
    serializer_class = ProductsSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['parent','created' , 'bar_code']
    def get_queryset(self):
        if 'search' in self.request.GET:
            print 'herrrrrrrrrrreeeee'
            objs = Products.objects.all()
            product = objs.filter(part_no__contains=str(self.request.GET['search']))
            product1  = objs.filter(replaced__icontains=str(self.request.GET['search']))
            return product | product1
        elif 'searchContains' in self.request.GET:
            productList = list(Inventory.objects.filter(product__part_no__icontains = self.request.GET['searchContains']).distinct().values_list('product',flat=True))
            return Products.objects.filter(pk__in=productList)
        elif 'searchBom' in self.request.GET and 'project' in self.request.GET:
            productList=list(BoM.objects.filter(products__part_no__icontains=str(self.request.GET['searchBom']),project__pk=self.request.GET['project']).distinct().values_list('products',flat=True))
            return Products.objects.filter(pk__in=productList)
        else:
            return Products.objects.all()




class ProductSheetViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.AllowAny , )
    queryset = ProductSheet.objects.all()
    serializer_class = ProductSheetSerializer



class ProjectsViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.AllowAny , )
    queryset = Projects.objects.all().order_by('-created')
    serializer_class = ProjectsSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['status','title','savedStatus','junkStatus','comm_nr','flag']

    def get_queryset(self):
        user = self.request.user
        divisionObj = user.designation.division
        if 'searchContains' in self.request.GET:
            objs = Projects.objects.all()
            product = objs.filter(title__contains=str(self.request.GET['searchContains']),division = divisionObj)
            return product
        if 'name' in self.request.GET:
            return Projects.objects.filter(comm_nr__icontains= str(self.request.GET['name']))
        else:
            return Projects.objects.filter(division = divisionObj).order_by('-created')

class VendorViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.AllowAny , )
    serializer_class = VendorSerializer
    def get_queryset(self):
        user = self.request.user
        divisionObj = user.designation.division
        if 'name' in self.request.GET:
            queryset = Vendor.objects.filter(name__icontains = str(self.request.GET['name']),division = divisionObj)
            return queryset

        else:
            queryset = Vendor.objects.filter(division = divisionObj)
            return queryset


class BoMViewSet(viewsets.ModelViewSet):
    permissions_classes  = (permissions.AllowAny , )
    queryset = BoM.objects.all()
    serializer_class = BoMSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['products','project']
    def get_queryset(self):
        user = self.request.user
        divisionObj = user.designation.division
        if 'searchBom' in self.request.GET and 'project' in self.request.GET:
            productList=BoM.objects.filter(products__part_no__icontains=str(self.request.GET['searchBom']),project__pk=self.request.GET['project'],division = divisionObj)
            return productList
        else:
            return BoM.objects.filter(division = divisionObj)

class StockCheckViewSet(viewsets.ModelViewSet):
    permissions_classes  = (permissions.AllowAny , )
    serializer_class = StockCheckSerializer
    def get_queryset(self):
        user = self.request.user
        divisionObj = user.designation.division
        queryset = StockCheck.objects.filter(division = divisionObj)
        return queryset

class StockCheckLogViewSet(viewsets.ModelViewSet):
    permissions_classes  = (permissions.AllowAny , )
    queryset = StockCheckLog.objects.all()
    serializer_class = StockCheckLogSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['products','project']

class StockSummaryReportViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.AllowAny , )
    serializer_class = StockSummaryReportSerializer
    def get_queryset(self):
        user = self.request.user
        divisionObj = user.designation.division
        queryset = StockSummaryReport.objects.filter(division = divisionObj)
        return queryset

class ProjectStockSummaryViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.AllowAny , )
    queryset = ProjectStockSummary.objects.all()
    serializer_class = ProjectStockSummarySerializer
    def get_queryset(self):
        user = self.request.user
        divisionObj = user.designation.division
        queryset = ProjectStockSummary.objects.filter(division = divisionObj)
        return queryset

class InvoiceViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.AllowAny , )
    queryset = Invoice.objects.all()
    serializer_class = InvoiceSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['project','invoiceNumber','flag']
    def get_queryset(self):
        user = self.request.user
        divisionObj = user.designation.division
        if 'invoiceno' in self.request.GET:
            return Invoice.objects.filter(invoiceNumber__icontains = str(self.request.GET['invoiceno']))
        else:
            queryset = Invoice.objects.filter(division = divisionObj)

            return queryset

class InvoiceQtyViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.AllowAny , )
    queryset = InvoiceQty.objects.all()
    serializer_class = InvoiceQtySerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['invoice']

class DeliveryChallanViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.AllowAny , )
    serializer_class = DeliveryChallanSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['materialIssue','challanNo','flag']
    def get_queryset(self):
        user = self.request.user
        divisionObj = user.designation.division
        if 'name' in self.request.GET:
            return  DeliveryChallan.objects.filter(customername__icontains = str(self.request.GET['name']), division = divisionObj)
        else:
            queryset = DeliveryChallan.objects.filter(division = divisionObj)
            return queryset

class StockCheckReportViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.AllowAny , )
    serializer_class = StockCheckReportSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['flag']
    def get_queryset(self):
        user = self.request.user
        divisionObj = user.designation.division
        if 'pk' in self.request.GET:
            return StockCheckReport.objects.filter(pk = int(self.request.GET['pk']),division = divisionObj)
        else:
            queryset = StockCheckReport.objects.filter(division = divisionObj)
            return queryset

class StockCheckItemViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.AllowAny , )
    queryset = StockCheckItem.objects.all()
    serializer_class = StockCheckItemSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['stockReport','flag']

class ProductsUploadAPIView(APIView):
    permission_classes = (permissions.IsAuthenticated ,)

    def post(self , request , format = None):

        excelFile = request.FILES['excelFile']

        print 'llllllllllll'
        wb = load_workbook(filename = BytesIO(excelFile.read()) ,  read_only=True)


        for ws in wb.worksheets:
            wsTitle = ws.title
            sheetObj = ProductSheet(file_name = excelFile.name,sheet=wsTitle)
            sheetObj.save()
            unSaved = []
            count = 0
            for i in range(2,ws.max_row+1):
                try:
                    print 'aaaaaaaaaaa'
                    try:
                        part_no = ws['A' + str(i)].value
                    except:
                        part_no = None



                    try:
                        description_1 = ws['B' + str(i)].value
                    except:
                        description_1 = None


                    try:
                        description_2 = ws['C' + str(i)].value
                    except:
                        description_2 = None


                    try:
                        weight = ws['D' + str(i)].value
                    except:
                        weight = None


                    try:
                        price = ws['E' + str(i)].value
                    except:
                        price = None


                    try:
                        parent_no = ws['F' + str(i)].value
                        parent = None
                        if parent_no:
                            par = Products.objects.get(part_no=parent_no)
                            print par ,'rrrrrrrrrrrrrr'
                            parent = par
                    except:
                        parent = None

                    try:
                        replaced = ws['F' + str(i)].value
                    except:
                        replaced = None


                    try:
                        customs_no = ws['G' + str(i)].value

                    except:
                        customs_no = None

                    try:
                        custom = ws['H' + str(i)].value
                    except:
                        custom = 0
                    try:
                        gst = ws['I' + str(i)].value
                    except:
                        gst = 18

                    try:
                        bar_code = ws['J' + str(i)].value

                    except:
                        bar_code = None


                    Products.objects.get_or_create(part_no=part_no, description_1=description_1,description_2=description_2,replaced=replaced,parent=parent,weight=weight, price=price,customs_no=customs_no,custom=custom,gst=gst,bar_code=bar_code)
                    count+=1
                except:
                    unSaved.append(part_no)
        data = {"count":count,"no":part_no}
        return Response(data,status=status.HTTP_200_OK)

class ProductsUpdateAPIView(APIView):
    permission_classes = (permissions.IsAuthenticated ,)

    def post(self , request , format = None):
        excelFile = request.FILES['excelFile']
        print 'llllllllllll'
        wb = load_workbook(filename = BytesIO(excelFile.read()) ,  read_only=True)
        count = 0
        notUpdated = 0
        notUpdated_partNo = []
        for ws in wb.worksheets:
            wsTitle = ws.title
            sheetObj = ProductSheet(file_name = excelFile.name,sheet=wsTitle)
            sheetObj.save()
            unSaved = []

            for i in range(2,ws.max_row+1):
                try:
                    try:
                        part_no = ws['A' + str(i)].value
                    except:
                        part_no = None

                    try:
                        obj = Products.objects.get(part_no=part_no)
                        try:
                            description_1 = ws['B' + str(i)].value
                        except:
                            description_1 = obj.description_1


                        try:
                            description_2 = ws['C' + str(i)].value
                        except:
                            description_2 = obj.description_2


                        try:
                            weight = ws['D' + str(i)].value
                        except:
                            weight = obj.weight


                        try:
                            price = ws['E' + str(i)].value
                        except:
                            price = obj.price


                        try:
                            parent_no = ws['F' + str(i)].value
                            parent = None
                            if parent_no:
                                par = Products.objects.get(part_no=parent_no)
                                print par ,'rrrrrrrrrrrrrr'
                                parent = par
                        except:
                            parent = obj.parent

                        try:
                            replaced = ws['F' + str(i)].value
                        except:
                            replaced = obj.replaced


                        try:
                            customs_no = ws['G' + str(i)].value

                        except:
                            customs_no = obj.customs_no

                        try:
                            custom = ws['H' + str(i)].value
                        except:
                            custom = obj.custom
                        try:
                            gst = ws['I' + str(i)].value
                        except:
                            gst = obj.gst

                        try:
                            bar_code = ws['J' + str(i)].value

                        except:
                            bar_code = obj.bar_code
                        print weight,'hhhhhhhhhhhhhhhhhhhh'

                        serializer = ProductsSerializer(obj, data={'description_1':description_1,'description_2':description_2,'replaced':replaced,'parent':parent,'weight':weight,'price':price,'customs_no':customs_no,'custom':custom,'gst':gst,'bar_code':bar_code})
                        if serializer.is_valid():

                            serializer.save()
                            count += 1
                            print weight,count,'ggggggggggggggggggg'
                        else:
                            notUpdated +=1
                            notUpdated_partNo.append(part_no)
                            print weight,'iiiiiiiiiiiiiiiiiii'

                    except:
                        notUpdated +=1
                        notUpdated_partNo.append(part_no)
                        print weight,'kkkkkkkkkkkkkkkkkk'

                except:
                    notUpdated +=1
                    notUpdated_partNo.append(part_no)
                    print weight,'llllllllllllllllllll'
        print notUpdated_partNo ,'notttttttttt'
        data = {"count":count,"notUpdated":notUpdated,'notUpdated_partNo':notUpdated_partNo}
        return Response(data,status=status.HTTP_200_OK)


class UploadSheetAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    permission_classes = (permissions.IsAuthenticated ,)
    def post(self , request , format = None):
        project = Projects.objects.get(pk=request.GET['projectpk'])
        projectPk = project.id
        userPk = request.GET['userpk']
        excelFile = request.FILES['excelFile']
        wb = load_workbook(filename = BytesIO(excelFile.read()) ,  read_only=True)
        products=[]
        productsExist = []
        for ws in wb.worksheets:
            wsTitle = ws.title
            sheetObj = ProductSheet(file_name = excelFile.name,sheet=wsTitle)
            sheetObj.save()
            unSaved = []
            count = 0
            notAdded = 0
            totInvoiceValue = 0
            incorrect_partNo = 0
            incorrect_Arr = []
            for i in range(2,ws.max_row+1):
                print 'aaaaaaaaaaa'
                try:
                    part_no = ws['A' + str(i)].value
                except:
                    part_no = None

                try:
                    quantity = ws['B' + str(i)].value
                except:
                    quantity = None

                try:
                    prodObj = Products.objects.get(part_no=str(part_no))
                    prodPk = prodObj.id
                    quotePrice = round(((project.profitMargin * prodObj.price) / 100 + prodObj.price),2)
                    inrPrice = round((prodObj.price * project.exRate),2)
                    invoiceValue = prodObj.price*quantity

                    print quotePrice,inrPrice
                    packing = round(((project.packing/ invoiceValue) * inrPrice),2)
                    insurance = round((((project.insurance  / invoiceValue) * inrPrice)),2)
                    freight = round((((project.freight / invoiceValue) * inrPrice)),2)
                    cif = round((inrPrice + packing + insurance + freight),2)
                    cifround = cif * quantity
                    customVal = round(((cif +((cif * project.assessableValue)/100))*(prodObj.custom)/100),2)
                    socialVal = round((customVal *0.1),2)
                    gstVal = round(((cif+customVal+socialVal)*(prodObj.gst)/100),2)
                    charge1 = round((inrPrice * (project.clearingCharges1 / (invoiceValue * project.exRate))),2)
                    charge2 = round((inrPrice * (project.clearingCharges2 / (invoiceValue * project.exRate))),2)
                    landng=cif+customVal+charge2+charge1+customVal
                    try:
                        bomExist = BoM.objects.get(products__pk=prodPk,project__pk=projectPk)
                        productsExist.append(bomExist.pk)
                        notAdded += 1
                    except:
                        bomExist = None


                        totInvoiceValue += invoiceValue
                        bomObj = BoM.objects.create(user=User.objects.get(pk=userPk), products=Products.objects.get(pk=prodPk),project=Projects.objects.get(pk=projectPk),quantity1=quantity,quantity2=quantity,price=prodObj.price, landed_price=landng,gst=prodObj.gst,custom=prodObj.custom,customs_no=prodObj.customs_no)
                        bomObj.save()
                        productsObj = ProductsSerializer(instance=Products.objects.get(pk=prodPk)).data
                        prjojectObj = ProjectsSerializer(instance=Projects.objects.get(pk=projectPk)).data


                        count+=1
                        products.append({'pk':bomObj.pk,'user':bomObj.user.pk,'products':productsObj,'project':prjojectObj,'quantity1':bomObj.quantity1,'quantity2':bomObj.quantity2,'price':bomObj.price,'landed_price':bomObj.landed_price,'gst':bomObj.gst,'gstVal':gstVal,'custom':bomObj.custom,'charge1':charge1,'charge2':charge2,'cif':cif,'customVal':customVal,'customer_price':0,'invoice_price':0,'customs_no':bomObj.customs_no,'freight':freight,'inrPrice':inrPrice,'insurance':insurance,'packing':packing,'quotePrice':quotePrice,'socialVal':socialVal})
                except:
                    print part_no,'partnoooooooooooo'

                    prodObj = None
                    incorrect_Arr.append(part_no)
                    incorrect_partNo += 1

        print products,'dddddddddddddddddd'
        data = {"count":count,'notAdded':notAdded,'inCorrect':incorrect_partNo,"product":products,'productExist':productsExist,'invoiceValue':totInvoiceValue,'incorrcetPartno':incorrect_Arr}
        return Response(data,status=status.HTTP_200_OK)


from reportlab.lib.styles import getSampleStyleSheet
from svglib.svglib import svg2rlg
from reportlab.lib.enums import TA_RIGHT, TA_CENTER
styles = getSampleStyleSheet()


def footer(canvas, doc):
    canvas.saveState()
    header_content = svg2rlg(os.path.join(globalSettings.BASE_DIR , 'static_shared','images' , 'Bruderer_Logo_svg.svg'))
    sx=sy=0.17
    header_content.width,header_content.height = header_content.minWidth()*sx, header_content.height*sy
    header_content.scale(sx,sy)
    header_content.hAlign = 'RIGHT'
    header_content.drawOn(canvas, 450,750)
    P = Paragraph("<para  align='center'><font size='8'>CIN No. U31900KA2001PTC099049<br/> Regd. Office. #17P, Sadaramangala Industrial Area,Whitefield Road,Kadugodi,Bangalore 560 048<br/> Phone : +91 80 2841 1049<br/> e-mail : info.in@bruderer.com  website : www.bruderer.com</font> </para>", styles['Normal'])
    w, h = P.wrap(doc.width, doc.bottomMargin)
    P.drawOn(canvas, doc.leftMargin, h-20)

    canvas.restoreState()


Round = lambda x, n: eval('"%.' + str(int(n)) + 'f" % ' + repr(x))
def purchaseOrder(response , project , purchaselist, multNumber,currencyTyp, request):
    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(response,pagesize=letter, topMargin=70,leftMargin=1.2*cm,rightMargin=1.2*cm)
    doc.request = request
    elements = []
    style_right = ParagraphStyle(name='right', parent=styles['Normal'], alignment=TA_RIGHT)
    summryHeader = Paragraph("""
    <para >
    <font size='14'>
    PURCHASE ORDER
    <br/>
    <br/>
    </font>
    </para>
    """ %(),styles['Normal'])
    summryHeader1 = Paragraph("""
    <para leftIndent = 10>
    <font size ='10'>
    <b>Purchase Order Ref No :</b> %s <br/>
    <b>Purchase Order Ref Date :</b> %s <br/>
    </font></para>
    """ %(project.poNumber , project.poDate),styles['Normal'])
    tdheader=[[summryHeader,' ',summryHeader1]]
    theader=Table(tdheader,colWidths=[3*inch , 1*inch , 3*inch])
    theader.hAlign = 'LEFT'
    elements.append(theader)
    summryParaSrc = Paragraph("""
    <para >
    <font size='10'>
    <b>%s</b> <br/>
    %s <br/>
    %s - %s<br/>
    %s <br/>
    %s<br/>
    </font>
    </para>
    """ %(project.vendor.name,project.vendor.street ,project.vendor.city,project.vendor.pincode,project.vendor.state , project.vendor.country),styles['Normal'])
    summryParaSrc1 = Paragraph("""
    <para leftIndent = 10>
    <font size ='10'>
    <b>Kind attn :</b> %s <br/>
    <b>Your Quote Ref :</b> %s <br/><br/><br/><br/>
    </font></para>
    """ %(project.vendor.personName,project.quote_ref),styles['Normal'])
    td=[[summryParaSrc,' ',summryParaSrc1]]
    t=Table(td,colWidths=[3*inch , 1*inch , 3*inch])
    t.hAlign = 'LEFT'
    elements.append(t)
    details = Paragraph("""
    <para >
    <font size='10'>
    <b>BILL TO AND SHIP TO : </b> <br/>
    <b>BRUDERER PRESSES INDIA PRIVATE LTD </b> <br/>
    #17P Sadaramangala Industrial Area, <br/>
    Whitefield Road, Kadugodi. <br/>
    Bangalore 560048 <br/>
    Karnataka <br/>
    India <br/><br/>
    <b>GST Number :  29AABCB6326Q1Z6 </b> <br/>
    Contact no. : +91 7975883145 <br/>
    E-Mail : Manjula.Shenoy@bruderer.com
    </font>
    </para>
    """ %(),styles['Normal'])
    details1 = Paragraph("""
    <para >
    <font size='10'>
    <b>Machine Type :</b> %s <br/>
    <b>Comm Nr :</b> %s<br/>
     <br/><br/><br/><br/><br/>
    </font>
    </para>
    """ %(project.machinemodel, project.comm_nr),styles['Normal'])
    td1=[[details,'',details1]]
    t=Table(td1,colWidths=[3.2*inch , 1*inch , 3*inch])
    t.hAlign = 'LEFT'
    elements.append(t)

    elements.append(Spacer(1,16))
    p14_02 =Paragraph("<para fontSize=8>INCO TERMS</para>",styles['Normal'])
    p14_03 =Paragraph("<para fontSize=8>SPECIAL INSTRUCTIONS</para>",styles['Normal'])
    p14_04 =Paragraph("<para fontSize=8>PAYMENT TERMS</para>",styles['Normal'])
    data6=[[p14_02,p14_03,p14_04]]
    special1 = "Delivery - " + str(project.date)
    special2 = "Shipment mode - " + project.shipmentMode
    special3 = project.shipmentDetails
    p15_02 =Paragraph("<para fontSize=8>{0}</para>".format(project.termspo),styles['Normal'])
    p15_03 =Paragraph("<para fontSize=8>{0}<br/>{1}<br/><b>{2}</b></para>".format(special1,special2,special3),styles['Normal'])
    p15_04 =Paragraph("<para fontSize=8>{0}</para>".format(project.paymentTerms1),styles['Normal'])
    data6+=[[p15_02,p15_03,p15_04]]
    t6=Table(data6)
    t6.hAlign = 'LEFT'
    t6.setStyle(TableStyle([('TEXTFONT', (0, 0), (-1, -1), 'Times-Bold'),('TEXTCOLOR',(0,0),(-1,-1),black),('ALIGN',(0,0),(-1,-1),'LEFT'),('VALIGN',(0,0),(-1,-1),'TOP'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black)]))
    elements.append(t6)

    elements.append(Spacer(1,10))
    if currencyTyp == 'INR':
        priceTitle = 'Unit price in INR'
        amountTitle = 'Amount in INR'
        totalTitle = 'Total in INR'
    else:
        priceTitle = 'Unit price in '+project.currency
        amountTitle = 'Amount in '+project.currency
        totalTitle = 'Total in '+project.currency
    p101_01 =Paragraph("<para fontSize=8>Sl. no</para>",styles['Normal'])
    p101_02 =Paragraph("<para fontSize=8>Part Number</para>",styles['Normal'])
    p101_03 =Paragraph("<para fontSize=8>Part Desc</para>",styles['Normal'])
    p101_07 =Paragraph("<para fontSize=8>HSN Code</para>",styles['Normal'])
    p101_04 =Paragraph("<para fontSize=8>Qty</para>",styles['Normal'])
    p101_05 =Paragraph("<para fontSize=8>{0}</para>".format(priceTitle),styles['Normal'])
    p101_06 =Paragraph("<para fontSize=8>{0}</para>".format(amountTitle),styles['Normal'])


    data5=[[p101_01,p101_02,p101_03,p101_07,p101_04,p101_05,p101_06]]

    grandTotal = 0
    data2 = []
    id=0
    for i in purchaselist:
        id+=1
        part_no = i.products.part_no
        desc = i.products.description_1
        hs = i.products.customs_no
        if currencyTyp == 'INR':
            price = i.landed_price
        else:
            price = i.price

        qty = i.quantity1
        amnt = round((price * qty),2)
        grandTotal +=amnt
        grandTotal = round(grandTotal,2)

        p12_01 = Paragraph("<para fontSize=8>{0}</para>".format(id),styles['Normal'])
        p12_02 =Paragraph("<para fontSize=8>{0}</para>".format(part_no),styles['Normal'])
        p12_03 =Paragraph("<para fontSize=8>{0}</para>".format(smart_str(desc)),styles['Normal'])
        p12_07 =Paragraph("<para fontSize=8>{0}</para>".format(smart_str(hs)),styles['Normal'])
        p12_04 =Paragraph("<para fontSize=8>{0}</para>".format(qty),styles['Normal'])
        p12_05 =Paragraph("<para fontSize=8> {:,.2f}</para>".format(price),style_right)
        p12_06 =Paragraph("<para fontSize=8> {:,.2f}</para>".format(amnt),style_right)
        data5.append([p12_01,p12_02,p12_03,p12_07,p12_04,p12_05,p12_06])


    grandTotal = round(grandTotal,2)
    p13_01 =Paragraph("<para fontSize=8>{0}</para>".format(''),styles['Normal'])
    p13_02 =Paragraph("<para fontSize=8>{0}</para>".format(''),styles['Normal'])
    p13_03 =Paragraph("<para fontSize=8>{0}</para>".format(''),styles['Normal'])
    p13_04 =Paragraph("<para fontSize=8>{0}</para>".format(''),styles['Normal'])
    p13_05 =Paragraph("<para fontSize=8></para>",styles['Normal'])
    p13_06 =Paragraph("<para fontSize=8> {0}</para>".format(totalTitle),styles['Normal'])
    p13_07 =Paragraph("<para fontSize=8>{:,.2f}</para>".format(grandTotal),style_right)

    data5+=[[p13_01,p13_02,p13_03,p13_04,p13_05,p13_06,p13_07]]
    t3=Table(data5,colWidths=(10*mm,None, 50*mm, None, None, None, None),repeatRows =1)
    t3.setStyle(TableStyle([('TEXTFONT', (0, 0), (-1, -1), 'Times-Bold'),('TEXTCOLOR',(0,0),(-1,-1),black),('ALIGN',(0,0),(-1,-1),'LEFT'),('VALIGN',(0,0),(-1,-1),'TOP'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black)]))

    elements.append(t3)
    elements.append(Spacer(1,8))
    elements.append(Paragraph("<para fontSize=8>Notes:</para>",styles['Normal']))

    try:
        datanotes = project.poNotes.replace('\n', '<br />')
    except:
        datanotes = project.poNotes
    try:
        datanotes =datanotes.encode('utf-8')
    except:
        datanotes = datanotes
    elements.append(Paragraph("<para fontSize=8>{0} </para>".format(datanotes),styles['Normal']))
    elements.append(Spacer(1,40))

    elements.append(Paragraph("<para fontSize=8>With Best Regards </para>",styles['Normal']))
    elements.append(Spacer(1,8))
    elements.append(Paragraph("<para fontSize=8>For BRUDERER PRESSES INDIA PVT LTD.,</para>",styles['Normal']))
    elements.append(Spacer(1,25))
    elements.append(Paragraph("<para fontSize=8>Authorised Signatory.</para>",styles['Normal']))

    def tableHeader(canvas, doc):
        canvas.saveState()
        header_content = svg2rlg(os.path.join(globalSettings.BASE_DIR , 'static_shared','images' , 'Bruderer_Logo_svg.svg'))
        sx=sy=0.17

        P = Paragraph("<para  align='center'><font size='8'>CIN No. U31900KA2001PTC099049<br/> Regd. Office. #17P, Sadaramangala Industrial Area,Whitefield Road,Kadugodi,Bangalore 560 048<br/> Phone : +91 80 2841 1049<br/> e-mail : info.in@bruderer.com  website : www.bruderer.com</font> </para>", styles['Normal'])
        w, h = P.wrap(doc.width, doc.bottomMargin)
        P.drawOn(canvas, doc.leftMargin, h-20)

        print canvas.getPageNumber(),'counttttttttttttttt'

        canvas.restoreState()

    frame1 = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height,  id='header')
    tableHeader = PageTemplate(id='Later', frames=frame1,onPage=tableHeader)
    doc.addPageTemplates([tableHeader])
    doc.build(elements, onLaterPages=tableHeader)

"""Excel Landing detail download"""

def landingDetails(response, project , purchaselist, request,workbook):

    hdFont = Font(size=12,bold=True)
    vertical_align = Alignment(wrap_text=True, horizontal='center',vertical='center')
    right_align = Alignment(wrap_text=True, horizontal='right')
    alphaChars = list(string.ascii_uppercase)
    print purchaselist, "bom objects"
    Sheet1 = workbook.active
    Sheet1.title = 'Landing Details'
    print workbook.sheetnames, ": sheetNames"
    Sheet1['A2'] = "Landing Details"
    Sheet1['A3'] = "Project Name"
    Sheet1["B3"] = project.title
    Sheet1['A4'] = "Comm nr"
    Sheet1['B4'] = project.comm_nr
    Sheet1['A5'] = "PO ref no :"
    Sheet1['B5'] = project.poNumber
    Sheet1['A6'] = "PO Date :"
    Sheet1['B6'] = project.poDate
    Sheet1['A7'] = "Invoice number :"
    Sheet1['B7'] = project.invoiceNumber
    Sheet1['A8'] = "BOE number :"
    Sheet1['B8'] = project.boeRefNumber

    def sheetStyle(cl,font,fill):
        if font == 'True':
            Sheet1[cl].font = hdFont
        if fill == 'True':
            Sheet1[cl].fill = PatternFill(start_color="48dce0", end_color="48dce0", fill_type = "solid")
        Sheet1[cl].alignment = vertical_align

    sheetStyle('A2','True','False')
    sheetStyle('A3','False','False')
    sheetStyle('B3','False','False')
    sheetStyle('A4','False','False')
    sheetStyle('B4','False','False')
    sheetStyle('A5','False','False')
    sheetStyle('B5','False','False')
    sheetStyle('A6','False','False')
    sheetStyle('B6','False','False')
    sheetStyle('A7','False','False')
    sheetStyle('B7','False','False')
    sheetStyle('A8','False','False')
    sheetStyle('B8','False','False')



    Sheet1.append([''])

    Sheet1.append(['','Ex work value in CHF','Value in INR','Factor'])
    for idx,i in enumerate(['','Ex work value in CHF','Value in INR','Factor']):
        cl = str(alphaChars[idx])+'10'
        if cl == 'A10':
            sheetStyle(cl,'False','False')
        else:
            sheetStyle(cl,'True','True')

    Sheet1.row_dimensions[10].height = 35
    addcount = 11

    Sheet1.append(['Ex works price', round(project.invoiceValue,2),round(project.invoiceValue*project.exRate,2)])
    Sheet1.append(['Packing',round(project.packing,2),round(project.packing*project.exRate,2),round((project.packing*100 / project.invoiceValue),2)])
    Sheet1.append(['Insurance',round(project.insurance,2),round(project.insurance*project.exRate,2),round((project.insurance*100 / project.invoiceValue),2)])
    Sheet1.append(['Freight',round(project.freight,2),round(project.freight*project.exRate,2),round((project.freight*100 / project.invoiceValue),2)])
    Sheet1.append(['Assessable Value','','',round(project.assessableValue,2)])
    Sheet1.append(['GST','','',18])
    Sheet1.append(['GST','','',28])
    Sheet1.append(['Clearing Charges 1','',round(project.clearingCharges1,2),round((project.clearingCharges1*100/(project.invoiceValue* project.exRate)),2)])
    Sheet1.append(['Clearing Charges 2','',round(project.clearingCharges2,2),round((project.clearingCharges2*100/(project.invoiceValue* project.exRate)),2)])
    Sheet1.append(['Profit margin ','','',round(project.profitMargin,2)])
    Sheet1.append([''])
    for i in range(10):
        Sheet1['B'+str(addcount)].number_format = '* #,##0.00'
        Sheet1['C'+str(addcount)].number_format = '* #,##0.00'
        Sheet1['D'+str(addcount)].number_format = '* #,##0.00'
        for idx in range(4):
            cl = str(alphaChars[idx])+str(addcount)
            sheetStyle(cl,'False','False')
        addcount = addcount+1
    Sheet1.append(['Ex Rate',project.exRate])
    Sheet1.append([''])
    for idx in range(2):
        cl = str(alphaChars[idx])+'22'
        sheetStyle(cl,'True','False')

    listArr = ["S.No",'Part No','Products','Weight',"Price({0})".format(project.currency),'Qty','HSN','QP({0})'.format(project.currency),'Ex work price(INR)','Packing','Insurance','Freight','CIF/Pc','Total CIF','CD %','CD Value','SWE','GST(%)', 'GST Value','CC 1','CC 2', 'Landing/Pc']
    Sheet1.append(["S.No",'Part No','Products','Weight',"Price({0})".format(project.currency),'Qty','HSN','QP({0})'.format(project.currency),'Ex work price(INR)','Packing','Insurance','Freight','CIF/Pc','Total CIF','CD %','CD Value','SWE','GST(%)', 'GST Value','CC 1','CC 2', 'Landing/Pc'])
    count = 24

    for idx,i in enumerate(listArr):
        cl = str(alphaChars[idx])+'24'
        sheetStyle(cl,'True','True')
    id = 0
    totprice = 0
    totquote = 0
    totinr = 0
    totpack = 0
    inspack = 0
    frepack = 0
    landtot = 0
    ciftot = 0
    cifroundtot = 0
    totcustomVal = 0
    totsocialVal = 0
    totgstVal = 0
    totcharge1Val = 0
    totcharge2Val = 0
    count = 25
    row = []
    for i in purchaselist:
        id+=1
        print i.price, 'iproce'
        pricetot = i.price*i.quantity1
        totprice+=pricetot
        quotePrice = ((project.profitMargin * i.price) / 100 + i.price)
        quotetot = quotePrice*i.quantity1
        totquote+=quotetot
        inrPrice = quotePrice * project.exRate
        inrtot = inrPrice*i.quantity1
        totinr+=inrtot
        packing = ((project.packing/ project.invoiceValue) * inrPrice)
        packingtot = packing*i.quantity1
        totpack+=packingtot
        insurance = (((project.insurance  / project.invoiceValue) * inrPrice))
        insurancetot = insurance*i.quantity1
        inspack+=insurancetot
        freight = (((project.freight / project.invoiceValue) * inrPrice))
        freighttot = freight*i.quantity1
        frepack+=freighttot
        cif = (inrPrice + packing + insurance + freight)
        ciftot+=cif
        cifround = cif * i.quantity1
        cifroundtot+=cifround
        customVal = ((cif +((cif * project.assessableValue)/100))*(i.custom)/100)
        customtot = customVal * i.quantity1
        totcustomVal +=customtot
        socialVal = (customVal *0.1)
        socialtot = socialVal * i.quantity1
        totsocialVal +=socialtot
        gstVal = ((cif+customVal+socialVal)*(i.gst)/100)
        gsttot = gstVal * i.quantity1
        totgstVal +=gsttot
        charge1 = (inrPrice * (project.clearingCharges1 / (project.invoiceValue * project.exRate)))
        charge1tot = charge1 * i.quantity1
        totcharge1Val +=charge1tot
        charge2 = (inrPrice * (project.clearingCharges2 / (project.invoiceValue * project.exRate)))
        charge2tot = charge2 * i.quantity1
        totcharge2Val +=charge2tot
        landng=i.landed_price*i.quantity1
        landtot+=landng
        row = [id,i.products.part_no,i.products.description_1,i.products.weight,i.price,i.quantity1,i.products.customs_no,round(quotePrice,2), round(inrPrice,2),round(packing,2),round(insurance,2),round(freight,2),round(cif,2),round(cif*i.quantity1,2),round(i.custom,2),round(customVal,2),round(socialVal,2),i.gst, round(gstVal,2),round(charge1,2),round(charge2,2), i.landed_price]
        Sheet1.append([id,i.products.part_no,i.products.description_1,i.products.weight,i.price,i.quantity1,i.products.customs_no,quotePrice,inrPrice,packing,insurance,freight,cif,cif*i.quantity1,i.custom,customVal,socialVal,i.gst,gstVal,charge1,charge2,i.landed_price ])
        Sheet1['E'+str(count)].number_format = '* #,##0.00'
        Sheet1['F'+str(count)].number_format = '* #,##0'
        Sheet1['H'+str(count)].number_format = '* #,##0.00'
        Sheet1['I'+str(count)].number_format = '* #,##0.00'
        Sheet1['J'+str(count)].number_format = '* #,##0.00'
        Sheet1['K'+str(count)].number_format = '* #,##0.00'
        Sheet1['L'+str(count)].number_format = '* #,##0.00'
        Sheet1['M'+str(count)].number_format = '* #,##0.00'
        Sheet1['N'+str(count)].number_format = '* #,##0.00'
        Sheet1['P'+str(count)].number_format = '* #,##0.00'
        Sheet1['Q'+str(count)].number_format = '* #,##0.00'
        Sheet1['S'+str(count)].number_format = '* #,##0.00'
        Sheet1['T'+str(count)].number_format = '* #,##0.00'
        Sheet1['U'+str(count)].number_format = '* #,##0.00'
        Sheet1['V'+str(count)].number_format = '* #,##0.00'
        for idx,i in enumerate(row):
            cl = str(alphaChars[idx])+str(count)
            sheetStyle(cl,'False','False')
        count +=1
    Sheet1.append(["Total",'','','',round(totprice,2),'','',round(totquote,2),round(totinr,2),round(totpack,2),round(inspack,2),round(frepack,2),round(ciftot,2),round(cifroundtot,2),'',round(totcustomVal,2),round(totsocialVal,2),'',round(totgstVal,2),round(totcharge1Val,2),round(totcharge2Val,2),round(landtot,2)])
    Sheet1['E'+str(count)].number_format = '* #,##0.00'
    Sheet1['H'+str(count)].number_format = '* #,##0.00'
    Sheet1['I'+str(count)].number_format = '* #,##0.00'
    Sheet1['J'+str(count)].number_format = '* #,##0.00'
    Sheet1['K'+str(count)].number_format = '* #,##0.00'
    Sheet1['L'+str(count)].number_format = '* #,##0.00'
    Sheet1['M'+str(count)].number_format = '* #,##0.00'
    Sheet1['N'+str(count)].number_format = '* #,##0.00'
    Sheet1['P'+str(count)].number_format = '* #,##0.00'
    Sheet1['Q'+str(count)].number_format = '* #,##0.00'
    Sheet1['S'+str(count)].number_format = '* #,##0.00'
    Sheet1['T'+str(count)].number_format = '* #,##0.00'
    Sheet1['U'+str(count)].number_format = '* #,##0.00'
    Sheet1['V'+str(count)].number_format = '* #,##0.00'
    for idx,i in enumerate(row):
        cl = str(alphaChars[idx])+str(count)
        sheetStyle(cl,'False','False')
    Sheet1.column_dimensions['A'].width = 20
    Sheet1.column_dimensions['B'].width = 20
    Sheet1.column_dimensions['C'].width = 40
    Sheet1.column_dimensions['D'].width = 20
    Sheet1.column_dimensions['E'].width = 20
    Sheet1.column_dimensions['F'].width = 20
    Sheet1.column_dimensions['G'].width = 20
    Sheet1.column_dimensions['H'].width = 20
    Sheet1.column_dimensions['I'].width = 20
    Sheet1.column_dimensions['J'].width = 20
    Sheet1.column_dimensions['K'].width = 20
    Sheet1.column_dimensions['L'].width = 20
    Sheet1.column_dimensions['M'].width = 20
    Sheet1.column_dimensions['N'].width = 20
    Sheet1.column_dimensions['O'].width = 20
    Sheet1.column_dimensions['P'].width = 20
    Sheet1.column_dimensions['Q'].width = 20
    Sheet1.column_dimensions['R'].width = 20
    Sheet1.column_dimensions['S'].width = 20
    Sheet1.column_dimensions['T'].width = 20
    Sheet1.column_dimensions['U'].width = 20
    Sheet1.column_dimensions['V'].width = 20
    response = HttpResponse(content=save_virtual_workbook(workbook),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment;filename="Landingdownload.xlsx"'
    return response


def landingDetailsPdf(response , project , purchaselist, request):
    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(response,pagesize=A2, topMargin=50,leftMargin=1.2*cm,rightMargin=1.2*cm)
    doc.request = request
    elements = []
    elements.append(Spacer(1,30))
    logo = svg2rlg(os.path.join(globalSettings.BASE_DIR , 'static_shared','images' , 'Bruderer_Logo_svg.svg'))
    sx=sy=0.17
    elements.append(Spacer(1,10))
    drawing = svg2rlg(os.path.join(globalSettings.BASE_DIR , 'static_shared','images' , 'anchor_icon.svg'))
    headerDetails = Paragraph("""
    <para>
    <font size ='14'>
    <b> LANDING DETAILS</b><br/>
    </font></para>
    """ %(),styles['Normal'])
    tdheader=[[headerDetails]]
    theader=Table(tdheader,colWidths=[3*inch])
    theader.hAlign = 'LEFT'
    elements.append(theader)
    elements.append(Spacer(1,10))
    details = Paragraph("""
    <para >
    <font size='10'>
    <b>Project Name - </b> %s<br/><br/>
    <b>Comm nr - </b> %s<br/><br/>
    <b>PO ref no - </b> %s <b>Date - </b> %s<br/><br/>
    <b>Invoice ref no - </b> %s <br/><br/>
    <b>BOE ref no - </b> %s </font>
    </para>
    """ %(project.title,project.comm_nr,project.poNumber,project.poDate,project.invoiceNumber,project.boeRefNumber),styles['Normal'])
    tdheader=[[details]]
    theader=Table(tdheader,colWidths=[3*inch])
    theader.hAlign = 'LEFT'
    elements.append(theader)
    elements.append(Spacer(1,10))
    data = []
    p1 =Paragraph("<para fontSize=8></para>",styles['Normal'])
    p2 =Paragraph("<para fontSize=8>Ex works value in {0}</para>".format(project.currency),styles['Normal'])
    p3 =Paragraph("<para fontSize=8>Value in INR</para>",styles['Normal'])
    p4 =Paragraph("<para fontSize=8>Factor</para>",styles['Normal'])
    data += [[p1,p2,p3,p4]]
    p01 =Paragraph("<para fontSize=8>Ex Works price </para>",styles['Normal'])
    p02 =Paragraph("<para fontSize=8>{0}</para>".format(round(project.invoiceValue,2)),styles['Normal'])
    p03 =Paragraph("<para fontSize=8>{0}</para>".format(round(project.invoiceValue*project.exRate,2)),styles['Normal'])
    p04 =Paragraph("<para fontSize=8></para>",styles['Normal'])
    data += [[p01,p02,p03,p04]]
    p11 =Paragraph("<para fontSize=8>Packing </para>",styles['Normal'])
    p12 =Paragraph("<para fontSize=8>{0}</para>".format(round(project.packing,2)),styles['Normal'])
    p13 =Paragraph("<para fontSize=8>{0}</para>".format(round(project.packing*project.exRate,2)),styles['Normal'])
    p14 =Paragraph("<para fontSize=8>{0}%</para>".format(round((project.packing*100 / project.invoiceValue),2)),styles['Normal'])
    data += [[p11,p12,p13,p14]]
    p21 =Paragraph("<para fontSize=8>Insurance </para>",styles['Normal'])
    p22 =Paragraph("<para fontSize=8>{0}</para>".format(round(project.insurance,2)),styles['Normal'])
    p23 =Paragraph("<para fontSize=8>{0}</para>".format(round(project.insurance*project.exRate,2)),styles['Normal'])
    p24 =Paragraph("<para fontSize=8>{0}%</para>".format(round((project.insurance*100 / project.invoiceValue),2)),styles['Normal'])
    data += [[p21,p22,p23,p24]]
    p31 =Paragraph("<para fontSize=8>Freight </para>",styles['Normal'])
    p32 =Paragraph("<para fontSize=8>{0}</para>".format(round(project.freight,2)),styles['Normal'])
    p33 =Paragraph("<para fontSize=8>{0}</para>".format(round(project.freight*project.exRate,2)),styles['Normal'])
    p34 =Paragraph("<para fontSize=8>{0}%</para>".format(round((project.freight*100 / project.invoiceValue),2)),styles['Normal'])
    data += [[p31,p32,p33,p34]]
    p41 =Paragraph("<para fontSize=8>Assessable Value </para>",styles['Normal'])
    p42 =Paragraph("<para fontSize=8></para>",styles['Normal'])
    p43 =Paragraph("<para fontSize=8></para>",styles['Normal'])
    p44 =Paragraph("<para fontSize=8>{0}%</para>".format(round(project.assessableValue,2)),styles['Normal'])
    data += [[p41,p42,p43,p44]]

    p71 =Paragraph("<para fontSize=8>Clearing charges - 1 </para>",styles['Normal'])
    p72 =Paragraph("<para fontSize=8></para>",styles['Normal'])
    p73 =Paragraph("<para fontSize=8> {:,}</para>".format(round(project.clearingCharges1,2)),styles['Normal'])
    p74 =Paragraph("<para fontSize=8>{0}%</para>".format(round((project.clearingCharges1*100/(project.invoiceValue* project.exRate)),2)),styles['Normal'])
    data += [[p71,p72,p73,p74]]
    p81 =Paragraph("<para fontSize=8>Clearing charges - 2 </para>",styles['Normal'])
    p82 =Paragraph("<para fontSize=8></para>",styles['Normal'])
    p83 =Paragraph("<para fontSize=8> {:,}</para>".format(round(project.clearingCharges2,2)),styles['Normal'])
    p84 =Paragraph("<para fontSize=8>{0}%</para>".format(round((project.clearingCharges2*100/(project.invoiceValue* project.exRate)),2)),styles['Normal'])
    data += [[p81,p82,p83,p84]]
    p91 =Paragraph("<para fontSize=8>Profit margin </para>",styles['Normal'])
    p92 =Paragraph("<para fontSize=8></para>",styles['Normal'])
    p93 =Paragraph("<para fontSize=8></para>",styles['Normal'])
    p94 =Paragraph("<para fontSize=8>{0}%</para>".format(round(project.profitMargin,2)),styles['Normal'])
    data += [[p91,p92,p93,p94]]
    t=Table(data,colWidths=(40*mm,35*mm,  35*mm, 35*mm))
    t.hAlign = 'LEFT'
    t.setStyle(TableStyle([('TEXTFONT', (0, 0), (-1, -1), 'Times-Bold'),('TEXTCOLOR',(0,0),(-1,-1),black),('ALIGN',(0,0),(-1,-1),'RIGHT'),('VALIGN',(0,0),(-1,-1),'TOP'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black)]))
    elements.append(t)
    elements.append(Spacer(1,30))
    print project.exRate,'aaaaaa'
    p101 =Paragraph("<para fontSize=8>Ex rate  </para>",styles['Normal'])
    p102 =Paragraph("<para fontSize=8>{0}</para>".format(project.exRate),styles['Normal'])
    data1 = [[p101,p102]]
    t1=Table(data1,colWidths=(40*mm,30*mm))
    t1.hAlign = 'LEFT'
    t1.setStyle(TableStyle([('TEXTFONT', (0, 0), (-1, -1), 'Times-Bold'),('TEXTCOLOR',(0,0),(-1,-1),black),('ALIGN',(0,0),(-1,-1),'RIGHT'),('VALIGN',(0,0),(-1,-1),'TOP'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black)]))
    elements.append(t1)
    elements.append(Spacer(1,30))
    data2=[]
    s01 =Paragraph("<para fontSize=8>S.No </para>",styles['Normal'])
    s02 =Paragraph("<para fontSize=8>Part No </para>",styles['Normal'])
    s03 =Paragraph("<para fontSize=8>Products </para>",styles['Normal'])
    s04 =Paragraph("<para fontSize=8>Weight </para>",styles['Normal'])
    s05 =Paragraph("<para fontSize=8>Price({0}) </para>".format(project.currency),styles['Normal'])
    s06 =Paragraph("<para fontSize=8>Qty </para>",styles['Normal'])
    s07 =Paragraph("<para fontSize=8>HSN </para>",styles['Normal'])
    s08 =Paragraph("<para fontSize=8>QP({0}) </para>".format(project.currency),styles['Normal'])
    s09 =Paragraph("<para fontSize=8>X-Work Price(INR) </para>",styles['Normal'])
    s10 =Paragraph("<para fontSize=8>Packing </para>",styles['Normal'])
    s11 =Paragraph("<para fontSize=8>Insurance </para>",styles['Normal'])
    s12 =Paragraph("<para fontSize=8>Freight </para>",styles['Normal'])
    s13 =Paragraph("<para fontSize=8>CIF/Pc </para>",styles['Normal'])
    s14 =Paragraph("<para fontSize=8>Total CIF </para>",styles['Normal'])
    s1c5 =Paragraph("<para fontSize=8>CD % </para>",styles['Normal'])
    s15 =Paragraph("<para fontSize=8>CD Value </para>",styles['Normal'])
    s16 =Paragraph("<para fontSize=8>SWE </para>",styles['Normal'])
    s1g7 =Paragraph("<para fontSize=8>GST% </para>",styles['Normal'])
    s17 =Paragraph("<para fontSize=8>GST </para>",styles['Normal'])
    s18 =Paragraph("<para fontSize=8>CC 1 </para>",styles['Normal'])
    s19 =Paragraph("<para fontSize=8>CC 2 </para>",styles['Normal'])
    s20 =Paragraph("<para fontSize=8>Landing/Pc </para>",styles['Normal'])
    data2 += [[s01,s02,s03,s04,s05,s06,s07,s08,s09,s10,s11,s12,s13,s14,s1c5,s15,s16,s1g7,s17,s18,s19,s20]]
    id = 0
    totprice = 0
    totquote = 0
    totinr = 0
    totpack = 0
    inspack = 0
    frepack = 0
    landtot = 0
    ciftot = 0
    cifroundtot = 0
    totcustomVal = 0
    totsocialVal = 0
    totgstVal = 0
    totcharge1Val = 0
    totcharge2Val = 0
    for i in purchaselist:
        id+=1
        pricetot = i.price*i.quantity1
        totprice+=pricetot
        quotePrice = round(((project.profitMargin * i.price) / 100 + i.price),2)
        quotetot = quotePrice*i.quantity1
        totquote+=quotetot
        inrPrice = round((quotePrice * project.exRate),2)
        inrtot = inrPrice*i.quantity1
        totinr+=inrtot
        packing = round(((project.packing/ project.invoiceValue) * inrPrice),2)
        packingtot = packing*i.quantity1
        totpack+=packingtot
        insurance = round((((project.insurance  / project.invoiceValue) * inrPrice)),2)
        insurancetot = insurance*i.quantity1
        inspack+=insurancetot
        freight = round((((project.freight / project.invoiceValue) * inrPrice)),2)
        freighttot = freight*i.quantity1
        frepack+=freighttot
        cif = round((inrPrice + packing + insurance + freight),2)
        ciftot+=cif
        cifround = cif * i.quantity1
        cifroundtot+=cifround
        customVal = round(((cif +((cif * project.assessableValue)/100))*(i.custom)/100),2)
        customtot = customVal * i.quantity1
        totcustomVal +=customtot
        socialVal = round((customVal *0.1),2)
        socialtot = socialVal * i.quantity1
        totsocialVal +=socialtot
        gstVal = round(((cif+customVal+socialVal)*(i.gst)/100),2)
        gsttot = gstVal * i.quantity1
        totgstVal +=gsttot
        charge1 = round((inrPrice * (project.clearingCharges1 / (project.invoiceValue * project.exRate))),2)
        charge1tot = charge1 * i.quantity1
        totcharge1Val +=charge1tot
        charge2 = round((inrPrice * (project.clearingCharges2 / (project.invoiceValue * project.exRate))),2)
        charge2tot = charge2 * i.quantity1
        totcharge2Val +=charge2tot
        landng=i.landed_price*i.quantity1
        landtot+=landng
        s21 =Paragraph("<para fontSize=8>{0} </para>".format(id),styles['Normal'])
        s22 =Paragraph("<para fontSize=8>{0} </para>".format(i.products.part_no),styles['Normal'])
        s23 =Paragraph("<para fontSize=8>{0} </para>".format(smart_str(i.products.description_1)),styles['BodyText'])
        s24 =Paragraph("<para fontSize=8>{0} </para>".format(i.products.weight),styles['Normal'])
        s25 =Paragraph("<para fontSize=8> {:,}</para>".format(round(i.price,2)),styles['Normal'])
        s26 =Paragraph("<para fontSize=8>{0} </para>".format(i.quantity1),styles['Normal'])
        s27 =Paragraph("<para fontSize=8>{0} </para>".format(i.products.customs_no),styles['Normal'])
        s28 =Paragraph("<para fontSize=8> {:,} </para>".format(round(quotePrice,2)),styles['Normal'])
        s29 =Paragraph("<para fontSize=8> {:,} </para>".format(round(inrPrice,2)),styles['Normal'])
        s30 =Paragraph("<para fontSize=8> {:,} </para>".format(round(packing,2)),styles['Normal'])
        s31 =Paragraph("<para fontSize=8> {:,}</para>".format(round(insurance,2)),styles['Normal'])
        s32 =Paragraph("<para fontSize=8> {:,} </para>".format(round(freight,2)),styles['Normal'])
        s33 =Paragraph("<para fontSize=8> {:,}</para>".format(round(cif,2)),styles['Normal'])
        s34 =Paragraph("<para fontSize=8> {:,} </para>".format(round(cif*i.quantity1,2)),styles['Normal'])
        s3c5 =Paragraph("<para fontSize=8>{0} </para>".format(round(i.custom,2)),styles['Normal'])
        s35 =Paragraph("<para fontSize=8> {:,} </para>".format(round(customVal,2)),styles['Normal'])
        s36 =Paragraph("<para fontSize=8> {:,} </para>".format(round(socialVal,2)),styles['Normal'])
        s3g7 =Paragraph("<para fontSize=8>{0} </para>".format(i.gst),styles['Normal'])
        s37 =Paragraph("<para fontSize=8> {:,} </para>".format(round(gstVal,2)),styles['Normal'])
        s38 =Paragraph("<para fontSize=8> {:,}</para>".format(round(charge1,2)),styles['Normal'])
        s39 =Paragraph("<para fontSize=8> {:,} </para>".format(round(charge2,2)),styles['Normal'])
        s40 =Paragraph("<para fontSize=8> {:,} </para>".format(i.landed_price),styles['Normal'])
        data2.append([s21,s22,s23,s24,s25,s26,s27,s28,s29,s30,s31,s32,s33,s34,s3c5,s35,s36,s3g7,s37,s38,s39,s40])
    s21 =Paragraph("<para fontSize=8> </para>",styles['Normal'])
    s22 =Paragraph("<para fontSize=8> </para>",styles['Normal'])
    s23 =Paragraph("<para fontSize=8> </para>",styles['Normal'])
    s24 =Paragraph("<para fontSize=8> </para>",styles['Normal'])
    s25 =Paragraph("<para fontSize=8>{:,}</para>".format(round(totprice,2)),styles['Normal'])
    s26 =Paragraph("<para fontSize=8> </para>",styles['Normal'])
    s27 =Paragraph("<para fontSize=8> </para>",styles['Normal'])
    s28 =Paragraph("<para fontSize=8> {:,} </para>".format(round(totquote,2)),styles['Normal'])
    s29 =Paragraph("<para fontSize=8> {:,}  </para>".format(round(totinr,2)),styles['Normal'])
    s30 =Paragraph("<para fontSize=8> {:,} </para>".format(round(totpack,2)),styles['Normal'])
    s31 =Paragraph("<para fontSize=8> {:,}</para>".format(round(inspack,2)),styles['Normal'])
    s32 =Paragraph("<para fontSize=8> {:,} </para>".format(round(frepack,2)),styles['Normal'])
    s33 =Paragraph("<para fontSize=8> {:,} </para>".format(round(ciftot,2)),styles['Normal'])
    s34 =Paragraph("<para fontSize=8> {:,} </para>".format(round(cifroundtot,2)),styles['Normal'])
    s3c5 =Paragraph("<para fontSize=8> </para>",styles['Normal'])
    s35 =Paragraph("<para fontSize=8> {:,}  </para>".format(round(totcustomVal,2)),styles['Normal'])
    s36 =Paragraph("<para fontSize=8> {:,} </para>".format(round(totsocialVal,2)),styles['Normal'])
    s3g7 =Paragraph("<para fontSize=8></para>",styles['Normal'])
    s37 =Paragraph("<para fontSize=8>{:,}</para>".format(round(totgstVal,2)),styles['Normal'])
    s38 =Paragraph("<para fontSize=8>{:,}</para>".format(round(totcharge1Val,2)),styles['Normal'])
    s39 =Paragraph("<para fontSize=8> {:,}</para>".format(round(totcharge2Val,2)),styles['Normal'])
    s40 =Paragraph("<para fontSize=8> {:,} </para>".format(round(landtot,2)),styles['Normal'])
    data2.append([s21,s22,s23,s24,s25,s26,s27,s28,s29,s30,s31,s32,s33,s34,s3c5,s35,s36,s3g7,s37,s38,s39,s40])
    t2=Table(data2,colWidths=(11*mm,23*mm, 35*mm, 20*mm, 20*mm, 8*mm, 20*mm,20*mm,20*mm,15*mm,15*mm,15*mm,20*mm,20*mm,10*mm,20*mm,15*mm,10*mm,20*mm,15*mm,15*mm,20*mm),repeatRows=1)
    t2.hAlign = 'LEFT'
    t2.setStyle(TableStyle([('TEXTFONT', (0, 0), (-1, -1), 'Times-Bold'),('TEXTCOLOR',(0,0),(-1,-1),black),('ALIGN',(0,0),(-1,-1),'RIGHT'),('VALIGN',(0,0),(-1,-1),'TOP'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black)]))
    elements.append(t2)

    def tableHeader(canvas, doc):
        canvas.saveState()
        canvas.restoreState()

    frame = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height,  id='normal')
    template = PageTemplate(id='footer', frames=frame, onPage=tableHeader)
    doc.addPageTemplates([template])
    doc.build(elements,onLaterPages=tableHeader)
    return response


def quotation(response , project , purchaselist , multNumber,typ,request):
    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(response,pagesize=letter, topMargin=70,leftMargin=1.2*cm,rightMargin=1.2*cm)
    width, height = letter
    doc.request = request
    elements = []
    style_right = ParagraphStyle(name='right', parent=styles['Normal'], alignment=TA_RIGHT)

    if typ == 'Invoice':
        summryHeader = Paragraph("""
        <para >
        <font size='14'>
        INVOICE
        <br/>
        <br/>
        </font>
        </para>
        """ %(),styles['Normal'])
    else :
        summryHeader = Paragraph("""
        <para >
        <font size='14'>
        QUOTATION
        <br/>
        <br/>
        </font>
        </para>
        """ %(),styles['Normal'])

    summryHeader1 = Paragraph("""
    <para leftIndent = 10>
    <font size ='10'>
    <b>Quote Ref No :</b> %s <br/>
    <b>Quote Ref Date :</b> %s <br/>
    </font></para>
    """ %(project.quoteRefNumber , project.quoteDate),styles['Normal'])
    tdheader=[[summryHeader,' ',summryHeader1]]
    theader=Table(tdheader,colWidths=[3*inch , 1*inch , 3*inch])
    theader.hAlign = 'LEFT'
    elements.append(theader)
    summryParaSrc = Paragraph("""
    <para >
    <font size='10'>
    <b>%s</b> <br/>
    %s <br/>
    %s - %s<br/>
    %s <br/>
    %s<br/>
    </font>
    </para>
    """ %(project.service.name,project.service.address.street ,project.service.address.city,project.service.address.pincode,project.service.address.state , project.service.address.country),styles['Normal'])
    summryParaSrc1 = Paragraph("""
    <para leftIndent = 10>
    <font size ='10'>
    <b>Kind attn :</b> %s <br/>
    <b>Contact no :</b> %s <br/>
    <b>E-Mail ID :</b> %s <br/>
    <b>Your Enquiry Ref :</b> %s <br/><br/>
    </font></para>
    """
    %(project.service.name , project.service.mobile , project.service.about,project.enquiry_ref),styles['Normal'])
    td=[[summryParaSrc,' ',summryParaSrc1]]
    t=Table(td,colWidths=[3*inch , 1*inch , 3*inch])
    t.hAlign = 'LEFT'
    elements.append(t)
    summrymachineDetails = Paragraph("""
    <para >
    <font size='10'>
    <b>Machine Type :</b> %s <br/>
    <b>Comm Nr :</b> %s
    </font>
    </para>
    """
     %(project.machinemodel, project.comm_nr),styles['Normal'])
    tmachine=[[summrymachineDetails,' ','']]
    tdmachine=Table(tmachine,colWidths=[3*inch , 1*inch , 3*inch])
    tdmachine.hAlign = 'LEFT'
    elements.append(tdmachine)
    elements.append(Spacer(1,10))
    if typ != 'Invoice':
        elements.append(Paragraph("<para fontSize=8>Dear Sir,</para>",styles['Normal']))
        elements.append(Spacer(1,10))
        elements.append(Paragraph("<para fontSize=8>We thank you for your enquiry and take pleasure in quoting as follows:</para>",styles['Normal']))
    if typ == 'INR'or typ == 'Invoice':
        data5 = []
        priceFormat = 'Unit price in INR'
        amountFormat = 'Amount in INR'
        p101_01 =Paragraph("<para fontSize=8>Sl. no</para>",styles['Normal'])
        p101_02 =Paragraph("<para fontSize=8>Part Number</para>",styles['Normal'])
        p101_03 =Paragraph("<para fontSize=8>Part Desc</para>",styles['Normal'])
        p101_09 =Paragraph("<para fontSize=8>HSN</para>",styles['Normal'])
        p101_04 =Paragraph("<para fontSize=8>Qty</para>",styles['Normal'])
        p101_05 =Paragraph("<para fontSize=8>{0}</para>".format(priceFormat),styles['Normal'])
        p101_06 =Paragraph("<para fontSize=8>{0}</para>".format(amountFormat),styles['Normal'])
        p101_07 =Paragraph("<para fontSize=8>GST</para>",styles['Normal'])
        p101_08 =Paragraph("<para fontSize=8>Total With GST</para>",styles['Normal'])

        data5+=[[p101_01,p101_02,p101_03,p101_09,p101_04,p101_05,p101_06,p101_07,p101_08]]

        gstValTotal = 0
        grandTotal = 0
        data2 = []
        id=0
        grandtotWeight=0
        for i in purchaselist:
            id+=1
            part_no = i.products.part_no
            desc = i.products.description_1
            custom_no = i.products.customs_no
            weight = i.products.weight
            basicprice = i.price
            landingPrice = i.landed_price
            pricesum =round((i.landed_price+(i.price*multNumber*project.profitMargin)/100),2)
            qty = i.quantity1
            amnt = round((pricesum * qty),2)
            grandTotal +=amnt
            gst = i.gst
            gstVal = round((((amnt *gst)/100)+amnt),2)
            if weight is not None:
                totweight = round((i.products.weight * qty),2)
                grandtotWeight+=totweight
            gstValTotal += gstVal
            p12_01 = Paragraph("<para fontSize=8>{0}</para>".format(id),styles['Normal'])
            p12_02 =Paragraph("<para fontSize=8>{0}</para>".format(part_no),styles['Normal'])
            p12_03 =Paragraph("<para fontSize=8>{0}</para>".format(smart_str(desc)),styles['Normal'])
            p12_09 =Paragraph("<para fontSize=8>{0}</para>".format(custom_no),styles['Normal'])
            p12_04 =Paragraph("<para fontSize=8>{0}</para>".format(qty),styles['Normal'])
            p12_05 =Paragraph("<para fontSize=8>{:,}</para>".format(round(pricesum,2)),style_right)
            p12_06 =Paragraph("<para fontSize=8>{:,}</para>".format(round(amnt,2)),style_right)
            p12_07 =Paragraph("<para fontSize=8>{0}%</para>".format(gst),styles['Normal'])
            p12_08 =Paragraph("<para fontSize=8>{:,}</para>".format(round(gstVal,2)),style_right)
            data5.append([p12_01,p12_02,p12_03,p12_09,p12_04,p12_05,p12_06,p12_07,p12_08])
        grandTotal = round(grandTotal,2)
        gstValTotal = round(gstValTotal,2)
        p13_01 =Paragraph("<para fontSize=8>{0}</para>".format(''),styles['Normal'])
        p13_02 =Paragraph("<para fontSize=8>{0}</para>".format(''),styles['Normal'])
        p13_03 =Paragraph("<para fontSize=8></para>",styles['Normal'])
        p13_09 =Paragraph("<para fontSize=8></para>",styles['Normal'])
        p13_04 =Paragraph("<para fontSize=8>{0}</para>".format(''),styles['Normal'])
        p13_05 =Paragraph("<para fontSize=8>Total in INR</para>",styles['Normal'])
        p13_06 =Paragraph("<para fontSize=8>{:,}</para>".format(round(grandTotal,2)),style_right)
        p13_07 =Paragraph("<para fontSize=8></para>",styles['Normal'])
        p13_08 =Paragraph("<para fontSize=8>{:,}</para>".format(round(gstValTotal,2)),style_right)
        data5+=[[p13_01,p13_02,p13_03,p13_09,p13_04,p13_05,p13_06,p13_07,p13_08]]
        t3=Table(data5,colWidths=(8*mm, None,  43*mm,None, 10*mm, None, None, None, None),repeatRows=1)
        t3.setStyle(TableStyle([('TEXTFONT', (0, 0), (-1, -1), 'Times-Bold'),('TEXTCOLOR',(0,0),(-1,-1),black),('ALIGN',(0,0),(-1,-1),'LEFT'),('VALIGN',(0,0),(-1,-1),'TOP'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black)]))
        elements.append(t3)
    else:
        data5 = []
        priceFormat = 'Unit price in ' + project.currency
        amountFormat = 'Amount in ' + project.currency
        p101_01 =Paragraph("<para fontSize=8>Sl. no</para>",styles['Normal'])
        p101_02 =Paragraph("<para fontSize=8>Part Number</para>",styles['Normal'])
        p101_03 =Paragraph("<para fontSize=8>Part Desc</para>",styles['Normal'])
        p101_07 =Paragraph("<para fontSize=8>HSN</para>",styles['Normal'])
        p101_04 =Paragraph("<para fontSize=8>Qty</para>",styles['Normal'])
        p101_05 =Paragraph("<para fontSize=8>{0}</para>".format(priceFormat),styles['Normal'])
        p101_06 =Paragraph("<para fontSize=8>{0}</para>".format(amountFormat),styles['Normal'])
        data5+=[[p101_01,p101_02,p101_03,p101_07,p101_04,p101_05,p101_06]]
        grandTotal = 0
        data2 = []
        id=0
        grandtotWeight = 0
        for i in purchaselist:
            id+=1
            part_no = i.products.part_no
            desc = i.products.description_1
            custom_no =  i.products.customs_no
            weight = i.products.weight
            basicprice = i.price
            pricesum = round((((i.price*project.profitMargin)/100)+i.price),2)
            price = round((pricesum * multNumber),2)
            qty = i.quantity1
            amnt = round((price * qty),2)
            if weight is not None:
                totweight = round((i.products.weight * qty),2)
                grandtotWeight+=totweight
            grandTotal +=amnt
            p12_01 = Paragraph("<para fontSize=8>{0}</para>".format(id),styles['Normal'])
            p12_02 =Paragraph("<para fontSize=8>{0}</para>".format(part_no),styles['Normal'])
            p12_03 =Paragraph("<para fontSize=8>{0}</para>".format(smart_str(desc)),styles['Normal'])
            p12_07 =Paragraph("<para fontSize=8>{0}</para>".format(custom_no),styles['Normal'])
            p12_04 =Paragraph("<para fontSize=8>{0}</para>".format(qty),styles['Normal'])
            p12_05 =Paragraph("<para fontSize=8> {:,}</para>".format(round(price,2)),style_right)
            p12_06 =Paragraph("<para fontSize=8>{:,}</para>".format(amnt),style_right)
            data5.append([p12_01,p12_02,p12_03,p12_07,p12_04,p12_05,p12_06])
        grandTotal = round(grandTotal,2)
        p13_01 =Paragraph("<para fontSize=8>{0}</para>".format(''),styles['Normal'])
        p13_02 =Paragraph("<para fontSize=8>{0}</para>".format(''),styles['Normal'])
        p13_03 =Paragraph("<para fontSize=8></para>",styles['Normal'])
        p13_07 =Paragraph("<para fontSize=8></para>",styles['Normal'])
        p13_04 =Paragraph("<para fontSize=8>{0}</para>".format(''),styles['Normal'])
        p13_05 =Paragraph("<para fontSize=8>Total in {0} </para>".format(project.currency),styles['Normal'])
        p13_06 =Paragraph("<para fontSize=8>{:,}</para>".format(round(grandTotal,2)),style_right)
        data5+=[[p13_01,p13_02,p13_03,p13_07,p13_04,p13_05,p13_06]]
        t3=Table(data5,colWidths=(10*mm,None,  60*mm,None, 10*mm, None, None),repeatRows=1)
        t3.hAlign = 'LEFT'
        t3.setStyle(TableStyle([('TEXTFONT', (0, 0), (-1, -1), 'Times-Bold'),('TEXTCOLOR',(0,0),(-1,-1),black),('ALIGN',(0,0),(-1,-1),'LEFT'),('VALIGN',(0,0),(-1,-1),'TOP'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black)]))
        elements.append(t3)

    elements.append(Spacer(1,16))
    p14_01 =Paragraph("<para fontSize=8>QUOTATION VALIDITY</para>",styles['Normal'])
    p14_02 =Paragraph("<para fontSize=8>INCO TERMS</para>",styles['Normal'])
    p14_03 =Paragraph("<para fontSize=8>DELIVERY</para>",styles['Normal'])
    p14_04 =Paragraph("<para fontSize=8>PAYMENT TERMS</para>",styles['Normal'])
    data6=[[p14_01,p14_02,p14_03,p14_04]]
    p15_01 =Paragraph("<para fontSize=8>{0}</para>".format(project.quoteValidity),styles['Normal'])
    p15_02 =Paragraph("<para fontSize=8>{0}</para>".format(project.terms),styles['Normal'])
    p15_03 =Paragraph("<para fontSize=8>{0}</para>".format(project.delivery),styles['Normal'])
    p15_04 =Paragraph("<para fontSize=8>{0}</para>".format(project.paymentTerms),styles['Normal'])
    data6+=[[p15_01,p15_02,p15_03,p15_04]]
    t6=Table(data6)
    t6.hAlign = 'LEFT'
    t6.setStyle(TableStyle([('TEXTFONT', (0, 0), (-1, -1), 'Times-Bold'),('TEXTCOLOR',(0,0),(-1,-1),black),('ALIGN',(0,0),(-1,-1),'LEFT'),('VALIGN',(0,0),(-1,-1),'TOP'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black)]))
    elements.append(t6)
    elements.append(Spacer(1,8))
    if typ == 'INR':
        p16_02 =Paragraph("<para fontSize=8>BANK DETAILS</para>",styles['Normal'])
        data8=[[p16_02]]
        p17_01 =Paragraph("<para fontSize=8>IDBI Bank Ltd<br/>Whitefield Branch<br/>Bangalore 560 066, Karnataka<br/>Account No. 1545102000003858 <br/>IFSC Code : IBKL0001545</para>",styles['Normal'])
        data8 +=[[p17_01]]
        t9=Table(data8,6*[3*inch])
        t9.hAlign = 'LEFT'
        t9.setStyle(TableStyle([('TEXTFONT', (0, 0), (-1, -1), 'Times-Bold'),('TEXTCOLOR',(0,0),(-1,-1),black),('ALIGN',(0,0),(-1,-1),'LEFT'),('VALIGN',(0,0),(-1,-1),'TOP'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black)]))
        elements.append(t9)
        elements.append(Spacer(1,8))
        p77_01 =Paragraph("<para fontSize=8>PAN NO. </para>",styles['Normal'])
        p77_02 =Paragraph("<para fontSize=8>GST NO. </para>",styles['Normal'])
        data8=[[p77_01,p77_02]]
        pan =Paragraph("<para fontSize=8>AABCB6326Q</para>",styles['Normal'])
        gst =Paragraph("<para fontSize=8>29AABCB6326Q1Z6</para>",styles['Normal'])
        data8 +=[[pan,gst]]
        t9=Table(data8,6*[1.5*inch])
        t9.hAlign = 'LEFT'
        t9.setStyle(TableStyle([('TEXTFONT', (0, 0), (-1, -1), 'Times-Bold'),('TEXTCOLOR',(0,0),(-1,-1),black),('ALIGN',(0,0),(-1,-1),'LEFT'),('VALIGN',(0,0),(-1,-1),'TOP'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black)]))
        elements.append(t9)
        elements.append(Spacer(1,8))
        elements.append(Paragraph("<para fontSize=8>Notes:</para>",styles['Normal']))
        try:
            datanotes = project.quoteNotes.replace('\n', '<br />')
        except:
            datanotes = project.quoteNotes
        try:
            datanotes =datanotes.encode('utf-8')
        except:
            datanotes = datanotes
        elements.append(Paragraph("<para fontSize=8>{0} </para>".format(datanotes),styles['Normal']))

        elements.append(Spacer(1,8))

    else:
        p16_01 =Paragraph("<para fontSize=8>PO to raised on</para>",styles['Normal'])
        p16_02 =Paragraph("<para fontSize=8>BANK DETAILS</para>",styles['Normal'])
        data8=[[p16_01,p16_02]]
        t8=Table(data8,6*[1.4*inch],1*[0.2*inch])
        t8.hAlign = 'LEFT'
        t8.setStyle(TableStyle([('TEXTFONT', (0, 0), (-1, -1), 'Times-Bold'),('TEXTCOLOR',(0,0),(-1,-1),black),('ALIGN',(0,0),(-1,-1),'LEFT'),('VALIGN',(0,0),(-1,-1),'TOP'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black)]))
        elements.append(t8)
        p17_01 =Paragraph("<para fontSize=8><b>BRUDERER AG</b><br/>Stanzautomaten<br/>Egnacherstrasse<br/>9320 FRASNACHT <br/>SWITZERLAND</para>",styles['Normal'])
        p17_02 =Paragraph("<para fontSize=8></para>",styles['Normal'])
        data9=[[p17_01,p17_02]]
        t9=Table(data9,6*[1.4*inch],1*[0.9*inch])
        t9.hAlign = 'LEFT'
        t9.setStyle(TableStyle([('TEXTFONT', (0, 0), (-1, -1), 'Times-Bold'),('TEXTCOLOR',(0,0),(-1,-1),black),('ALIGN',(0,0),(-1,-1),'LEFT'),('VALIGN',(0,0),(-1,-1),'TOP'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black)]))
        elements.append(t9)
        elements.append(Spacer(1,8))
        elements.append(Paragraph("<para fontSize=8>Notes:</para>",styles['Normal']))
        try:
            datanotes = project.quoteNotes.replace('\n', '<br />')
        except:
            datanotes = project.quoteNotes
        try:
            datanotes =datanotes.encode('utf-8')
        except:
            datanotes = datanotes
        elements.append(Paragraph("<para fontSize=8>{0} </para>".format(datanotes),styles['Normal']))

        elements.append(Spacer(1,8))

    if typ != 'Invoice':
        para1 =Paragraph("<para fontSize=8>We hope that this quotation will meet your requirement and would be glad to receive your firm order. For further information please do not hesitate to cotnact us any time. </para>",styles['Normal'])
        para2 = Paragraph("<para fontSize=8>For BRUDERER PRESSES INDIA PVT LTD.,<br/><br/><br/><br/>Authorised Signatory.</para>",styles['Normal'])
        data = [[para1],[para2]]
        t9=Table(data,colWidths=192*mm)
        elements.append(t9)

    def tableHeader(canvas, doc):
        canvas.saveState()
        P = Paragraph("<para  align='center'><font size='8'>CIN No. U31900KA2001PTC099049<br/> Regd. Office. #17P, Sadaramangala Industrial Area,Whitefield Road,Kadugodi,Bangalore 560 048<br/> Phone : +91 80 2841 1049<br/> e-mail : info.in@bruderer.com  website : www.bruderer.com</font> </para>", styles['Normal'])
        w, h = P.wrap(doc.width, doc.bottomMargin)
        P.drawOn(canvas, doc.leftMargin, h-20)
        print canvas.getPageNumber(),'counttttttttttttttt'
        canvas.restoreState()
    frame = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height,  id='normal')
    template = PageTemplate(id='footer', frames=frame, onPage=tableHeader)
    doc.addPageTemplates([template])
    doc.build(elements, onLaterPages=tableHeader)

def grn(response , project , purchaselist , request):
    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(response,pagesize=letter, topMargin=0.2*cm,leftMargin=0.1*cm,rightMargin=0.1*cm)
    doc.request = request
    elements = []
    p1 = Paragraph("<para alignment='center'fontSize=15  ><b> Goods Received Note </b></para>",styles['Normal'])
    elements.append(p1)
    elements.append(Spacer(1, 10))
    addrdetails = Paragraph("""
    <para >
    <b>%s</b><br/>
    %s %s - %s<br/>
    %s - %s<br/>
    </para>
    """ %(project.vendor.name ,project.vendor.street,project.vendor.city,project.vendor.pincode,project.vendor.state,project.vendor.country),styles['Normal'])
    td=[[addrdetails]]
    t=Table(td,colWidths=[4*inch])
    t.hAlign = 'LEFT'
    elements.append(t)
    elements.append(Spacer(1,10))
    details = Paragraph("""
    <para >
    <b>Project Title - </b>%s<br/>
    <b>Comm nr - </b>%s <br/>
    <b>PO ref no - </b>%s<br/>
    <b>GRN Date - </b>%s<br/>
    </para>
    """ %(project.title,project.comm_nr ,project.poNumber,project.grnDate),styles['Normal'])
    td=[[details]]
    t=Table(td,colWidths=[4*inch])
    t.hAlign = 'LEFT'
    elements.append(t)
    p9_01 =Paragraph("<para fontSize=8>Sl. no</para>",styles['Normal'])
    p9_02 =Paragraph("<para fontSize=8>Part Number</para>",styles['Normal'])
    p9_03 =Paragraph("<para fontSize=8>Quantity</para>",styles['Normal'])
    data2=[[p9_01,p9_02,p9_03]]
    t2=Table(data2,3*[2.8*inch],1*[0.2*inch])
    t2.setStyle(TableStyle([('TEXTFONT', (0, 0), (-1, -1), 'Times-Bold'),('TEXTCOLOR',(0,0),(-1,-1),black),('ALIGN',(0,0),(-1,-1),'LEFT'),('VALIGN',(0,0),(-1,-1),'TOP'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black)]))
    data3 = []
    id=0
    for i in purchaselist:
        id+=1
        part_no = i.products.part_no
        quanty = i.quantity2
        p10_01 =Paragraph("<para fontSize=8>{0}</para>".format(id),styles['Normal'])
        p10_02 =Paragraph("<para fontSize=8>{0}</para>".format(part_no),styles['Normal'])
        p10_03 =Paragraph("<para fontSize=8>{0}</para>".format(quanty),styles['Normal'])
        data3.append([p10_01,p10_02,p10_03])
    t3=Table(data3,3*[2.8*inch],id*[0.2*inch])
    t3.setStyle(TableStyle([('TEXTFONT', (0, 0), (-1, -1), 'Times-Bold'),('TEXTCOLOR',(0,0),(-1,-1),black),('ALIGN',(0,0),(-1,-1),'LEFT'),('VALIGN',(0,0),(-1,-1),'TOP'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black)]))
    elements.append(t2)
    elements.append(t3)
    doc.build(elements)

class GetPurchaseAPIView(APIView):
    def get(self , request , format = None):
        project = Projects.objects.get(pk = request.GET['project'])
        purchaselist = BoM.objects.filter(project = request.GET['project'])
        try:
            if 'typ' in request.GET:
                currencyTyp = request.GET['typ']
                if currencyTyp == 'INR':
                    multNumber = float(project.exRate)

                else:
                    multNumber = 1
                    currencyTyp = request.GET['typ']
            else:
                multNumber = 1
                currencyTyp = request.GET['typ']
        except:
            multNumber = 1
            currencyTyp = ''

        print multNumber,'aaaaaaaaaaaaa'
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment;filename="PurchaseOrderdownload.pdf"'
        purchaseOrder(response , project , purchaselist,multNumber,currencyTyp, request)
        return response

class GetLandingAPIView(APIView):
    def get(self , request , format = None):
        project = Projects.objects.get(pk = request.GET['project'])
        purchaselist = BoM.objects.filter(project = request.GET['project'])
        if request.GET['type'] == 'pdf':
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = 'attachment;filename="Landingdownload.pdf"'
            return landingDetailsPdf(response , project , purchaselist, request)
        if request.GET['type'] == 'excel':
            workbook = Workbook()
            response = HttpResponse(content=save_virtual_workbook(workbook),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment;filename="Landingdownload.xlsx"'
            return landingDetails(response,project , purchaselist, request,workbook)

class PurchasewiseReportAPIView(APIView):
    def get(self , request , format = None):
        comm_nr = request.GET['commnr']
        supplierObj = Vendor.objects.get(pk = int(request.GET['supplier']))
        workbook = Workbook()
        Sheet1 = workbook.active
        hdFont = Font(size=12,bold=True)
        vertical_align = Alignment(wrap_text=True, horizontal='center',vertical='center')
        right_align = Alignment(wrap_text=True, horizontal='right')
        alphaChars = list(string.ascii_uppercase)
        supplier_hd = ['Supplier',supplierObj.name,'','','','','']
        Sheet1.append(supplier_hd)
        for idx,i in enumerate(supplier_hd):
            cl = str(alphaChars[idx])+'1'
            Sheet1[cl].font = hdFont
            Sheet1[cl].alignment = vertical_align
        comm_hd = ['Comm_nr',comm_nr,'','','','','']
        Sheet1.append(comm_hd)
        for idx,i in enumerate(comm_hd):
            cl = str(alphaChars[idx])+'2'
            Sheet1[cl].font = hdFont
            Sheet1[cl].alignment = vertical_align
        Sheet1.append(['','','','','','',''])
        hd = ['PO Ref','Part Number','Description','Qty','Price/each','Amount','Invoice']
        Sheet1.append(hd)
        for idx,i in enumerate(hd):
            cl = str(alphaChars[idx])+'4'
            Sheet1[cl].fill = PatternFill(start_color="48dce0", end_color="48dce0", fill_type = "solid")
            Sheet1[cl].font = hdFont
            Sheet1[cl].alignment = vertical_align

        bomObjs = BoM.objects.filter(project__comm_nr = comm_nr,project__vendor = int(supplierObj.pk),project__flag = request.GET['flag'],project__status__exact='ongoing')

        count = 5
        if len(bomObjs)>0:
            for idx,i in enumerate(bomObjs):
                Sheet1.append([i.project.title,i.products.part_no,i.products.description_1,i.quantity2,i.price,i.quantity2*i.price,i.project.invoiceNumber])
                for idx,i in enumerate(hd):
                    cl = str(alphaChars[idx])+str(count)
                    Sheet1[cl].alignment = vertical_align
                Sheet1['D'+str(count)].number_format = '* #,##0'
                Sheet1['E'+str(count)].number_format = '* #,##0'
                Sheet1['F'+str(count)].number_format = '* #,##0'
                count =count+1
        else:
            Sheet1.append(['-  ','-  ','-  ','-  ','-  ','-  ','-  '])
            for idx,i in enumerate(hd):
                cl = str(alphaChars[idx])+'5'
                Sheet1[cl].alignment = vertical_align
                Sheet1[cl].font = hdFont

        Sheet1.column_dimensions['A'].width = 20
        Sheet1.column_dimensions['B'].width = 20
        Sheet1.column_dimensions['C'].width = 40
        Sheet1.column_dimensions['D'].width = 15
        Sheet1.column_dimensions['E'].width = 15
        Sheet1.column_dimensions['F'].width = 15
        Sheet1.column_dimensions['G'].width = 15
        response = HttpResponse(content=save_virtual_workbook(workbook),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=purchasewisereport.xlsx'
        return response


class ConsumptionewiseReportAPIView(APIView):
    def get(self , request , format = None):
        comm_nr = request.GET['commnr']
        supplierFlag = request.GET['all']
        workbook = Workbook()
        Sheet1 = workbook.active
        hdFont = Font(size=12,bold=True)
        vertical_align = Alignment(wrap_text=True, horizontal='center',vertical='center')
        right_align = Alignment(wrap_text=True, horizontal='right')
        alphaChars = list(string.ascii_uppercase)
        comm_hd = ['Comm_nr',comm_nr,'','','','','','','']
        Sheet1.append(comm_hd)
        for idx,i in enumerate(comm_hd):
            cl = str(alphaChars[idx])+'1'
            Sheet1[cl].font = hdFont
            Sheet1[cl].alignment = vertical_align
        Sheet1.append(['','','','','','','',''])
        hd = ['Supplier','PO Ref','Part Number','Description','Consumed Qty','Price(Landed cost)','Amount','Invoice Ref','Picked from comm_nr']
        Sheet1.append(hd)
        for idx,i in enumerate(hd):
            cl = str(alphaChars[idx])+'3'
            Sheet1[cl].fill = PatternFill(start_color="48dce0", end_color="48dce0", fill_type = "solid")
            Sheet1[cl].font = hdFont
            Sheet1[cl].alignment = vertical_align

        if supplierFlag == 'True':
            bomObjs = BoM.objects.filter(project__flag = request.GET['flag'] )
            print bomObjs,'ghghghghgh'
            matObjects = MaterialIssueMain.objects.filter(project__comm_nr = comm_nr,project__flag = request.GET['flag'] )
            print matObjects,'hjjjjjjjjjjjj'
            projectsObjs = Projects.objects.filter(flag = request.GET['flag']).filter(Q(status='ongoing')| Q(savedStatus=True)|Q(junkStatus=True))
            count = 4
            data = []
            if len(matObjects)>0:
                print matObjects, 'matObjects'
                for matIss in matObjects:
                    matList = matIss.materialIssue.all()
                    print matList, 'matObjects'
                    for item in matList:
                        stock =  ast.literal_eval(item.stock)
                        for it in stock:
                            print it["project"],it["product"], 'issued from'

                            itemList = []
                            flag = False
                            i = projectsObjs.get(pk=it["project"])
                            if len(bomObjs)>0:

                                bom = bomObjs.get(project = it["project"],products = it["product"])
                                print bom , 'bomfilter'
                                print i.vendor.name , 'vendor name'
                                print bom.landed_price, 'landed price'
                                itemList.append(i.vendor.name)
                                itemList.append(i.title)
                                itemList.append(bom.products.part_no)
                                itemList.append(bom.products.description_1)
                                itemList.append(it['addedqty'])
                                itemList.append(round(bom.landed_price))
                                itemList.append(round(it['addedqty']*bom.landed_price))
                                itemList.append(i.invoiceNumber)
                                itemList.append(it['comm_nr'])

                                if len(data)>0:
                                    for id, i in enumerate(data):
                                        print i[1], i[2],'data items'
                                        if i[1] == itemList[1] and i[2] == itemList[2]:
                                            print data[id][4], 'data[id][4]'
                                            data[id][4] += itemList[4]
                                            data[id][6]  += itemList[6]

                                            print data[id][4], 'data[id][4]'
                                            flag = True
                                    if(flag == False):
                                        print itemList[6],'itemmmmmmmmmmmmmmmmmmm'
                                        data.append(itemList)
                                else:
                                    print itemList, 'ListItem'
                                    data.append(itemList)

        else:
            print request.GET['supplier'],'supplierrrrrrrrrr'
            supplierObj = Vendor.objects.get(pk = int(request.GET['supplier']))
            bomObjs = BoM.objects.filter(project__vendor = int(supplierObj.pk) ,project__flag = request.GET['flag']).filter(Q(project__status__exact='ongoing')| Q(project__savedStatus__exact=True)|Q(project__junkStatus__exact=True))
            print bomObjs,'ghghghghgh'
            matObjects = MaterialIssueMain.objects.filter(project__comm_nr = comm_nr,project__flag = request.GET['flag'])
            print matObjects,'hjjjjjjjjjjjj'
            projectsObjs = Projects.objects.filter(vendor = int(supplierObj.pk),flag = request.GET['flag']).filter(Q(status='ongoing')| Q(savedStatus=True)|Q(junkStatus=True))
            projectList = [items.pk for items in projectsObjs]
            print projectList,'jkkkkkkkkkkkkkkk'
            count = 4
            data = []
            if len(matObjects)>0:
                print matObjects, 'matObjects'
                for matIss in matObjects:
                    matList = matIss.materialIssue.all()
                    print matList, 'matObjects'
                    for item in matList:
                        stock =  ast.literal_eval(item.stock)
                        for it in stock:
                            print it["project"], 'issued from'
                            if it["project"] in projectList:
                                print it, projectList
                                itemList = []
                                flag = False
                                i = projectsObjs.get(pk=it["project"])
                                if len(bomObjs)>0:
                                    bom = bomObjs.get(project = it["project"],products = it["product"])
                                    print i.vendor.name , 'vendor name'
                                    print bom.landed_price, 'landed price'
                                    itemList.append(i.vendor.name)
                                    itemList.append(i.title)
                                    itemList.append(bom.products.part_no)
                                    itemList.append(bom.products.description_1)
                                    itemList.append(it['addedqty'])
                                    itemList.append(round(bom.landed_price))
                                    itemList.append(round(it['addedqty']*bom.landed_price))
                                    itemList.append(i.invoiceNumber)
                                    itemList.append(it['comm_nr'])

                                    if len(data)>0:
                                        for id, i in enumerate(data):
                                            print i[1], i[2],'data items'
                                            if i[1] == itemList[1] and i[2] == itemList[2]:
                                                print data[id][4], 'data[id][4]'
                                                data[id][4] += itemList[4]
                                                data[id][6]  += itemList[6]
                                                print data[id][4], 'data[id][4]'
                                                flag = True
                                        if(flag == False):
                                            data.append(itemList)
                                    else:
                                        print itemList, 'ListItem'
                                        data.append(itemList)

        print data, 'data'
        if len(data)==0:
            data.append([0,0,0,0,0,0,0,0,0])

        rows = count
        for i in data:
            Sheet1.append(i)
            Sheet1['E'+str(rows)].number_format = '* #,##0'
            Sheet1['F'+str(rows)].number_format = '* #,##0.00'
            Sheet1['G'+str(rows)].number_format = '* #,##0.00'
            rows = rows +1
        for idx,i in enumerate(data):
            for idx,i in enumerate(i):
                cl = str(alphaChars[idx])+str(count)
                Sheet1[cl].alignment = vertical_align
            count =count+1
        Sheet1.column_dimensions['A'].width = 20
        Sheet1.column_dimensions['B'].width = 20
        Sheet1.column_dimensions['C'].width = 20
        Sheet1.column_dimensions['D'].width = 40
        Sheet1.column_dimensions['E'].width = 15
        Sheet1.column_dimensions['F'].width = 30
        Sheet1.column_dimensions['G'].width = 15
        Sheet1.column_dimensions['H'].width = 15
        Sheet1.column_dimensions['I'].width = 30
        response = HttpResponse(content=save_virtual_workbook(workbook),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=consumtionwisereport.xlsx'
        return response


class StockAvailableReportAPIView(APIView):
    def get(self , request , format = None):

        supplierObj = Vendor.objects.get(pk = int(request.GET['supplier']))
        workbook = Workbook()
        Sheet1 = workbook.active

        hdFont = Font(size=12,bold=True)
        vertical_align = Alignment(wrap_text=True, horizontal='center',vertical='center')
        right_align = Alignment(wrap_text=True, horizontal='right')
        alphaChars = list(string.ascii_uppercase)

        comm_hd = ['Supplier',supplierObj.name,'','','','','','']
        Sheet1.append(comm_hd)
        for idx,i in enumerate(comm_hd):
            cl = str(alphaChars[idx])+'1'
            Sheet1[cl].font = hdFont
            Sheet1[cl].alignment = vertical_align
        Sheet1.append(['','','','','','','',''])
        hd = ['PO Ref','Part Number','Description','Available Qty','Price(Landed cost)','Amount',' comm_nr','Invoice Ref']

        Sheet1.append(hd)
        for idx,i in enumerate(hd):
            cl = str(alphaChars[idx])+'3'
            Sheet1[cl].fill = PatternFill(start_color="48dce0", end_color="48dce0", fill_type = "solid")
            Sheet1[cl].font = hdFont
            Sheet1[cl].alignment = vertical_align

        bomObjs = BoM.objects.filter(project__vendor = int(supplierObj.pk),project__flag =request.GET['flag']  ).filter(Q(project__status__exact='ongoing')| Q(project__savedStatus__exact=True)|Q(project__junkStatus__exact=True))
        print bomObjs,'ghghghghgh'
        invObjs = Inventory.objects.filter(project__vendor = int(supplierObj.pk),project__flag =request.GET['flag'])
        print invObjs,'hjjjjjjjjjjjj'
        projectsObjs = Projects.objects.filter(vendor = int(supplierObj.pk),flag =request.GET['flag']).filter(Q(status='ongoing')| Q(savedStatus=True)|Q(junkStatus=True),flag=request.GET['flag'])

        count = 4
        data = []
        if len(invObjs)>0:
            print invObjs, 'invObjs'
            for inv in invObjs:
                itemList = []
                flag = False
                i = projectsObjs.get(pk=inv.project.pk)
                if len(bomObjs)>0:
                    if inv.qty>0:
                        bom = bomObjs.get(project = inv.project.pk,products = inv.product.pk)

                        itemList.append(i.title)
                        itemList.append(bom.products.part_no)
                        itemList.append(bom.products.description_1)
                        itemList.append(inv.qty)
                        itemList.append(round(bom.landed_price))
                        itemList.append(round(inv.qty*bom.landed_price))
                        itemList.append(i.comm_nr)
                        itemList.append(i.invoiceNumber)


                        data.append(itemList)

        print data, 'data'
        if len(data)==0:
            data.append([' - ',' - ',' - ',' - ',' - ',' - ',' - ',' - ',' - '])

        rows = count
        for i in data:
            Sheet1.append(i)
            Sheet1['D'+str(rows)].number_format = '* #,##0'
            Sheet1['E'+str(rows)].number_format = '* #,##0.00'
            Sheet1['F'+str(rows)].number_format = '* #,##0.00'
            rows = rows +1
        for idx,i in enumerate(data):
            for idx,i in enumerate(hd):
                cl = str(alphaChars[idx])+str(count)
                Sheet1[cl].alignment = vertical_align
            count =count+1
        Sheet1.column_dimensions['A'].width = 20
        Sheet1.column_dimensions['B'].width = 20
        Sheet1.column_dimensions['C'].width = 40
        Sheet1.column_dimensions['D'].width = 20
        Sheet1.column_dimensions['E'].width = 20
        Sheet1.column_dimensions['F'].width = 30
        Sheet1.column_dimensions['G'].width = 15
        Sheet1.column_dimensions['H'].width = 15

        response = HttpResponse(content=save_virtual_workbook(workbook),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=stockAvailableReport.xlsx'
        return response


class GrnAPIView(APIView):
    def get(self , request , format = None):
        project = Projects.objects.get(pk = request.GET['project'])
        purchaselist = BoM.objects.filter(project = request.GET['project'])
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment;filename="Grndownload.pdf"'
        grn(response , project , purchaselist , request)
        return response

class QuotationAPIView(APIView):
    def get(self , request , format = None):
        project = Projects.objects.get(pk = request.GET['project'])
        try:
            if 'typ' in request.GET:
                currencyTyp = request.GET['typ']
                if currencyTyp == 'INR':
                    multNumber = float(project.exRate)

                else:
                    multNumber = 1
                    currencyTyp = request.GET['typ']
            else:
                multNumber = 1
                currencyTyp = request.GET['typ']
        except:
            multNumber = 1
            currencyTyp = ''

        print multNumber,'aaaaaaaaaaaaa'

        purchaselist = BoM.objects.filter(project = request.GET['project'])
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment;filename="Quotationdownload.pdf"'
        quotation(response , project , purchaselist  ,multNumber,currencyTyp,request)
        return response
from reportlab.platypus.flowables import HRFlowable

def materialIssued(response , value ,projectPk, request):
    if value !='':
        invdata = MaterialIssueMain.objects.get(pk = request.GET['value'])
    elif projectPk!='':
        data = MaterialIssueMain.objects.filter(project__id = request.GET['projectPk'])
        invdata = data[0]

    styles = getSampleStyleSheet()
    style_right = ParagraphStyle(name='right', parent=styles['Normal'], alignment=TA_RIGHT)

    doc = SimpleDocTemplate(response,pagesize=letter, topMargin=0.2*cm,leftMargin=0.2*cm,rightMargin=0.2*cm)
    doc.request = request
    elements = []

    p1 = Paragraph("<para alignment='center' fontSize=15  ><b> MATERIAL ISSUE NOTE </b></para>",styles['Normal'])

    elements.append(p1)
    elements.append(Spacer(1,15))
    cuss_no = invdata.project.comm_nr
    projecttitle =invdata.project.title
    customer =invdata.project.service.name
    dated = invdata.created.date()
    p0_01 =Paragraph("<para fontSize=10>Comm nr</para>",styles['Normal'])
    p0_02 =Paragraph(str(cuss_no),styles['Normal'])
    p2_01 =Paragraph("<para fontSize=10>Customer</para>",styles['Normal'])
    p2_02 =Paragraph(str(customer),styles['Normal'])
    p3_01 =Paragraph("<para fontSize=10>Date of issue</para>",styles['Normal'])
    p3_02 =Paragraph(str(dated),styles['Normal'])
    p3_03 =Paragraph("<para fontSize=10>Po Ref</para>",styles['Normal'])
    p3_04 =Paragraph(str(invdata.project.title),styles['Normal'])
    p3_05 =Paragraph("<para fontSize=10>MaterialIssue Ref</para>",styles['Normal'])
    p3_06 =Paragraph(str(invdata.pk),styles['Normal'])
    data1=[[p0_01,p0_02],[p2_01,p2_02],[p3_01,p3_02],[p3_03,p3_04],[p3_05,p3_06]]
    rheights=5*[0.2*inch] #[1.1*inch,1.1*inch]
    cwidths=2*inch,5.5*inch
    t1=Table(data1,rowHeights=rheights,colWidths=cwidths)
    elements.append(t1)
    elements.append(Spacer(1,40))
    p4_00 =Paragraph("<para fontSize=6 align=center><b>SI No</b></para>",styles['Normal'])
    p4_01 =Paragraph("<para fontSize=6 align=center><b>Part number</b></para>",styles['Normal'])
    p4_02 =Paragraph("<para fontSize=6 align=center><b>Part description </b></para>",styles['BodyText'])
    p4_03 =Paragraph("<para fontSize=6 align=center><b>Qty</b></para>",styles['Normal'])
    p4_04 =Paragraph("<para fontSize=6 align=center><b>Stock value / unit </b></para>",styles['Normal'])
    p4_05 =Paragraph("<para fontSize=6 align=center><b>Stock value consumed for the comm nr</b></para>",styles['Normal'])
    data2= [[p4_00,p4_01,p4_02,p4_03,p4_04,p4_05]]
    grandtotal = 0
    count=1
    if value !='':
        for i in list(invdata.materialIssue.values()):
            product = Products.objects.get(pk = i['product_id'])
            partno = product.part_no
            description = product.description_1
            qty = i['qty']
            qdata = qty
            price = i['price']
            pdata = price
            total = qty*price
            tdata = total
            grandtotal+=total
            gtotal = grandtotal
            print i,'ffffffffff'
            p6_00 =Paragraph(str(count),styles['Normal'])
            p6_01 =Paragraph(partno,styles['Normal'])
            p6_02 =Paragraph(description,styles['BodyText'])
            p6_03 =Paragraph("{:,}".format(qdata),styles['Normal'])
            p6_04 =Paragraph("{:,}".format(round(pdata,2)),style_right)
            p6_05 =Paragraph("{:,}".format(round(tdata,2)),style_right)
            data2+=[[p6_00,p6_01,p6_02,p6_03,p6_04,p6_05]]
            count += 1
    else:
        print data,'aaaaaaaaaaaaaaaa'
        for i in data:
            print i.materialIssue,'aaaaaaaaaaaaaaaaafffffffffffff'
            for j in list(i.materialIssue.values()):
                product = Products.objects.get(pk = j['product_id'])
                partno = product.part_no
                description = product.description_1
                qty = j['qty']
                qdata = qty
                price = j['price']
                pdata = price
                total = qty*price
                tdata = total
                grandtotal+=total
                gtotal = grandtotal
                p6_00 =Paragraph(count,styles['Normal'])
                p6_01 =Paragraph(partno,styles['Normal'])
                p6_02 =Paragraph(description,styles['BodyText'])
                p6_03 =Paragraph("{:,}".format(qdata),styles['Normal'])
                p6_04 =Paragraph("{:,}".format(round(pdata,2)),style_right)
                p6_05 =Paragraph("{:,}".format(round(tdata,2)),style_right)
                data2+=[[p6_00,p6_01,p6_02,p6_03,p6_04,p6_05]]
                count += 1
    p7_00 =Paragraph("<para fontSize=8 ></para>",styles['Normal'])
    p7_01 =Paragraph("<para fontSize=8 ></para>",styles['Normal'])
    p7_02 =Paragraph("<para fontSize=8 ></para>",styles['BodyText'])
    p7_03 =Paragraph("<para fontSize=8 ></para>",styles['Normal'])
    p7_04 =Paragraph("<para fontSize=8 ><b>Total</b></para>",style_right)
    p7_05 =Paragraph("{:,}".format(round(gtotal,2)),style_right)
    data2+=[[p7_00,p7_01,p7_02,p7_03,p7_04,p7_05]]
    cwidth=0.4*inch,1.4*inch,2.5*inch,0.4*inch,1.2*inch,1.5*inch
    t2=Table(data2,colWidths=cwidth)
    t2.setStyle(TableStyle([('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black)]))
    elements.append(t2)
    doc.build(elements)

class InventoryViewSet(viewsets.ModelViewSet):
    permissions_classes  = (permissions.AllowAny , )
    serializer_class = InventorySerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['product','project']
    def get_queryset(self):
        user = self.request.user
        divisionObj = user.designation.division
        queryset = Inventory.objects.filter(division = divisionObj)
        return queryset

class MaterialIssueViewSet(viewsets.ModelViewSet):
    permissions_classes  = (permissions.AllowAny , )
    serializer_class = MaterialIssueSerializer
    def get_queryset(self):
        user = self.request.user
        divisionObj = user.designation.division
        queryset = MaterialIssue.objects.filter(division = divisionObj)
        return queryset

class MaterialIssueMainViewSet(viewsets.ModelViewSet):
    permissions_classes  = (permissions.AllowAny , )
    serializer_class = MaterialIssueMainSerializer
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['project','created']

    def get_queryset(self):
        user = self.request.user
        divisionObj = user.designation.division
        if 'search' in self.request.GET:
            return MaterialIssueMain.objects.filter(project__title__icontains=self.request.GET['search'],division = divisionObj).order_by('-id')
        else:
            return MaterialIssueMain.objects.filter(division = divisionObj)


class MaterialIssuedNoteAPIView(APIView):
    def get(self , request , format = None):
        if 'value' in request.GET:
            value = request.GET['value']
        else:
            value = ''
        if 'projectPk' in request.GET:
            projectPk = request.GET['projectPk']
        else:
            projectPk = ''

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment;filename="Quotationdownload.pdf"'
        materialIssued(response , value , projectPk,request)
        return response

class ProductInventoryAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    def get(self , request , format = None):
        user = self.request.user
        divisionObj = user.designation.division
        total = 0
        toReturn = []

        if 'flag' in request.GET and 'search' in request.GET:
            productlist = Inventory.objects.filter( (Q(product__part_no__icontains=request.GET['search']) | Q(product__description_1__icontains=request.GET['search'])),Q(project__flag=request.GET['flag']),division = divisionObj)
        elif 'searchQuery' in request.GET:
            productlist = Inventory.objects.filter(Q(project__flag=request.GET['flag']),Q(product__part_no__icontains=request.GET['searchQuery']),division = divisionObj)
        else:
            productlist = Inventory.objects.filter(project__flag=request.GET['flag'],division = divisionObj)
        productsList=list(productlist.values('product').distinct().values('product__pk','product__description_1','product__part_no','product__description_2','product__weight','product__price','product__bar_code'))
        for i in productsList:
            totalprice = 0
            totalqty = 0
            totalVal =0
            totalSum = 0
            data = list(productlist.filter(product=i['product__pk']).values())
            print data,'dataaaa'
            for k in data:

                prjojectObj = Projects.objects.get(pk=k['project_id'])
                k['commnr']=prjojectObj.comm_nr
                k['po_no']=prjojectObj.poNumber
                print k['rate'] ,k['qty']
                print type(k['rate']) ,type(k['qty'])
                if k['rate']:
                    rt = k['rate']
                else:
                    rt = 0
                if k['qty']:
                    qt = k['qty']
                else:
                    qt = 0
                if qt>0:
                    totalVal = rt * qt
                    totalprice += rt
                    totalqty += qt
                    totalSum+=totalVal

            toReturn.append({'productPk':i['product__pk'],'productDesc':i['product__description_1'],'productPartno':i['product__part_no'],'productDesc2':i['product__description_2'],'productBarCode':i['product__bar_code'],'weight':i['product__weight'],'price':i['product__price'],'data':data,'totalprice':totalprice,'totalqty':totalqty,'totalVal':totalSum})
            total+=totalSum
        if 'offset' in request.GET:
            offset = int(request.GET['offset'])
            limit = offset + int(request.GET['limit'])
            print offset,limit
            returnData ={'data' :toReturn[offset : limit],'total':total }
        else:
            returnData ={'data' :toReturn,'total':total }
            print returnData,'aaaaaaaaa'
        return Response(returnData,status=status.HTTP_200_OK)


def stock(response, request):
    print 'aaaaaaaaaaa'
    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(response,pagesize=letter, topMargin=0.2*cm,leftMargin=0.1*cm,rightMargin=0.1*cm)
    doc.request = request
    elements = []
    elements.append(Spacer(1,30))
    logo = svg2rlg(os.path.join(globalSettings.BASE_DIR , 'static_shared','images' , 'Bruderer_Logo_svg.svg'))
    sx=sy=0.17
    logo.width,logo.height = logo.minWidth()*sx, logo.height*sy
    logo.scale(sx,sy)
    logo.hAlign = 'RIGHT'
    elements.append(logo)
    elements.append(Spacer(1,10))
    drawing = svg2rlg(os.path.join(globalSettings.BASE_DIR , 'static_shared','images' , 'anchor_icon.svg'))
    productlist = Inventory.objects.filter(project__flag=request.GET['flag'])
    productsList = list(productlist.values('product').distinct().values('product__pk','product__description_1','product__part_no','product__description_2','product__weight','product__price','product__bar_code'))
    toReturn=[]
    for i in productsList:
        data = list(productlist.filter(product=i['product__pk']).values('product__part_no','project__vendor__name','qty','addedqty','rate','project__invoiceNumber'))
        print data
        toReturn.append({'productPk':i['product__pk'],'productDesc':i['product__description_1'],'productPartno':i['product__part_no'],'productDesc2':i['product__description_2'],'productBarCode':i['product__bar_code'],'weight':i['product__weight'],'price':i['product__price'],'data':data})
    dtime = datetime.datetime.now()
    dt = dtime.date()
    headerDetails = Paragraph("""
    <para leftIndent = 10>
    <font size ='14'>
    <b> STOCK DETAILS</b><br/></font><br/>
    <b> Date : </b>%s<br/>
    </para>
    """ %(str(dt)),styles['Normal'])
    tdheader=[[headerDetails]]
    theader=Table(tdheader,colWidths=[3*inch])
    theader.hAlign = 'LEFT'
    elements.append(theader)
    elements.append(Spacer(1,10))
    data = []
    p1 =Paragraph("<para fontSize=8><b>Part No</b> </para>",styles['Normal'])
    p2 =Paragraph("<para fontSize=8><b>Description</b></para>",styles['Normal'])
    p3 =Paragraph("<para fontSize=8><b>Quantity</b></para>",styles['Normal'])
    p4 =Paragraph("<para fontSize=8><b>Price in INR</b></para>",styles['Normal'])
    p5 =Paragraph("<para fontSize=8><b>Total in INR</b></para>",styles['Normal'])
    p6 =Paragraph("<para fontSize=8><b>Supplier</b></para>",styles['Normal'])
    p7 =Paragraph("<para fontSize=8><b>Invoice No</b></para>",styles['Normal'])
    data += [[p1,p2,p3,p4,p5,p6,p7]]
    grandtot = 0
    for a in toReturn:
            qtyCount = 0
            p01 =Paragraph("<para fontSize=8>{0} </para>".format(a['productPartno']),styles['Normal'])
            p02 =Paragraph("<para fontSize=8>{0}</para>".format(smart_str(a['productDesc'])),styles['Normal'])

            datasmall = []
            qdata = []
            rdata = []
            tdata = []
            sdata = []
            idata = []
            for d in a['data']:
                qtyCount = d['qty']
                if d['qty']>0:
                    qty = Paragraph("""
                    <para align="center">
                    <font size ='8'>
                    %s\n
                    </font>
                    </para>
                    """ %(d['qty']),styles['Normal'])
                    tdheader=[[qty]]
                    theader=Table(tdheader)
                    qdata.append(theader)

                    price = Paragraph("<para align='right'><font size ='8' > {:,} \n</font> </para>".format(int(d['rate'])),styles['Normal'])
                    trateheader=[[price]]
                    trateheader=Table(trateheader)
                    rdata.append(trateheader)
                    totalVal = 0
                    totalVal = d['qty'] * d['rate']
                    grandtot+=totalVal

                    total = Paragraph("<para align='right'><font size ='8' > {:,} \n</font> </para>".format(int(totalVal)),styles['Normal'])
                    ttotalheader=[[total]]
                    ttotalheader=Table(ttotalheader)
                    tdata.append(ttotalheader)
                    supplier = Paragraph("<para ><font size ='8' > {0} \n</font> </para>".format(d['project__vendor__name']),styles['Normal'])
                    supplierheader=[[supplier]]
                    supplierheader=Table(supplierheader)
                    sdata.append(supplierheader)
                    invoiceid = Paragraph("<para><font size ='8' > {0} \n</font> </para>".format(d['project__invoiceNumber']),styles['Normal'])
                    invoiceheader=[[invoiceid]]
                    invoiceheader=Table(invoiceheader)
                    idata.append(invoiceheader)
            if len(qdata) > 0:
                p03 =qdata
            else:
                p03 = Paragraph("<para align='center'><font size ='8' > 0 </font> </para>",styles['Normal'])
            if len(rdata) > 0:
                p04 =rdata
            else:
                p04 = Paragraph("<para align='right'><font size ='8' > 0 &nbsp;&nbsp;</font> </para>",styles['Normal'])
            if len(tdata) > 0:
                p05 = tdata
            else:
                p05 = Paragraph("<para align='right'><font size ='8' > 0 &nbsp;&nbsp;</font> </para>",styles['Normal'])
            if len(sdata) > 0:
                p06 = sdata
            else:
                p06 = Paragraph("<para align='right'><font size ='8' >  &nbsp;&nbsp;</font> </para>",styles['Normal'])
            if len(idata) > 0:
                p07 = idata
            else:
                p07 = Paragraph("<para align='right'><font size ='8' >  &nbsp;&nbsp;</font> </para>",styles['Normal'])

            if qtyCount > 0:
                data.append([p01,p02,p03,p04,p05,p06,p07])
    p11 =Paragraph("<para fontSize=8> </para>",styles['Normal'])
    p12 =Paragraph("<para fontSize=8></para>",styles['Normal'])
    p13 =Paragraph("<para fontSize=8></para>",styles['Normal'])
    p14 =Paragraph("<para fontSize=8><b>Total in INR</b></para>",styles['Normal'])
    p15 =Paragraph("<para fontSize=8 align='right'><b>{:,}</b></para>".format(int(grandtot)),styles['Normal'])
    p16 =Paragraph("<para fontSize=8></para>",styles['Normal'])
    p17 =Paragraph("<para fontSize=8></para>",styles['Normal'])
    data += [[p11,p12,p13,p14,p15,p16,p17]]

    t=Table(data,colWidths=(28*mm,37*mm,20*mm, 30*mm, 30*mm,40*mm,30*mm))
    t.setStyle(TableStyle([('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black)]))
    elements.append(t)
    doc.build(elements)

class StockDownloadAPIView(APIView):
    def get(self , request , format = None):
        projectObjs = Projects.objects.all()
        workbook = Workbook()
        toReturn = workbook.active
        hdFont = Font(size=12,bold=True)
        alphaChars = list(string.ascii_uppercase)
        hd = ['Part No','Description','Quantity','Price in INR','Total in Inr','Supplier','Invoice No','Comm_nr','PO Ref']
        vertical_align = Alignment(wrap_text=True, horizontal='center',vertical='center')
        dtime = datetime.datetime.now()
        dt = dtime.date()
        toReturn.append([])
        toReturn.append([])
        toReturn.append(['STOCK DETAILS'])
        toReturn['A3'].font = hdFont
        toReturn['A3'].alignment = vertical_align
        toReturn.append([])
        toReturn.append(['Date'])
        toReturn['A5'].font = hdFont
        toReturn['A5'].alignment = vertical_align
        toReturn.append([dt])
        toReturn['A6'].font = hdFont
        toReturn['A6'].alignment = vertical_align
        toReturn.append([])
        toReturn.append(hd)
        for idx,i in enumerate(hd):
            cl = str(alphaChars[idx])+'8'
            toReturn[cl].fill = PatternFill(start_color="48dce0", end_color="48dce0", fill_type = "solid")
            toReturn[cl].font = hdFont
            toReturn[cl].alignment = vertical_align
        productlist = Inventory.objects.filter(project__flag=request.GET['flag'])
        productsList = list(productlist.values('product').distinct().values('product__pk','product__description_1','product__part_no','product__description_2','product__weight','product__price','product__bar_code','project__pk'))
        stockData=[]

        for i in productsList:
            _project = projectObjs.get(pk=i['project__pk'])
            data = list(productlist.filter(product=i['product__pk']).filter(project=i['project__pk']).values('product__part_no','project__vendor__name','qty','addedqty','rate','project__invoiceNumber'))
            stockData.append({'productPk':i['product__pk'],'productDesc':i['product__description_1'],'productPartno':i['product__part_no'],'productDesc2':i['product__description_2'],'productBarCode':i['product__bar_code'],'weight':i['product__weight'],'price':i['product__price'],'data':data,'commnr':_project.comm_nr,'poTitle':_project.title})

        rowCount = 9
        grandtot = 0
        for a in stockData:
            qtyCount = 0
            part_no =smart_str(a['productPartno'])
            description =smart_str(a['productDesc'])
            data = []
            datasmall = []
            qdata = []
            rdata = []
            tdata = []
            sdata = []
            idata = []
            commr = a['commnr']
            pono = a['poTitle']
            count = 0
            for d in a['data']:
                qtyCount = d['qty']
                if d['qty']>0:
                    qty = d['qty']
                    qdata.append(qty)
                    price = int(d['rate'])
                    rdata.append(price)
                    totalVal = 0
                    totalVal = d['qty'] * d['rate']
                    grandtot+=totalVal
                    total = int(totalVal)
                    tdata.append(total)
                    supplier = d['project__vendor__name']
                    sdata.append(supplier)
                    invoiceid = d['project__invoiceNumber']
                    idata.append(invoiceid)
                    toReturn.append([part_no,description,qty,price,total,supplier,invoiceid,commr,pono])
                    for idx,i in enumerate(hd):
                        cl = str(alphaChars[idx])+str(rowCount)
                        toReturn[cl].alignment = vertical_align
                        toReturn['C'+str(rowCount)].number_format = '* #,##0'
                        toReturn['D'+str(rowCount)].number_format = '* #,##0'
                        toReturn['E'+str(rowCount)].number_format = '* #,##0'
                        toReturn['C'+str(rowCount)].alignment = vertical_align
                        toReturn['D'+str(rowCount)].alignment = vertical_align
                        toReturn['E'+str(rowCount)].alignment = vertical_align
                    count += 1
                    rowCount += 1
        grandtot = round(grandtot,2)
        toReturn.append(['','','','Total',grandtot,'',''])
        toReturn['E'+str(rowCount)].number_format = '* #,##0'
        toReturn['E'+str(rowCount)].alignment = vertical_align
        for idx,i in enumerate(hd):
            cl = str(alphaChars[idx])+str(rowCount)
            toReturn[cl].alignment = vertical_align
            toReturn[cl].font = hdFont
        toReturn.column_dimensions['A'].width = 20
        toReturn.column_dimensions['B'].width = 40
        toReturn.column_dimensions['C'].width = 15
        toReturn.column_dimensions['D'].width = 20
        toReturn.column_dimensions['E'].width = 20
        toReturn.column_dimensions['F'].width = 20
        toReturn.column_dimensions['G'].width = 20
        toReturn.column_dimensions['H'].width = 15
        toReturn.column_dimensions['I'].width = 15
        response = HttpResponse(content=save_virtual_workbook(workbook),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=inventory.xlsx'
        return response
import json, ast
class OrderAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    def post(self , request , format = None):
        user =  User.objects.get(pk=request.data["user"])
        # user1 = self.request.user
        divisionObj = user.designation.division
        project = Projects.objects.get(pk=request.data["project"])
        if type(request.data["products"])==unicode:
            prodList = ast.literal_eval(request.data["products"])
        else:
            prodList = request.data["products"]
        orderlist =[]
        for i in prodList:
            prodListQty = i['prodQty']
            invlist = Inventory.objects.filter(product=i['pk'])
            listData = []
            stockList = []
            price = 0
            totalqty = 0
            prodListTot = 0
            for j in invlist:
                totalqty += j.qty
                print totalqty,'aaaaaaaaaaa'
                if prodListQty>totalqty:
                    prodListTot = totalqty
                else:
                    prodListTot = prodListQty
            prodListQty = prodListTot
            if prodListQty>0:
                for p in invlist:
                    if p.qty>0:
                            if prodListQty>p.qty and prodListQty>0 and p.qty>0:
                                stockVal = []
                                print 'aaaaaaaaaaaaaa',p.qty,prodListQty, p.product.part_no

                                stockVal.append({'part_no':p.product.part_no,'qty': p.qty,'inventory':p.pk,'project':p.project.pk,'savedqty':p.addedqty,'product':p.product.pk,'addedqty':p.qty,'comm_nr':p.project.comm_nr})
                                data = {
                                'qty': p.qty,
                                'product' :Products.objects.get(pk=i['pk']),
                                'price' : p.rate,
                                'stock': stockVal,
                                'division':divisionObj,
                                }
                                orderObj = MaterialIssue.objects.create(**data)
                                orderObj.save()
                                orderlist.append(orderObj)
                                prodListQty = prodListQty - p.qty
                                p.qty = 0
                                p.save()
                            elif prodListQty<p.qty and prodListQty>0 and p.qty>0:
                                stockVal = []
                                print 'bbbbbbbbbbbb',p.qty,prodListQty,p.pk
                                print p.product.pk , 'p.product.pk'
                                print p.project, 'p.project.pk'

                                stockVal.append({'part_no':p.product.part_no,'qty': p.qty,'inventory':p.pk,'project':p.project.pk,'savedqty':p.addedqty,'product':p.product.pk,'addedqty':prodListQty,'comm_nr':p.project.comm_nr})
                                print stockVal
                                data = {
                                'qty': prodListQty,
                                'product' :Products.objects.get(pk=i['pk']),
                                'price' : p.rate,
                                'stock': stockVal,
                                'division':divisionObj,
                                }
                                orderObj = MaterialIssue.objects.create(**data)
                                orderObj.save()
                                orderlist.append(orderObj)
                                print p.qty,prodListQty
                                p.qty = p.qty - prodListQty
                                print p.qty
                                p.save()
                                prodListQty = 0
                                print prodListQty,'pppppppggggggg'
                            elif prodListQty==p.qty and prodListQty>0 and p.qty>0:
                                stockVal = []
                                print 'ccccccccccccccccccccccc',p.qty,prodListQty,p.pk

                                stockVal.append({'part_no':p.product.part_no,'qty': p.qty,'inventory':p.pk,'project':p.project.pk,'savedqty':p.addedqty,'product':p.product.pk,'addedqty':prodListQty,'comm_nr':p.project.comm_nr})
                                data = {
                                'qty': prodListQty,
                                'product' :Products.objects.get(pk=i['pk']),
                                'price' : p.rate,
                                'stock': stockVal,
                                'division':divisionObj,
                                }
                                # prodListQty = prodListQty - p.qty
                                orderObj = MaterialIssue.objects.create(**data)
                                orderObj.save()
                                orderlist.append(orderObj)
                                p.qty = 0
                                p.save()
                                prodListQty = 0
                                print prodListQty,'pppppppkkkkkkkkkkkgg'

        dataVal = {
            "user" : user,
            "project" : project,
            "division" : divisionObj,
        }
        materialIssueObj = MaterialIssueMain.objects.create(**dataVal)
        for i in orderlist:
            materialIssueObj.materialIssue.add(i)
        materialIssueObj.save()
        print materialIssueObj,'aaaaaaaaa'
        return Response(materialIssueObj.pk,status=status.HTTP_200_OK)

import requests
class EmailApi(APIView):
    permission_classes = (permissions.AllowAny ,)
    def post(self, request, format=None):
        email=[]
        projectPk = request.data['pkValue']
        link = request.data['link']
        linkUrl = link['origin'] + '/login?next=/approve/?project=' + str(projectPk)
        project = Projects.objects.get(pk=projectPk)
        productDetails = BoM.objects.filter(project__id = projectPk)
        totalprice = 0
        totalqty = 0
        totalcustomerPrice = 0
        for i in productDetails:
            totalprice+=i.price
            totalqty+=i.quantity1
            totalcustomerPrice+=i.landed_price
        ctx = {
            'recieverName' : 'admin',
            'productDetails' : productDetails,
            'project':project,
            'link':linkUrl,
            'totalPrice':totalprice,
            'totalQty' : totalqty,
            'totalCustomerPrice' : totalcustomerPrice,
            'message' : 'Please click on the below link to change the status'

        }
        # email.append('Gopinath.Anandan@bruderer.com')
        email.append('pankajjoshi300.p@gmail.com')
        email_subject = 'Approval'
        email_body = get_template('app.approval.email.html').render(ctx)
        msg = EmailMessage(email_subject, email_body, to= email , from_email= 'ankita.k@cioc.in' )
        msg.content_subtype = 'html'
        msg.send()
        return Response(status = status.HTTP_200_OK)

class CalculateAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    def post(self , request , format = None):
        packing =0
        insurance =0
        freight =0
        assessableValue =0
        gst1 =0
        gst2 =0
        clearingCharges1 =0
        clearingCharges2 =0
        project = Projects.objects.get(pk=request.data['projectPK'])
        invoiceValue = request.data['invoiceValue']
        packing = request.data['packing']
        insurance = request.data['insurance']
        freight = request.data['freight']
        assessableValue = request.data['assessableValue']
        gst1 = request.data['gst1']
        gst2 = request.data['gst2']
        clearingCharges1 = request.data['clearingCharges1']
        clearingCharges2 = request.data['clearingCharges2']
        project.invoiceValue = invoiceValue
        if packing > 0:
            packingPer = round((float(packing) / float(invoiceValue))*100, 2)
        else:
            packingPer = 0
        project.packing = float(packing)
        if insurance > 0:
            insurancePer = round((float(insurance) / float(invoiceValue))*100, 2)
        else:
            insurancePer = 0
        project.insurance = insurance
        if freight > 0:
            freightPer = round((float(freight) / float(invoiceValue))*100, 2)
        else:
            freightPer = 0
        project.freight =  freight
        if assessableValue > 0:
            assessableValuePer = round((float(assessableValue) / float(invoiceValue))*100, 2)
        else:
            assessableValuePer = 0
        project.assessableValue = assessableValue
        if gst1 > 0:
            gst1Per = round((float(gst1) / float(invoiceValue))*100, 2)
        else:
            gst1Per = 0
        project.gst1 = gst1
        if gst2 > 0:
            gst2Per = round((float(gst2) / float(invoiceValue))*100, 2)
        else:
            gst2Per = 0
        project.gst2 = gst2
        if clearingCharges1 > 0:
            clearingCharges1Per = round((float(clearingCharges1) / float(invoiceValue))*100, 2)
        else:
            clearingCharges1Per = 0
        project.clearingCharges1 = clearingCharges1
        if clearingCharges2 > 0:
            clearingCharges2Per = round((float(clearingCharges2) / float(invoiceValue))*100, 2)
        else:
            clearingCharges2Per = 0
        project.clearingCharges2 = clearingCharges2
        project.save()

        bomData = BoM.objects.filter(project__id=request.data['projectPK'])
        for i in bomData:
            print i.invoice_price,'qqqqqqqqqqqqqq'
            packingTotal = round((float(i.invoice_price)*packingPer)/100, 2)
            insuranceTotal = round((float(i.invoice_price)*insurancePer)/100, 2)
            freightTotal = round((float(i.invoice_price)*freightPer)/100, 2)
            assessableValueTotal = round((float(i.invoice_price)*assessableValuePer)/100, 2)
            gst1Total = round((float(i.invoice_price)*gst1Per)/100, 2)
            gst2Total = round((float(i.invoice_price)*gst2Per)/100, 2)
            clearingCharges1Total = round((float(i.invoice_price)*clearingCharges1Per)/100, 2)
            clearingCharges2Total = round((float(i.invoice_price)*clearingCharges2Per)/100, 2)
            total = packingTotal + insuranceTotal + freightTotal + assessableValueTotal + gst1Total + gst2Total + clearingCharges1Total + clearingCharges2Total
            i.landed_price = round(total + i.invoice_price,2)
            i.save()
            print i.landed_price,'ggggggggggggggggg'
        return Response(status = status.HTTP_200_OK)

class GetMaterialAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    def get(self , request , format = None):
        toReturn = []
        print  request.GET
        if 'search' in request.GET:
            materialList = MaterialIssueMain.objects.filter( Q(project__title__icontains=request.GET['search']))
        else:
            materialList = MaterialIssueMain.objects.all()
        materialsList = list(materialList.values('project').distinct().values('project__pk','project__title','project__comm_nr'))

        for i in materialsList:
            datamaterial = list(materialList.filter(project=i['project__pk']))
            totalVal = 0
            total = 0
            lisData = []
            for k in datamaterial:
                print k.project.title
                data = list(k.materialIssue.values())
                for m in data:
                    price=m['price']
                    print price,'priccceee'
                    qty=m['qty']
                    print qty,'qtttyyyyy'
                    total=price*qty
                    print total,'ttoootttaaalll'
                    totalVal+=total
                    lisData.append(m)
                tot=round(totalVal,2)
                print tot,'grrandtoott'
            toReturn.append({'projectPk':i['project__pk'],'projectTittle':i['project__title'],'projectComm':i['project__comm_nr'],'data':lisData,'totalprice':tot})

        if 'offset' in request.GET:
            offset = int(request.GET['offset'])
            limit = offset + int(request.GET['limit'])
            returnData =toReturn[offset : limit]
        else:
            returnData =toReturn
        return Response(returnData,status=status.HTTP_200_OK)



from django.http import HttpResponse
# from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook

import ast
import json
from openpyxl.styles import PatternFill , Font
import string

def num_to_col_letters(num):
    letters = ''
    while num:
        mod = (num - 1) % 26
        letters += chr(mod + 65)
        num = (num - 1) // 26
    return ''.join(reversed(letters))

def comma_me(amount):
    orig = amount
    new = re.sub("^(-?\d+)(\d{3})", '\g<1>,\g<2>', amount)
    if orig == new:
        return new
    else:
        return comma_me(new)


class DownloadProjectSCExcelReponse(APIView):
    permission_classes = (permissions.IsAuthenticated,)
    def get(self , request , format = None):
        print "hello therer how are you....."
        workbook = Workbook()

        projectObj = Projects.objects.filter(Q(status='ongoing')| Q(savedStatus=True)|Q(junkStatus=True),flag=request.GET['flag'])
        projectsObj = list(projectObj.values('comm_nr').distinct())
        sendData =[]
        hdFont = Font(size=10,bold=True)
        alphaChars = list(string.ascii_uppercase)
        materialObjs=MaterialIssueMain.objects.all()
        center_align = Alignment(wrap_text=True, horizontal='center')
        vertical_align = Alignment(wrap_text=True, horizontal='center',vertical='center')
        right_align = Alignment(wrap_text=True, horizontal='right')
        for idx,p in enumerate(projectsObj):
            if idx==0:
                Sheet1 = workbook.active
                Sheet1.title = p['comm_nr']
                comm_nr = p['comm_nr']

            if idx>0:
                Sheet1 = workbook.create_sheet(p['comm_nr'])
                comm_nr = p['comm_nr']

            totalConsumedInThis =0
            totalConsumedIntoOthers = 0
            totalConsumedFromOthers = 0
            totalValueThis = 0
            print 'totalConsumedInThis and idx :', totalConsumedInThis, idx
            hd = ["Supplier", "Part No",'Description','Qty','Landed Cost','Stock inward value','Stock Consumed in '+str(comm_nr),'Stock Value Consumed in '+ str(comm_nr),'Stock Consumed for other comm nrs','Stock value Consumed for other comm nrs','Stock (Qty) available','Stock (value) available','Stock Consumed In']
            count=1
            hdWidth = [10,10,10]
            Sheet1.append(hd)
            Sheet1.row_dimensions[1].height = 35

            for idx,i in enumerate(hd):
                cl = str(alphaChars[idx])+'1'
                Sheet1[cl].fill = PatternFill(start_color="48dce0", end_color="48dce0", fill_type = "solid")
                Sheet1[cl].font = hdFont
                Sheet1[cl].alignment = vertical_align
            projData = projectObj.filter(comm_nr__exact=p['comm_nr'])
            bomObj=BoM.objects.filter(project__flag=request.GET['flag'],project__comm_nr__exact=p['comm_nr']).filter(Q(project__status__exact='ongoing')| Q(project__savedStatus__exact=True)|Q(project__junkStatus__exact=True))
            clIdx = 2
            for j in bomObj:
                bomMat = MaterialIssueMain.objects.all()
                bomMat1 = bomMat.filter(project__comm_nr =j.project.comm_nr)
                bomMatQty = 0
                stockConsumed = 0
                totalValueThis += round(float(j.quantity2) * float(j.landed_price),2)
                print 'totalValueThis :', totalValueThis
                total = 0
                for mat in bomMat1:
                    matList = mat.materialIssue.filter(product__id=j.products.pk)
                    for item in matList:
                        stock =  ast.literal_eval(item.stock)
                        for it in stock:
                            if it["project"] == j.project.pk:
                                bomMatQty += int(it['addedqty'])

                bomMat2 = bomMat.filter(~Q(project__comm_nr =j.project.comm_nr))
                otherQty = 0
                consumedIntoOthers = 0
                val = []
                for mat in bomMat2:
                    matList = mat.materialIssue.filter(product__id=j.products.pk)
                    for item in matList:
                        stock =  ast.literal_eval(item.stock)
                        for it in stock:
                            if it["project"] == j.project.pk:
                                commision_n = mat.project.comm_nr
                                value = '(quantity : ' + str(it['addedqty']) + ', comm_nr : ' + commision_n + ')'
                                totalConsumedIntoOthers += round(float(it['addedqty']) * float(j.landed_price),2)
                                consumedIntoOthers = round(float(it['addedqty']) * float(j.landed_price),2)
                                val.append(value)
                                otherQty = it['addedqty']

                stockConsumed = bomMatQty
                print stockConsumed, "stockConsumed"
                total = round(float(bomMatQty) * float(j.landed_price),2)
                totalConsumedInThis += total
                val = json.dumps(val)
                count+=1
                lan_price = round(j.landed_price)
                tot_price = round(total)
                inward_price = round(j.quantity2*j.landed_price)
                consumedIntoOthers_price = round(consumedIntoOthers)
                stock_price = round((j.quantity2*j.landed_price)-(total+consumedIntoOthers))
                stock_qty = round(j.quantity2-(otherQty + stockConsumed))
                if (j.quantity2*j.landed_price)-(total+consumedIntoOthers) == 0:
                    stock_price = '-'
                if consumedIntoOthers == 0:
                    consumedIntoOthers_price = '-'
                if total == 0:
                    tot_price = '-'
                if j.quantity2-(otherQty + stockConsumed) == 0:
                    stock_qty = '-'
                if otherQty == 0:
                    otherQty = '-'
                if stockConsumed == 0:
                    stockConsumed = '-'

                Sheet1.append([j.project.vendor.name, j.products.part_no,j.products.description_1,j.quantity2,lan_price,inward_price,stockConsumed,tot_price,otherQty,consumedIntoOthers_price,stock_qty,stock_price,val])
                coloumnD =[j.project.vendor.name, j.products.part_no,j.products.description_1,j.quantity2,j.landed_price,inward_price,stockConsumed,total,otherQty,consumedIntoOthers_price,stock_qty,stock_price,val]
                for idx,i in enumerate(coloumnD):
                    alpha = str(alphaChars[idx])
                    cl = str(alphaChars[idx])+str(clIdx)
                    if alpha=='D':
                        Sheet1[cl].alignment = center_align
                    if alpha=='E' or alpha=='F' or alpha=='G' or alpha=='H' or alpha=='I' or alpha=='J' or alpha=='K' or alpha=='L':
                        Sheet1[cl].alignment = right_align

                Sheet1['E'+str(clIdx)].number_format = '* #,##0'
                Sheet1['F'+str(clIdx)].number_format = '* #,##0'
                Sheet1['G'+str(clIdx)].number_format = '* #,##0'
                Sheet1['H'+str(clIdx)].number_format = '* #,##0'
                Sheet1['I'+str(clIdx)].number_format = '* #,##0'
                Sheet1['J'+str(clIdx)].number_format = '* #,##0'
                Sheet1['K'+str(clIdx)].number_format = '* #,##0'
                Sheet1['L'+str(clIdx)].number_format = '* #,##0'
                clIdx = clIdx+1

            Sheet1.append(['', '','','',' ',' ',' '])
            hd2= ['Supplier','Part No','Description','Quantity','landed Cost','total','consumed From']
            Sheet1.append(hd2)
            count = count+2
            for idx,i in enumerate(hd2):
                alpha = str(alphaChars[idx])
                cl = str(alphaChars[idx])+str(count)
                Sheet1[cl].fill = PatternFill(start_color="48dce0", end_color="48dce0", fill_type = "solid")
                Sheet1[cl].font = hdFont
                Sheet1[cl].alignment = vertical_align

            matConssumed = MaterialIssueMain.objects.filter(project__comm_nr=p['comm_nr'],project__flag=request.GET['flag'])
            index = count+1
            clIdx1 = index
            for item2 in matConssumed:
                matList2 = item2.materialIssue.all()

                for item3 in matList2:

                    item4 =  ast.literal_eval(item3.stock)
                    for stock2 in item4:
                        inval = []
                        suply = []
                        if stock2['comm_nr'] != p['comm_nr']:
                            if stock2['addedqty']>0:
                                sup = projectObj.get(pk=stock2['project'])
                                invalue = '(quantity : ' + str(stock2['addedqty']) + ', comm_nr : ' + stock2['comm_nr'] + ')'
                                supplier = sup.vendor.name
                                inval.append(str(invalue))
                                suply.append(str(supplier))
                            inval = json.dumps(inval)
                            suply = str(suply).replace("[", " ")
                            suply = str(suply).replace("]", " ")
                            if len(inval)>5:
                                tot = item3.qty*item3.price
                                totalConsumedFromOthers += tot
                                count = count+1
                                lan_price = item3.price
                                tot_price = tot
                                Sheet1.append([str(suply),item3.product.part_no, item3.product.description_1,item3.qty,lan_price,tot_price,inval])
                                coloumnD = [str(suply),item3.product.part_no, item3.product.description_1,item3.qty,item3.price,tot,inval]
                                for idx,i in enumerate(coloumnD):
                                    alpha = str(alphaChars[idx])
                                    cl = str(alphaChars[idx])+str(clIdx1)
                                    if alpha=='D':
                                        Sheet1[cl].alignment = center_align
                                    if alpha=='E' or alpha=='F':
                                        Sheet1[cl].alignment = right_align
                                Sheet1['D'+str(clIdx1)].number_format = '* #,##0'
                                Sheet1['E'+str(clIdx1)].number_format = '* #,##0'
                                Sheet1['F'+str(clIdx1)].number_format = '* #,##0'
                                clIdx1 = clIdx1+1
            Sheet1.append(['', '','','','','',''])
            hd3= ['','','Stock Status','','Value','','']
            Sheet1.append(hd3)
            count = count+2
            for idx,i in enumerate(hd3):
                cl = str(alphaChars[idx])+str(count)
                Sheet1[cl].fill = PatternFill(start_color="48dce0", end_color="48dce0", fill_type = "solid")
                Sheet1[cl].font = hdFont
                Sheet1[cl].alignment = vertical_align
            totalAvailable = totalValueThis - (totalConsumedInThis + totalConsumedIntoOthers)
            print 'totalAvailable :', totalAvailable
            Sheet1.append(['', '','Total consumption in {0} from {1}'.format(p['comm_nr'],p['comm_nr']),'',round(totalConsumedInThis),'',''])
            Sheet1.append(['', '','Total consumption into other commission number from {0}'.format(p['comm_nr']),'',round(totalConsumedIntoOthers) ,'',''])
            Sheet1.append(['', '','Total consumption from {0} '.format(p['comm_nr']),'',round(totalConsumedIntoOthers+totalConsumedInThis),'',''])
            Sheet1.append(['', '','Total consumption from other commission number into {0}'.format(p['comm_nr']),'',round(totalConsumedFromOthers),'',''])
            Sheet1.append(['', '','Total stock available in {0} '.format(p['comm_nr']),'',round(totalAvailable),'',''])
            Sheet1['E'+str(count+1)].alignment = right_align
            Sheet1['E'+str(count+2)].alignment = right_align
            Sheet1['E'+str(count+3)].alignment = right_align
            Sheet1['E'+str(count+4)].alignment = right_align
            Sheet1['E'+str(count+5)].alignment = right_align
            Sheet1['E'+str(count+1)].number_format = '* #,##0'
            Sheet1['E'+str(count+2)].number_format = '* #,##0'
            Sheet1['E'+str(count+3)].number_format = '* #,##0'
            Sheet1['E'+str(count+4)].number_format = '* #,##0'
            Sheet1['E'+str(count+5)].number_format = '* #,##0'
            Sheet1.column_dimensions['A'].width = 20
            Sheet1.column_dimensions['B'].width = 20
            Sheet1.column_dimensions['C'].width = 40
            Sheet1.column_dimensions['D'].width = 10
            Sheet1.column_dimensions['E'].width = 15
            Sheet1.column_dimensions['F'].width = 15
            Sheet1.column_dimensions['G'].width = 15
            Sheet1.column_dimensions['H'].width = 15
            Sheet1.column_dimensions['I'].width = 15
            Sheet1.column_dimensions['J'].width = 15
            Sheet1.column_dimensions['K'].width = 15
            Sheet1.column_dimensions['L'].width = 15
            Sheet1.column_dimensions['M'].width = 30
        response = HttpResponse(content=save_virtual_workbook(workbook),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=stockConsumed.xlsx'
        return response

class CreateStockReportDataAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    def get(self , request , format = None):
        print  request.GET
        toRet = {'status':'Invalid Data'}
        dtime = datetime.datetime.now()
        dt = dtime.date()
        prodObj = Products.objects.filter(created__lte=dtime)
        stockTotal = 0
        for i in prodObj:
            invtObjs = Inventory.objects.filter(product=i,created__lte=dtime,project__flag=request.GET['flag'])
            if invtObjs.count()>0:
                total = invtObjs.aggregate(total=Sum(F('qty') * F('rate'),output_field=FloatField())).get('total',0)
                stockTotal += total
        print 'total valueeeeeeeeeee',stockTotal
        if stockTotal>0:
            try:
                ssReportObj = StockSummaryReport.objects.get(dated=dt)
                ssReportObj.stockValue = stockTotal
                ssReportObj.save()
            except:
                ssReportObj = StockSummaryReport.objects.create(dated=dt,stockValue=stockTotal)
            projectsObjs=Projects.objects.filter(Q(status='approved')|Q(status='ongoing'),flag=request.GET['flag'],savedStatus=False,junkStatus=False,created__lte=dtime)
            print projectsObjs.count()
            projStackSummary = []
            for i in projectsObjs:
                matIssMainObjs = MaterialIssueMain.objects.filter(project=i,created__lte=dtime)
                if matIssMainObjs.count()>0:
                    vl = 0
                    for j in matIssMainObjs:
                        tot = 0
                        matIssueObjs = j.materialIssue.all()
                        tot = matIssueObjs.aggregate(total=Sum(F('qty') * F('price'),output_field=FloatField())).get('total',0)
                        vl += tot
                    try:
                        pObj=ProjectStockSummary.objects.get(stockReport=ssReportObj,title=i.title,comm_nr=i.comm_nr,flag=request.GET['flag'])
                        pObj.value=vl
                        pObj.save()
                    except:
                        projStackSummary.append(ProjectStockSummary(stockReport=ssReportObj,value=vl,title=i.title,comm_nr=i.comm_nr,flag=request.GET['flag']))
                else:
                    try:
                        pObj=ProjectStockSummary.objects.get(stockReport=ssReportObj,flag=request.GET['flag'])
                    except:
                        projStackSummary.append(ProjectStockSummary(stockReport=ssReportObj,value=0,title=i.title,comm_nr=i.comm_nr,flag=request.GET['flag']))

            print len(projStackSummary)
            ProjectStockSummary.objects.bulk_create(projStackSummary)
            toRet['status'] = 'Successfully Saved'
        else:
            toRet['status'] = 'No Data Exists'
        return Response(toRet,status=status.HTTP_200_OK)

class DownloadStockReportAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    def get(self , request , format = None):
        print  request.GET
        workbook = Workbook()
        toReturn = workbook.active
        dtObj = datetime.datetime.now()
        if dtObj.month > 3:
            fstDate = date(dtObj.year,4,1)
            lstDate = fstDate + relativedelta(years=1)
        else:
            fstDate = date(dtObj.year-1,4,1)
            lstDate = fstDate + relativedelta(years=1)
        print fstDate , lstDate

        reportsObj = StockSummaryReport.objects.filter(dated__range=[fstDate,lstDate]).order_by('dated')
        print reportsObj.count()
        hdFont = Font(size=12,bold=True)
        vertical_align = Alignment(wrap_text=True, horizontal='center',vertical='center')
        right_align = Alignment(wrap_text=True, horizontal='right')
        alphaChars = list(string.ascii_uppercase)
        rptPkList = list(reportsObj.values_list('pk',flat=True))
        projStokObjs = ProjectStockSummary.objects.filter(stockReport__in=rptPkList,flag=request.GET['flag'])
        print projStokObjs,rptPkList,'changeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeee'
        unqTitles = list(projStokObjs.values_list('title',flat=True).distinct())
        hd = ['Date','Stock value at warehouse']
        titles = list(projStokObjs.values_list('title',flat=True).distinct().values('title','comm_nr'))
        titl = []
        for p in titles:
            print p['comm_nr'],'ssssssssss'
            titl.append(p['title']+' - ' +p['comm_nr'])
        hdWidth = [10,10,10]
        hd += titl
        toReturn.append(hd)
        for idx,i in enumerate(hd):
            cl =  num_to_col_letters(idx+1)+'1'
            toReturn[cl].fill = PatternFill(start_color="48dce0", end_color="48dce0", fill_type = "solid")
            toReturn[cl].font = hdFont
            toReturn[cl].alignment = vertical_align
        toReturn.row_dimensions[1].height = 28
        count = 2
        for i in reportsObj:
            sam = []
            sam.append(str(i.dated))
            stockValue = '{:,}'.format(int(i.stockValue))
            sam.append(stockValue)
            for j in unqTitles:
                try:
                    p = projStokObjs.get(stockReport=i,title=j)
                    value = '{:,}'.format(int(p.value))
                    sam.append(value)
                except:
                    sam.append(0)
            toReturn.append(sam)
            for idx,i in enumerate(sam):
                cl = num_to_col_letters(idx+1)+'1'
                print cl
                toReturn[cl].fill = PatternFill(start_color="48dce0", end_color="48dce0", fill_type = "solid")
                toReturn[cl].font = hdFont
            for idx,i in enumerate(sam):
                cl = num_to_col_letters(idx+1)+str(count)
                print cl
                toReturn[cl].alignment = right_align
            count = count +1
        if toReturn.max_column <= len(string.ascii_uppercase):
            for character in string.ascii_uppercase[0:toReturn.max_column]:
                toReturn.column_dimensions[character].width = 20
        response = HttpResponse(content=save_virtual_workbook(workbook),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=stockSummary.xlsx'
        return response

class DownloadInvoiceReportAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    def get(self , request , format = None):
        projectObj = Projects.objects.filter(Q(status='approved')|Q(status='ongoing'),junkStatus=False,savedStatus=False,flag=request.GET['flag'])
        workbook = Workbook()
        toReturn = workbook.active
        hdFont = Font(size=12,bold=True)
        alphaChars = list(string.ascii_uppercase)
        hd = ['Purchase Order Ref','Supplier','Invoice No.','BOE']
        hdWidth = [10,10,10]
        toReturn.append(hd)
        for idx,i in enumerate(hd):
            cl = str(alphaChars[idx])+'1'
            toReturn[cl].fill = PatternFill(start_color="48dce0", end_color="48dce0", fill_type = "solid")
            toReturn[cl].font = hdFont
        for p in projectObj:
            sam = []
            sam.append(p.poNumber)
            sam.append(p.vendor.name)
            sam.append(p.invoiceNumber)
            sam.append(p.boeRefNumber)
            toReturn.append(sam)
        toReturn.column_dimensions['A'].width = 20
        toReturn.column_dimensions['B'].width = 20
        toReturn.column_dimensions['C'].width = 20
        toReturn.column_dimensions['D'].width = 20
        response = HttpResponse(content=save_virtual_workbook(workbook),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Invoice_BOE.xlsx'
        return response

class GetCmrListAPIView(APIView):
    renderer_classes = (JSONRenderer,)
    def get(self , request , format = None):
        if 'flag' in request.GET:
            projObj = Projects.objects.filter(savedStatus=False,junkStatus=False,flag=request.GET['flag']).values_list('pk',flat=True)
            cmrList = list(Projects.objects.filter(savedStatus=False,junkStatus=False,flag=request.GET['flag']).values_list('comm_nr',flat=True).distinct())

        return Response(cmrList,status=status.HTTP_200_OK)

class ProjectProductAPIView(APIView):
    def get(self , request , format = None):
        projectObj = Projects.objects.filter(savedStatus=False,junkStatus=False,comm_nr=request.GET['comm'])
        print projectObj,'projjjjjjjjjjjjj'
        toReturn = {}
        for i in projectObj:
            bomlist = []
            bomObj = BoM.objects.filter(project=i)
            for j in bomObj:

                qtPrice = round((((i.profitMargin*j.price)/100.0) + j.price),2)
                wkPrice = round(j.price*i.exRate,2)
                try:
                    packingCost = round((i.packing/i.invoiceValue)*wkPrice,2)
                except:
                    packingCost = 0
                try:
                    insurance = round((i.insurance/i.invoiceValue)*wkPrice,2)
                except:
                    insurance = 0
                try:
                    freight = round((i.freight/i.invoiceValue)*wkPrice,2)
                except:
                    freight = 0
                cifPc = round(wkPrice+packingCost+insurance+freight,2)
                totcif = round(cifPc*j.quantity1,2)
                cdVal = round((cifPc+((cifPc*i.assessableValue)/100))*(j.custom/100),2)
                swe = round(cdVal*0.1,2)
                gstVal = round((cifPc+cdVal+swe)*j.gst/100,2)
                try:
                    cc1 = round(wkPrice*(i.clearingCharges1/(i.invoiceValue*i.exRate)),2)
                except:
                    cc1 = 0
                try:
                    cc2 = round(wkPrice*(i.clearingCharges2/(i.invoiceValue*i.exRate)),2)
                except:
                    cc2 = 0

                bomlist.append({'productDesc1':j.products.description_1,'productDesc2':j.products.description_2,'partNo':j.products.part_no,'weight':j.products.weight,'qty1':j.quantity1,'hsn':j.products.customs_no,'price':j.price,'qty2':j.quantity2,'qtPrice':qtPrice,'wkPrice':wkPrice,'packingCost':packingCost,'insurance':insurance,'freight':freight,'cifPc':cifPc,'totcif':totcif,'cdperc':j.custom,'cdVal':cdVal,'swe':swe,'gst':j.gst,'gstVal':gstVal,'cc1':cc1,'cc2':cc2,'landingCost':j.landed_price,})

            toReturn[i.pk]=bomlist
        return Response(toReturn,status=status.HTTP_200_OK)

class CancelMaterialAPIView(APIView):
    def post(self , request , format = None):
        toReturn = []
        materialObj = MaterialIssueMain.objects.get(pk=request.data['pkData'])
        obj = materialObj.materialIssue.all()
        for i in obj:
            for j in ast.literal_eval(i.stock):
                print j['inventory']
                invobj = Inventory.objects.get(pk=j['inventory'])
                qty = 0
                addedqty = 0
                qty = invobj.qty
                qty+= j['addedqty']
                invobj.qty = qty
                addedqty = invobj.addedqty
                addedqty+= j['addedqty']
                invobj.addedqty = addedqty
                invobj.save()
            i.delete()
        materialObj.delete()
        return Response(status=status.HTTP_200_OK)

def deliveryChallan(response , value , request):
    print value,'llllllllllll'
    materialdata = DeliveryChallan.objects.get(pk=value)
    styles = getSampleStyleSheet()
    style_right = ParagraphStyle(name='right', parent=styles['Normal'], alignment=TA_RIGHT)
    doc = SimpleDocTemplate(response,pagesize=letter, topMargin=0.2*cm,leftMargin=1*cm,rightMargin=1*cm)
    doc.request = request
    elements = []
    try:
        heading = materialdata.heading.replace('\n', '<br />')
    except:
        heading = materialdata.heading
    p1 = Paragraph("<para alignment='center' fontSize=10  ><b> {0}</b></para>".format(heading),styles['Normal'])
    elements.append(p1)
    elements.append(Spacer(1,15))
    detail = Paragraph("""
    <para align="left">
    <font size='8'>
    No : %s <br/>
    Date : %s
    </font>
    </para>
    """ %(materialdata.challanNo,materialdata.challanDate),styles['Normal'])
    detail1 = Paragraph("""
    <para align="left">
    <font size ='8'>
    <b>Our GST IN: 29AABCB6326Q1Z6</b> <br/><br/>

    </font></para>
    """ %(),styles['Normal'])
    td=[[detail,detail1]]
    t=Table(td)
    elements.append(t)
    try:
        address = materialdata.customeraddress.replace('\n', '<br />')
    except:
        address = materialdata.customeraddress
    detailaddress = Paragraph("""
    <para align="left">
    <font size ='8'><br/>
    <b>To,</b> <br/>
    %s <br/>
    GST IN : %s
    </font></para>
    """ %(address,materialdata.customergst),styles['Normal'])
    detailval = Paragraph("""
    <para align="left">
    <font size ='8'><br/>
    Delivery through : %s <br/><br/>
    Kind Attn : %s <br/>
    Your reference : %s <br/>
    </font></para>
    """ %(materialdata.deliveryThr,materialdata.customername,materialdata.refNo),styles['Normal'])
    td1=[[detailaddress,detailval]]
    t1=Table(td1)
    t.hAlign = 'LEFT'
    elements.append(t1)
    elements.append(Spacer(1,15))
    p4_00 =Paragraph("<para fontSize=6 align=center><b>Sl. No</b></para>",styles['Normal'])
    p4_01 =Paragraph("<para fontSize=6 align=center><b>Part number</b></para>",styles['Normal'])
    p4_02 =Paragraph("<para fontSize=6 align=center><b>Part description</b></para>",styles['BodyText'])
    p4_03 =Paragraph("<para fontSize=6 align=center><b>Quantity</b></para>",styles['Normal'])
    p4_04 =Paragraph("<para fontSize=6 align=center><b>HSN</b></para>",styles['Normal'])
    data2= [[p4_00,p4_01,p4_02,p4_03,p4_04]]
    indx = 0
    for i in materialdata.materialIssue.materialIssue.all():
        indx+=1
        p6_00 =Paragraph("<para fontSize=6>{0}</para>".format(indx),styles['Normal'])
        p6_01 =Paragraph("<para fontSize=6>{0}</para>".format(i.product.part_no),styles['Normal'])
        p6_02 =Paragraph("<para fontSize=6>{0}</para>".format(i.product.description_1),styles['BodyText'])
        p6_03 =Paragraph("<para fontSize=6 align='center'>{0}</para>".format(i.qty),styles['Normal'])
        p6_04 =Paragraph("<para fontSize=6>{0}</para>".format(i.product.customs_no),styles['Normal'])
        data2+=[[p6_00,p6_01,p6_02,p6_03,p6_04]]
    t2=Table(data2,colWidths=(15*mm,38*mm,80*mm,25*mm,38*mm),repeatRows=1)
    t2.setStyle(TableStyle([('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black)]))
    elements.append(t2)
    elements.append(Spacer(1,15))
    footer1 = Paragraph("""
    <para align="left">
    <font size ='8'>
    Approximate value of goods in INR : %s <br/>
    <b> NOT FOR SALE</b>
    </font></para>
    """ %(materialdata.apprx),styles['Normal'])
    td3=[[footer1]]
    t3=Table(td3)
    elements.append(t3)
    elements.append(Spacer(1,15))
    try:
        notes = materialdata.notes.replace('\n', '<br />')
    except:
        notes = materialdata.notes
    try:
        notes =notes.encode('utf-8')
    except:
        notes =notes
    footer2 = Paragraph("""
    <para align="left">
    <font size ='8'>
    Notes :  <br/>
    %s <br/>
    </font></para>
    """ %(notes),styles['Normal'])
    td4=[[footer2]]
    t4=Table(td4)
    elements.append(t4)
    footer3 = Paragraph("""
    <para align="left">
    <font size ='8'>
    For BRUDERER PRESSES INDIA PRIVATE LIMITED <br/> <br/>
    Authorised Signatory <br/>
    </font></para>
    """ %(),styles['Normal'])
    td5=[[footer3]]
    t5=Table(td5)
    elements.append(t5)
    doc.build(elements)

class DeliveryChallanNoteAPIView(APIView):
    def get(self , request , format = None):
        print request.GET,'aaaaaa'
        value = request.GET['value']
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment;filename="Challandownload.pdf"'
        deliveryChallan(response , value ,request)
        return response

class PageNumCanvas(canvas.Canvas):
    #----------------------------------------------------------------------
    def __init__(self, *args, **kwargs):
        """Constructor"""
        canvas.Canvas.__init__(self, *args, **kwargs)
        self.pages = []
    #----------------------------------------------------------------------
    def showPage(self):
        """
        On a page break, add information to the list
        """
        self.pages.append(dict(self.__dict__))
        self._startPage()
    #----------------------------------------------------------------------
    def save(self):
        """
        Add the page number to each page (page x of y)
        """
        page_count = len(self.pages)

        for page in self.pages:
            self.__dict__.update(page)
            self.draw_page_number(page_count)
            canvas.Canvas.showPage(self)

        canvas.Canvas.save(self)
    #----------------------------------------------------------------------
    def draw_page_number(self, page_count):
        """
        Add
        the page number
        """

        styles = getSampleStyleSheet()
        text = "<font size='10'><b>&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; BRUDERER PRESSES INDIA PVT. LTD.</b><br/>No.17P, Sadaramangala Industrial Area,Whitefield Road,<br/>&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;Kadugodi,Bangalore 560 048, KARNATAKA<br/>&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;Phone : 080-28411049<br/>&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;GSTIN NO : 29AABCB6326Q1Z6</font>"
        logo ="""

        <font size ='12'>
        <img height="90" width="220" src="static_shared/images/image6.png"/> </font>

        """
        text1 = "<font size='8'>Page %s of %s</font>" % (
            self._pageNumber, page_count)
        p = Paragraph(text1, styles['Normal'])
        p.wrapOn(self, 50 * mm, 10 * mm)
        p.drawOn(self, 120 * mm, 8 * mm)
        l= Paragraph(logo , styles['Normal'])
        p = Paragraph(text , styles['Normal'])
        p.wrapOn(self , 100*mm , 50*mm)
        p.drawOn(self , 95*mm , 189*mm)
        l.wrapOn(self , 100*mm , 50*mm)
        l.drawOn(self , 190*mm , 184*mm)

class PageNumCanvas1(canvas.Canvas):
    #----------------------------------------------------------------------
    def __init__(self, *args, **kwargs):
        """Constructor"""
        canvas.Canvas.__init__(self, *args, **kwargs)
        self.pages = []
    #----------------------------------------------------------------------
    def showPage(self):
        """
        On a page break, add information to the list
        """
        self.pages.append(dict(self.__dict__))
        self._startPage()
    #----------------------------------------------------------------------
    def save(self):
        """
        Add the page number to each page (page x of y)
        """
        page_count = len(self.pages)

        for page in self.pages:
            self.__dict__.update(page)
            self.draw_page_number(page_count)
            canvas.Canvas.showPage(self)

        canvas.Canvas.save(self)
    #----------------------------------------------------------------------
    def draw_page_number(self, page_count):
        """
        Add
        the page number
        """

        styles = getSampleStyleSheet()
        text = "<font size='8'><b>&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; BRUDERER PRESSES INDIA PVT. LTD.</b><br/>No.17P, Sadaramangala Industrial Area,Whitefield Road,<br/>&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;Kadugodi,Bangalore 560 048, KARNATAKA<br/>&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;Phone : 080-28411049<br/>&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;GSTIN NO : 29AABCB6326Q1Z6</font>"
        logo ="""

        <font size ='12'>
        <img height="60" width="160" src="static_shared/images/image6.png"/> </font>

        """
        text1 = "<font size='8'>Page %s of %s</font>" % (
            self._pageNumber, page_count)
        p = Paragraph(text1, styles['Normal'])
        p.wrapOn(self, 50 * mm, 10 * mm)
        p.drawOn(self, 120 * mm, 8 * mm)
        l= Paragraph(logo , styles['Normal'])
        p = Paragraph(text , styles['Normal'])
        p.wrapOn(self , 100*mm , 50*mm)
        p.drawOn(self , 78*mm , 250*mm)
        l.wrapOn(self , 100*mm , 50*mm)
        l.drawOn(self , 155*mm , 254*mm)

def invoice(response, pkVal  , request,flag):
    print flag , "flag value "
    if flag == "False":
        styles = getSampleStyleSheet()
        style_right = ParagraphStyle(name='right', parent=styles['Normal'], alignment=TA_RIGHT)
        doc = SimpleDocTemplate(response,pagesize=landscape(letter), topMargin=3*cm,leftMargin=1.1*cm,rightMargin=1*cm)
        doc.page_height = landscape(letter)
        doc.page_width = landscape(letter)
        doc.request = request
        elements = []
        inv = Invoice.objects.get(pk=pkVal)
        invdetails = InvoiceQty.objects.filter(invoice__id=pkVal)
        print invdetails,'aaaaaaaaaa'
        #
        headerDetails = Paragraph("""
        <para align="center">
        <font size ='8'>
        <b> BRUDERER PRESSES INDIA PVT. LTD.</b><br/>
        No.17P, Sadaramangala Industrial Area,Whitefield Road,<br/>
        Kadugodi,Bangalore 560 048, KARNATAKA<br/>
        Phone : 080-28411049<br/>
        GSTIN NO : 29AABCB6326Q1Z6</font>
        </para>
        """ %(),styles['Normal'])
        headerTitle = Paragraph("""
        <para align='center'>
        <font size ='12'>
        <b> Tax Invoice</b></font>
        </para>
        """ %(),styles['Normal'])
        tdheader=[[headerTitle]]
        t2=Table(tdheader,colWidths=(254*mm))
        t2.setStyle(TableStyle([('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black)]))
        elements.append(t2)
        date = inv.invoiceDate
        print date,"-"
        date1 = date.strftime('%d /%m /%Y')
        print date1,"-"
        detail01 = Paragraph("""
        <para>
        Invoice No : %s
        </para>
        """ %(inv.invoiceNumber),styles['Normal'])
        detail02 = Paragraph("""
        <para >
        Invoice Date : %s
        </para>
        """ %(date1),styles['Normal'])
        tdata=[detail01],[detail02]
        tdheader=[tdata]
        t3=Table(tdheader,colWidths=(127*mm,127*mm))
        t3.setStyle(TableStyle([('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black)]))
        elements.append(t3)
        detail11 = Paragraph("""
        <para>
        Customer PO Ref :
        </para>
        """ %(),styles['Normal'])
        detail12 = Paragraph("""
        <para >
        %s
        </para>
        """ %(inv.poNumber),styles['Normal'])
        detail13 = Paragraph("""
        <para >
        Insurance :
        </para>
        """ %(),styles['Normal'])
        detail14 = Paragraph("""
        <para >
            %s
        </para>
        """ %(inv.insuranceNumber),styles['Normal'])
        t1data=[detail11],[detail12],[detail13],[detail14]
        td1header=[t1data]
        detail21 = Paragraph("""
        <para>
        Transporter Name :
        </para>
        """ %(),styles['Normal'])
        detail22 = Paragraph("""
        <para >
        %s
        </para>
        """ %(inv.transporter),styles['Normal'])
        detail23 = Paragraph("""
        <para >
        LR No :
        </para>
        """ %(),styles['Normal'])
        detail24 = Paragraph("""
        <para >
        %s
        </para>
        """ %(inv.lrNo),styles['Normal'])
        t2data=[detail21],[detail22],[detail23],[detail24]
        td1header+=[t2data]
        t4=Table(td1header,colWidths=(37*mm,90*mm,37*mm,90*mm))
        t4.setStyle(TableStyle([('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black)]))
        elements.append(t4)
        detail31 = Paragraph("""
        <para align='center'>
        <b>Bill to Party</b>
        </para>
        """ %(),styles['Normal'])
        detail32 = Paragraph("""
        <para align='center'>
        <b>Ship to Party</b>
        </para>
        """ %(),styles['Normal'])
        t2data=[detail31],[detail32]
        td2header=[t2data]
        t5=Table(td2header,colWidths=(127*mm,127*mm))
        t5.setStyle(TableStyle([('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black)]))
        elements.append(t5)
        detail41 = Paragraph("""
        <para>
        Name :
        </para>
        """ %(),styles['Normal'])
        detail42 = Paragraph("""
        <para >
        %s
        </para>
        """ %(inv.billName),styles['Normal'])
        detail43 = Paragraph("""
        <para >
        Name :
        </para>
        """ %(),styles['Normal'])
        detail44 = Paragraph("""
        <para >
        %s
        </para>
        """ %(inv.shipName),styles['Normal'])
        t4data=[detail41],[detail42],[detail43],[detail44]
        td4header=[t4data]
        t6=Table(td4header,colWidths=(37*mm,90*mm,37*mm,90*mm))
        t6.setStyle(TableStyle([('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black)]))
        elements.append(t6)
        inv.billAddress = json.loads(inv.billAddress)
        inv.shipAddress = json.loads(inv.shipAddress)
        print inv.billAddress,'billadddddddddddddddddddd'
        billaddr = inv.billAddress['street'] +'<br/>' + inv.billAddress['city'] +'<br/>'+str(inv.billAddress['pincode'])
        shipaddr = inv.shipAddress['street'] +'<br/>' + inv.shipAddress['city']+'<br/>'+str(inv.shipAddress['pincode'])
        detail51 = Paragraph("""
        <para>
        Address : <br/> %s
        </para>
        """ %(billaddr),styles['Normal'])
        detail52 = Paragraph("""
        <para>
        Address : <br/> %s
        </para>
        """ %(shipaddr),styles['Normal'])
        t5data=[detail51],[detail52]
        td5header=[t5data]
        detail61 = Paragraph("""
        <para>
        GSTIN : %s
        </para>
        """ %(inv.billGst),styles['Normal'])
        detail62 = Paragraph("""
        <para>
        GSTIN : %s
        </para>
        """ %(inv.billGst),styles['Normal'])
        t6data=[detail61],[detail62]
        td5header+=[t6data]
        t7=Table(td5header,colWidths=(127*mm,127*mm))
        t7.setStyle(TableStyle([('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black)]))
        elements.append(t7)
        detail51 = Paragraph("""
        <para>
        State :  %s
        </para>
        """ %(inv.billState),styles['Normal'])
        detail52 = Paragraph("""
        <para>
        Code :  %s
        </para>
        """ %(inv.billCode),styles['Normal'])
        detail53 = Paragraph("""
        <para>
        State :  %s
        </para>
        """ %(inv.shipState),styles['Normal'])
        detail54 = Paragraph("""
        <para>
        Code :  %s
        </para>
        """ %(inv.shipCode),styles['Normal'])
        t6data=[detail51],[detail52] ,[detail53],[detail54]
        td6header=[t6data]
        t8=Table(td6header,colWidths=(37*mm,90*mm,37*mm,90*mm))
        t8.setStyle(TableStyle([('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black)]))
        elements.append(t8)
        data2=[]
        s01 =Paragraph("<para fontSize=8>S.No </para>",styles['Normal'])
        s02 =Paragraph("<para fontSize=8>Part No & Product Description </para>",styles['BodyText'])
        s03 =Paragraph("<para fontSize=8>Qty </para>",styles['Normal'])
        s04 =Paragraph("<para fontSize=8>Rate </para>",styles['Normal'])
        s05 =Paragraph("<para fontSize=8>Taxable Value </para>",styles['Normal'])
        s20 =Paragraph("<para fontSize=8>HSN Code </para>",styles['Normal'])
        s06 =Paragraph("<para fontSize=8>CGST Rate (%) </para>",styles['Normal'])
        s07 =Paragraph("<para fontSize=8>CGST Amount </para>",styles['Normal'])
        s08 =Paragraph("<para fontSize=8>SGST Rate (%)</para>",styles['Normal'])
        s09 =Paragraph("<para fontSize=8>SGST Amount </para>",styles['Normal'])
        s10 =Paragraph("<para fontSize=8>IGST Rate (%)  </para>",styles['Normal'])
        s11 =Paragraph("<para fontSize=8>IGST Amount </para>",styles['Normal'])
        s12 =Paragraph("<para fontSize=8>Total </para>",styles['Normal'])
        data2 += [[s01,s02,s03,s04,s05,s20,s06,s07,s08,s09,s10,s11,s12]]
        id = 0
        cgsttot = 0
        taxable = 0
        igsttot = 0
        sgsttot = 0
        grandtot =0
        for i in invdetails:
            id+=1
            cgsttot += i.cgstVal
            taxable +=i.taxableprice
            igsttot += float(i.igstVal)
            sgsttot += i.sgstVal
            grandtot +=i.total
            print grandtot
            s21 =Paragraph("<para fontSize=8>{0} </para>".format(id),styles['Normal'])
            s22 =Paragraph("<para fontSize=8>{0} <br/> {1} </para>".format(i.part_no,smart_str(i.description_1)),styles['BodyText'])
            s23 =Paragraph("<para fontSize=8 alignment='center'>{0} </para>".format(i.qty),styles['Normal'])
            s24 =Paragraph("<para fontSize=8  alignment='right'>{:,} </para>".format(round(i.price,2)),styles['Normal'])
            s25 =Paragraph("<para fontSize=8 alignment='right'> {:,}</para>".format(round(i.taxableprice,2)),styles['Normal'])
            s40 =Paragraph("<para fontSize=8 alignment='right'> {0}</para>".format(i.customs_no),styles['Normal'])
            s26 =Paragraph("<para fontSize=8 alignment='right'>{0} </para>".format(i.cgst),styles['Normal'])
            s27 =Paragraph("<para fontSize=8 alignment='right'>{:,} </para>".format(round(i.cgstVal,2)),styles['Normal'])
            s28 =Paragraph("<para fontSize=8 alignment='right'> {0} </para>".format(i.sgst),styles['Normal'])
            s29 =Paragraph("<para fontSize=8 alignment='right'> {:,} </para>".format(round(i.sgstVal,2)),styles['Normal'])
            s30 =Paragraph("<para fontSize=8 alignment='right'> {0} </para>".format(i.igst),styles['Normal'])
            s31 =Paragraph("<para fontSize=8 alignment='right'> {:,}</para>".format(round(i.igstVal,2)),styles['Normal'])
            s32 =Paragraph("<para fontSize=8 alignment='right'> {:,} </para>".format(round(i.total,2)),styles['Normal'])
            data2.append([s21,s22,s23,s24,s25,s40,s26,s27,s28,s29,s30,s31,s32])
        s21 =Paragraph("<para fontSize=8> </para>",styles['Normal'])
        s22 =Paragraph("<para fontSize=8><b>Total in INR</b></para>",styles['Normal'])
        s23 =Paragraph("<para fontSize=8> </para>",styles['Normal'])
        s24 =Paragraph("<para fontSize=8> </para>",styles['Normal'])
        s25 =Paragraph("<para fontSize=8  alignment='right'><b>{:,}</b></para>".format(round(taxable,2)),styles['Normal'])
        s40 =Paragraph("<para fontSize=8> </para>",styles['Normal'])
        s26 =Paragraph("<para fontSize=8> </para>",styles['Normal'])
        s27 =Paragraph("<para fontSize=8 alignment='right'><b>{:,}</b> </para>".format(round(cgsttot,2)),styles['Normal'])
        s28 =Paragraph("<para fontSize=8>  </para>",styles['Normal'])
        s29 =Paragraph("<para fontSize=8 alignment='right'> <b>{:,}</b>  </para>".format(round(sgsttot,2)),styles['Normal'])
        s30 =Paragraph("<para fontSize=8>  </para>",styles['Normal'])
        s31 =Paragraph("<para fontSize=8 alignment='right'><b> {:,}</b></para>".format(round(igsttot,2)),styles['Normal'])
        s32 =Paragraph("<para fontSize=8 alignment='right'><b> {:,}</b> </para>".format(round(grandtot,2)),styles['Normal'])
        data2.append([s21,s22,s23,s24,s25,s40,s26,s27,s28,s29,s30,s31,s32])
        t9=Table(data2,colWidths=(8*mm,56.3*mm,12*mm,20*mm,20*mm,20*mm,16*mm,16*mm,16*mm,16*mm,16*mm,15.8*mm,22*mm))
        t9.hAlign = 'LEFT'
        t9.setStyle(TableStyle([('TEXTFONT', (0, 0), (-1, -1), 'Times-Bold'),('TEXTCOLOR',(0,0),(-1,-1),black),('ALIGN',(0,0),(-1,-1),'RIGHT'),('VALIGN',(0,0),(-1,-1),'TOP'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black)]))
        elements.append(t9)
        gtotalText = num2words(int(grandtot), to='cardinal', lang='en_IN')
        print gtotalText
        s41 =Paragraph("<para fontSize=8> Rupees {0} </para>".format(gtotalText),styles['Normal'])
        datawords =[[s41]]
        t10=Table(datawords,colWidths=(254*mm))
        t10.hAlign = 'LEFT'
        t10.setStyle(TableStyle([('TEXTFONT', (0, 0), (-1, -1), 'Times-Bold'),('TEXTCOLOR',(0,0),(-1,-1),black),('ALIGN',(0,0),(-1,-1),'RIGHT'),('VALIGN',(0,0),(-1,-1),'TOP'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black)]))
        elements.append(t10)
        try:
            invtrms = inv.invoiceTerms.replace('\n', '<br />')
        except:
            invtrms = inv.invoiceTerms
        packing = inv.packing
        dataFooter = []
        s51 =Paragraph("<para fontSize=8>Payment Terms : {0} </para>".format(invtrms),styles['Normal'])
        s52 =Paragraph("<para fontSize=8>Packing : {0} </para>".format(packing),styles['Normal'])
        s53 =Paragraph("<para fontSize=6 alignment='center'> Certified that the particulars given above are true and correct <br/></para><para fontSize=10> For BRUDERER PRESSES INDIA PVT.LTD.</para>",styles['Normal'])
        dataFooter =[[s51,s52,s53]]
        s61 =Paragraph("<para fontSize=8>Bank Details : IDBI Bank Ltd., Whitefield Branch,<br/>Bangalore - 560 066, Karnataka<br/>Account No. 1545102000003858 <br/>IFSC Code : IBKL0001545</para>",styles['Normal'])
        s62 =Paragraph("<para fontSize=8> </para>",styles['Normal'])
        s63 =Paragraph("<para fontSize=8> </para>",styles['Normal'])
        dataFooter +=[[s61,s62,s63]]
        t11=Table(dataFooter,colWidths=(85*mm,84*mm,85*mm))
        t11.setStyle(TableStyle([('TEXTFONT', (0, 0), (-1, -1), 'Times-Bold'),('TEXTCOLOR',(0,0),(-1,-1),black),('ALIGN',(0,0),(-1,-1),'RIGHT'),('VALIGN',(0,0),(-1,-1),'TOP'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black), ('LINEABOVE', (1,1), (-1,-1), 0.25, colors.white),]))
        elements.append(t11)

        doc.build(elements , canvasmaker = PageNumCanvas,)


#-----------------------------------------landscape-------------------------------------------------



    else:
        styles = getSampleStyleSheet()
        style_right = ParagraphStyle(name='right', parent=styles['Normal'], alignment=TA_RIGHT)
        doc = SimpleDocTemplate(response,pagesize=letter, topMargin=3*cm,leftMargin=0.1*cm,rightMargin=0.1*cm)
        doc.request = request
        elements = []
        inv = Invoice.objects.get(pk=pkVal)
        invdetails = InvoiceQty.objects.filter(invoice__id=pkVal)
        headerTitle = Paragraph("""
        <para align='center'>
        <font size ='12'>
        <b> Tax Invoice</b></font>
        </para>
        """ %(),styles['Normal'])
        tdheader=[[headerTitle]]
        t2=Table(tdheader,colWidths=(205*mm))
        t2.setStyle(TableStyle([('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black)]))
        elements.append(t2)
        date = inv.invoiceDate
        print date,"-"
        date1 = date.strftime('%d /%m /%Y')
        print date1,"-"
        detail01 = Paragraph("""
        <para>
        Invoice No : %s
        </para>
        """ %(inv.invoiceNumber),styles['Normal'])
        detail02 = Paragraph("""
        <para >
        Invoice Date : %s
        </para>
        """ %(date1),styles['Normal'])
        tdata=[detail01],[detail02]
        tdheader=[tdata]
        t3=Table(tdheader,colWidths=(102.5*mm,102.5*mm))
        t3.setStyle(TableStyle([('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black)]))
        elements.append(t3)
        detail11 = Paragraph("""
        <para>
        Customer PO Ref :
        </para>
        """ %(),styles['Normal'])
        detail12 = Paragraph("""
        <para >
        %s
        </para>
        """ %(inv.poNumber),styles['Normal'])
        detail13 = Paragraph("""
        <para >
        Insurance :
        </para>
        """ %(),styles['Normal'])
        detail14 = Paragraph("""
        <para >
            %s
        </para>
        """ %(inv.insuranceNumber),styles['Normal'])
        t1data=[detail11],[detail12],[detail13],[detail14]
        td1header=[t1data]
        detail21 = Paragraph("""
        <para>
        Transporter Name :
        </para>
        """ %(),styles['Normal'])
        detail22 = Paragraph("""
        <para >
        %s
        </para>
        """ %(inv.transporter),styles['Normal'])
        detail23 = Paragraph("""
        <para >
        LR No :
        </para>
        """ %(),styles['Normal'])
        detail24 = Paragraph("""
        <para >
        %s
        </para>
        """ %(inv.lrNo),styles['Normal'])
        t2data=[detail21],[detail22],[detail23],[detail24]
        td1header+=[t2data]
        t4=Table(td1header,colWidths=(30*mm,72.5*mm,30*mm,72.5*mm))
        t4.setStyle(TableStyle([('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black)]))
        elements.append(t4)
        detail31 = Paragraph("""
        <para align='center'>
        <b>Bill to Party</b>
        </para>
        """ %(),styles['Normal'])
        detail32 = Paragraph("""
        <para align='center'>
        <b>Ship to Party</b>
        </para>
        """ %(),styles['Normal'])
        t2data=[detail31],[detail32]
        td2header=[t2data]
        t5=Table(td2header,colWidths=(102.5*mm,102.5*mm))
        t5.setStyle(TableStyle([('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black)]))
        elements.append(t5)
        detail41 = Paragraph("""
        <para>
        Name :
        </para>
        """ %(),styles['Normal'])
        detail42 = Paragraph("""
        <para >
        %s
        </para>
        """ %(inv.billName),styles['Normal'])
        detail43 = Paragraph("""
        <para >
        Name :
        </para>
        """ %(),styles['Normal'])
        detail44 = Paragraph("""
        <para >
        %s
        </para>
        """ %(inv.shipName),styles['Normal'])
        t4data=[detail41],[detail42],[detail43],[detail44]
        td4header=[t4data]
        t6=Table(td4header,colWidths=(30*mm,72.5*mm,30*mm,72.5*mm))
        t6.setStyle(TableStyle([('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black)]))
        elements.append(t6)
        inv.billAddress = json.loads(inv.billAddress)
        inv.shipAddress = json.loads(inv.shipAddress)
        print inv.billAddress,'billadddddddddddddddddddd'
        billaddr = inv.billAddress['street'] +'<br />' + inv.billAddress['city'] +'<br />'+str(inv.billAddress['pincode'])
        shipaddr = inv.shipAddress['street'] +'<br />' + inv.shipAddress['city']+'<br />'+str(inv.shipAddress['pincode'])
        detail51 = Paragraph("""
        <para>
        Address : <br/> %s
        </para>
        """ %(billaddr),styles['Normal'])
        detail52 = Paragraph("""
        <para>
        Address : <br/> %s
        </para>
        """ %(shipaddr),styles['Normal'])
        t5data=[detail51],[detail52]
        td5header=[t5data]
        detail61 = Paragraph("""
        <para>
        GSTIN : %s
        </para>
        """ %(inv.billGst),styles['Normal'])
        detail62 = Paragraph("""
        <para>
        GSTIN : %s
        </para>
        """ %(inv.billGst),styles['Normal'])
        t6data=[detail61],[detail62]
        td5header+=[t6data]
        t7=Table(td5header,colWidths=(102.5*mm,102.5*mm))
        t7.setStyle(TableStyle([('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black)]))
        elements.append(t7)
        detail51 = Paragraph("""
        <para>
        State :  %s
        </para>
        """ %(inv.billState),styles['Normal'])
        detail52 = Paragraph("""
        <para>
        Code :  %s
        </para>
        """ %(inv.billCode),styles['Normal'])
        detail53 = Paragraph("""
        <para>
        State :  %s
        </para>
        """ %(inv.shipState),styles['Normal'])
        detail54 = Paragraph("""
        <para>
        Code :  %s
        </para>
        """ %(inv.shipCode),styles['Normal'])
        t6data=[detail51],[detail52] ,[detail53],[detail54]
        td6header=[t6data]
        t8=Table(td6header,colWidths=(30*mm,72.5*mm,30*mm,72.5*mm))
        t8.setStyle(TableStyle([('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black)]))
        elements.append(t8)
        data2=[]
        s01 =Paragraph("<para fontSize=8>S.No </para>",styles['Normal'])
        s02 =Paragraph("<para fontSize=8>Part No & Product Description </para>",styles['BodyText'])
        s03 =Paragraph("<para fontSize=8>Qty </para>",styles['Normal'])
        s04 =Paragraph("<para fontSize=8>Rate </para>",styles['Normal'])
        s05 =Paragraph("<para fontSize=8>Taxable Value </para>",styles['Normal'])
        s20 =Paragraph("<para fontSize=8>HSN Code </para>",styles['Normal'])
        s06 =Paragraph("<para fontSize=8>CGST Rate (%) </para>",styles['Normal'])
        s07 =Paragraph("<para fontSize=8>CGST Amount </para>",styles['Normal'])
        s08 =Paragraph("<para fontSize=8>SGST Rate (%)</para>",styles['Normal'])
        s09 =Paragraph("<para fontSize=8>SGST Amount </para>",styles['Normal'])
        s10 =Paragraph("<para fontSize=8>IGST Rate (%)  </para>",styles['Normal'])
        s11 =Paragraph("<para fontSize=8>IGST Amount </para>",styles['Normal'])
        s12 =Paragraph("<para fontSize=8>Total </para>",styles['Normal'])

        data2 += [[s01,s02,s03,s04,s05,s20,s06,s07,s08,s09,s10,s11,s12]]
        id = 0
        cgsttot = 0
        taxable = 0
        igsttot = 0
        sgsttot = 0
        grandtot =0

        for i in invdetails:
            id+=1
            cgsttot += i.cgstVal
            taxable +=i.taxableprice
            igsttot += float(i.igstVal)
            sgsttot += i.sgstVal
            grandtot +=i.total
            print grandtot
            s21 =Paragraph("<para fontSize=8>{0} </para>".format(id),styles['Normal'])
            s22 =Paragraph("<para fontSize=8>{0} <br/> {1} </para>".format(i.part_no,smart_str(i.description_1)),styles['BodyText'])
            s23 =Paragraph("<para fontSize=8 alignment='center'>{0} </para>".format(i.qty),styles['Normal'])
            s24 =Paragraph("<para fontSize=8  alignment='right'>{:,} </para>".format(round(i.price,2)),styles['Normal'])
            s25 =Paragraph("<para fontSize=8 alignment='right'> {:,}</para>".format(round(i.taxableprice,2)),styles['Normal'])
            s40 =Paragraph("<para fontSize=8 alignment='right'> {0}</para>".format(i.customs_no),styles['Normal'])
            s26 =Paragraph("<para fontSize=8 alignment='right'>{0} </para>".format(i.cgst),styles['Normal'])
            s27 =Paragraph("<para fontSize=8 alignment='right'>{:,} </para>".format(round(i.cgstVal,2)),styles['Normal'])
            s28 =Paragraph("<para fontSize=8 alignment='right'> {0} </para>".format(i.sgst),styles['Normal'])
            s29 =Paragraph("<para fontSize=8 alignment='right'> {:,} </para>".format(round(i.sgstVal,2)),styles['Normal'])
            s30 =Paragraph("<para fontSize=8 alignment='right'> {0} </para>".format(i.igst),styles['Normal'])
            s31 =Paragraph("<para fontSize=8 alignment='right'> {:,}</para>".format(round(i.igstVal,2)),styles['Normal'])
            s32 =Paragraph("<para fontSize=8 alignment='right'> {:,} </para>".format(round(i.total,2)),styles['Normal'])
            data2.append([s21,s22,s23,s24,s25,s40,s26,s27,s28,s29,s30,s31,s32])
        s21 =Paragraph("<para fontSize=8> </para>",styles['Normal'])
        s22 =Paragraph("<para fontSize=8><b>Total in INR</b></para>",styles['Normal'])
        s23 =Paragraph("<para fontSize=8> </para>",styles['Normal'])
        s24 =Paragraph("<para fontSize=8> </para>",styles['Normal'])
        s25 =Paragraph("<para fontSize=8  alignment='right'><b>{:,}</b></para>".format(round(taxable,2)),styles['Normal'])
        s40 =Paragraph("<para fontSize=8> </para>",styles['Normal'])
        s26 =Paragraph("<para fontSize=8> </para>",styles['Normal'])
        s27 =Paragraph("<para fontSize=8 alignment='right'><b>{:,}</b> </para>".format(round(cgsttot,2)),styles['Normal'])
        s28 =Paragraph("<para fontSize=8>  </para>",styles['Normal'])
        s29 =Paragraph("<para fontSize=8 alignment='right'> <b>{:,}</b>  </para>".format(round(sgsttot,2)),styles['Normal'])
        s30 =Paragraph("<para fontSize=8>  </para>",styles['Normal'])
        s31 =Paragraph("<para fontSize=8 alignment='right'><b> {:,}</b></para>".format(round(igsttot,2)),styles['Normal'])
        s32 =Paragraph("<para fontSize=8 alignment='right'><b> {:,}</b> </para>".format(round(grandtot,2)),styles['Normal'])
        data2.append([s21,s22,s23,s24,s25,s40,s26,s27,s28,s29,s30,s31,s32])
        t9=Table(data2,colWidths=(8*mm,31*mm,12*mm,12*mm,16*mm,16*mm,15*mm,15*mm,15*mm,15*mm,15*mm,15*mm,20*mm))
        t9.setStyle(TableStyle([('TEXTFONT', (0, 0), (-1, -1), 'Times-Bold'),('TEXTCOLOR',(0,0),(-1,-1),black),('ALIGN',(0,0),(-1,-1),'RIGHT'),('VALIGN',(0,0),(-1,-1),'TOP'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black)]))
        elements.append(t9)
        gtotalText = num2words(int(grandtot), to='cardinal', lang='en_IN')
        print gtotalText
        s41 =Paragraph("<para fontSize=8> Rupees {0} </para>".format(gtotalText),styles['Normal'])
        datawords =[[s41]]
        t10=Table(datawords,colWidths=(205*mm))
        t10.setStyle(TableStyle([('TEXTFONT', (0, 0), (-1, -1), 'Times-Bold'),('TEXTCOLOR',(0,0),(-1,-1),black),('ALIGN',(0,0),(-1,-1),'RIGHT'),('VALIGN',(0,0),(-1,-1),'TOP'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black)]))
        elements.append(t10)
        try:
            invtrms = inv.invoiceTerms.replace('\n', '<br />')
        except:
            invtrms = inv.invoiceTerms
        packing = inv.packing
        dataFooter = []
        s51 =Paragraph("<para fontSize=8>Payment Terms : {0} </para>".format(invtrms),styles['Normal'])
        s52 =Paragraph("<para fontSize=8>Packing : {0} </para>".format(packing),styles['Normal'])
        s53 =Paragraph("<para fontSize=6 alignment='center'> Certified that the particulars given above are true and correct <br/></para><para fontSize=10> For BRUDERER PRESSES INDIA PVT.LTD.</para>",styles['Normal'])
        dataFooter =[[s51,s52,s53]]
        s61 =Paragraph("<para fontSize=8>Bank Details : IDBI Bank Ltd., Whitefield Branch,<br/>Bangalore - 560 066, Karnataka<br/>Account No. 1545102000003858 <br/>IFSC Code : IBKL0001545</para>",styles['Normal'])
        s62 =Paragraph("<para fontSize=8> </para>",styles['Normal'])
        s63 =Paragraph("<para fontSize=8> </para>",styles['Normal'])
        dataFooter +=[[s61,s62,s63]]
        t11=Table(dataFooter,colWidths=(68.33*mm,68.33*mm,68.33*mm))
        t11.setStyle(TableStyle([('TEXTFONT', (0, 0), (-1, -1), 'Times-Bold'),('TEXTCOLOR',(0,0),(-1,-1),black),('ALIGN',(0,0),(-1,-1),'RIGHT'),('VALIGN',(0,0),(-1,-1),'TOP'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black), ('LINEABOVE', (1,1), (-1,-1), 0.25, colors.white),]))
        elements.append(t11)
        doc.build(elements,canvasmaker = PageNumCanvas1)

class InvoiceDownloadAPIView(APIView):
    def get(self , request , format = None):
        pkVal = request.GET['pkVal']
        flag = request.GET['potrait']
        print flag, "flag value invoice"
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment;filename="Invoicedownload.pdf"'
        invoice(response, pkVal ,request,flag)
        return response

class StockReportAPIView(APIView):
    def get(self , request , format = None):
        data ={}
        toReturn = []
        today_min = datetime.datetime.combine(datetime.date.today(), datetime.time.min)
        today_max = datetime.datetime.combine(datetime.date.today(), datetime.time.max)
        try:
            obj = StockCheckReport.objects.get(created__range=(today_min, today_max),flag=request.GET['flag'])
            print obj,'hhhhhhhhhhhhhhhhhh'
            count = 1
            data = {'pk':obj.pk,'created':obj.created,'user':obj.user.pk}
        except:
            count = 0
        toReturn.append({'count':count,'data':data})
        return Response(toReturn,status=status.HTTP_200_OK)


def stockSheet(response, value, created,flag, request):
    styles = getSampleStyleSheet()
    style_right = ParagraphStyle(name='right', parent=styles['Normal'], alignment=TA_RIGHT)
    doc = SimpleDocTemplate(response,pagesize=letter, topMargin=0.2*cm,leftMargin=2*cm,rightMargin=0.1*cm)
    doc.request = request
    elements = []
    dateVal =  created.split('T')[0]
    obj = StockCheckItem.objects.filter(stockReport__id=int(value),flag=flag)
    objStock = StockCheckItem.objects.filter(stockReport__id=int(value),matching = True,flag=flag)
    ntMatchingStock = StockCheckItem.objects.filter(stockReport__id=int(value),matching = False,flag=flag)
    invobj = Inventory.objects.filter(project__flag=flag)
    invobjList = list(invobj.values('product').distinct().values('product__pk','product__description_1','product__part_no','product__description_2','product__weight','product__price','product__bar_code'))
    header = Paragraph("""
    <para align='left'>
    <font size="10"><b>Stock Report</b></font><br/><br/>
    <font size="8"><b>Dated : </b> %s </font>
    </para>
    """ %(dateVal),styles['Normal'])
    tdata=[header]
    tdheader=[tdata]
    t=Table(tdheader)
    t.setStyle(TableStyle([('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE')]))
    elements.append(t)
    if len(objStock)>0:
        relatedheader = Paragraph("""
        <para align='left'>
        <font size="8"><b> Matching Items</b></font>
        </para>
        """ %(),styles['Normal'])
        t1data=[relatedheader]
        td1header=[t1data]
        t1=Table(td1header)
        t1.setStyle(TableStyle([('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE')]))
        elements.append(t1)
        data2=[]
        i = 0
        s00 =Paragraph("<para fontSize=8>S.No </para>",styles['Normal'])
        s01 =Paragraph("<para fontSize=8>Part No </para>",styles['Normal'])
        s02 =Paragraph("<para fontSize=8>Product Description </para>",styles['BodyText'])
        s03 =Paragraph("<para fontSize=8>Qty </para>",styles['Normal'])
        data2+=[[s01,s02,s03]]
        for p in objStock:
            i+=1
            s10 =Paragraph("<para fontSize=8 alignment='center'> {0} </para>".format(i),styles['Normal'])
            s11 =Paragraph("<para fontSize=8 alignment='left'> {0} </para>".format(smart_str(p.product.part_no)),styles['Normal'])
            s12 =Paragraph("<para fontSize=8 alignment='left'> {0} </para>".format(smart_str(p.product.description_1)),styles['BodyText'])
            s13 =Paragraph("<para fontSize=8 alignment='center'> {0}</para>".format(p.qty),styles['Normal'])
            data2.append([s11,s12,s13])
        t2=Table(data2,colWidths=(30*mm,80*mm,12*mm))
        t2.hAlign = 'LEFT'
        t2.setStyle(TableStyle([('TEXTFONT', (0, 0), (-1, -1), 'Times-Bold'),('TEXTCOLOR',(0,0),(-1,-1),black),('ALIGN',(0,0),(-1,-1),'RIGHT'),('VALIGN',(0,0),(-1,-1),'TOP'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black)]))
        elements.append(t2)
        elements.append(Spacer(1,15))
    if len(ntMatchingStock)>0:
        ntmatheader = Paragraph("""
        <para align='left'>
        <font size="8"><b> Not Matching Items</b></font>
        </para>
        """ %(),styles['Normal'])
        t3data=[ntmatheader]
        td3header=[t3data]
        t3=Table(td3header)
        t3.setStyle(TableStyle([('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE')]))
        elements.append(t3)
        data3=[]
        i = 0
        s20 =Paragraph("<para fontSize=8>S.No </para>",styles['Normal'])
        s21 =Paragraph("<para fontSize=8>Part No </para>",styles['Normal'])
        s22 =Paragraph("<para fontSize=8>Product Description </para>",styles['BodyText'])
        s23 =Paragraph("<para fontSize=8>Qty </para>",styles['Normal'])
        data3+=[[s21,s22,s23]]
        for j in ntMatchingStock:
            i+=1
            s30 =Paragraph("<para fontSize=8 alignment='center'> {0} </para>".format(i),styles['Normal'])
            s31 =Paragraph("<para fontSize=8 alignment='left'> {0} </para>".format(smart_str(j.product.part_no)),styles['Normal'])
            s32 =Paragraph("<para fontSize=8 alignment='left'> {0} </para>".format(smart_str(j.product.description_1)),styles['BodyText'])
            s33 =Paragraph("<para fontSize=8 alignment='center'> {0}</para>".format(j.qty),styles['Normal'])
            data3.append([s31,s32,s33])
        t4=Table(data3,colWidths=(30*mm,80*mm,12*mm))
        t4.hAlign = 'LEFT'
        t4.setStyle(TableStyle([('TEXTFONT', (0, 0), (-1, -1), 'Times-Bold'),('TEXTCOLOR',(0,0),(-1,-1),black),('ALIGN',(0,0),(-1,-1),'RIGHT'),('VALIGN',(0,0),(-1,-1),'TOP'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black)]))
        elements.append(t4)
        elements.append(Spacer(1,15))
    missingheader = Paragraph("""
    <para align='left'>
    <font size="8"><b> Missing Items</b></font>
    </para>
    """ %(),styles['Normal'])
    t5data=[missingheader]
    td5header=[t5data]
    t5=Table(td5header)
    t5.setStyle(TableStyle([('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE')]))
    elements.append(t5)
    data5=[]
    i = 0
    s41 =Paragraph("<para fontSize=8>Part No </para>",styles['Normal'])
    s42 =Paragraph("<para fontSize=8>Product Description </para>",styles['BodyText'])
    s43 =Paragraph("<para fontSize=8>Qty </para>",styles['Normal'])
    s44 =Paragraph("<para fontSize=8>Qty in Inventory </para>",styles['Normal'])
    data5+=[[s41,s42,s43,s44]]
    for q in ntMatchingStock:
        try:
            count = 0
            tot = 0
            matobj = Inventory.objects.filter(product__id = q.product.pk,created__lt=created,project__flag=flag)
            count =  matobj.aggregate(total=Sum(F('qty'),output_field=PositiveIntegerField())).get('total',0)
            tot = count - q.qty
            i+=1
            s50 =Paragraph("<para fontSize=8 alignment='center'> {0} </para>".format(i),styles['Normal'])
            s51 =Paragraph("<para fontSize=8 alignment='left'> {0} </para>".format(smart_str(q.product.part_no)),styles['Normal'])
            s52 =Paragraph("<para fontSize=8 alignment='left'> {0} </para>".format(smart_str(q.product.description_1)),styles['BodyText'])
            s53 =Paragraph("<para fontSize=8 alignment='center'> {0}</para>".format(tot),styles['Normal'])
            s54 =Paragraph("<para fontSize=8 alignment='center'> {0}</para>".format(count),styles['Normal'])
            data5.append([s51,s52,s53,s54])
        except:
            pass
    for k in invobjList:
        objts = obj.filter(product_id=k['product__pk'])
        if(len(objts)>0):
            pass
        else:
            count = 0
            matobj = Inventory.objects.filter(product__id = k['product__pk'],created__lt=created,project__flag=flag)
            count =  matobj.aggregate(total=Sum(F('qty'),output_field=PositiveIntegerField())).get('total',0)
            if count>0:
                s50 =Paragraph("<para fontSize=8 alignment='center'> {0} </para>".format(i),styles['Normal'])
                s51 =Paragraph("<para fontSize=8 alignment='left'> {0} </para>".format(smart_str(k['product__part_no'])),styles['Normal'])
                s52 =Paragraph("<para fontSize=8 alignment='left'> {0} </para>".format(smart_str(k['product__description_1'])),styles['BodyText'])
                s53 =Paragraph("<para fontSize=8 alignment='center'> {0}</para>".format(count),styles['Normal'])
                s54 =Paragraph("<para fontSize=8 alignment='center'> {0}</para>".format(count),styles['Normal'])
                data5.append([s51,s52,s53,s54])
    t6=Table(data5,colWidths=(30*mm,80*mm,12*mm,30*mm))
    t6.hAlign = 'LEFT'
    t6.setStyle(TableStyle([('TEXTFONT', (0, 0), (-1, -1), 'Times-Bold'),('TEXTCOLOR',(0,0),(-1,-1),black),('ALIGN',(0,0),(-1,-1),'RIGHT'),('VALIGN',(0,0),(-1,-1),'TOP'),('BOX',(0,0),(-1,-1),0.25,colors.black),('INNERGRID', (0,0), (-1,-1), 0.25, colors.black)]))
    elements.append(t6)
    doc.build(elements)

class StockSheetAPIView(APIView):
    def get(self , request , format = None):
        value = request.GET['value']
        created = request.GET['created']
        flag = request.GET['flag']
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment;filename="sheetdownload.pdf"'
        stockSheet(response, value, created,flag, request)
        return response

class AddInventoryAPIView(APIView):
    permission_classes = (permissions.IsAuthenticated ,)
    def post(self , request , format = None):
        user = self.request.user
        print user,'addinventory.......................llllllllllllllllllllllllllllllllllllllllllllllllllllllllllllllll'
        divisionObj = user.designation.division
        try:
            invObj = Inventory.objects.get(project__id=request.data['project'],product__id=request.data['product'],division = divisionObj)
            invObj.division = divisionObj
            print invObj,'addinventory.......................dddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddd'
            invObj.save()
            print invObj,"addinventry..........................ssssssssssssssssssssssssssssssssssssssssssssssssssssss"
            if invObj.qty>0:
                if invObj.qty==invObj.addedqty:
                    invObj.qty = request.data['qty']
                    invObj.addedqty = request.data['qty']
                    invObj.division = divisionObj
                    invObj.save()
                    bomObj = BoM.objects.get(project__id=request.data['project'],products__id=request.data['product'],division = divisionObj)
                    bomObj.quantity2 = request.data['qty']
                    bomObj.division = divisionObj
                    bomObj.save()
                    msg = "Saved"
                    typ = "success"
                else:
                    msg = "Product Already Used from Inventory"
                    typ = "err"
        except:
            msg = "Product Not Found in PO"
            typ = "err"
        toreturn ={"msg":msg,"typ":typ}
        return Response(toreturn,status=status.HTTP_200_OK)

class BulkCreateInventoryAPIView(APIView):
    permission_classes = (permissions.IsAuthenticated ,)
    def post(self , request , format = None):
        print request.data
        if 'project' in request.data:
            bomData = BoM.objects.filter(project=int(request.data['project']))
            invData = []
            projObj = Projects.objects.get(pk=int(request.data['project']))
            for i in bomData:
                invObj = Inventory(project=projObj,product=i.products,qty=i.quantity2,addedqty=i.quantity2,rate=i.landed_price,division=i.division)
                invData.append(invObj)
            print len(invData),'inventory objects will create'
            Inventory.objects.bulk_create(invData)
            toreturn ={"msg":'Success'}
        else:
            toreturn ={"msg":'Error'}
        return Response(toreturn,status=status.HTTP_200_OK)
class complaintManagementViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.AllowAny , )
    serializer_class = complaintManagementSerializer
    def get_queryset(self):
        user = self.request.user
        divisionObj = user.designation.division
        if 'complaintRef' in self.request.GET:
            queryset = ComplaintManagement.objects.filter(complaintRef__icontains = self.request.GET['complaintRef'], division = divisionObj)
        elif 'complaintId' in self.request.GET:
            queryset = ComplaintManagement.objects.filter(pk = int(self.request.GET['complaintId']),division = divisionObj)
            return queryset
        else:
            queryset = ComplaintManagement.objects.filter(division = divisionObj)
        return queryset

class complaintEmailApi(APIView):
    permission_classes = (permissions.AllowAny,)
    def post(self,request,format = None):
        email = []
        objPk = request.data['pkValue']
        link = request.data['link']
        linkUrl = link['origin'] + '/login?next='
        print linkUrl,'llllllllllllllllllllllllllllll'
        complaintObj = ComplaintManagement.objects.get(pk = objPk)
        customer = complaintObj.customer.personName
        machine = complaintObj.machine
        errorCode = complaintObj.errorCode
        date = complaintObj.date
        registeredBy = complaintObj.registeredBy.username
        description = complaintObj.description
        user = self.request.user
        emailObj = user.email
        ctx = {
            'recieverName' : 'admin',
            'link':linkUrl,
            'customer':customer,
            'machine':machine,
            'errorCode':errorCode,
            'date':date,
            'registeredBy':registeredBy,
            'description':description,
            'message' : 'Please click on the below link to change the status'

        }
        # email.append(emailObj)
        email.append('pankajjoshi300.p@gmail.com')
        email_subject = 'Approval'
        email_body = get_template('app.complaintApproval.email.html').render(ctx)
        msg = EmailMessage(email_subject, email_body, to= email , from_email= 'pankajoshi300@gmail.com' )
        msg.content_subtype = 'html'
        msg.send()
        return Response(status = status.HTTP_200_OK)
from django.core.files.storage import FileSystemStorage


def complaintPdf(request):
    complaintObj = ComplaintManagement.objects.get(pk = request.GET.get('pk'))
    print complaintObj,'oooooooooddddddddddfffffffffff'
    doc=SimpleDocTemplate("/tmp/somefilename.pdf")
    styles=getSampleStyleSheet()
    story=[]
    style = styles["Normal"]
    styleN = styles['Normal']
    styleN.alignment = TA_LEFT
    styleN.fontSize = 8
    styleC = styles["Normal"]
    styleC = styles['Title']
    styleC.alignment = TA_CENTER
    styleC.fontSize = 12
    styleT = styles['Title']
    styleT.alignment = TA_CENTER
    styleT.fontSize = 8
    styleB = styles["Title"]
    styleB.alignment =  TA_CENTER
    styleB.fontSize = 8

    tableSpecs = [
        ('INNERGRID', (0,0), (-1,-1), 0.25, colors.black),
        ('BOX', (0,0), (-1,-1), 1, colors.black),
        ]
    emptyLine = Paragraph('', styleT)
    sheetName = Paragraph("BRUDERER PRESSES INDIA PVT LTD", styleC)
    addressData = [

        [Paragraph('Complaint No:', styleT), Paragraph(str(complaintObj.pk), styleN),Paragraph('Complaint Registered on:', styleT), Paragraph(str(complaintObj.date), styleN)],
        [Paragraph('Customer  :', styleT), Paragraph(str(complaintObj.customer.personName), styleN),Paragraph('Customer Representative :', styleT), Paragraph(str(complaintObj.contact), styleN)],

        [Paragraph('Complaint reference :', styleT), Paragraph(str(complaintObj.complaintRef), styleN),Paragraph('Machine :', styleT), Paragraph(str(complaintObj.machine), styleN)],

        [Paragraph('Complaint type :', styleT), Paragraph(str(complaintObj.complaintType), styleN),Paragraph('Whether machine running:', styleT), Paragraph(str(complaintObj.machineRunning), styleN)],

        [Paragraph('Complaint registered by :', styleT), Paragraph(str(complaintObj.registeredBy.username), styleN),Paragraph('Complaint   closed by  :', styleT), Paragraph(str(complaintObj.closedBy.username), styleN)],

        [Paragraph('Refurbished by BIND :', styleT), Paragraph(str(complaintObj.RefurbishedBind), styleN),Paragraph('Service Report number :', styleT), Paragraph(str(complaintObj.serviceReportNo), styleN)],

        [Paragraph('Error Code :', styleT), Paragraph(str(complaintObj.errorCode), styleN),Paragraph('Complaint closed :', styleT), Paragraph(str(complaintObj.is_CloseApproved), styleN)],

        [Paragraph('Complaint closed on :', styleT), Paragraph(str(complaintObj.closedDate), styleN),Paragraph('Nature of complaint :', styleT), Paragraph(str(complaintObj.attr1), styleN)],

        [Paragraph('Interim action :', styleT), Paragraph(str(complaintObj.attr2), styleN),Paragraph('Disposition :', styleT), Paragraph(str(complaintObj.attr3), styleN)],

        ]
    emptyLine1 = Paragraph('', styleT)

    colwidths = (120, 140,140,140)
    address=Table(addressData, colwidths, hAlign='CENTER', style=tableSpecs)
    story.append(sheetName)
    story.append(address)
    doc.build(story)
    fs = FileSystemStorage("/tmp")
    with fs.open("somefilename.pdf") as pdf:
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="'+str(complaintObj.customer.personName).upper()+'".pdf"'

        return response


class getProjObjViewset(APIView):
    renderer_classes = (JSONRenderer,)
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['comm_nr']
    def get(self , request , format = None):

        toRet = []
        projObj =  Projects.objects.all()
        commNo = projObj.values_list('comm_nr', flat = True).distinct()
        for c in commNo:
            data = {'comm_nr' : c}
            data['totalProjects'] = projObj.filter(comm_nr = c).count()
            data['approved'] = projObj.filter(comm_nr = c, status = 'approved').count()
            data['pending'] = projObj.filter(comm_nr = c).exclude(status = 'approved').count()
            toRet.append(data)
        return Response(toRet,status = status.HTTP_200_OK)
